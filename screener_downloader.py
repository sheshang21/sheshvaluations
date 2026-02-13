"""
Screener.in Auto Downloader with Excel Converter (ENHANCED)
============================================================
Downloads Excel from Screener.in, removes blank columns, converts to template format
Enhanced with better error handling, debugging, and multiple fallback strategies
"""

import requests
from bs4 import BeautifulSoup
import pickle
import os
import re
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import time


class ScreenerDownloader:
    """Downloads Excel files from Screener.in with authentication"""
    
    def __init__(self, cookies_path="screener_cookies.pkl"):
        """
        Initialize downloader with cookies
        
        Args:
            cookies_path: Path to pickled cookies file
        """
        self.cookies_path = cookies_path
        self.session = requests.Session()
        self.debug_mode = True  # Enable detailed logging
        self._load_cookies()
        
    def _load_cookies(self):
        """Load cookies from pickle file"""
        if not os.path.exists(self.cookies_path):
            raise FileNotFoundError(f"Cookies file not found: {self.cookies_path}")
        
        with open(self.cookies_path, 'rb') as f:
            cookies = pickle.load(f)
        
        # Convert to requests cookies
        if isinstance(cookies, list):
            for cookie in cookies:
                self.session.cookies.set(cookie['name'], cookie['value'])
        elif isinstance(cookies, dict):
            for name, value in cookies.items():
                self.session.cookies.set(name, value)
        
        if self.debug_mode:
            print(f"✓ Loaded {len(self.session.cookies)} cookies")
    
    def download_excel(self, company_symbol, output_path=None, use_consolidated=False, use_id_url=False):
        """
        Download Excel file from Screener.in by clicking Export button
        
        Args:
            company_symbol: Company symbol (e.g., 'HONASA') or ID number (e.g., '1285886')
            output_path: Where to save the file (optional)
            use_consolidated: Use consolidated financials (default: False)
            use_id_url: Use ID-based URL format /company/id/NUMBER/ (default: False)
            
        Returns:
            str: Path to downloaded file or None if failed
        """
        # Construct URL based on flags
        if use_id_url:
            url_suffix = "consolidated/" if use_consolidated else ""
            company_url = f"https://www.screener.in/company/id/{company_symbol}/{url_suffix}"
        else:
            url_suffix = "consolidated/" if use_consolidated else ""
            company_url = f"https://www.screener.in/company/{company_symbol}/{url_suffix}"
        
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
            'Accept-Language': 'en-US,en;q=0.5',
            'Referer': 'https://www.screener.in/'
        }
        
        try:
            print(f"[Step 1/4] Accessing: {company_url}")
            response = self.session.get(company_url, headers=headers, timeout=20)
            
            if response.status_code != 200:
                print(f"❌ Error: Could not access page (Status: {response.status_code})")
                print(f"Response text preview: {response.text[:500]}")
                return None
            
            print(f"✓ Page loaded successfully ({len(response.content)} bytes)")
            
            soup = BeautifulSoup(response.content, 'html.parser')
            
            # ENHANCED: Try multiple methods to find export button
            print(f"[Step 2/4] Looking for export button...")
            export_button = None
            formaction = None
            
            # Method 1: Look for button with formaction
            export_button = soup.find('button', {'formaction': re.compile(r'/user/company/export/\d+/')})
            if export_button:
                formaction = export_button.get('formaction')
                print(f"✓ Found export button (Method 1): {formaction}")
            
            # Method 2: Look for export link
            if not export_button:
                export_link = soup.find('a', href=re.compile(r'/user/company/export/\d+/'))
                if export_link:
                    formaction = export_link.get('href')
                    print(f"✓ Found export link (Method 2): {formaction}")
            
            # Method 3: Search all buttons and forms
            if not formaction:
                print("Searching all buttons and forms...")
                all_buttons = soup.find_all('button')
                for btn in all_buttons:
                    if 'export' in str(btn).lower():
                        fa = btn.get('formaction')
                        if fa:
                            formaction = fa
                            print(f"✓ Found export button (Method 3): {formaction}")
                            break
            
            # Method 4: Extract company ID from page and construct export URL
            if not formaction:
                print("Trying to extract company ID from page...")
                # Look for company ID in various places
                company_id = None
                
                # Check meta tags
                for meta in soup.find_all('meta'):
                    content = meta.get('content', '')
                    if re.search(r'/company/\w+/\d+/', content):
                        match = re.search(r'/company/\w+/(\d+)/', content)
                        if match:
                            company_id = match.group(1)
                            break
                
                # Check all links
                if not company_id:
                    for link in soup.find_all('a', href=True):
                        href = link['href']
                        if '/company/' in href and re.search(r'\d{6,}', href):
                            match = re.search(r'(\d{6,})', href)
                            if match:
                                company_id = match.group(1)
                                break
                
                if company_id:
                    formaction = f"/user/company/export/{company_id}/"
                    print(f"✓ Constructed export URL (Method 4): {formaction}")
            
            if not formaction:
                print("❌ Error: Could not find export button or link")
                print("Available buttons:", [btn.get('formaction') for btn in soup.find_all('button') if btn.get('formaction')])
                
                # Save debug HTML
                debug_path = "debug_screener_page.html"
                with open(debug_path, 'w', encoding='utf-8') as f:
                    f.write(str(soup.prettify()))
                print(f"Debug: Saved page HTML to {debug_path}")
                
                return None
            
            # Get CSRF token
            print(f"[Step 3/4] Getting CSRF token...")
            csrf_token = None
            
            # Method 1: From form
            if export_button:
                form = export_button.find_parent('form')
                if form:
                    csrf_input = form.find('input', {'name': 'csrfmiddlewaretoken'})
                    if csrf_input:
                        csrf_token = csrf_input.get('value')
                        print(f"✓ Found CSRF token from form")
            
            # Method 2: From cookies
            if not csrf_token:
                csrf_token = self.session.cookies.get('csrftoken', '')
                if csrf_token:
                    print(f"✓ Found CSRF token from cookies")
            
            # Method 3: From page meta
            if not csrf_token:
                csrf_meta = soup.find('meta', {'name': 'csrf-token'})
                if csrf_meta:
                    csrf_token = csrf_meta.get('content', '')
                    print(f"✓ Found CSRF token from meta tag")
            
            if not csrf_token:
                print("⚠️ Warning: No CSRF token found, proceeding anyway...")
                csrf_token = ''
            
            # POST to export URL
            export_url = f"https://www.screener.in{formaction}"
            
            post_headers = {
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
                'Accept': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
                'Referer': company_url,
                'Origin': 'https://www.screener.in',
                'X-CSRFToken': csrf_token
            }
            
            post_data = {
                'csrfmiddlewaretoken': csrf_token,
                'next': f'/company/id/{company_symbol}/{url_suffix}' if use_id_url else f'/company/{company_symbol}/{url_suffix}'
            }
            
            print(f"[Step 4/4] Downloading from: {export_url}")
            
            # Add small delay to avoid rate limiting
            time.sleep(1)
            
            download_response = self.session.post(export_url, headers=post_headers, data=post_data, timeout=30, allow_redirects=True)
            
            if download_response.status_code != 200:
                print(f"❌ Error: Download failed (Status: {download_response.status_code})")
                print(f"Response headers: {dict(download_response.headers)}")
                print(f"Response preview: {download_response.text[:500]}")
                return None
            
            # Check if response is actually Excel
            content_type = download_response.headers.get('content-type', '')
            if 'excel' not in content_type.lower() and 'spreadsheet' not in content_type.lower():
                print(f"⚠️ Warning: Unexpected content type: {content_type}")
                # Try anyway - sometimes it still works
            
            # Save file
            if output_path is None:
                output_path = f"{company_symbol}_screener.xlsx"
            
            with open(output_path, 'wb') as f:
                f.write(download_response.content)
            
            if os.path.exists(output_path) and os.path.getsize(output_path) > 0:
                size_kb = os.path.getsize(output_path) / 1024
                print(f"✅ Downloaded successfully: {output_path} ({size_kb:.1f} KB)")
                
                # Verify it's a valid Excel file
                try:
                    test_wb = load_workbook(output_path, read_only=True)
                    test_wb.close()
                    print(f"✓ Excel file verified")
                except Exception as e:
                    print(f"❌ Error: Downloaded file is not valid Excel: {e}")
                    return None
                
                return output_path
            else:
                print("❌ Error: File empty or not created")
                return None
                
        except Exception as e:
            print(f"❌ Error downloading: {e}")
            import traceback
            traceback.print_exc()
            return None
    
    def remove_empty_year_columns(self, excel_path):
        """
        Remove columns from Data Sheet that don't have data in financial rows
        
        Args:
            excel_path: Path to Excel file
            
        Returns:
            bool: True if successful
        """
        try:
            print("[Cleanup] Removing empty columns...")
            wb = load_workbook(excel_path)
            
            if 'Data Sheet' not in wb.sheetnames:
                print("⚠️ Warning: Data Sheet not found, skipping cleanup")
                wb.close()
                return True  # Don't fail if sheet structure is different
            
            ws = wb['Data Sheet']
            
            # Find P&L section
            pl_date_row = None
            for i in range(1, 50):
                val = ws.cell(i, 1).value
                if val and 'Report Date' in str(val):
                    pl_date_row = i
                    break
            
            if not pl_date_row:
                print("⚠️ Warning: Could not find Report Date row, skipping cleanup")
                wb.close()
                return True
            
            # Check which columns have actual year data
            cols_to_delete = []
            for col in range(2, ws.max_column + 1):
                date_val = ws.cell(pl_date_row, col).value
                
                # Check if this column has a valid date
                has_valid_date = False
                if date_val:
                    if hasattr(date_val, 'year'):
                        has_valid_date = True
                    else:
                        if re.search(r'20\d{2}', str(date_val)):
                            has_valid_date = True
                
                if not has_valid_date:
                    cols_to_delete.append(col)
                    continue
                
                # Also check if column has any financial data (check Sales row)
                sales_row = pl_date_row + 1
                has_data = False
                for check_row in range(sales_row, sales_row + 10):
                    if check_row <= ws.max_row:
                        val = ws.cell(check_row, col).value
                        if val and val != 0:
                            has_data = True
                            break
                
                if not has_data:
                    cols_to_delete.append(col)
            
            # Delete empty columns (in reverse to maintain indices)
            for col in reversed(cols_to_delete):
                ws.delete_cols(col)
                if self.debug_mode:
                    print(f"  Deleted empty column {col}")
            
            wb.save(excel_path)
            wb.close()
            
            if cols_to_delete:
                print(f"✓ Removed {len(cols_to_delete)} empty columns")
            else:
                print(f"✓ No empty columns found")
            
            return True
            
        except Exception as e:
            print(f"⚠️ Warning: Cleanup failed: {e}")
            # Don't fail the entire process if cleanup fails
            return True
    
    def convert_to_template(self, screener_excel_path, output_path=None):
        """
        Convert Screener Excel to exact template format
        
        Args:
            screener_excel_path: Path to downloaded Excel
            output_path: Where to save template (optional)
            
        Returns:
            str: Path to template file or None if failed
        """
        try:
            print("[Conversion] Converting to template format...")
            
            # Load source workbook
            src_wb = load_workbook(screener_excel_path, data_only=True)
            
            if 'Data Sheet' not in src_wb.sheetnames:
                print("❌ Error: Data Sheet not found in source file")
                print(f"Available sheets: {src_wb.sheetnames}")
                src_wb.close()
                return None
            
            src_ws = src_wb['Data Sheet']
            
            # Find P&L and Balance Sheet date rows
            pl_date_row = None
            bs_date_row = None
            
            for i in range(1, min(100, src_ws.max_row + 1)):
                val = src_ws.cell(i, 1).value
                if val:
                    val_str = str(val).strip()
                    if 'Report Date' in val_str and pl_date_row is None:
                        pl_date_row = i
                    elif 'Report Date' in val_str and pl_date_row is not None and bs_date_row is None:
                        bs_date_row = i
                        break
            
            if not pl_date_row or not bs_date_row:
                print(f"❌ Error: Could not find date rows (P&L: {pl_date_row}, BS: {bs_date_row})")
                src_wb.close()
                return None
            
            print(f"✓ Found P&L at row {pl_date_row}, Balance Sheet at row {bs_date_row}")
            
            # Extract dates from P&L section
            dates = []
            date_cols = []
            for col in range(2, src_ws.max_column + 1):
                date_val = src_ws.cell(pl_date_row, col).value
                if date_val:
                    dates.append(date_val)
                    date_cols.append(col)
            
            if not dates:
                print("❌ Error: No dates found")
                src_wb.close()
                return None
            
            print(f"✓ Found {len(dates)} year columns")
            
            # Create new workbook with exact template structure
            new_wb = Workbook()
            new_wb.remove(new_wb.active)  # Remove default sheet
            
            # ============== BALANCE SHEET (First Sheet) ==============
            bs_ws = new_wb.create_sheet("Balance Sheet", 0)
            
            # Row 1: Title
            bs_ws['A1'] = 'BALANCE SHEET'
            
            # Row 2: Report Date header + dates
            bs_ws['A2'] = 'Report Date'
            for idx, date_val in enumerate(dates, start=2):
                bs_ws.cell(2, idx).value = date_val
            
            # Define Balance Sheet items IN EXACT ORDER from target
            bs_items = [
                'Equity Capital',
                'Reserves',
                'Borrowings',
                'Other Liabilities',
                'Total',  # First Total = Total Liabilities
                'Net Block',
                'Capital Work in Progress',
                'Investments',
                'Other Assets',
                'Total',  # Second Total = Total Assets
                'Receivables',
                'Inventory',
                'Cash & Bank',
                'No. of Equity Shares',
                'New Bonus Shares',
                'Face value'
            ]
            
            # Copy Balance Sheet data
            current_target_row = 3
            total_count = 0
            
            for item_name in bs_items:
                bs_ws.cell(current_target_row, 1).value = item_name
                
                # Find this item in source
                for src_row in range(bs_date_row + 1, min(bs_date_row + 30, src_ws.max_row + 1)):
                    src_item = src_ws.cell(src_row, 1).value
                    if src_item:
                        src_item_str = str(src_item).strip()
                        
                        # Handle "Total" - need to track which one
                        if item_name == 'Total':
                            if src_item_str == 'Total':
                                total_count += 1
                                # First Total = row 7, Second Total = row 12
                                if (current_target_row == 7 and total_count == 1) or \
                                   (current_target_row == 12 and total_count == 2):
                                    # Copy data
                                    for idx, src_col in enumerate(date_cols, start=2):
                                        val = src_ws.cell(src_row, src_col).value
                                        bs_ws.cell(current_target_row, idx).value = val
                                    break
                        else:
                            # Normal item matching
                            if src_item_str == item_name:
                                # Copy data
                                for idx, src_col in enumerate(date_cols, start=2):
                                    val = src_ws.cell(src_row, src_col).value
                                    bs_ws.cell(current_target_row, idx).value = val
                                break
                
                current_target_row += 1
            
            # ============== PROFIT AND LOSS ACCOUNT (Second Sheet) ==============
            pl_ws = new_wb.create_sheet("Profit and Loss Account", 1)
            
            # Row 1: Title
            pl_ws['A1'] = 'PROFIT & LOSS'
            
            # Row 2: Report Date header + dates
            pl_ws['A2'] = 'Report Date'
            for idx, date_val in enumerate(dates, start=2):
                pl_ws.cell(2, idx).value = date_val
            
            # Define P&L items IN EXACT ORDER from target
            pl_items = [
                'Sales',
                'Raw Material Cost',
                'Change in Inventory',
                'Power and Fuel',
                'Other Mfr. Exp',
                'Employee Cost',
                'Selling and admin',
                'Other Expenses',
                'Other Income',
                'Depreciation',
                'Interest',
                'Profit before tax',
                'Tax',
                'Net profit',
                'Dividend Amount'
            ]
            
            # Copy P&L data
            current_target_row = 3
            
            for item_name in pl_items:
                pl_ws.cell(current_target_row, 1).value = item_name
                
                # Find this item in source
                for src_row in range(pl_date_row + 1, min(pl_date_row + 30, src_ws.max_row + 1)):
                    src_item = src_ws.cell(src_row, 1).value
                    if src_item and str(src_item).strip() == item_name:
                        # Copy data
                        for idx, src_col in enumerate(date_cols, start=2):
                            val = src_ws.cell(src_row, src_col).value
                            pl_ws.cell(current_target_row, idx).value = val
                        break
                
                current_target_row += 1
            
            # Save
            if output_path is None:
                base_name = os.path.splitext(os.path.basename(screener_excel_path))[0]
                output_path = f"{base_name}_template.xlsx"
            
            new_wb.save(output_path)
            new_wb.close()
            src_wb.close()
            
            print(f"✅ Template created: {output_path}")
            print(f"  - Sheet 1: Balance Sheet ({len(bs_items)} items)")
            print(f"  - Sheet 2: Profit and Loss Account ({len(pl_items)} items)")
            
            return output_path
            
        except Exception as e:
            print(f"❌ Error converting: {e}")
            import traceback
            traceback.print_exc()
            return None
    
    def auto_download_and_convert(self, company_symbol, output_dir=".", keep_original=False, use_consolidated=False, use_id_url=False):
        """
        Complete workflow: download, clean, convert to ready-to-use format
        
        Args:
            company_symbol: Company symbol (e.g., 'HONASA') or ID number (e.g., '1285886')
            output_dir: Directory for output files
            keep_original: Keep original downloaded Excel
            use_consolidated: Use consolidated financials
            use_id_url: Use ID-based URL format /company/id/NUMBER/
            
        Returns:
            str: Path to ready-to-use Excel file or None if failed
        """
        try:
            print(f"\n{'='*70}")
            print(f"AUTO DOWNLOAD AND CONVERT WORKFLOW")
            print(f"Company: {company_symbol}")
            print(f"Consolidated: {use_consolidated}")
            print(f"ID URL: {use_id_url}")
            print(f"{'='*70}\n")
            
            os.makedirs(output_dir, exist_ok=True)
            
            # Download Excel
            original_path = os.path.join(output_dir, f"{company_symbol}_original.xlsx")
            downloaded_path = self.download_excel(company_symbol, original_path, use_consolidated, use_id_url)
            
            if not downloaded_path:
                print("\n" + "="*70)
                print("❌ WORKFLOW FAILED: Could not download Excel file")
                print("="*70)
                return None
            
            # Remove empty columns
            self.remove_empty_year_columns(downloaded_path)
            
            # Convert to template format
            template_path = os.path.join(output_dir, f"{company_symbol}_template.xlsx")
            converted_path = self.convert_to_template(downloaded_path, template_path)
            
            # Clean up original if not needed
            if not keep_original and os.path.exists(downloaded_path):
                os.remove(downloaded_path)
                print(f"✓ Removed original file")
            
            if converted_path:
                print("\n" + "="*70)
                print("✅ WORKFLOW COMPLETED SUCCESSFULLY")
                print(f"Template file: {converted_path}")
                print("="*70)
            else:
                print("\n" + "="*70)
                print("❌ WORKFLOW FAILED: Could not convert to template")
                print("="*70)
            
            return converted_path
            
        except Exception as e:
            print(f"\n{'='*70}")
            print(f"❌ WORKFLOW ERROR: {e}")
            print(f"{'='*70}")
            import traceback
            traceback.print_exc()
            return None


def download_screener_data(company_symbol, cookies_path="screener_cookies.pkl", output_dir="."):
    """
    Convenience function for quick downloads
    
    Args:
        company_symbol: Company symbol
        cookies_path: Path to cookies file
        output_dir: Output directory
        
    Returns:
        str: Path to template file
    """
    downloader = ScreenerDownloader(cookies_path)
    return downloader.auto_download_and_convert(company_symbol, output_dir)


if __name__ == "__main__":
    print("Screener Downloader Module (Enhanced)")
    print("Usage: from screener_downloader import download_screener_data")
    print("\nTest download:")
    
    # Test with a known working company
    test_symbol = "RELIANCE"
    print(f"\nTesting download for {test_symbol}...")
    
    try:
        result = download_screener_data(test_symbol, output_dir="./test_output")
        if result:
            print(f"\n✅ Test successful! File created: {result}")
        else:
            print(f"\n❌ Test failed!")
    except Exception as e:
        print(f"\n❌ Test error: {e}")
