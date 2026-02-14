import streamlit as st
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import time
import random
import requests
from bs4 import BeautifulSoup
import re
from io import StringIO
import yfinance as yf
import plotly.graph_objects as go
import plotly.express as px
from plotly.subplots import make_subplots

# Stock Price Comparison Module
try:
    from stock_price_comparison import (
        get_stock_comparison_data_listed,
        get_stock_comparison_data_screener
    )
    STOCK_COMPARISON_AVAILABLE = True
except ImportError as e:
    STOCK_COMPARISON_AVAILABLE = False
    STOCK_COMPARISON_ERROR = str(e)

# Screener Excel Mode Module
try:
    from screener_excel_mode import (
        parse_screener_excel_to_dataframes,
        get_value_from_screener_df,
        detect_screener_year_columns,
        extract_screener_financials,
        get_screener_shares_outstanding,
        calculate_screener_ddm_valuation,
        calculate_screener_rim_valuation,
        generate_screener_valuation_excel,
        display_screener_financial_summary,
        display_screener_ddm_results,
        display_screener_rim_results,
        fetch_ticker_data_for_screener
    )
    SCREENER_MODE_AVAILABLE = True
except ImportError as e:
    SCREENER_MODE_AVAILABLE = False
    SCREENER_MODE_ERROR = str(e)

# Screener Auto Download Module
try:
    from screener_auto_download_streamlit import integrate_with_existing_upload_section
    AUTO_DOWNLOAD_AVAILABLE = True
except ImportError as e:
    AUTO_DOWNLOAD_AVAILABLE = False
    AUTO_DOWNLOAD_ERROR = str(e)

# Indian Stock Market APIs (fallback for Yahoo Finance)
SCREENER_IMPORT_ERROR = None
try:
    from utils_indian_apis import get_indian_stock_data, get_nse_quote, get_screener_data, fetch_screener_financials
    INDIAN_APIS_AVAILABLE = True
except ImportError as e:
    INDIAN_APIS_AVAILABLE = False
    SCREENER_IMPORT_ERROR = f"ImportError: {e}"
    # Define robust embedded screener with better parsing
    def fetch_screener_financials(symbol, num_years=5):
        """Robust Screener.in scraper with detailed logging and Streamlit Cloud compatibility"""
        import time as _time
        import random as _random
        try:
            from bs4 import BeautifulSoup
            import requests
            import streamlit as st
            from requests.adapters import HTTPAdapter
            from urllib3.util.retry import Retry
            import os
            
            # Disable proxy that may be blocking screener.in
            os.environ.pop('HTTP_PROXY', None)
            os.environ.pop('HTTPS_PROXY', None)
            os.environ.pop('http_proxy', None)
            os.environ.pop('https_proxy', None)
            
            headers = {
                'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 Chrome/120.0.0.0 Safari/537.36',
                'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
                'Accept-Language': 'en-US,en;q=0.5',
                'Connection': 'keep-alive'
            }
            
            # Set up session with retries
            session = requests.Session()
            session.trust_env = False  # Bypass proxy blocking screener.in
            retry_strategy = Retry(
                total=3,
                backoff_factor=1,
                status_forcelist=[429, 500, 502, 503, 504],
                allowed_methods=["HEAD", "GET", "OPTIONS"]
            )
            adapter = HTTPAdapter(max_retries=retry_strategy)
            session.mount("https://", adapter)
            session.mount("http://", adapter)
            
            urls_to_try = [f"https://www.screener.in/company/{symbol}/consolidated/", f"https://www.screener.in/company/{symbol}/"]
            
            soup = None
            connection_error = False
            last_error = None
            
            for url in urls_to_try:
                try:
                    _time.sleep(_random.uniform(1.5, 3.0))
                    st.info(f"üîç Attempting to fetch from: {url}")
                    
                    # Try with SSL verification first
                    try:
                        resp = session.get(url, headers=headers, timeout=30, verify=True)
                    except requests.exceptions.SSLError:
                        st.warning("‚ö†Ô∏è SSL verification failed, retrying without verification...")
                        resp = session.get(url, headers=headers, timeout=30, verify=False)
                    
                    if resp.status_code == 200:
                        soup = BeautifulSoup(resp.content, 'lxml')
                        st.success(f"‚úÖ Successfully connected to Screener.in")
                        break
                    else:
                        st.warning(f"‚ö†Ô∏è Received status code {resp.status_code}, trying next URL...")
                        
                except requests.exceptions.ConnectionError as e:
                    connection_error = True
                    last_error = str(e)
                    st.error(f"‚ùå CONNECTION ERROR: Cannot reach www.screener.in")
                    
                    # Check if it's Streamlit Cloud specific issue
                    if "Connection refused" in str(e) or "Errno 111" in str(e):
                        st.error("üî¥ **STREAMLIT CLOUD NETWORK RESTRICTION DETECTED**")
                        st.markdown("---")
                        st.markdown("### ‚úÖ Recommended Solutions:")
                        st.markdown("""
                        **Option 1: Use Screener Excel Mode (Recommended)**
                        1. Visit [screener.in/company/{}/consolidated/](https://www.screener.in/company/{}/consolidated/)
                        2. Click the **Export** button to download Excel file
                        3. Return to this app and select **"Screener Excel Mode"**
                        4. Upload the downloaded Excel file
                        
                        **Option 2: Use Yahoo Finance Mode**
                        - For listed companies with NSE/BSE tickers
                        - Select "Listed Company (Yahoo Finance)" mode
                        
                        **Option 3: Deploy Elsewhere**
                        - Deploy on Heroku, Railway, or your own server
                        - These platforms typically have fewer network restrictions
                        
                        **Option 4: Upgrade Streamlit Cloud**
                        - Streamlit Cloud Teams/Enterprise may have better network access
                        """.format(symbol, symbol))
                        st.markdown("---")
                    return None
                    
                except requests.exceptions.Timeout:
                    st.warning(f"‚ö†Ô∏è Timeout accessing {url} (30s), trying next URL...")
                    continue
                    
                except Exception as e:
                    st.warning(f"‚ö†Ô∏è Error accessing {url}: {type(e).__name__}: {str(e)}")
                    last_error = str(e)
                    continue
            
            if soup is None:
                if not connection_error:
                    st.error(f"‚ùå Could not access Screener.in page for {symbol}")
                    if last_error:
                        st.error(f"Last error: {last_error}")
                    st.info("üí° Try using **Screener Excel Mode** instead - upload a manually downloaded file from screener.in")
                return None
            
            company_name = soup.find('h1').get_text(strip=True) if soup.find('h1') else symbol
            
            # Try to extract current price from page
            current_price = 0
            try:
                # Look for price in the top metrics section
                price_elem = soup.find('span', class_='number')
                if price_elem:
                    price_text = price_elem.get_text(strip=True).replace(',', '').replace('‚Çπ', '').strip()
                    current_price = float(price_text)
                    st.write(f"‚úì Found Current Price: ‚Çπ{current_price:.2f}")
            except:
                pass
            
            # More flexible table parsing
            def parse_row(table, keywords, debug_name=""):
                if table is None:
                    return []
                for tr in table.find_all('tr'):
                    cells = tr.find_all(['td', 'th'])
                    if not cells:
                        continue
                    label = cells[0].get_text(strip=True).lower().replace('\xa0', ' ').replace('‚Äì', '-').replace('+', '').strip()
                    
                    # Try each keyword
                    for kw in keywords:
                        if kw.lower() in label:
                            values = []
                            for cell in cells[1:]:
                                raw = cell.get_text(strip=True).replace(',', '').replace('\xa0', '')
                                try:
                                    values.append(float(raw))
                                except:
                                    values.append(0.0)
                            if debug_name and values:
                                st.write(f"‚úì Found {debug_name}: {len(values)} years")
                            return values
                return []
            
            # Get ALL tables on page
            all_tables = soup.find_all('table')
            st.info(f"üìä Found {len(all_tables)} tables on Screener.in page")
            
            # Find tables by their parent section IDs (more reliable)
            pl_table = None
            bs_table = None
            
            # Method 1: Find by section ID
            pl_section = soup.find('section', {'id': 'profit-loss'})
            if pl_section:
                # Find the main results table (not segment table)
                main_table = pl_section.find('div', {'data-result-table': ''})
                if main_table:
                    pl_table = main_table.find('table')
                    if pl_table:
                        st.write("‚úì Found P&L table by section ID and data-result-table")
            
            bs_section = soup.find('section', {'id': 'balance-sheet'})
            if bs_section:
                main_table = bs_section.find('div', {'data-result-table': ''})
                if main_table:
                    bs_table = main_table.find('table')
                    if bs_table:
                        st.write("‚úì Found Balance Sheet table by section ID and data-result-table")
            
            # Method 2: Fallback - look for class="data-table"
            if pl_table is None or bs_table is None:
                st.warning("‚ö†Ô∏è Using fallback method - searching by class")
                data_tables = soup.find_all('table', class_='data-table')
                
                for idx, table in enumerate(data_tables):
                    # Check what section this table is in
                    parent_section = table.find_parent('section')
                    if parent_section and parent_section.get('id'):
                        section_id = parent_section.get('id')
                        st.write(f"  Table #{idx+1} is in section: {section_id}")
                        
                        if 'profit' in section_id or 'loss' in section_id:
                            if pl_table is None:
                                pl_table = table
                                st.write(f"‚úì Using Table #{idx+1} as P&L")
                        elif 'balance' in section_id or 'sheet' in section_id:
                            if bs_table is None:
                                bs_table = table
                                st.write(f"‚úì Using Table #{idx+1} as Balance Sheet")
            
            if pl_table is None or bs_table is None:
                st.error(f"‚ùå Could not locate financial tables. Found {len(all_tables)} total tables.")
                if pl_table is None:
                    st.error("  Missing: P&L table")
                if bs_table is None:
                    st.error("  Missing: Balance Sheet table")
                return None
            
            st.write("### üìã Parsing P&L Statement")
            # More flexible parsing - return first match found
            def parse_row_flexible(table, keywords, debug_name=""):
                if table is None:
                    return []
                
                for tr in table.find_all('tr'):
                    cells = tr.find_all(['td', 'th'])
                    if not cells:
                        continue
                    
                    # Get label from first cell
                    first_cell = cells[0]
                    label = first_cell.get_text(strip=True).lower()
                    
                    # Remove special characters
                    label = label.replace('\xa0', ' ').replace('‚Äì', '-').replace('+', '').replace('&amp;', '').replace('  ', ' ').strip()
                    
                    # Try each keyword
                    for kw in keywords:
                        if kw.lower() in label:
                            values = []
                            for cell in cells[1:]:
                                raw = cell.get_text(strip=True).replace(',', '').replace('\xa0', '')
                                try:
                                    val = float(raw)
                                    values.append(val)
                                except:
                                    values.append(0.0)
                            
                            # Only return if we found actual non-zero values
                            if values and any(v != 0 for v in values):
                                if debug_name:
                                    st.write(f"‚úì Found {debug_name}: {len(values)} years - Label: '{label[:60]}'")
                                return values
                
                return []
            
            raw_revenue = parse_row_flexible(pl_table, ['revenue'], "Revenue")
            raw_expenses = parse_row_flexible(pl_table, ['expenses'], "Expenses")
            raw_operating_profit = parse_row_flexible(pl_table, ['financing profit', 'operating profit'], "Financing/Operating Profit")
            raw_other_income = parse_row_flexible(pl_table, ['other income'], "Other Income")
            raw_interest = parse_row_flexible(pl_table, ['interest'], "Interest")
            raw_depreciation = parse_row_flexible(pl_table, ['depreciation'], "Depreciation")
            raw_pbt = parse_row_flexible(pl_table, ['profit before tax'], "PBT")
            raw_tax = []
            raw_tax_pct = parse_row_flexible(pl_table, ['tax %'], "Tax %")
            raw_net_profit = parse_row_flexible(pl_table, ['net profit'], "Net Profit")
            raw_eps = parse_row_flexible(pl_table, ['eps in rs'], "EPS")
            
            st.write("### üè¶ Parsing Balance Sheet")
            # Equity & Liabilities
            raw_equity_capital = parse_row_flexible(bs_table, ['equity capital'], "Equity Capital")
            raw_reserves = parse_row_flexible(bs_table, ['reserves'], "Reserves")
            raw_borrowing = parse_row_flexible(bs_table, ['borrowing'], "Borrowing")
            raw_other_liabilities = parse_row_flexible(bs_table, ['other liabilities'], "Other Liabilities")
            raw_trade_payables = parse_row_flexible(bs_table, ['trade payables'], "Trade Payables")
            raw_advance_customers = parse_row_flexible(bs_table, ['advance from customers'], "Advance from Customers")
            
            # Assets - main items
            raw_fixed_assets = parse_row_flexible(bs_table, ['fixed assets'], "Fixed Assets")
            raw_gross_block = parse_row_flexible(bs_table, ['gross block'], "Gross Block")
            raw_accumulated_dep = parse_row_flexible(bs_table, ['accumulated depreciation'], "Accumulated Depreciation")
            raw_cwip = parse_row_flexible(bs_table, ['cwip'], "CWIP")
            raw_investments = parse_row_flexible(bs_table, ['investments'], "Investments")
            
            # Current Assets
            raw_trade_receivables = parse_row_flexible(bs_table, ['trade receivables'], "Trade Receivables")
            raw_cash = parse_row_flexible(bs_table, ['cash equivalents'], "Cash")
            raw_inventory = parse_row_flexible(bs_table, ['inventories'], "Inventory")
            raw_loans_advances = parse_row_flexible(bs_table, ['loans n advances'], "Loans & Advances")
            raw_other_assets = parse_row_flexible(bs_table, ['other assets'], "Other Assets")
            
            # SECOND PASS: If main items are missing, look for them as nested items
            # This catches items that appear AFTER expandable section headers
            if not raw_trade_receivables:
                st.info("  üîç Trade receivables not found as main item, searching nested items...")
                # Look for rows that come after "Other Assets" or similar
                found_nested = False
                in_other_assets_section = False
                for tr in bs_table.find_all('tr'):
                    cells = tr.find_all(['td', 'th'])
                    if not cells:
                        continue
                    label = cells[0].get_text(strip=True).lower()
                    
                    # Check if we're in "Other Assets" section
                    if 'other asset' in label:
                        in_other_assets_section = True
                        continue
                    
                    # If we're in the section and find trade receivables
                    if in_other_assets_section and ('trade receivable' in label or 'receivable' in label):
                        values = []
                        for cell in cells[1:]:
                            raw = cell.get_text(strip=True).replace(',', '').replace('\xa0', '')
                            try:
                                values.append(float(raw))
                            except:
                                values.append(0.0)
                        if values and any(v != 0 for v in values):
                            raw_trade_receivables = values
                            st.write(f"  ‚úì Found nested Trade Receivables: {len(values)} years")
                            found_nested = True
                            break
                    
                    # Stop if we hit next major section
                    if in_other_assets_section and ('total' in label or 'fixed asset' in label or 'investment' in label):
                        break
            
            if not raw_inventory:
                st.info("  üîç Inventory not found as main item, searching nested items...")
                in_other_assets_section = False
                for tr in bs_table.find_all('tr'):
                    cells = tr.find_all(['td', 'th'])
                    if not cells:
                        continue
                    label = cells[0].get_text(strip=True).lower()
                    
                    if 'other asset' in label:
                        in_other_assets_section = True
                        continue
                    
                    if in_other_assets_section and ('inventor' in label or 'stock' in label):
                        values = []
                        for cell in cells[1:]:
                            raw = cell.get_text(strip=True).replace(',', '').replace('\xa0', '')
                            try:
                                values.append(float(raw))
                            except:
                                values.append(0.0)
                        if values and any(v != 0 for v in values):
                            raw_inventory = values
                            st.write(f"  ‚úì Found nested Inventory: {len(values)} years")
                            break
                    
                    if in_other_assets_section and ('total' in label or 'fixed asset' in label):
                        break
            
            raw_total_assets = parse_row_flexible(bs_table, ['total assets', 'total asset'], "Total Assets")
            raw_total_liabilities = parse_row_flexible(bs_table, ['total liabilities', 'total liability'], "Total Liabilities")
            
            # Determine main receivables based on what's available
            st.write("### üîç Determining Working Capital Items")
            
            # Use trade receivables if available
            raw_receivables = raw_trade_receivables if raw_trade_receivables and sum(raw_trade_receivables) > 0 else []
            
            # If no trade receivables but has loans & advances (NBFC), use that
            if not raw_receivables and raw_loans_advances and sum(raw_loans_advances) > 0:
                raw_receivables = raw_loans_advances
                st.info("  üè¶ Using Loans & Advances as Receivables (NBFC detected)")
            
            # Use trade payables if available
            raw_payables = raw_trade_payables if raw_trade_payables and sum(raw_trade_payables) > 0 else []
            
            # If no trade payables but has advances from customers, use that
            if not raw_payables and raw_advance_customers and sum(raw_advance_customers) > 0:
                raw_payables = raw_advance_customers
                st.info("  üì¶ Using Advances from Customers as Payables")
            
            # If no payables at all, use other liabilities
            if not raw_payables and raw_other_liabilities and sum(raw_other_liabilities) > 0:
                raw_payables = raw_other_liabilities
                st.info("  üìä Using Other Liabilities as Payables")
            
            # Show what we found
            if raw_receivables:
                st.write(f"  ‚úì Receivables: {len(raw_receivables)} years")
            else:
                st.warning("  ‚ö†Ô∏è No Receivables found")
            
            if raw_payables:
                st.write(f"  ‚úì Payables: {len(raw_payables)} years")
            else:
                st.warning("  ‚ö†Ô∏è No Payables found")
            
            if raw_inventory:
                st.write(f"  ‚úì Inventory: {len(raw_inventory)} years")
            else:
                st.info("  ‚ÑπÔ∏è No Inventory (may be service/NBFC company)")
            
            # Check what we got
            items_found = sum([
                1 if raw_revenue else 0,
                1 if raw_net_profit else 0,
                1 if raw_equity_capital else 0,
                1 if raw_reserves else 0,
                1 if raw_borrowing else 0,
                1 if raw_fixed_assets else 0,
                1 if raw_operating_profit or raw_pbt else 0
            ])
            
            st.info(f"üìä Extracted {items_found}/7 key line items successfully")
            
            # Show what's missing
            missing_items = []
            if not raw_revenue:
                missing_items.append("Revenue")
            if not raw_net_profit:
                missing_items.append("Net Profit")
            if not raw_equity_capital:
                missing_items.append("Equity Capital")
            if not raw_reserves:
                missing_items.append("Reserves")
            if not raw_borrowing:
                missing_items.append("Borrowings")
            if not raw_fixed_assets:
                missing_items.append("Fixed Assets")
            if not (raw_operating_profit or raw_pbt):
                missing_items.append("Operating Profit/PBT")
            
            if missing_items:
                st.warning(f"‚ö†Ô∏è Missing items: {', '.join(missing_items)}")
            
            if items_found < 3:
                st.error("‚ùå Insufficient data extracted from Screener.in. Too few key line items found.")
                st.warning("üí° This can happen if:\n- Company has limited financial history\n- Screener.in changed their HTML structure\n- Company uses non-standard accounting labels")
                return None
            
            def pad(lst, n):
                lst = [v for v in lst if v is not None]
                if len(lst) < n:
                    lst = [0.0] * (n - len(lst)) + lst
                return lst[-n:]
            
            n = num_years
            revenue = pad(raw_revenue, n)
            expenses = pad(raw_expenses, n)
            operating_profit = pad(raw_operating_profit, n)
            other_income = pad(raw_other_income, n)
            interest = pad(raw_interest, n)
            depreciation = pad(raw_depreciation, n)
            pbt = pad(raw_pbt, n)
            tax = pad(raw_tax, n)
            tax_pct = pad(raw_tax_pct, n)
            net_profit = pad(raw_net_profit, n)
            eps = pad(raw_eps, n)
            
            equity_capital = pad(raw_equity_capital, n)
            reserves = pad(raw_reserves, n)
            borrowing = pad(raw_borrowing, n)
            other_liabilities = pad(raw_other_liabilities, n)
            payables = pad(raw_payables, n)
            receivables = pad(raw_receivables, n)
            fixed_assets = pad(raw_fixed_assets, n)
            gross_block = pad(raw_gross_block, n)
            accumulated_dep = pad(raw_accumulated_dep, n)
            cwip = pad(raw_cwip, n)
            cash_vals = pad(raw_cash, n)
            inventory_vals = pad(raw_inventory, n)
            investments = pad(raw_investments, n)
            other_assets = pad(raw_other_assets, n)
            total_assets = pad(raw_total_assets, n)
            
            # Calculate shares from EPS
            shares = 0
            for i in range(n - 1, -1, -1):
                if eps[i] != 0 and net_profit[i] != 0:
                    shares = int((net_profit[i] * 10_000_000) / eps[i])
                    st.success(f"‚úÖ Calculated shares: {shares:,} (from Year {i+1} EPS: ‚Çπ{eps[i]:.2f})")
                    break
            
            # If EPS method failed, try NSEPy
            if shares == 0:
                st.warning("‚ö†Ô∏è Could not calculate from EPS. Trying NSEPy...")
                try:
                    # Try to get from NSE
                    import requests
                    nse_url = f"https://www.nseindia.com/api/quote-equity?symbol={symbol}"
                    headers = {
                        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
                        'Accept': '*/*',
                        'Accept-Language': 'en-US,en;q=0.9',
                    }
                    # Get session cookies first
                    session = requests.Session()
                    session.get("https://www.nseindia.com", headers=headers, timeout=5)
                    resp = session.get(nse_url, headers=headers, timeout=5)
                    if resp.status_code == 200:
                        data = resp.json()
                        # NSE gives "issuedSize" in number of shares
                        shares = data.get('securityInfo', {}).get('issuedSize', 0)
                        if shares > 0:
                            st.success(f"‚úÖ Fetched shares from NSE: {shares:,}")
                        else:
                            st.info("üí° NSE data available but no issuedSize found")
                    else:
                        st.info(f"üí° NSE API returned status {resp.status_code}")
                except Exception as e:
                    st.info(f"üí° NSE fetch failed: {str(e)[:100]}")
            
            if shares == 0:
                st.warning("‚ö†Ô∏è Could not calculate shares outstanding. Will need manual input.")
            
            CR_TO_LAC = 10.0
            from datetime import datetime as _dt
            current_year = _dt.now().year
            years_labels = [str(current_year - i) for i in range(n)]
            
            financials_out = {
                'years': years_labels,
                'revenue': [], 'cogs': [], 'opex': [], 'ebitda': [], 'depreciation': [],
                'ebit': [], 'interest': [], 'interest_income': [], 'tax': [], 'nopat': [],
                'fixed_assets': [], 'inventory': [], 'receivables': [], 'payables': [],
                'cash': [], 'equity': [], 'st_debt': [], 'lt_debt': []
            }
            
            for i in range(n - 1, -1, -1):
                rev = revenue[i] * CR_TO_LAC
                dep_val = depreciation[i] * CR_TO_LAC
                int_val = interest[i] * CR_TO_LAC
                other_inc = other_income[i] * CR_TO_LAC
                
                # Use operating profit if available, else derive from PBT
                if operating_profit[i] != 0:
                    ebit_val = operating_profit[i] * CR_TO_LAC
                    ebitda_val = ebit_val + dep_val
                else:
                    pbt_val = pbt[i] * CR_TO_LAC
                    ebitda_val = pbt_val + int_val + dep_val
                    ebit_val = ebitda_val - dep_val
                
                # COGS and OpEx estimation
                cogs_val = rev * 0.55 if rev > 0 else 0.0
                opex_val = max(0, rev - cogs_val - ebitda_val)
                if opex_val < 0 and expenses[i] != 0:
                    total_exp = expenses[i] * CR_TO_LAC
                    cogs_val = total_exp * 0.65
                    opex_val = total_exp * 0.35
                
                # Tax calculation
                tax_val = tax[i] * CR_TO_LAC if tax[i] != 0 else ebit_val * 0.25
                t_rate = min(0.35, tax_val / ebit_val) if ebit_val != 0 else 0.25
                nopat_val = ebit_val * (1 - t_rate)
                
                # Balance sheet
                eq_val = (equity_capital[i] + reserves[i]) * CR_TO_LAC
                fa_val = fixed_assets[i] * CR_TO_LAC
                pay_val = payables[i] * CR_TO_LAC
                rec_val = receivables[i] * CR_TO_LAC
                cash_val = cash_vals[i] * CR_TO_LAC
                inv_val = inventory_vals[i] * CR_TO_LAC
                borrow_val = borrowing[i] * CR_TO_LAC
                st_debt_val = borrow_val * 0.30
                lt_debt_val = borrow_val * 0.70
                
                for key, val in zip(
                    ['revenue', 'cogs', 'opex', 'ebitda', 'depreciation', 'ebit', 'interest', 'interest_income', 'tax', 'nopat', 'fixed_assets', 'inventory', 'receivables', 'payables', 'cash', 'equity', 'st_debt', 'lt_debt'],
                    [rev, cogs_val, opex_val, ebitda_val, dep_val, ebit_val, int_val, other_inc, tax_val, nopat_val, fa_val, inv_val, rec_val, pay_val, cash_val, eq_val, st_debt_val, lt_debt_val]
                ):
                    financials_out[key].append(val)
            
            st.success(f"‚úÖ Successfully parsed financials for {company_name}")
            return {'financials': financials_out, 'shares': shares, 'company_name': company_name, 'current_price': current_price}
            
        except Exception as e:
            st.error(f"‚ùå Scraper error: {str(e)}")
            import traceback
            st.code(traceback.format_exc())
            return None
except Exception as e:
    INDIAN_APIS_AVAILABLE = False
    SCREENER_IMPORT_ERROR = f"Exception: {e}"
    fetch_screener_financials = None

# ReportLab imports for PDF generation
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, PageBreak, Table, TableStyle, Image as RLImage, KeepTogether
from reportlab.pdfgen import canvas
from reportlab.lib.colors import HexColor
import tempfile
import os
from PIL import Image as PILImage
import io


# ================================
# AGGRESSIVE RATE LIMIT PREVENTION WITH DATA CACHING
# ================================

# Global cache for yfinance DATA (not just ticker objects)
# This caches .info, .financials, .balance_sheet, .cashflow to prevent duplicate API calls
_TICKER_DATA_CACHE = {}
_CACHE_TIMESTAMP = {}
CACHE_DURATION = 300  # 5 minutes cache

class CachedTickerData:
    """
    Wrapper class that caches yfinance ticker data properties
    This prevents rate limits by caching .info, .financials, etc.
    """
    def __init__(self, ticker_symbol, force_refresh=False):
        self.symbol = ticker_symbol
        self._ticker_obj = None
        self._info = None
        self._financials = None
        self._balance_sheet = None
        self._cashflow = None
        self._history = None
        self._loaded = False
        
        # Load from cache or fetch new
        if not force_refresh and ticker_symbol in _TICKER_DATA_CACHE:
            cache_age = time.time() - _CACHE_TIMESTAMP.get(ticker_symbol, 0)
            if cache_age < CACHE_DURATION:
                cached_data = _TICKER_DATA_CACHE[ticker_symbol]
                self._info = cached_data.get('info')
                self._financials = cached_data.get('financials')
                self._balance_sheet = cached_data.get('balance_sheet')
                self._cashflow = cached_data.get('cashflow')
                self._history = cached_data.get('history')
                self._loaded = True
    
    def _ensure_loaded(self):
        """Lazy load ticker data on first access"""
        if not self._loaded:
            try:
                self._ticker_obj = yf.Ticker(self.symbol)
                # Fetch all data at once to minimize API calls
                self._info = self._ticker_obj.info
                self._financials = self._ticker_obj.financials
                self._balance_sheet = self._ticker_obj.balance_sheet
                self._cashflow = self._ticker_obj.cashflow
                self._loaded = True
                
                # Cache the data
                _TICKER_DATA_CACHE[self.symbol] = {
                    'info': self._info,
                    'financials': self._financials,
                    'balance_sheet': self._balance_sheet,
                    'cashflow': self._cashflow,
                    'history': self._history
                }
                _CACHE_TIMESTAMP[self.symbol] = time.time()
            except Exception as e:
                # Return empty data on error
                self._info = {}
                self._financials = pd.DataFrame()
                self._balance_sheet = pd.DataFrame()
                self._cashflow = pd.DataFrame()
                raise e
    
    @property
    def info(self):
        """Get cached info or fetch if needed"""
        if self._info is None:
            self._ensure_loaded()
        return self._info if self._info is not None else {}
    
    @property
    def financials(self):
        """Get cached financials or fetch if needed"""
        if self._financials is None:
            self._ensure_loaded()
        return self._financials if self._financials is not None else pd.DataFrame()
    
    @property
    def balance_sheet(self):
        """Get cached balance sheet or fetch if needed"""
        if self._balance_sheet is None:
            self._ensure_loaded()
        return self._balance_sheet if self._balance_sheet is not None else pd.DataFrame()
    
    @property
    def cashflow(self):
        """Get cached cashflow or fetch if needed"""
        if self._cashflow is None:
            self._ensure_loaded()
        return self._cashflow if self._cashflow is not None else pd.DataFrame()
    
    def history(self, *args, **kwargs):
        """Get historical data - cache by parameters"""
        cache_key = f"{self.symbol}_history_{str(args)}_{str(kwargs)}"
        
        if cache_key in _TICKER_DATA_CACHE:
            cache_age = time.time() - _CACHE_TIMESTAMP.get(cache_key, 0)
            if cache_age < CACHE_DURATION:
                return _TICKER_DATA_CACHE[cache_key]
        
        # Fetch new history
        if self._ticker_obj is None:
            self._ticker_obj = yf.Ticker(self.symbol)
        
        hist_data = self._ticker_obj.history(*args, **kwargs)
        
        # Cache it
        _TICKER_DATA_CACHE[cache_key] = hist_data
        _CACHE_TIMESTAMP[cache_key] = time.time()
        
        return hist_data

def get_cached_ticker(ticker_symbol, force_refresh=False):
    """
    Get cached ticker data wrapper
    This returns a CachedTickerData object that caches .info, .financials, etc.
    to prevent rate limits from repeated API calls
    """
    return CachedTickerData(ticker_symbol, force_refresh)

def clear_ticker_cache():
    """Clear the ticker data cache"""
    global _TICKER_DATA_CACHE, _CACHE_TIMESTAMP
    _TICKER_DATA_CACHE = {}
    _CACHE_TIMESTAMP = {}


# ================================

# ================================

# Helper function for ticker exchange suffix
def get_ticker_with_exchange(ticker, exchange):
    """Add exchange suffix to ticker"""
    ticker = ticker.strip().upper()
    if '.NS' in ticker or '.BO' in ticker:
        return ticker  # Already has suffix
    return f"{ticker}.{exchange}"

# PDF EXPORT FUNCTIONS (EMBEDDED)
# ================================
class PageNumCanvas(canvas.Canvas):
    """Custom canvas to add page numbers and headers"""
    
    def __init__(self, *args, **kwargs):
        canvas.Canvas.__init__(self, *args, **kwargs)
        self.pages = []
        
    def showPage(self):
        self.pages.append(dict(self.__dict__))
        self._startPage()
        
    def save(self):
        page_count = len(self.pages)
        for page_num, page in enumerate(self.pages, 1):
            self.__dict__.update(page)
            if page_num > 1:  # Skip page number on cover
                self.draw_page_number(page_num - 1, page_count - 1)
            canvas.Canvas.showPage(self)
        canvas.Canvas.save(self)
        
    def draw_page_number(self, page_num, page_count):
        """Add page number at bottom"""
        self.setFont("Helvetica", 9)
        self.setFillColor(colors.grey)
        self.drawRightString(
            7.5*inch, 0.5*inch,
            f"Page {page_num} of {page_count}"
        )


def save_plotly_as_image(fig, width=1400, height=600):
    """Convert plotly figure to image bytes for PDF"""
    try:
        img_bytes = fig.to_image(format="png", width=width, height=height, scale=2)
        return img_bytes
    except Exception as e:
        print(f"Error converting chart: {e}")
        return None


def create_fair_value_chart(fair_values, current_price=None):
    """Create the fair value comparison bar chart"""
    methods = list(fair_values.keys())
    values = list(fair_values.values())
    
    colors_map = {
        'DCF': '#06A77D',
        'P/E': '#2E86AB',
        'P/B': '#4ECDC4',
        'P/S': '#FF6B6B',
        'EV/EBITDA': '#95E1D3',
        'DDM': '#F38181',
        'Residual Income': '#AA96DA'
    }
    
    bar_colors = [colors_map.get(m.split()[0], '#A8DADC') for m in methods]
    
    fig = go.Figure()
    
    fig.add_trace(go.Bar(
        x=methods,
        y=values,
        marker=dict(color=bar_colors, line=dict(color='white', width=2)),
        text=[f"‚Çπ{v:.2f}" for v in values],
        textposition='outside',
        showlegend=False
    ))
    
    avg_value = np.mean(values)
    fig.add_hline(y=avg_value, line_dash="dash", line_color="red",
                  annotation_text=f"Average: ‚Çπ{avg_value:.2f}")
    
    if current_price and current_price > 0:
        fig.add_hline(y=current_price, line_dash="dot", line_color="blue",
                      annotation_text=f"Current: ‚Çπ{current_price:.2f}")
    
    fig.update_layout(
        title="Fair Value Comparison - All Methods",
        xaxis_title="Valuation Method",
        yaxis_title="Fair Value (‚Çπ)",
        height=500,
        showlegend=False,
        plot_bgcolor='white',
        font=dict(size=12)
    )
    
    fig.update_xaxes(tickangle=-45)
    
    return fig


def create_peer_heatmap(peer_data, target_ticker):
    """Create peer comparison heatmap"""
    if peer_data.empty:
        return None
    
    # Prepare data for heatmap
    metrics = ['pe', 'pb', 'ps', 'ev_ebitda']
    display_names = ['P/E', 'P/B', 'P/S', 'EV/EBITDA']
    
    heat_data = []
    companies = []
    
    for _, row in peer_data.iterrows():
        companies.append(row.get('name', row.get('ticker', 'Unknown')))
        heat_data.append([row.get(m, 0) for m in metrics])
    
    heat_array = np.array(heat_data).T
    
    # Normalize to percentiles
    heat_normalized = np.zeros_like(heat_array)
    for i in range(len(heat_array)):
        heat_normalized[i] = np.argsort(np.argsort(heat_array[i])) / (len(heat_array[i]) - 1) * 100
    
    fig = go.Figure(data=go.Heatmap(
        z=heat_normalized,
        x=companies,
        y=display_names,
        text=heat_array,
        texttemplate='%{text:.2f}x',
        textfont={"size": 10},
        colorscale=[
            [0, '#06A77D'],
            [0.5, '#F4D35E'],
            [1, '#D62828']
        ],
        colorbar=dict(title="Percentile")
    ))
    
    fig.update_layout(
        title="Peer Valuation Multiples Heatmap",
        height=400,
        xaxis_title="Company",
        yaxis_title="Multiple"
    )
    
    return fig


def create_spider_chart(target_multiples, peer_medians):
    """Create spider/radar chart comparing target vs peers"""
    categories = list(target_multiples.keys())
    
    fig = go.Figure()
    
    fig.add_trace(go.Scatterpolar(
        r=list(target_multiples.values()),
        theta=categories,
        fill='toself',
        name='Target Company',
        line=dict(color='#D62828', width=2)
    ))
    
    fig.add_trace(go.Scatterpolar(
        r=list(peer_medians.values()),
        theta=categories,
        fill='toself',
        name='Peer Median',
        line=dict(color='#2E86AB', width=2)
    ))
    
    fig.update_layout(
        polar=dict(radialaxis=dict(visible=True)),
        showlegend=True,
        title="Target vs Peer Median - Valuation Multiples",
        height=500
    )
    
    return fig


def generate_professional_pdf(data_package, output_path=None):
    """
    Generate professional PDF report
    
    Parameters:
    -----------
    data_package : dict
        {
            'company_name': str,
            'ticker': str,
            'current_price': float,
            'financials': dict,
            'dcf_results': dict,
            'fair_values': dict,
            'peer_data': pd.DataFrame,
            'comp_results': dict (optional)
        }
    
    Returns:
    --------
    str : Path to generated PDF
    """
    
    # Extract data
    company_name = data_package.get('company_name', 'Company')
    ticker = data_package.get('ticker', 'TICKER')
    current_price = data_package.get('current_price', 0)
    financials = data_package.get('financials', {})
    dcf_results = data_package.get('dcf_results', {})
    fair_values = data_package.get('fair_values', {})
    peer_data = data_package.get('peer_data', pd.DataFrame())
    comp_results = data_package.get('comp_results')
    
    # Create output path
    if not output_path:
        temp_dir = tempfile.gettempdir()
        output_path = os.path.join(temp_dir, f"{ticker}_Professional_Report_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf")
    
    # Create PDF document
    doc = SimpleDocTemplate(
        output_path,
        pagesize=letter,
        rightMargin=0.75*inch,
        leftMargin=0.75*inch,
        topMargin=1*inch,
        bottomMargin=0.75*inch
    )
    
    # Styles
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(
        name='CoverTitle',
        parent=styles['Heading1'],
        fontSize=36,
        textColor=HexColor('#2E86AB'),
        alignment=TA_CENTER,
        fontName='Helvetica-Bold',
        spaceAfter=20
    ))
    styles.add(ParagraphStyle(
        name='SectionHeader',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=HexColor('#2E86AB'),
        spaceAfter=12,
        spaceBefore=12
    ))
    styles.add(ParagraphStyle(
        name='SubHeader',
        parent=styles['Heading2'],
        fontSize=14,
        textColor=HexColor('#06A77D'),
        spaceAfter=8
    ))
    
    story = []
    
    # ==========================================
    # COVER PAGE
    # ==========================================
    story.append(Spacer(1, 2*inch))
    story.append(Paragraph("DCF VALUATION REPORT", styles['CoverTitle']))
    story.append(Spacer(1, 0.5*inch))
    story.append(Paragraph(f"<b>{company_name}</b>", styles['Heading1']))
    story.append(Paragraph(f"Ticker: {ticker}", styles['Normal']))
    story.append(Spacer(1, 1*inch))
    story.append(Paragraph(
        f"<para align=center>Report Generated: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}</para>",
        styles['Normal']
    ))
    story.append(Spacer(1, 0.5*inch))
    story.append(Paragraph(
        "<para align=center><i>Professional Discounted Cash Flow Analysis</i></para>",
        styles['Normal']
    ))
    story.append(PageBreak())
    
    # ==========================================
    # EXECUTIVE SUMMARY
    # ==========================================
    story.append(Paragraph("EXECUTIVE SUMMARY", styles['SectionHeader']))
    story.append(Spacer(1, 0.2*inch))
    
    avg_fv = np.mean([v for v in fair_values.values() if v and v > 0]) if fair_values else 0
    upside = ((avg_fv - current_price) / current_price * 100) if current_price > 0 else 0
    
    # Key metrics table
    summary_data = [
        ['Metric', 'Value', 'Details'],
        ['Current Market Price', f"‚Çπ{current_price:.2f}" if current_price > 0 else "N/A (Unlisted)", 
         'Latest trading price' if current_price > 0 else 'Private company'],
        ['DCF Fair Value', f"‚Çπ{dcf_results.get('fair_value_per_share', 0):.2f}", 
         'Intrinsic value from DCF model'],
        ['Average Fair Value', f"‚Çπ{avg_fv:.2f}", 
         f'Average of {len(fair_values)} valuation methods'],
        ['Upside Potential', f"{upside:+.1f}%", 
         'Potential gain/loss from current price']
    ]
    
    summary_table = Table(summary_data, colWidths=[2*inch, 1.5*inch, 3*inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), HexColor('#2E86AB')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, HexColor('#F5F5F5')])
    ]))
    
    story.append(summary_table)
    story.append(Spacer(1, 0.3*inch))
    
    # Recommendation
    if upside > 15:
        recommendation = "STRONG BUY"
        rec_color = HexColor('#06A77D')
    elif upside > 0:
        recommendation = "BUY / HOLD"
        rec_color = HexColor('#F4D35E')
    else:
        recommendation = "HOLD / SELL"
        rec_color = HexColor('#D62828')
    
    rec_style = ParagraphStyle(
        'Recommendation',
        parent=styles['Normal'],
        fontSize=16,
        textColor=rec_color,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold'
    )
    
    story.append(Paragraph(f"<b>Investment Recommendation: {recommendation}</b>", rec_style))
    story.append(PageBreak())
    
    # ==========================================
    # HISTORICAL FINANCIALS
    # ==========================================
    if financials and 'years' in financials:
        story.append(Paragraph("HISTORICAL FINANCIAL ANALYSIS", styles['SectionHeader']))
        story.append(Spacer(1, 0.2*inch))
        
        years = financials['years']
        
        # Financial metrics table
        fin_data = [['Metric'] + [str(y) for y in years]]
        
        metrics_map = {
            'Revenue (‚Çπ Lacs)': 'revenue',
            'EBITDA (‚Çπ Lacs)': 'ebitda',
            'EBIT (‚Çπ Lacs)': 'ebit',
            'NOPAT (‚Çπ Lacs)': 'nopat',
            'Free Cash Flow (‚Çπ Lacs)': 'fcf',
            'CAPEX (‚Çπ Lacs)': 'capex'
        }
        
        for label, key in metrics_map.items():
            if key in financials:
                values = financials[key]
                row = [label] + [f"‚Çπ{v:,.0f}" for v in values]
                fin_data.append(row)
        
        # Growth rates
        if 'revenue' in financials and len(financials['revenue']) > 1:
            revenues = financials['revenue']
            growth_rates = []
            for i in range(1, len(revenues)):
                growth = ((revenues[i] - revenues[i-1]) / revenues[i-1]) * 100
                growth_rates.append(f"{growth:.1f}%")
            fin_data.append(['Revenue Growth (YoY)'] + ['‚Äî'] + growth_rates)
        
        fin_table = Table(fin_data, colWidths=[2.5*inch] + [1.2*inch] * len(years))
        fin_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), HexColor('#2E86AB')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, HexColor('#F5F5F5')])
        ]))
        
        story.append(fin_table)
        story.append(PageBreak())
    
    # ==========================================
    # DCF VALUATION BREAKDOWN
    # ==========================================
    if dcf_results:
        story.append(Paragraph("DCF VALUATION METHODOLOGY", styles['SectionHeader']))
        story.append(Spacer(1, 0.2*inch))
        
        story.append(Paragraph("Key Assumptions", styles['SubHeader']))
        
        assumptions_data = [
            ['Parameter', 'Value', 'Description'],
            ['WACC', f"{dcf_results.get('wacc', 0)*100:.2f}%", 'Weighted Average Cost of Capital'],
            ['Terminal Growth Rate', f"{dcf_results.get('terminal_growth_rate', 0)*100:.2f}%", 'Perpetual growth assumption'],
            ['Forecast Period', f"{dcf_results.get('forecast_years', 5)} years", 'Explicit forecast period'],
            ['Tax Rate', f"{dcf_results.get('tax_rate', 0)*100:.2f}%", 'Corporate tax rate']
        ]
        
        assumptions_table = Table(assumptions_data, colWidths=[2*inch, 1.5*inch, 3*inch])
        assumptions_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), HexColor('#06A77D')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, HexColor('#F5F5F5')])
        ]))
        
        story.append(assumptions_table)
        story.append(Spacer(1, 0.3*inch))
        
        # Valuation results
        story.append(Paragraph("Valuation Results", styles['SubHeader']))
        
        valuation_data = [
            ['Component', 'Value (‚Çπ Lacs)', 'Per Share (‚Çπ)'],
            ['Enterprise Value', f"{dcf_results.get('enterprise_value', 0):,.2f}", '‚Äî'],
            ['Less: Net Debt', f"{dcf_results.get('net_debt', 0):,.2f}", '‚Äî'],
            ['Equity Value', f"{dcf_results.get('equity_value', 0):,.2f}", '‚Äî'],
            ['Shares Outstanding', f"{dcf_results.get('shares', 0):,.0f}", '‚Äî'],
            ['Fair Value per Share', '‚Äî', f"‚Çπ{dcf_results.get('fair_value_per_share', 0):.2f}"]
        ]
        
        valuation_table = Table(valuation_data, colWidths=[2.5*inch, 2*inch, 2*inch])
        valuation_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), HexColor('#2E86AB')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('BACKGROUND', (0, -1), (-1, -1), HexColor('#E8F4F8')),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, HexColor('#F5F5F5')])
        ]))
        
        story.append(valuation_table)
        story.append(PageBreak())
    
    # ==========================================
    # FAIR VALUE COMPARISON CHART
    # ==========================================
    if fair_values:
        story.append(Paragraph("FAIR VALUE ANALYSIS", styles['SectionHeader']))
        story.append(Spacer(1, 0.2*inch))
        
        # Create and add fair value chart
        fv_chart = create_fair_value_chart(fair_values, current_price)
        fv_img = save_plotly_as_image(fv_chart, width=1400, height=600)
        
        if fv_img:
            img = RLImage(io.BytesIO(fv_img), width=6.5*inch, height=3*inch)
            story.append(KeepTogether([img]))
            story.append(Spacer(1, 0.2*inch))
        
        # Summary statistics
        values = [v for v in fair_values.values() if v and v > 0]
        stats_data = [
            ['Statistic', 'Value'],
            ['Minimum', f"‚Çπ{min(values):.2f}"],
            ['Maximum', f"‚Çπ{max(values):.2f}"],
            ['Average', f"‚Çπ{np.mean(values):.2f}"],
            ['Median', f"‚Çπ{np.median(values):.2f}"],
            ['Std Deviation', f"‚Çπ{np.std(values):.2f}"]
        ]
        
        stats_table = Table(stats_data, colWidths=[3*inch, 3*inch])
        stats_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), HexColor('#06A77D')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, HexColor('#F5F5F5')])
        ]))
        
        story.append(stats_table)
        story.append(PageBreak())
    
    # ==========================================
    # PEER COMPARISON (if available)
    # ==========================================
    if comp_results and not peer_data.empty:
        story.append(Paragraph("PEER COMPARISON ANALYSIS", styles['SectionHeader']))
        story.append(Spacer(1, 0.2*inch))
        
        # Peer heatmap
        heatmap = create_peer_heatmap(peer_data, ticker)
        if heatmap:
            heatmap_img = save_plotly_as_image(heatmap, width=1400, height=500)
            if heatmap_img:
                img = RLImage(io.BytesIO(heatmap_img), width=6.5*inch, height=2.5*inch)
                story.append(KeepTogether([img]))
                story.append(Spacer(1, 0.3*inch))
        
        # Spider chart if we have target multiples
        if comp_results.get('target'):
            target = comp_results['target']
            target_multiples = {
                'P/E': target.get('pe', 0),
                'P/B': target.get('pb', 0),
                'P/S': target.get('ps', 0),
                'EV/EBITDA': target.get('ev_ebitda', 0)
            }
            
            peer_medians = {
                'P/E': comp_results['multiples_stats'].get('pe', {}).get('median', 0),
                'P/B': comp_results['multiples_stats'].get('pb', {}).get('median', 0),
                'P/S': comp_results['multiples_stats'].get('ps', {}).get('median', 0),
                'EV/EBITDA': comp_results['multiples_stats'].get('ev_ebitda', {}).get('median', 0)
            }
            
            spider = create_spider_chart(target_multiples, peer_medians)
            spider_img = save_plotly_as_image(spider, width=1400, height=600)
            
            if spider_img:
                img = RLImage(io.BytesIO(spider_img), width=6.5*inch, height=3*inch)
                story.append(KeepTogether([img]))
        
        story.append(PageBreak())
    
    # ==========================================
    # CONCLUSION
    # ==========================================
    story.append(Paragraph("CONCLUSION & RECOMMENDATIONS", styles['SectionHeader']))
    story.append(Spacer(1, 0.2*inch))
    
    conclusion_text = f"""
    Based on comprehensive discounted cash flow analysis and relative valuation methodologies, 
    the estimated fair value for <b>{company_name}</b> ({ticker}) is <b>‚Çπ{avg_fv:.2f}</b> per share.
    """
    
    if current_price > 0:
        conclusion_text += f"""
        <br/><br/>
        The current market price of ‚Çπ{current_price:.2f} represents a <b>{upside:+.1f}%</b> 
        {'discount to' if upside > 0 else 'premium to'} our fair value estimate.
        <br/><br/>
        <b>Investment Recommendation: {recommendation}</b>
        """
    
    conclusion_text += """
    <br/><br/>
    <i>Risk Factors:</i>
    <br/>
    ‚Ä¢ Market volatility and macroeconomic conditions
    <br/>
    ‚Ä¢ Changes in industry dynamics and competitive landscape
    <br/>
    ‚Ä¢ Execution risks in business strategy
    <br/>
    ‚Ä¢ Regulatory and policy changes
    <br/><br/>
    <i>Disclaimer: This valuation is for informational purposes only and should not be 
    considered as investment advice. Please consult with a qualified financial advisor 
    before making investment decisions.</i>
    """
    
    story.append(Paragraph(conclusion_text, styles['Normal']))
    
    # Build PDF
    doc.build(story, canvasmaker=PageNumCanvas)
    
    return output_path


# Convenience function for easy import
def export_to_pdf(data_package):
    """
    Simple wrapper function for PDF generation
    
    Usage:
        from pdf_exporter import export_to_pdf
        
        pdf_path = export_to_pdf({
            'company_name': 'Reliance Industries',
            'ticker': 'RELIANCE',
            'current_price': 2500,
            'financials': financials_dict,
            'dcf_results': dcf_dict,
            'fair_values': {'DCF': 2600, 'P/E': 2550},
            'peer_data': peer_df,
            'comp_results': comp_dict
        })
    
    Returns:
        str: Path to generated PDF file
    """
    return generate_professional_pdf(data_package)



# STREAMLIT UI
# ================================

# Initialize session state for rate limiting
if 'last_yahoo_request' not in st.session_state:
    st.session_state.last_yahoo_request = 0

if 'yahoo_request_count' not in st.session_state:
    st.session_state.yahoo_request_count = 0

if 'session_start_time' not in st.session_state:
    import time
    st.session_state.session_start_time = time.time()

# Reset counter every hour
import time
if time.time() - st.session_state.session_start_time > 3600:
    st.session_state.yahoo_request_count = 0
    st.session_state.session_start_time = time.time()

# Fix text truncation in metrics and throughout the app
st.markdown("""
    <style>
    /* Fix metric value truncation */
    [data-testid="stMetricValue"] {
        width: fit-content;
        white-space: nowrap;
        overflow: visible !important;
        text-overflow: clip !important;
    }
    
    /* Fix metric label truncation */
    [data-testid="stMetricLabel"] {
        width: fit-content;
        white-space: nowrap;
        overflow: visible !important;
        text-overflow: clip !important;
    }
    
    /* Fix metric container */
    [data-testid="metric-container"] {
        overflow: visible !important;
        width: fit-content;
        min-width: max-content;
    }
    
    /* Fix metric delta truncation */
    [data-testid="stMetricDelta"] {
        white-space: nowrap;
        overflow: visible !important;
    }
    
    /* General fix for all div elements in columns */
    [data-testid="column"] > div {
        overflow: visible !important;
    }
    
    /* Ensure proper spacing for columns */
    [data-testid="column"] {
        overflow: visible !important;
        min-width: fit-content;
    }
    </style>
    """, unsafe_allow_html=True)


# UTILITY FUNCTIONS
# ================================


# Auto peer fetching utility - MANDATORY
try:
    from utils_peer_fetcher import get_industry_peers
    PEER_FETCHER_AVAILABLE = True
except Exception as e:
    PEER_FETCHER_AVAILABLE = False
    print(f"[DCF] ERROR: Peer fetcher not available: {e}")

def sanitize_value(val):
    """Convert string values to float, handling various formats"""
    if pd.isna(val) or val == '' or val == '-':
        return 0.0
    try:
        return float(str(val).replace(',', ''))
    except:
        return 0.0

def safe_extract(data, key, year, default=0.0):
    """
    ROBUST EXTRACTOR: Safely extract value from DataFrame/dict, handling None/NaN/missing data
    
    This is THE solution to the VEDL/TATAMOTORS None-value problem.
    Returns 0.0 instead of None to prevent NaN cascading through calculations.
    
    Args:
        data: DataFrame or dict containing financial data
        key: Row index or dictionary key to extract
        year: Column name (for DataFrame) or nested key
        default: Default value if extraction fails (default: 0.0)
    
    Returns:
        float: Extracted value or default (never None or NaN)
    """
    try:
        if isinstance(data, pd.DataFrame):
            if key in data.index and year in data.columns:
                val = data.loc[key, year]
                # Handle None, NaN, inf
                if val is None or pd.isna(val) or np.isinf(val):
                    return default
                return abs(float(val))
            return default
        elif isinstance(data, dict):
            val = data.get(key, default)
            if val is None or (isinstance(val, float) and (np.isnan(val) or np.isinf(val))):
                return default
            return abs(float(val))
        return default
    except Exception:
        return default

def safe_divide(numerator, denominator, default=0.0):
    """
    ROBUST DIVISION: Safely divide two numbers, handling None/NaN/zero division
    
    Args:
        numerator: Number to divide
        denominator: Number to divide by
        default: Default value if division fails (default: 0.0)
    
    Returns:
        float: Result of division or default (never None or NaN)
    """
    try:
        # Handle None inputs
        if numerator is None or denominator is None:
            return default
        
        # Convert to float and check for NaN/inf
        num = float(numerator)
        den = float(denominator)
        
        if np.isnan(num) or np.isinf(num) or np.isnan(den) or np.isinf(den):
            return default
        
        # Handle zero denominator
        if den == 0:
            return default
        
        result = num / den
        
        # Check result validity
        if np.isnan(result) or np.isinf(result):
            return default
        
        return result
    except Exception:
        return default

def ensure_valid_number(val, default=0.0):
    """
    ROBUST VALIDATOR: Ensure a value is a valid number (not None, NaN, or inf)
    
    Args:
        val: Value to validate
        default: Default value if validation fails (default: 0.0)
    
    Returns:
        float: Valid number or default (never None or NaN)
    """
    try:
        if val is None:
            return default
        
        num = float(val)
        
        if np.isnan(num) or np.isinf(num):
            return default
        
        return num
    except Exception:
        return default

def parse_excel_to_dataframes(excel_file):
    """Parse Excel file and extract Balance Sheet and P&L as DataFrames"""
    try:
        # Read both sheets from Excel
        df_bs = pd.read_excel(excel_file, sheet_name='BalanceSheet', header=None)
        df_pl = pd.read_excel(excel_file, sheet_name='Profit&Loss', header=None)
        
        # Set first column as 'Item' and rest as year columns
        # Column 0 = Item names, Columns 1,2,3... = Year data
        
        # Extract header row (first row contains year numbers)
        bs_years = df_bs.iloc[0, 1:].values  # Get year values from first row
        pl_years = df_pl.iloc[0, 1:].values
        
        # Create column names: 'Item' for first column, '_XX' for year columns
        bs_columns = ['Item'] + [f'_{int(year)}' if pd.notna(year) else f'_col{i}' 
                                  for i, year in enumerate(bs_years, 1)]
        pl_columns = ['Item'] + [f'_{int(year)}' if pd.notna(year) else f'_col{i}' 
                                  for i, year in enumerate(pl_years, 1)]
        
        # Remove header row and set column names
        df_bs = df_bs.iloc[1:].copy()  # Skip header row
        df_bs.columns = bs_columns
        
        df_pl = df_pl.iloc[1:].copy()
        df_pl.columns = pl_columns
        
        # Reset index
        df_bs = df_bs.reset_index(drop=True)
        df_pl = df_pl.reset_index(drop=True)
        
        # Convert year columns to numeric (sanitize values)
        for col in df_bs.columns[1:]:  # Skip 'Item' column
            df_bs[col] = df_bs[col].apply(sanitize_value)
        
        for col in df_pl.columns[1:]:
            df_pl[col] = df_pl[col].apply(sanitize_value)
        
        # Remove rows where Item is NaN or empty
        df_bs = df_bs[df_bs['Item'].notna() & (df_bs['Item'] != '')]
        df_pl = df_pl[df_pl['Item'].notna() & (df_pl['Item'] != '')]
        
        return df_bs, df_pl
        
    except Exception as e:
        st.error(f"Error parsing Excel: {str(e)}")
        return None, None

def get_value_from_df(df, item_name, year_col):
    """Extract value from DataFrame by item name (case-insensitive partial match)"""
    if df is None or df.empty:
        return 0.0
    
    item_name_lower = item_name.lower()
    mask = df['Item'].str.lower().str.contains(item_name_lower, na=False, regex=False)
    matching = df[mask]
    
    if not matching.empty and year_col in matching.columns:
        return matching.iloc[0][year_col]
    return 0.0

def detect_year_columns(df):
    """Detect year columns dynamically (columns starting with _)"""
    if df is None or df.empty:
        return []
    
    year_cols = [col for col in df.columns if col.startswith('_') and col != 'Item']
    # Sort by numeric value after underscore
    year_cols.sort(key=lambda x: int(x[1:]) if x[1:].isdigit() else 0)
    return year_cols

# ================================
# BUSINESS MODEL CLASSIFICATION (RULEBOOK COMPLIANT)
# ================================

def classify_business_model(financials, income_stmt=None, balance_sheet=None):
    """
    Classify company as OPERATING or INTEREST-DOMINANT per Rulebook Section 2.
    
    Returns:
        dict: {
            'type': 'OPERATING' or 'INTEREST_DOMINANT',
            'criteria_met': list of criteria that triggered classification,
            'metrics': dict of calculated ratios
        }
    """
    criteria_met = []
    metrics = {}
    
    # Calculate 3-year averages for classification
    try:
        # Method 1: From extracted financials dict
        if financials and 'revenue' in financials and 'interest' in financials:
            revenues = financials['revenue']
            interest_expenses = financials['interest']
            
            # Get interest income if available (from additional fields)
            interest_income = financials.get('interest_income', [0] * len(revenues))
            
            # Calculate averages
            avg_revenue = np.mean(revenues) if revenues else 0
            avg_interest_expense = np.mean(interest_expenses) if interest_expenses else 0
            avg_interest_income = np.mean(interest_income) if interest_income else 0
            
            # Criterion 1: Interest Income / Total Revenue ‚â• 50%
            if avg_revenue > 0:
                interest_income_ratio = (avg_interest_income / avg_revenue) * 100
                metrics['interest_income_ratio'] = interest_income_ratio
                if interest_income_ratio >= 50:
                    criteria_met.append(f"Interest Income / Revenue = {interest_income_ratio:.1f}% (‚â•50%)")
            
            # Criterion 2: Interest Expense / Total Expenses ‚â• 40%
            # Total Expenses = COGS + Opex + Interest + Depreciation
            total_expenses = []
            for i in range(len(revenues)):
                exp = (financials.get('cogs', [0]*len(revenues))[i] + 
                       financials.get('opex', [0]*len(revenues))[i] + 
                       interest_expenses[i] +
                       financials.get('depreciation', [0]*len(revenues))[i])
                total_expenses.append(exp)
            
            avg_total_expenses = np.mean(total_expenses) if total_expenses else 0
            if avg_total_expenses > 0:
                interest_expense_ratio = (avg_interest_expense / avg_total_expenses) * 100
                metrics['interest_expense_ratio'] = interest_expense_ratio
                if interest_expense_ratio >= 40:
                    criteria_met.append(f"Interest Expense / Total Expenses = {interest_expense_ratio:.1f}% (‚â•40%)")
        
        # Method 2: Check raw statements for Net Interest Income presence
        if income_stmt is not None:
            nii_fields = ['Net Interest Income', 'Net Interest Margin', 'Interest Income Net']
            for field in nii_fields:
                if field in income_stmt.index:
                    criteria_met.append(f"Presence of '{field}' line item")
                    break
        
        # Method 3: Balance sheet structure check
        if balance_sheet is not None:
            # Check for lending business indicators
            lending_indicators = ['Loans', 'Advances', 'Loans And Advances', 'Net Loans']
            financial_assets = ['Financial Assets', 'Investment Securities']
            
            for indicator in lending_indicators + financial_assets:
                if indicator in balance_sheet.index:
                    # Check if it's a significant portion
                    if 'Total Assets' in balance_sheet.index:
                        try:
                            asset_val = abs(balance_sheet.loc[indicator, balance_sheet.columns[0]])
                            total_assets = abs(balance_sheet.loc['Total Assets', balance_sheet.columns[0]])
                            if total_assets > 0 and (asset_val / total_assets) > 0.5:
                                criteria_met.append(f"Balance Sheet dominated by {indicator} ({asset_val/total_assets*100:.1f}% of assets)")
                        except:
                            pass
    
    except Exception as e:
        st.warning(f"Classification warning: {str(e)}")
    
    # Decision: INTEREST-DOMINANT if 2+ criteria met
    is_interest_dominant = len(criteria_met) >= 2
    
    classification = {
        'type': 'INTEREST_DOMINANT' if is_interest_dominant else 'OPERATING',
        'criteria_met': criteria_met,
        'metrics': metrics
    }
    
    return classification

def validate_fcff_eligibility(classification):
    """
    Check if FCFF DCF is valid per Rulebook Section 3.
    Returns: (is_valid: bool, reason: str)
    """
    if classification['type'] == 'INTEREST_DOMINANT':
        return False, "üö´ FCFF DCF is NOT VALID for Interest-Dominant entities. Debt is operating raw material, not financing."
    
    return True, "‚úÖ FCFF DCF is valid for Operating Companies"

def show_classification_warning(classification):
    """Display business model classification and restrictions"""
    if classification['type'] == 'INTEREST_DOMINANT':
        st.error("""
        üö´ **INTEREST-DOMINANT ENTITY DETECTED**
        
        This company derives significant income from interest operations (lending/banking).
        
        **Why FCFF DCF is Invalid:**
        - Interest expense = Operating Cost (like COGS), not financing cost
        - Interest income = Revenue
        - Debt = Operating raw material (inventory equivalent)
        - EBIT/NOPAT/WACC are economically meaningless
        
        **Criteria Met:**
        """)
        for criterion in classification['criteria_met']:
            st.write(f"  ‚Ä¢ {criterion}")
        
        st.info("""
        **Recommended Valuation Methods:**
        - ‚úÖ Residual Income Model (preferred)
        - ‚úÖ Dividend Discount Model
        - ‚úÖ P/B with ROE analysis
        - ‚úÖ Relative valuation (P/E, P/B)
        
        ‚ùå FCFF DCF is blocked to prevent economically invalid valuation.
        """)
        
        return True  # Should stop execution
    
    else:
        st.success(f"""
        ‚úÖ **OPERATING COMPANY CLASSIFICATION**
        
        FCFF DCF valuation is appropriate for this company.
        """)
        
        if classification['criteria_met']:
            with st.expander("‚ÑπÔ∏è Classification Details"):
                st.write("The following interest-related metrics were detected but did not exceed thresholds:")
                for criterion in classification['criteria_met']:
                    st.write(f"  ‚Ä¢ {criterion}")
        
        return False  # Can continue

# ================================
# BANK VALUATION METHODS
# ================================

def calculate_residual_income_model(financials, shares, cost_of_equity, terminal_growth=3.5, projection_years=5, assumed_roe=None, dcf_projections=None):
    """
    Residual Income Model - Suitable for both banks and non-banking companies
    RI = Net Income - (Cost of Equity √ó Book Value of Equity)
    Value = Book Value + PV(Future Residual Income)
    
    Particularly effective for:
    - Companies with stable book value
    - Banks and financial institutions
    - Asset-heavy businesses
    
    Args:
        terminal_growth: Terminal growth rate (default 3.5%)
        projection_years: Years to project (default 5)
        assumed_roe: Override ROE if provided
        dcf_projections: DCF projections dict with 'nopat' key - if provided, uses projected NOPAT as Net Income (NO DUPLICATION!)
    """
    try:
        # Validation checks with specific error messages
        if 'equity' not in financials or len(financials['equity']) == 0:
            return {
                'error': True,
                'reason': 'No equity (book value) data available',
                'suggestion': 'RIM requires balance sheet equity data. Use DCF or DDM instead.'
            }
        
        # Get latest data - USE NEWEST (index 0, not -1)
        latest_equity = financials['equity'][0] * 100000  # Convert from Lacs to Rupees
        
        if latest_equity <= 0:
            return {
                'error': True,
                'reason': f'Company has negative or zero book value (‚Çπ{latest_equity:,.0f})',
                'suggestion': 'RIM requires positive equity. This company may be distressed. Use DCF instead.'
            }
        
        # Calculate average ROE
        net_incomes = []
        equities = []
        for i in range(len(financials['years'])):
            # Approximate net income from NOPAT (banks don't really have NOPAT, using as proxy)
            ni = financials['nopat'][i] * 100000
            eq = financials['equity'][i] * 100000
            net_incomes.append(ni)
            equities.append(eq)
        
        avg_net_income = np.mean(net_incomes)
        
        if avg_net_income <= 0:
            return {
                'error': True,
                'reason': f'Company has negative average net income (‚Çπ{avg_net_income:,.0f} Lacs)',
                'suggestion': 'RIM requires profitable companies. This company is loss-making. Use DCF or asset-based valuation.'
            }
        
        # Use provided ROE or calculate
        if assumed_roe:
            roe = assumed_roe
        else:
            roe = (avg_net_income / latest_equity * 100) if latest_equity > 0 else 15
        
        if roe < 0:
            return {
                'error': True,
                'reason': f'Company has negative ROE ({roe:.1f}%)',
                'suggestion': 'RIM requires positive ROE. Company is destroying shareholder value. Use DCF instead.'
            }
        
        # Calculate historical book value growth rate using CAGR (data newest to oldest)
        if len(equities) > 1 and equities[-1] > 0 and equities[0] > 0:
            num_years = len(equities) - 1
            # Start = oldest (last), End = newest (first)
            bv_growth = ((equities[0] / equities[-1]) ** (1 / num_years) - 1) * 100
            bv_growth = max(-50, min(bv_growth, 150))  # Allow up to 150% growth
        else:
            bv_growth = 10.0  # Default
        
        # Project N years of residual income
        # ‚úÖ USE EXISTING DCF PROJECTIONS IF AVAILABLE - NO DUPLICATION!
        projections = []
        current_bv = latest_equity
        
        if dcf_projections and 'nopat' in dcf_projections and len(dcf_projections['nopat']) >= projection_years:
            # Use projected NOPAT from DCF as Net Income (already calculated!)
            for year in range(1, projection_years + 1):
                projected_nopat_lacs = dcf_projections['nopat'][year-1]  # 0-indexed
                current_ni = projected_nopat_lacs * 100000  # Convert to Rupees
                
                # Book value grows with retained earnings
                # Simplified: BV grows at historical rate
                current_bv = current_bv * (1 + bv_growth / 100)
                
                # Residual income = NI - (Ke √ó BV)
                ri = current_ni - (cost_of_equity / 100 * current_bv)
                
                # Present value
                pv_ri = ri / ((1 + cost_of_equity / 100) ** year)
                projections.append({
                    'year': year,
                    'book_value': current_bv,
                    'net_income': current_ni,
                    'residual_income': ri,
                    'pv_ri': pv_ri,
                    'source': 'DCF Projected NOPAT'
                })
        else:
            # Fallback: use ROE-based projection
            current_ni = avg_net_income
            for year in range(1, projection_years + 1):
                # Growth in book value
                current_bv = current_bv * (1 + bv_growth / 100)
                current_ni = current_bv * (roe / 100)
                
                # Residual income = NI - (Ke √ó BV)
                ri = current_ni - (cost_of_equity / 100 * current_bv)
                
                # Present value
                pv_ri = ri / ((1 + cost_of_equity / 100) ** year)
                projections.append({
                    'year': year,
                'book_value': current_bv,
                'net_income': current_ni,
                'residual_income': ri,
                'pv_ri': pv_ri,
                'source': 'ROE-based projection'
            })
        
        # Terminal value of residual income
        if cost_of_equity / 100 > terminal_growth / 100:
            terminal_ri = projections[-1]['residual_income'] * (1 + terminal_growth / 100) / (cost_of_equity / 100 - terminal_growth / 100)
            pv_terminal_ri = terminal_ri / ((1 + cost_of_equity / 100) ** 5)
        else:
            pv_terminal_ri = 0
        
        # Total value = Current BV + Sum of PV(RI) + PV(Terminal RI)
        sum_pv_ri = sum([p['pv_ri'] for p in projections])
        total_equity_value = latest_equity + sum_pv_ri + pv_terminal_ri
        
        value_per_share = total_equity_value / shares if shares > 0 else 0
        book_value_per_share = latest_equity / shares if shares > 0 else 0
        current_eps = avg_net_income / shares if shares > 0 else 0
        
        return {
            'method': 'Residual Income Model',
            'current_book_value': latest_equity,
            'book_value_per_share': book_value_per_share,
            'current_eps': current_eps,
            'roe': roe,
            'bv_growth': bv_growth,
            'cost_of_equity': cost_of_equity,
            'terminal_growth': terminal_growth,
            'projections': projections,
            'sum_pv_ri': sum_pv_ri,
            'terminal_ri_pv': pv_terminal_ri,
            'total_equity_value': total_equity_value,
            'value_per_share': value_per_share,
            'using_dcf_projections': bool(dcf_projections)
        }
    except Exception as e:
        return {
            'error': True,
            'reason': f'Calculation error: {str(e)}',
            'suggestion': 'Check if financial data is complete and valid. Use DCF as primary valuation method.',
            'technical_details': str(e)
        }

def calculate_dividend_discount_model(financials, shares, cost_of_equity, ticker=None, div_growth_override=None, payout_ratio_override=None, dcf_projections=None):
    """
    Dividend Discount Model (Gordon Growth Model)
    Value = D1 / (Ke - g)
    
    Args:
        div_growth_override: Override dividend growth rate
        payout_ratio_override: Override payout ratio
        dcf_projections: DCF projections dict with 'nopat' key - if provided, uses projected NOPAT for dividend projections (NO DUPLICATION!)
    """
    try:
        # Try to fetch actual dividend data from yfinance
        actual_dividends = []
        div_growth_calculated = None
        payout_ratio_calculated = None
        
        if ticker:
            try:
                stock = get_cached_ticker(get_ticker_with_exchange(ticker, exchange_suffix))
                dividends_hist = stock.dividends
                
                if not dividends_hist.empty and len(dividends_hist) > 0:
                    # Get annual dividends for last 3 years
                    dividends_by_year = dividends_hist.resample('Y').sum()
                    if len(dividends_by_year) >= 2:
                        recent_divs = dividends_by_year[-3:].values
                        actual_dividends = recent_divs.tolist()
                        
                        # Calculate growth rate using CAGR (data is newest to oldest)
                        if len(actual_dividends) >= 2 and actual_dividends[-1] > 0 and actual_dividends[0] > 0:
                            num_years = len(actual_dividends) - 1
                            # Start = oldest (last), End = newest (first)
                            div_growth_calculated = ((actual_dividends[0] / actual_dividends[-1]) ** (1 / num_years) - 1) * 100
                            div_growth_calculated = max(-50, min(div_growth_calculated, 150))  # Allow up to 150%
                        
                        # Calculate payout ratio from actual data - USE NEWEST nopat (index 0)
                        latest_div = actual_dividends[-1] if actual_dividends else 0
                        latest_ni = financials['nopat'][0] * 100000
                        if latest_ni > 0 and latest_div > 0:
                            payout_ratio_calculated = (latest_div * shares) / latest_ni
                            payout_ratio_calculated = max(0.1, min(payout_ratio_calculated, 0.9))
            except Exception as e:
                pass
        
        # Calculate average earnings
        net_incomes = []
        for i in range(len(financials['years'])):
            ni = financials['nopat'][i] * 100000
            net_incomes.append(ni)
        
        avg_net_income = np.mean(net_incomes)
        
        # Use overrides or calculated or default values
        if payout_ratio_override:
            payout_ratio = payout_ratio_override / 100
        elif payout_ratio_calculated:
            payout_ratio = payout_ratio_calculated
        else:
            payout_ratio = 0.40
        
        if div_growth_override:
            div_growth = div_growth_override
        elif div_growth_calculated:
            div_growth = div_growth_calculated
        else:
            div_growth = 8.0
        
        # Calculate dividends
        total_dividends = avg_net_income * payout_ratio
        dps = total_dividends / shares if shares > 0 else 0
        
        # If we have actual dividends, use the latest as current DPS
        if actual_dividends:
            dps = actual_dividends[-1]
        
        # Next year dividend
        d1 = dps * (1 + div_growth / 100)
        
        # DDM valuation
        if cost_of_equity <= div_growth:
            return None
        
        value_per_share = d1 / ((cost_of_equity - div_growth) / 100)
        
        # 5-year dividend projection
        # ‚úÖ USE EXISTING DCF PROJECTIONS IF AVAILABLE - NO DUPLICATION!
        projections_list = []
        
        if dcf_projections and 'nopat' in dcf_projections and len(dcf_projections['nopat']) > 0:
            # Use projected NOPAT from DCF (already calculated!)
            for year_idx, projected_nopat_lacs in enumerate(dcf_projections['nopat'], 1):
                projected_ni = projected_nopat_lacs * 100000  # Convert to Rupees
                projected_dividend = projected_ni * payout_ratio
                projected_dps = projected_dividend / shares if shares > 0 else 0
                pv_div = projected_dps / ((1 + cost_of_equity / 100) ** year_idx)
                projections_list.append({
                    'year': year_idx,
                    'dividend': projected_dps,
                    'pv_dividend': pv_div,
                    'source': 'DCF Projected NOPAT'
                })
        else:
            # Fallback: use growth-based projection
            current_div = dps
            for year in range(1, 6):
                current_div = current_div * (1 + div_growth / 100)
                pv_div = current_div / ((1 + cost_of_equity / 100) ** year)
                projections_list.append({
                    'year': year,
                    'dividend': current_div,
                    'pv_dividend': pv_div,
                    'source': 'Growth-based projection'
                })
        
        return {
            'method': 'Dividend Discount Model',
            'current_dps': dps,
            'payout_ratio': payout_ratio * 100,
            'dividend_growth': div_growth,
            'required_return': cost_of_equity,
            'next_year_dps': d1,
            'projections': projections_list,
            'value_per_share': value_per_share,
            'using_actual_data': bool(actual_dividends),
            'using_dcf_projections': bool(dcf_projections),
            'historical_dividends': actual_dividends if actual_dividends else None
        }
    except Exception as e:
        st.error(f"DDM error: {str(e)}")
        return None

def calculate_pb_roe_valuation(financials, shares, cost_of_equity, assumed_roe=None):
    """
    P/B with ROE Analysis
    Fair P/B = ROE / Cost of Equity
    
    Args:
        assumed_roe: Override ROE if provided
    """
    try:
        # Latest book value - USE NEWEST (index 0)
        latest_equity = financials['equity'][0] * 100000
        book_value_per_share = latest_equity / shares if shares > 0 else 0
        
        # Calculate ROE
        net_incomes = []
        equities = []
        for i in range(len(financials['years'])):
            ni = financials['nopat'][i] * 100000
            eq = financials['equity'][i] * 100000
            net_incomes.append(ni)
            equities.append(eq)
        
        avg_net_income = np.mean(net_incomes)
        avg_equity = np.mean(equities)
        
        # Use provided ROE or calculate
        if assumed_roe:
            roe = assumed_roe
        else:
            roe = (avg_net_income / avg_equity * 100) if avg_equity > 0 else 15
        
        # Fair P/B ratio
        fair_pb = roe / cost_of_equity
        
        # Fair value per share
        value_per_share = book_value_per_share * fair_pb
        
        return {
            'method': 'P/B with ROE Analysis',
            'book_value_per_share': book_value_per_share,
            'roe': roe,
            'cost_of_equity': cost_of_equity,
            'fair_pb_ratio': fair_pb,
            'value_per_share': value_per_share,
            'historical_roe': [(net_incomes[i] / equities[i] * 100) for i in range(len(net_incomes))]
        }
    except Exception as e:
        st.error(f"P/B ROE error: {str(e)}")
        return None


def get_auto_peers_or_default(ticker):
    """Get auto peers or fallback to defaults - MANDATORY"""
    if PEER_FETCHER_AVAILABLE:
        try:
            print(f"[DCF] Auto-fetching peers for {ticker}...")
            peers = get_industry_peers(ticker, max_peers=10, exclude_self=True)
            if peers:
                # VERIFY: ticker not in peer list
                ticker_upper = ticker.upper()
                peers = [p for p in peers if p.upper() != ticker_upper]
                peer_str = ",".join(peers)
                print(f"[DCF] ‚úÖ Auto-fetched {len(peers)} peers: {peers[:5]}")
                return peer_str
            else:
                print("[DCF] ‚ö†Ô∏è No peers found, using defaults")
        except Exception as e:
            print(f"[DCF] ‚ö†Ô∏è Error fetching peers: {e}")
    return "HDFCBANK,ICICIBANK,SBIN,AXISBANK,KOTAKBANK"

def calculate_relative_valuation(ticker, financials, shares, peer_tickers=None, exchange_suffix="NS"):
    """
    Relative Valuation using peer multiples
    P/E and P/B comparisons with actual peer data
    
    Handles rate limiting gracefully with retries and delays
    """
    import time
    import random
    
    try:
        if not ticker:
            return None
        
        # Get stock info with rate limit handling
        max_retries = 3
        retry_delay = 2  # seconds
        
        for attempt in range(max_retries):
            try:
                stock = get_cached_ticker(get_ticker_with_exchange(ticker, exchange_suffix))
                info = stock.info
                # Robust price fetching - try multiple methods
                current_price = info.get('currentPrice', 0)
                if not current_price or current_price == 0:
                    current_price = info.get('regularMarketPrice', 0)
                if not current_price or current_price == 0:
                    try:
                        hist = stock.history(period='1d')
                        if not hist.empty:
                            current_price = hist['Close'].iloc[-1]
                    except:
                        pass
                break
            except Exception as e:
                if "rate" in str(e).lower() or "429" in str(e):
                    if attempt < max_retries - 1:
                        wait_time = retry_delay * (attempt + 1) + random.uniform(0.5, 1.5)
                        st.warning(f"‚è≥ Rate limit hit. Waiting {wait_time:.1f}s... (Attempt {attempt + 1}/{max_retries})")
                        time.sleep(wait_time)
                    else:
                        st.error("‚ùå Rate limit exceeded. Using fallback calculations.")
                        current_price = 0
                else:
                    raise
        
        # Calculate company metrics - USE NEWEST (index 0)
        latest_ni = financials['nopat'][0] * 100000
        eps = latest_ni / shares if shares > 0 else 0
        
        # USE NEWEST equity (index 0)
        latest_equity = financials['equity'][0] * 100000
        bvps = latest_equity / shares if shares > 0 else 0
        
        current_pe = current_price / eps if eps > 0 else 0
        current_pb = current_price / bvps if bvps > 0 else 0
        
        # Fetch peer multiples with rate limiting
        peer_pe_list = []
        peer_pb_list = []
        peer_data = []
        
        # Default bank peers if none provided
        if not peer_tickers:
            peer_tickers = get_auto_peers_or_default(ticker)
        
        peers = [t.strip() for t in peer_tickers.split(',') if t.strip()]
        
        st.info(f"üìä Fetching data for {len(peers[:10])} peer companies...")
        
        for i, peer in enumerate(peers[:10]):  # Limit to 10 peers
            try:
                # Add delay between requests to avoid rate limiting
                if i > 0:
                    time.sleep(random.uniform(1.0, 1.5))  # Reduced from 2-3s since caching prevents duplicates
                
                peer_stock = get_cached_ticker(get_ticker_with_exchange(peer, exchange_suffix))
                peer_info = peer_stock.info
                
                peer_pe = peer_info.get('trailingPE', 0)
                peer_pb = peer_info.get('priceToBook', 0)
                # Robust price fetching for peers
                peer_price = peer_info.get('currentPrice', 0)
                if not peer_price or peer_price == 0:
                    peer_price = peer_info.get('regularMarketPrice', 0)
                if not peer_price or peer_price == 0:
                    try:
                        hist = peer_stock.history(period='1d')
                        if not hist.empty:
                            peer_price = hist['Close'].iloc[-1]
                    except:
                        pass
                
                if peer_pe and peer_pe > 0 and peer_pe < 100:  # Sanity check
                    peer_pe_list.append(peer_pe)
                
                if peer_pb and peer_pb > 0 and peer_pb < 20:  # Sanity check
                    peer_pb_list.append(peer_pb)
                
                if peer_price > 0:
                    peer_data.append({
                        'ticker': peer,
                        'price': peer_price,
                        'pe': peer_pe if peer_pe else 'N/A',
                        'pb': peer_pb if peer_pb else 'N/A'
                    })
            except Exception as e:
                if "rate" in str(e).lower() or "429" in str(e):
                    st.warning(f"‚è≥ Rate limit hit on peer {peer}. Skipping...")
                    time.sleep(2)  # Longer delay after rate limit
                continue
        
        # Calculate sector averages
        if peer_pe_list:
            sector_avg_pe = np.median(peer_pe_list)  # Use median to avoid outliers
            sector_low_pe = np.percentile(peer_pe_list, 25)
            sector_high_pe = np.percentile(peer_pe_list, 75)
            st.success(f"‚úÖ Fetched {len(peer_pe_list)} peer P/E ratios")
        else:
            st.warning("‚ö†Ô∏è Using default industry P/E multiples (no peer data available)")
            sector_avg_pe = 20  # Fallback
            sector_low_pe = 15
            sector_high_pe = 25
        
        if peer_pb_list:
            sector_avg_pb = np.median(peer_pb_list)
            sector_low_pb = np.percentile(peer_pb_list, 25)
            sector_high_pb = np.percentile(peer_pb_list, 75)
        else:
            sector_avg_pb = 3  # Fallback
            sector_low_pb = 2
            sector_high_pb = 4
        
        # Fair value based on sector multiples
        fair_value_pe = eps * sector_avg_pe
        fair_value_pb = bvps * sector_avg_pb
        
        # Conservative and aggressive estimates
        conservative_value = eps * sector_low_pe
        aggressive_value = eps * sector_high_pe
        
        return {
            'method': 'Relative Valuation',
            'current_price': current_price,
            'eps': eps,
            'bvps': bvps,
            'current_pe': current_pe,
            'current_pb': current_pb,
            'sector_avg_pe': sector_avg_pe,
            'sector_avg_pb': sector_avg_pb,
            'sector_low_pe': sector_low_pe,
            'sector_high_pe': sector_high_pe,
            'sector_low_pb': sector_low_pb,
            'sector_high_pb': sector_high_pb,
            'fair_value_pe_based': fair_value_pe,
            'fair_value_pb_based': fair_value_pb,
            'conservative_value': conservative_value,
            'aggressive_value': aggressive_value,
            'avg_fair_value': (fair_value_pe + fair_value_pb) / 2,
            'peer_count': len(peer_pe_list),
            'peer_data': peer_data,
            'rate_limited': len(peer_pe_list) == 0  # Flag if we got rate limited
        }
    except Exception as e:
        error_msg = str(e)
        if "rate" in error_msg.lower() or "429" in error_msg:
            st.error("‚ùå **Rate Limit Exceeded**")
            st.info("""
            **Why this happens:**
            - Yahoo Finance limits requests to prevent abuse
            - Too many requests in short time
            
            **Solutions:**
            1. Wait 5-10 minutes and try again
            2. Reduce number of peer companies
            3. Use the app during off-peak hours
            4. Consider using cached/default multiples
            
            **Alternative:** Use P/B ROE model or DDM for valuation instead
            """)
        else:
            st.error(f"Relative valuation error: {error_msg}")
        return None

# ================================
# YAHOO FINANCE SCRAPING WITH CACHING AND RETRY
# ================================

@st.cache_data(ttl=3600)  # Cache for 1 hour
def fetch_yahoo_financials_cached(ticker, exchange_suffix="NS"):
    """Cached wrapper for Yahoo Finance fetch"""
    return fetch_yahoo_financials_internal(ticker, exchange_suffix)

def fetch_yahoo_financials(ticker, exchange_suffix="NS"):
    """
    Fetch financial statements from Yahoo Finance with comprehensive error handling
    
    Features:
    - Retry logic with exponential backoff
    - Rate limit detection
    - Session caching
    - Automatic delays
    """
    import time
    import random
    
    max_retries = 3
    base_delay = 3  # seconds
    
    for attempt in range(max_retries):
        try:
            # Add delay before request (except first attempt)
            if attempt > 0:
                delay = base_delay * (2 ** attempt) + random.uniform(1, 3)
                st.warning(f"‚è≥ Rate limit detected. Waiting {delay:.1f} seconds... (Attempt {attempt + 1}/{max_retries})")
                time.sleep(delay)
            elif attempt == 0 and hasattr(st.session_state, 'last_yahoo_request'):
                # Add small delay between different requests
                elapsed = time.time() - st.session_state.last_yahoo_request
                if elapsed < 1:
                    time.sleep(1 - elapsed)
            
            # Try cached version first
            if attempt == 0:
                try:
                    return fetch_yahoo_financials_cached(ticker, exchange_suffix)
                except:
                    pass  # If cache fails, continue to direct fetch
            
            # Direct fetch
            result = fetch_yahoo_financials_internal(ticker, exchange_suffix)
            
            # Track request time and count
            st.session_state.last_yahoo_request = time.time()
            st.session_state.yahoo_request_count += 1
            
            return result
            
        except Exception as e:
            error_str = str(e).lower()
            
            if any(keyword in error_str for keyword in ['rate', '429', 'too many', 'limit']):
                if attempt < max_retries - 1:
                    continue  # Try again with delay
                else:
                    # All retries exhausted
                    st.error("‚ùå **Yahoo Finance Rate Limit Exceeded**")
                    st.warning("""
                    **Too many requests to Yahoo Finance. Please try one of these options:**
                    
                    1. ‚è∞ **Wait 10-15 minutes** and try again
                    2. üåê **Use a different network** (mobile hotspot, VPN)
                    3. üîÑ **Restart your Streamlit app** to clear session
                    4. üìä **Use the Excel upload feature** for unlisted companies instead
                    
                    **Why this happens:**
                    Yahoo Finance limits free API requests to prevent abuse. This is a Yahoo limitation, not an issue with our app.
                    """)
                    return None, "Rate limit exceeded. Please try again later or use alternative data sources."
            else:
                # Other error
                if attempt < max_retries - 1:
                    continue
                else:
                    return None, f"Error fetching data: {str(e)}"
    
    return None, "Failed to fetch data after multiple retries"

def fetch_yahoo_financials_internal(ticker, exchange_suffix="NS"):
    """Internal function for actual Yahoo Finance fetch"""
    try:
        stock = get_cached_ticker(get_ticker_with_exchange(ticker, exchange_suffix))
        
        # Get financial statements
        income_stmt = stock.financials
        balance_sheet = stock.balance_sheet
        cash_flow = stock.cashflow
        
        # Get company info
        info = stock.info
        
        if income_stmt.empty or balance_sheet.empty:
            return None, "No financial data available"
        
        # Get shares outstanding - ROBUST MULTI-METHOD APPROACH
        shares = 0
        
        # Method 1: Direct from info
        shares = info.get('sharesOutstanding', 0)
        
        # Method 2: Implied shares outstanding
        if shares == 0 or shares is None:
            shares = info.get('impliedSharesOutstanding', 0)
        
        # Method 3: From balance sheet (Total Common Stock / Par Value)
        if (shares == 0 or shares is None) and not balance_sheet.empty:
            try:
                # Try to get from Common Stock or Share Capital
                for row_name in ['Common Stock', 'Share Capital', 'Ordinary Shares Capital', 'Common Stock Equity']:
                    if row_name in balance_sheet.index:
                        common_stock = balance_sheet.loc[row_name].iloc[0]
                        # Typically par value is ‚Çπ1, ‚Çπ2, ‚Çπ5, or ‚Çπ10
                        # Try different par values to estimate
                        for par_value in [1, 2, 5, 10]:
                            estimated_shares = abs(common_stock) / par_value
                            # Sanity check: shares should be reasonable (between 1M and 100B)
                            if 1_000_000 < estimated_shares < 100_000_000_000:
                                shares = estimated_shares
                                break
                        if shares > 0:
                            break
            except:
                pass
        
        # Method 4: Calculate from market cap and price
        if (shares == 0 or shares is None) and 'marketCap' in info and 'currentPrice' in info:
            market_cap = info.get('marketCap', 0)
            current_price = info.get('currentPrice', 0)
            if market_cap > 0 and current_price > 0:
                shares = market_cap / current_price
        
        # Method 5: From enterprise value and price
        if (shares == 0 or shares is None) and 'enterpriseValue' in info and 'currentPrice' in info:
            ev = info.get('enterpriseValue', 0)
            price = info.get('currentPrice', 0)
            if ev > 0 and price > 0:
                # Rough estimate assuming EV ‚âà Market Cap for many companies
                shares = ev / price
        
        # Final check
        if shares is None:
            shares = 0
        
        shares_source = "Unknown"
        if shares > 0:
            if info.get('sharesOutstanding', 0) > 0:
                shares_source = "Direct (sharesOutstanding)"
            elif info.get('impliedSharesOutstanding', 0) > 0:
                shares_source = "Implied shares"
            elif 'marketCap' in info and 'currentPrice' in info:
                shares_source = "Calculated from Market Cap"
            else:
                shares_source = "Estimated from Balance Sheet"
        
        return {
            'income_statement': income_stmt,
            'balance_sheet': balance_sheet,
            'cash_flow': cash_flow,
            'info': info,
            'shares': shares,
            'shares_source': shares_source  # Track how shares were obtained
        }, None
        
    except Exception as e:
        return None, f"Error fetching data: {str(e)}"

def extract_financials_listed(yahoo_data, num_years=3):
    """
    Extract financial metrics from Yahoo Finance OR Screener.in data
    
    Args:
        yahoo_data: Dictionary containing either Yahoo Finance data OR Screener.in data
        num_years: Number of historical years to extract (default 3)
    
    Returns:
        Dictionary with financial metrics in ‚Çπ Lacs
    """
    
    # CHECK IF THIS IS SCREENER.IN DATA
    if '_screener_financials' in yahoo_data:
        # SCREENER.IN DATA PATH
        st.info("üìä Using Screener.in financial data")
        financials = yahoo_data['_screener_financials']
        
        # BUGFIX: Validate that all required fields exist and have data
        required_fields = [
            'years', 'revenue', 'cogs', 'opex', 'ebitda', 'depreciation',
            'ebit', 'interest', 'tax', 'nopat', 'fixed_assets', 'inventory',
            'receivables', 'payables', 'cash', 'equity', 'st_debt', 'lt_debt'
        ]
        
        missing_fields = []
        empty_fields = []
        for field in required_fields:
            if field not in financials:
                missing_fields.append(field)
            elif isinstance(financials[field], list) and len(financials[field]) == 0:
                empty_fields.append(field)
            elif isinstance(financials[field], list) and all(v == 0 for v in financials[field]):
                empty_fields.append(field + " (all zeros)")
        
        if missing_fields:
            st.error(f"‚ùå Missing financial fields from Screener.in: {', '.join(missing_fields)}")
            st.info("üí° Try using Yahoo Finance data source instead")
        
        if empty_fields:
            st.warning(f"‚ö†Ô∏è Empty or zero financial fields from Screener.in: {', '.join(empty_fields)}")
            st.info("üí° This may be due to incomplete data on Screener.in for this stock")
            
            # DEBUG: Show actual values for troubleshooting
            with st.expander("üîç Debug: View Raw Screener Data"):
                debug_data = {
                    'Fixed Assets': financials.get('fixed_assets', []),
                    'Inventory': financials.get('inventory', []),
                    'Receivables': financials.get('receivables', []),
                    'Payables': financials.get('payables', []),
                    'Equity': financials.get('equity', []),
                    'ST Debt': financials.get('st_debt', []),
                    'LT Debt': financials.get('lt_debt', [])
                }
                st.json(debug_data)
        
        # Screener data is already in the correct format (‚Çπ Lacs)
        # Just ensure we only return the requested number of years
        if num_years < len(financials['years']):
            # Truncate to requested years (newest first)
            for key in financials.keys():
                if isinstance(financials[key], list):
                    financials[key] = financials[key][:num_years]
        
        return financials
    
    # YAHOO FINANCE DATA PATH (ORIGINAL CODE)
    try:
        income_stmt = yahoo_data['income_statement']
        balance_sheet = yahoo_data['balance_sheet']
        cash_flow = yahoo_data['cash_flow']
        
        # Get last N years (columns are sorted newest to oldest)
        years = income_stmt.columns[:min(num_years, len(income_stmt.columns))]
        
        opex_methods_used = []  # Track which method was used for each year
        
        financials = {
            'years': [str(y.year) for y in years],
            'revenue': [],
            'cogs': [],
            'opex': [],
            'ebitda': [],
            'depreciation': [],
            'ebit': [],
            'interest': [],
            'interest_income': [],  # Added for business classification
            'tax': [],
            'nopat': [],
            'fixed_assets': [],
            'inventory': [],
            'receivables': [],
            'payables': [],
            'cash': [],
            'equity': [],
            'st_debt': [],
            'lt_debt': [],
        }
        
        for year in years:
            # Income Statement - Values are already in the correct currency
            # ROBUST: Use safe_extract to handle None values
            revenue = safe_extract(income_stmt, 'Total Revenue', year, default=0)
            cogs = safe_extract(income_stmt, 'Cost Of Revenue', year, default=0)
            
            # Try to get Operating Expenses directly from various fields
            opex = 0
            opex_method = "None"
            
            # Method 1: Try direct operating expense fields
            opex_fields = [
                'Operating Expense',
                'Total Operating Expenses',
                'Operating Expenses',
                'Selling General And Administration',
                'Selling General Administrative',
                'SG&A Expense'
            ]
            
            for field in opex_fields:
                if field in income_stmt.index:
                    opex = abs(income_stmt.loc[field, year])
                    opex_method = f"Method 1: Direct field '{field}'"
                    break
            
            # Method 2: If not found, try to calculate from Gross Profit - Operating Income
            if opex == 0:
                gross_profit = 0
                if 'Gross Profit' in income_stmt.index:
                    gross_profit = abs(income_stmt.loc['Gross Profit', year])
                elif revenue > 0 and cogs > 0:
                    gross_profit = revenue - cogs
                
                operating_income = abs(income_stmt.loc['Operating Income', year]) if 'Operating Income' in income_stmt.index else 0
                
                if gross_profit > 0 and operating_income > 0:
                    opex = gross_profit - operating_income
                    opex_method = "Method 2: Gross Profit - Operating Income"
            
            # Method 3: If still not found, try SG&A + R&D + Other
            if opex == 0:
                sga = abs(income_stmt.loc['Selling General And Administration', year]) if 'Selling General And Administration' in income_stmt.index else 0
                rd = abs(income_stmt.loc['Research And Development', year]) if 'Research And Development' in income_stmt.index else 0
                other_opex = abs(income_stmt.loc['Other Operating Expenses', year]) if 'Other Operating Expenses' in income_stmt.index else 0
                opex = sga + rd + other_opex
                if opex > 0:
                    opex_method = f"Method 3: SG&A({sga/100000:.2f}) + R&D({rd/100000:.2f}) + Other({other_opex/100000:.2f})"
            
            # Get EBITDA - ROBUST: Handle None values
            ebitda = (
                safe_extract(income_stmt, 'EBITDA', year) or
                safe_extract(income_stmt, 'Normalized EBITDA', year) or
                max(0, revenue - cogs - opex)  # Calculate if not available
            )
            ebitda = ensure_valid_number(ebitda, default=0)
            
            # Get depreciation separately for projections - ROBUST
            depreciation = (
                safe_extract(income_stmt, 'Reconciled Depreciation', year) or
                safe_extract(cash_flow, 'Depreciation And Amortization', year) or
                safe_extract(income_stmt, 'Depreciation', year) or
                0
            )
            
            # Fallback: Calculate from Operating Income vs EBITDA or use revenue-based estimate
            if depreciation == 0:
                operating_income = safe_extract(income_stmt, 'Operating Income', year, default=0)
                if ebitda > operating_income and operating_income > 0:
                    depreciation = ebitda - operating_income
                else:
                    depreciation = revenue * 0.02 if revenue > 0 else 0  # 2% of revenue fallback
            
            depreciation = ensure_valid_number(depreciation, default=0)
            
            # Final safety check: If opex is still 0 or unreasonable, derive from EBITDA
            if opex == 0 or opex < 0:
                opex = revenue - cogs - ebitda
                opex_method = "Method 4 (Fallback): Revenue - COGS - EBITDA"
                if opex < 0:
                    opex = revenue * 0.15  # Assume 15% of revenue as default
                    opex_method = "Method 5 (Default): 15% of Revenue"
            
            # EBIT
            ebit = ebitda - depreciation
            
            # Interest Expense - ROBUST: Handle None values
            interest = (
                safe_extract(income_stmt, 'Interest Expense', year) or
                safe_extract(income_stmt, 'Interest Expense Non Operating', year) or
                0
            )
            
            # For banks: check if Net Interest Income is negative (then it's an expense)
            if interest == 0 and 'Net Interest Income' in income_stmt.index:
                net_int = safe_extract(income_stmt, 'Net Interest Income', year, default=0)
                if net_int < 0:
                    interest = abs(net_int)
            
            interest = ensure_valid_number(interest, default=0)
            
            # Interest Income (for business classification) - ROBUST
            interest_income = (
                safe_extract(income_stmt, 'Interest Income', year) or
                safe_extract(income_stmt, 'Interest And Dividend Income', year) or
                0
            )
            
            # For banks: Net Interest Income is primary revenue if positive
            if interest_income == 0 and 'Net Interest Income' in income_stmt.index:
                net_int = safe_extract(income_stmt, 'Net Interest Income', year, default=0)
                if net_int > 0:
                    interest_income = abs(net_int)
            
            interest_income = ensure_valid_number(interest_income, default=0)
            
            # Tax - ROBUST
            tax = safe_extract(income_stmt, 'Tax Provision', year, default=0)
            tax = ensure_valid_number(tax, default=0)
            
            # NOPAT (using 25% tax as default) - ROBUST calculation
            ebt = ebit - interest
            if ebt > 0 and tax > 0:
                tax_rate_effective = safe_divide(tax, ebt, default=0.25)
                tax_rate_effective = min(max(tax_rate_effective, 0), 0.35)  # Clamp between 0 and 35%
            else:
                tax_rate_effective = 0.25
            
            nopat = ebit * (1 - tax_rate_effective)
            nopat = ensure_valid_number(nopat, default=0)
            
            # Balance Sheet - Values are already in the correct currency
            # ROBUST: Use safe_extract for all balance sheet items to handle None values
            total_assets = safe_extract(balance_sheet, 'Total Assets', year, default=0)
            
            # Fixed Assets
            fixed_assets = (
                safe_extract(balance_sheet, 'Net PPE', year) or
                safe_extract(balance_sheet, 'Gross PPE', year) or
                safe_extract(balance_sheet, 'Properties', year) or
                (total_assets * 0.3 if total_assets > 0 else 0)  # Fallback: 30% of total assets
            )
            
            # Current Assets - ROBUST: These can be None for many companies
            inventory = safe_extract(balance_sheet, 'Inventory', year, default=0)
            
            receivables = (
                safe_extract(balance_sheet, 'Receivables', year) or
                safe_extract(balance_sheet, 'Accounts Receivable', year) or
                safe_extract(balance_sheet, 'Gross Accounts Receivable', year) or
                0
            )
            
            cash = (
                safe_extract(balance_sheet, 'Cash And Cash Equivalents', year) or
                safe_extract(balance_sheet, 'Cash Cash Equivalents And Short Term Investments', year) or
                0
            )
            
            # Liabilities - ROBUST: These can also be None
            payables = (
                safe_extract(balance_sheet, 'Payables', year) or
                safe_extract(balance_sheet, 'Accounts Payable', year) or
                safe_extract(balance_sheet, 'Payables And Accrued Expenses', year) or
                0
            )
            
            # Debt - ROBUST: Handle None debt values
            st_debt = (
                safe_extract(balance_sheet, 'Current Debt', year) or
                safe_extract(balance_sheet, 'Current Debt And Capital Lease Obligation', year) or
                0
            )
            
            lt_debt = (
                safe_extract(balance_sheet, 'Long Term Debt', year) or
                safe_extract(balance_sheet, 'Long Term Debt And Capital Lease Obligation', year) or
                0
            )
            
            # Equity - ROBUST: Critical for WACC calculation
            equity = (
                safe_extract(balance_sheet, 'Stockholders Equity', year) or
                safe_extract(balance_sheet, 'Total Equity Gross Minority Interest', year) or
                safe_extract(balance_sheet, 'Common Stock Equity', year) or
                0
            )
            
            # Convert to Lacs (divide by 100,000)
            # ROBUST: Ensure all values are valid numbers before storage
            # Yahoo Finance data is in actual currency (Rupees for Indian stocks)
            financials['revenue'].append(ensure_valid_number(revenue / 100000, 0))
            financials['cogs'].append(ensure_valid_number(cogs / 100000, 0))
            financials['opex'].append(ensure_valid_number(opex / 100000, 0))
            financials['ebitda'].append(ensure_valid_number(ebitda / 100000, 0))
            financials['depreciation'].append(ensure_valid_number(depreciation / 100000, 0))
            financials['ebit'].append(ensure_valid_number(ebit / 100000, 0))
            financials['interest'].append(ensure_valid_number(interest / 100000, 0))
            financials['interest_income'].append(ensure_valid_number(interest_income / 100000, 0))
            financials['tax'].append(ensure_valid_number(tax / 100000, 0))
            financials['nopat'].append(ensure_valid_number(nopat / 100000, 0))
            
            financials['fixed_assets'].append(ensure_valid_number(fixed_assets / 100000, 0))
            financials['inventory'].append(ensure_valid_number(inventory / 100000, 0))
            financials['receivables'].append(ensure_valid_number(receivables / 100000, 0))
            financials['payables'].append(ensure_valid_number(payables / 100000, 0))
            financials['cash'].append(ensure_valid_number(cash / 100000, 0))
            financials['equity'].append(ensure_valid_number(equity / 100000, 0))
            financials['st_debt'].append(ensure_valid_number(st_debt / 100000, 0))
            financials['lt_debt'].append(ensure_valid_number(lt_debt / 100000, 0))
            
            # Track which method was used for opex
            opex_methods_used.append(f"{year.year}: {opex_method}")
        
        
        return financials
        
    except Exception as e:
        st.error(f"Error extracting financials: {str(e)}")
        import traceback
        st.error(traceback.format_exc())
        return None

def get_stock_beta(ticker, market_ticker=None, period_years=3):
    """Calculate beta using regression of stock returns vs market returns
    
    Automatically determines market index:
    - NSE (.NS) -> NIFTY (^NSEI)
    - BSE (.BO) -> SENSEX (^BSESN)
    """
    try:
        # Ensure ticker has exchange suffix
        if '.NS' not in ticker and '.BO' not in ticker:
            # No suffix, assume NSE
            ticker = ticker + '.NS'
        
        # Determine market index based on ticker suffix
        if market_ticker is None:
            if '.NS' in ticker:
                market_ticker = '^NSEI'  # NIFTY 50 for NSE
            elif '.BO' in ticker:
                market_ticker = '^BSESN'  # SENSEX for BSE
            else:
                market_ticker = '^BSESN'  # Default to SENSEX
        
        end_date = datetime.now()
        start_date = end_date - timedelta(days=period_years*365)
        
        # Download stock data - ticker now has suffix
        stock = yf.download(ticker, start=start_date, end=end_date, progress=False)
        market = yf.download(market_ticker, start=start_date, end=end_date, progress=False)
        
        if stock.empty or market.empty:
            st.warning(f"‚ö†Ô∏è No data for {ticker} - using default Œ≤=1.0")
            return 1.0
        
        # Calculate returns
        stock_returns = stock['Close'].pct_change().dropna()
        market_returns = market['Close'].pct_change().dropna()
        
        # Align data
        aligned = pd.concat([stock_returns, market_returns], axis=1, join='inner')
        aligned.columns = ['stock', 'market']
        aligned = aligned.dropna()
        
        if len(aligned) < 2:
            st.warning(f"‚ö†Ô∏è Insufficient data for {ticker} - using default Œ≤=1.0")
            return 1.0
        
        # Calculate beta using covariance method
        covariance = aligned['stock'].cov(aligned['market'])
        market_variance = aligned['market'].var()
        
        if market_variance == 0:
            return 1.0
        
        beta = covariance / market_variance
        
        # Display which index was used
        index_name = "NIFTY 50" if market_ticker == '^NSEI' else "SENSEX"
        st.caption(f"   (vs {index_name})")
        
        # Clamp between 0.1 and 3.0
        return max(0.1, min(beta, 3.0))
        
    except Exception as e:
        st.warning(f"Could not calculate beta for {ticker}: {str(e)}")
        return 1.0

def get_risk_free_rate(custom_ticker=None):
    """
    Get risk-free rate by calculating historical CAGR from Yahoo Finance ticker data.
    
    For bond indices/yields: Uses the closing value directly if it's already a percentage
    For stocks/indices: Calculates CAGR from price history
    
    Returns:
        tuple: (rate, debug_messages_list)
    """
    from datetime import datetime, timedelta
    import yfinance as yf
    import pandas as pd
    
    ticker = custom_ticker if custom_ticker else '^TNX'
    debug = []
    
    debug.append(f"üîç **DEBUG**: Starting RF rate fetch for ticker: `{ticker}`")
    
    try:
        # Fetch maximum available historical data
        end_date = datetime.now()
        start_date = end_date - timedelta(days=365*20)  # Try to get 20 years
        
        debug.append(f"üìÖ **DEBUG**: Requesting data from {start_date.date()} to {end_date.date()}")
        
        # Use yf.Ticker with period='max' to force maximum historical data
        ticker_obj = yf.Ticker(ticker)
        
        # Try multiple methods to get data
        debug.append(f"üîÑ **ATTEMPT 1**: Using history(period='max')...")
        gsec_data = ticker_obj.history(period='max')
        
        if len(gsec_data) < 2:
            debug.append(f"   ‚ö†Ô∏è Only got {len(gsec_data)} rows, trying history(start=...)...")
            gsec_data = ticker_obj.history(start=start_date, end=end_date)
        
        if len(gsec_data) < 2:
            debug.append(f"   ‚ö†Ô∏è Still only {len(gsec_data)} rows, trying download()...")
            # Fallback to yf.download
            gsec_data = yf.download(ticker, start=start_date, end=end_date, progress=False)
            # Convert multi-level columns if needed
            if isinstance(gsec_data.columns, pd.MultiIndex):
                gsec_data.columns = gsec_data.columns.get_level_values(0)
        
        # If still no data and ticker looks like Indian G-Sec, try NSEpy
        if len(gsec_data) < 2 and ('NIFTY' in ticker.upper() or 'GSEC' in ticker.upper() or 'GS' in ticker.upper()):
            debug.append(f"   ‚ö†Ô∏è Still only {len(gsec_data)} rows, trying NSEpy for Indian G-Sec...")
            try:
                from nsepy import get_history
                from datetime import date
                
                # NSEpy requires specific index names
                # Try common G-Sec index names
                index_names = ['NIFTY GS 10YR', 'NIFTY GS 10YR IDX', 'NIFTYGS10YR']
                
                for index_name in index_names:
                    try:
                        debug.append(f"   üîÑ Trying NSEpy with index: '{index_name}'...")
                        nsepy_data = get_history(
                            symbol=index_name,
                            start=start_date.date(),
                            end=end_date.date(),
                            index=True
                        )
                        
                        if not nsepy_data.empty and len(nsepy_data) > 1:
                            debug.append(f"   ‚úÖ NSEpy SUCCESS with '{index_name}': {len(nsepy_data)} rows")
                            gsec_data = nsepy_data
                            break
                    except Exception as e:
                        debug.append(f"   ‚ö†Ô∏è NSEpy failed for '{index_name}': {str(e)[:50]}")
                        continue
                        
            except ImportError:
                debug.append(f"   ‚ö†Ô∏è NSEpy not available (install with: pip install nsepy)")
            except Exception as e:
                debug.append(f"   ‚ö†Ô∏è NSEpy error: {str(e)[:100]}")
        
        debug.append(f"üìä **DEBUG**: Downloaded {len(gsec_data)} rows of data")
        
        # Show actual date range and data
        if len(gsec_data) > 0:
            debug.append(f"üìã **DEBUG**: Date range: {gsec_data.index[0].date()} to {gsec_data.index[-1].date()}")
            debug.append(f"üìã **DEBUG**: Columns available: {list(gsec_data.columns)}")
            debug.append(f"üìã **DEBUG**: First close: {gsec_data['Close'].iloc[0]:.2f}")
            debug.append(f"üìã **DEBUG**: Last close: {gsec_data['Close'].iloc[-1]:.2f}")
        
        if gsec_data.empty:
            debug.append(f"‚ùå **DEBUG**: No data returned for ticker {ticker}")
            debug.append("‚ö†Ô∏è This ticker may not exist or may not have historical data")
            debug.append("üí° **TRY THESE WORKING TICKERS:**")
            debug.append("   - ^TNX (US 10-Year Treasury Yield)")
            debug.append("   - ^IRX (US 13-Week T-Bill)")
            debug.append("   - RELIANCE.NS (will calculate CAGR)")
            fallback = 6.83
            debug.append(f"‚ö†Ô∏è Using fallback: {fallback}%")
            return fallback, debug
        
        if len(gsec_data) < 2:
            debug.append(f"‚ö†Ô∏è **DEBUG**: Only {len(gsec_data)} rows returned (need at least 2)")
            debug.append(f"üí° **SUGGESTION**: Try ticker.history() with different parameters")
            fallback = 6.83
            debug.append(f"‚ö†Ô∏è Using fallback: {fallback}%")
            return fallback, debug
        
        # Extract close prices
        close_prices = gsec_data['Close'].dropna()
        
        if len(close_prices) < 2:
            debug.append(f"‚ö†Ô∏è **DEBUG**: Insufficient valid close prices: {len(close_prices)}")
            fallback = 6.83
            debug.append(f"‚ö†Ô∏è Using fallback: {fallback}%")
            return fallback, debug
        
        # Get first and last prices
        first_price = float(close_prices.iloc[0])
        last_price = float(close_prices.iloc[-1])
        
        # Calculate number of years
        first_date = close_prices.index[0]
        last_date = close_prices.index[-1]
        days_diff = (last_date - first_date).days
        years = days_diff / 365.25
        
        debug.append(f"üìä **DEBUG**: Price change: {first_price:.2f} ‚Üí {last_price:.2f}")
        debug.append(f"üìä **DEBUG**: Time period: {years:.2f} years ({days_diff} days)")
        
        # CRITICAL LOGIC: Determine if this is a yield/rate or a price
        # If values are typically < 100 and relatively stable, it's likely a yield already
        # If values are > 100 or show significant growth, calculate CAGR
        
        avg_price = close_prices.mean()
        price_volatility = close_prices.std() / avg_price if avg_price > 0 else 0
        
        debug.append(f"üìä **DEBUG**: Average value: {avg_price:.2f}")
        debug.append(f"üìä **DEBUG**: Volatility (std/mean): {price_volatility:.2%}")
        
        # Decision logic
        if avg_price < 50 and price_volatility < 0.5:
            # Likely already a yield/rate (e.g., ^TNX returns 4.5 for 4.5%)
            debug.append(f"üí° **INTERPRETATION**: Values appear to be yields/rates (avg={avg_price:.2f} < 50)")
            
            # Use recent average (last 90 days or all available)
            days_to_use = min(90, len(close_prices))
            recent_prices = close_prices.iloc[-days_to_use:]
            avg_rate = recent_prices.mean()
            latest_rate = last_price
            
            debug.append(f"üìä **RESULT**: Using average of last {days_to_use} days")
            debug.append(f"üìä **RESULT**: Avg: {avg_rate:.2f}% | Latest: {latest_rate:.2f}%")
            
            return round(avg_rate, 2), debug
            
        else:
            # Calculate CAGR (for stocks, indices, etc.)
            debug.append(f"üí° **INTERPRETATION**: Values appear to be prices (avg={avg_price:.2f}), calculating CAGR")
            
            if first_price <= 0 or years <= 0:
                debug.append(f"‚ùå **ERROR**: Cannot calculate CAGR (first_price={first_price}, years={years})")
                fallback = 6.83
                debug.append(f"‚ö†Ô∏è Using fallback: {fallback}%")
                return fallback, debug
            
            # CAGR formula: ((End/Start)^(1/Years) - 1) * 100
            cagr = ((last_price / first_price) ** (1 / years) - 1) * 100
            
            debug.append(f"üìä **CAGR CALCULATION**:")
            debug.append(f"   - Formula: ({last_price:.2f}/{first_price:.2f})^(1/{years:.2f}) - 1")
            debug.append(f"   - Result: {cagr:.2f}%")
            
            # Sanity check for CAGR
            if cagr < -50 or cagr > 200:
                debug.append(f"‚ö†Ô∏è **WARNING**: CAGR {cagr:.2f}% seems extreme")
                debug.append(f"   This might indicate data quality issues")
            
            debug.append(f"‚úÖ **SUCCESS**: Using {cagr:.2f}% as historical return rate")
            
            return round(cagr, 2), debug
        
    except Exception as e:
        debug.append(f"‚ùå **DEBUG**: Exception occurred: {type(e).__name__}")
        debug.append(f"‚ùå **DEBUG**: Error message: {str(e)}")
        import traceback
        debug.append(f"‚ùå **DEBUG**: Traceback:\n```\n{traceback.format_exc()}\n```")
        fallback = 6.83
        debug.append(f"‚ö†Ô∏è Using fallback: {fallback}%")
        return fallback, debug

def get_market_return():
    """Calculate market return from Sensex historical data"""
    try:
        end_date = datetime.now()
        start_date = datetime.now() - timedelta(days=20*365)  # 20 years for better data
        
        sensex = yf.download('^BSESN', start=start_date, end=end_date, progress=False)
        
        if not sensex.empty and len(sensex) > 252:  # At least 1 year of data
            # Calculate CAGR
            start_price = float(sensex['Close'].iloc[0])
            end_price = float(sensex['Close'].iloc[-1])
            num_years = len(sensex) / 252  # 252 trading days per year
            
            if start_price > 0 and num_years > 0:
                cagr = ((end_price / start_price) ** (1 / num_years) - 1) * 100
                st.info(f"üìä Sensex CAGR (last {num_years:.1f} years): {cagr:.2f}%")
                return max(8.0, min(cagr, 150.0))  # Clamp between 8% and 150%
    except Exception as e:
        st.warning(f"Could not fetch Sensex data: {str(e)}")
    
    # Fallback
    st.warning("‚ö†Ô∏è Using fallback market return of 12%")
    return 12.0

# ================================
# ADVANCED CHARTING FUNCTIONS
# ================================

def create_waterfall_chart(valuation):
    """Create waterfall chart showing DCF value buildup"""
    fig = go.Figure(go.Waterfall(
        name = "DCF Waterfall",
        orientation = "v",
        measure = ["relative", "relative", "total"],
        x = ["PV of Projected FCFF", "PV of Terminal Value", "Enterprise Value"],
        textposition = "outside",
        text = [f"‚Çπ{valuation['sum_pv_fcff']:.2f}L", 
                f"‚Çπ{valuation['pv_terminal_value']:.2f}L",
                f"‚Çπ{valuation['enterprise_value']:.2f}L"],
        y = [valuation['sum_pv_fcff'], valuation['pv_terminal_value'], 0],
        connector = {"line":{"color":"rgb(63, 63, 63)"}},
    ))
    
    fig.update_layout(
        title = "DCF Valuation Waterfall",
        showlegend = False,
        height = 500,
        yaxis_title="Value (‚Çπ Lacs)"
    )
    
    return fig

def create_fcff_projection_chart(projections):
    """Create detailed FCFF projection chart with components"""
    years = [f"Year {y}" for y in projections['year']]
    
    fig = make_subplots(
        rows=2, cols=2,
        subplot_titles=('FCFF Projection', 'Revenue & EBITDA', 'Working Capital Changes', 'Capex & Depreciation'),
        specs=[[{"secondary_y": False}, {"secondary_y": False}],
               [{"secondary_y": False}, {"secondary_y": False}]]
    )
    
    # FCFF Projection
    fig.add_trace(
        go.Bar(name='FCFF', x=years, y=projections['fcff'], marker_color='#2E86AB'),
        row=1, col=1
    )
    
    # Revenue & EBITDA
    fig.add_trace(
        go.Scatter(name='Revenue', x=years, y=projections['revenue'], mode='lines+markers', line=dict(color='#06A77D', width=3)),
        row=1, col=2
    )
    fig.add_trace(
        go.Scatter(name='EBITDA', x=years, y=projections['ebitda'], mode='lines+markers', line=dict(color='#F77F00', width=3)),
        row=1, col=2
    )
    
    # Working Capital
    fig.add_trace(
        go.Bar(name='Œî WC', x=years, y=projections['delta_wc'], marker_color='#D62828'),
        row=2, col=1
    )
    
    # Capex & Depreciation
    fig.add_trace(
        go.Bar(name='Capex', x=years, y=projections['capex'], marker_color='#F77F00'),
        row=2, col=2
    )
    fig.add_trace(
        go.Bar(name='Depreciation', x=years, y=projections['depreciation'], marker_color='#06A77D'),
        row=2, col=2
    )
    
    fig.update_layout(height=800, showlegend=True, title_text="Comprehensive Financial Projections")
    fig.update_xaxes(title_text="Year", row=2, col=1)
    fig.update_xaxes(title_text="Year", row=2, col=2)
    fig.update_yaxes(title_text="‚Çπ Lacs", row=1, col=1)
    fig.update_yaxes(title_text="‚Çπ Lacs", row=1, col=2)
    fig.update_yaxes(title_text="‚Çπ Lacs", row=2, col=1)
    fig.update_yaxes(title_text="‚Çπ Lacs", row=2, col=2)
    
    return fig

def create_sensitivity_heatmap(projections, wacc_range, g_range, num_shares):
    """Create sensitivity analysis heatmap"""
    last_fcff = projections['fcff'][-1]
    n = len(projections['fcff'])
    
    # Create matrix
    matrix = []
    for w in wacc_range:
        row = []
        for g in g_range:
            if g >= w - 0.1:
                row.append(None)
            else:
                try:
                    fcff_n1 = last_fcff * (1 + g/100)
                    tv = fcff_n1 / ((w/100) - (g/100))
                    pv_tv = tv / ((1 + w/100) ** n)
                    
                    # Calculate sum_pv_fcff (approximate from first calc)
                    sum_pv_fcff = sum([projections['fcff'][i] / ((1 + w/100) ** (i+1)) for i in range(len(projections['fcff']))])
                    
                    ev = sum_pv_fcff + pv_tv
                    eq_val = ev * 100000
                    fv = eq_val / num_shares if num_shares > 0 else 0
                    row.append(fv)
                except:
                    row.append(None)
        matrix.append(row)
    
    fig = go.Figure(data=go.Heatmap(
        z=matrix,
        x=[f"{g:.1f}%" for g in g_range],
        y=[f"{w:.1f}%" for w in wacc_range],
        colorscale='RdYlGn',
        text=[[f"‚Çπ{val:.1f}" if val else "N/A" for val in row] for row in matrix],
        texttemplate="%{text}",
        textfont={"size":10},
        colorbar=dict(title="Fair Value ‚Çπ")
    ))
    
    fig.update_layout(
        title='Sensitivity Analysis: Fair Value per Share',
        xaxis_title='Terminal Growth Rate (g)',
        yaxis_title='WACC',
        height=600
    )
    
    return fig

def create_historical_financials_chart(financials, reverse_years=False):
    """
    Create comprehensive historical financials overview
    
    Args:
        financials: Financial data dict
        reverse_years: If True, reverse the years for chronological display (used in Screener mode)
    """
    years = financials['years']
    
    # Reverse years and all data arrays for chronological display if needed
    if reverse_years:
        years = list(reversed(years))
        # Create reversed version of financials dict for plotting
        financials_plot = {}
        for key, value in financials.items():
            if isinstance(value, list) and key != 'years':
                financials_plot[key] = list(reversed(value))
            else:
                financials_plot[key] = value
        financials_plot['years'] = years
    else:
        financials_plot = financials
    
    fig = make_subplots(
        rows=3, cols=2,
        subplot_titles=('Revenue & EBITDA Trend', 'Profitability Margins', 
                       'Balance Sheet Health', 'Cash Flow Quality',
                       'Working Capital Efficiency', 'Leverage Ratios'),
        specs=[[{"secondary_y": True}, {"secondary_y": False}],
               [{"secondary_y": True}, {"secondary_y": False}],
               [{"secondary_y": False}, {"secondary_y": False}]]
    )
    
    # Revenue & EBITDA
    fig.add_trace(
        go.Bar(name='Revenue', x=years, y=financials_plot['revenue'], marker_color='#06A77D'),
        row=1, col=1, secondary_y=False
    )
    fig.add_trace(
        go.Scatter(name='EBITDA Margin %', x=years, 
                  y=[(financials_plot['ebitda'][i]/financials_plot['revenue'][i]*100) if financials_plot['revenue'][i] > 0 else 0 
                     for i in range(len(years))],
                  mode='lines+markers', line=dict(color='#F77F00', width=3), marker=dict(size=10)),
        row=1, col=1, secondary_y=True
    )
    
    # Profitability Margins
    ebitda_margins = [(financials_plot['ebitda'][i]/financials_plot['revenue'][i]*100) if financials_plot['revenue'][i] > 0 else 0 
                      for i in range(len(years))]
    ebit_margins = [(financials_plot['ebit'][i]/financials_plot['revenue'][i]*100) if financials_plot['revenue'][i] > 0 else 0 
                    for i in range(len(years))]
    
    fig.add_trace(
        go.Scatter(name='EBITDA Margin', x=years, y=ebitda_margins, mode='lines+markers', line=dict(width=3)),
        row=1, col=2
    )
    fig.add_trace(
        go.Scatter(name='EBIT Margin', x=years, y=ebit_margins, mode='lines+markers', line=dict(width=3)),
        row=1, col=2
    )
    
    # Balance Sheet
    fig.add_trace(
        go.Bar(name='Equity', x=years, y=financials_plot['equity'], marker_color='#06A77D'),
        row=2, col=1, secondary_y=False
    )
    total_debt = [financials_plot['st_debt'][i] + financials_plot['lt_debt'][i] for i in range(len(years))]
    fig.add_trace(
        go.Bar(name='Debt', x=years, y=total_debt, marker_color='#D62828'),
        row=2, col=1, secondary_y=False
    )
    fig.add_trace(
        go.Scatter(name='Debt/Equity', x=years,
                  y=[(total_debt[i]/financials_plot['equity'][i]) if financials_plot['equity'][i] > 0 else 0 
                     for i in range(len(years))],
                  mode='lines+markers', line=dict(color='#2E86AB', width=3)),
        row=2, col=1, secondary_y=True
    )
    
    # Cash Flow Quality (NOPAT vs EBIT)
    fig.add_trace(
        go.Bar(name='EBIT', x=years, y=financials_plot['ebit'], marker_color='#F77F00'),
        row=2, col=2
    )
    fig.add_trace(
        go.Bar(name='NOPAT', x=years, y=financials_plot['nopat'], marker_color='#06A77D'),
        row=2, col=2
    )
    
    # Working Capital Components
    fig.add_trace(
        go.Bar(name='Inventory', x=years, y=financials_plot['inventory'], marker_color='#2E86AB'),
        row=3, col=1
    )
    fig.add_trace(
        go.Bar(name='Receivables', x=years, y=financials_plot['receivables'], marker_color='#06A77D'),
        row=3, col=1
    )
    fig.add_trace(
        go.Bar(name='Payables', x=years, y=financials_plot['payables'], marker_color='#D62828'),
        row=3, col=1
    )
    
    # Leverage Ratios
    debt_to_ebitda = [(total_debt[i]/financials_plot['ebitda'][i]) if financials_plot['ebitda'][i] > 0 else 0 
                      for i in range(len(years))]
    interest_coverage = [(financials_plot['ebit'][i]/financials_plot['interest'][i]) if financials_plot['interest'][i] > 0 else 0 
                         for i in range(len(years))]
    
    fig.add_trace(
        go.Scatter(name='Debt/EBITDA', x=years, y=debt_to_ebitda, mode='lines+markers', line=dict(width=3)),
        row=3, col=2
    )
    fig.add_trace(
        go.Scatter(name='Interest Coverage', x=years, y=interest_coverage, mode='lines+markers', line=dict(width=3)),
        row=3, col=2
    )
    
    fig.update_layout(height=1200, showlegend=True, title_text="Historical Financial Analysis Dashboard")
    fig.update_yaxes(title_text="‚Çπ Lacs", row=1, col=1, secondary_y=False)
    fig.update_yaxes(title_text="Margin %", row=1, col=1, secondary_y=True)
    fig.update_yaxes(title_text="Margin %", row=1, col=2)
    fig.update_yaxes(title_text="‚Çπ Lacs", row=2, col=1, secondary_y=False)
    fig.update_yaxes(title_text="Ratio", row=2, col=1, secondary_y=True)
    fig.update_yaxes(title_text="‚Çπ Lacs", row=2, col=2)
    fig.update_yaxes(title_text="‚Çπ Lacs", row=3, col=1)
    fig.update_yaxes(title_text="Ratio", row=3, col=2)
    
    return fig

def create_wacc_breakdown_chart(wacc_details):
    """Create visual breakdown of WACC components"""
    labels = ['Cost of Equity (Ke)', 'After-tax Cost of Debt (Kd)']
    values = [wacc_details['ke'], wacc_details['kd_after_tax']]
    weights = [wacc_details['we'], wacc_details['wd']]
    contributions = [wacc_details['ke'] * wacc_details['we'] / 100, 
                    wacc_details['kd_after_tax'] * wacc_details['wd'] / 100]
    
    fig = make_subplots(
        rows=1, cols=2,
        specs=[[{"type": "pie"}, {"type": "bar"}]],
        subplot_titles=('Capital Structure Weights', 'WACC Components Contribution')
    )
    
    # Capital structure pie
    fig.add_trace(
        go.Pie(labels=['Equity', 'Debt'], values=[wacc_details['we'], wacc_details['wd']],
               marker_colors=['#06A77D', '#D62828']),
        row=1, col=1
    )
    
    # WACC contribution bar
    fig.add_trace(
        go.Bar(name='Contribution to WACC', x=labels, y=contributions,
               marker_color=['#06A77D', '#D62828'],
               text=[f"{c:.2f}%" for c in contributions],
               textposition='auto'),
        row=1, col=2
    )
    
    fig.update_layout(height=400, showlegend=True, title_text=f"WACC Breakdown (Total: {wacc_details['wacc']:.2f}%)")
    
    return fig

def create_bank_valuation_comparison_chart(valuations_dict):
    """Create comparison chart for multiple bank valuation methods"""
    methods = []
    values = []
    colors = []
    
    color_map = {
        'Residual Income Model': '#2E86AB',
        'Dividend Discount Model': '#06A77D',
        'P/B with ROE Analysis': '#F77F00',
        'Relative Valuation (P/E)': '#D62828',
        'Relative Valuation (P/B)': '#9D4EDD'
    }
    
    for method, val_data in valuations_dict.items():
        if val_data and 'value_per_share' in val_data:
            methods.append(method)
            values.append(val_data['value_per_share'])
            colors.append(color_map.get(method, '#888888'))
    
    fig = go.Figure()
    
    fig.add_trace(go.Bar(
        x=methods,
        y=values,
        marker_color=colors,
        text=[f"‚Çπ{v:.2f}" for v in values],
        textposition='auto',
    ))
    
    if values:
        avg_value = np.mean(values)
        fig.add_hline(y=avg_value, line_dash="dash", line_color="red",
                     annotation_text=f"Average: ‚Çπ{avg_value:.2f}",
                     annotation_position="right")
    
    fig.update_layout(
        title="Bank Valuation Methods Comparison",
        xaxis_title="Valuation Method",
        yaxis_title="Fair Value per Share (‚Çπ)",
        height=500,
        showlegend=False
    )
    
    return fig

def create_price_vs_value_gauge(current_price, fair_value):
    """Create gauge chart showing current price vs fair value"""
    # Check for invalid fair value (zero, negative, or unrealistic)
    if fair_value <= 0:
        return None
    
    ratio = (current_price / fair_value) * 100
    
    fig = go.Figure(go.Indicator(
        mode = "gauge+number+delta",
        value = current_price,
        domain = {'x': [0, 1], 'y': [0, 1]},
        title = {'text': f"Current Price vs Fair Value (‚Çπ{fair_value:.2f})", 'font': {'size': 20}},
        delta = {'reference': fair_value, 'valueformat': '.2f'},
        gauge = {
            'axis': {'range': [None, max(current_price, fair_value) * 1.5], 'tickformat': '‚Çπ.2f'},
            'bar': {'color': "#2E86AB"},
            'bgcolor': "white",
            'borderwidth': 2,
            'bordercolor': "gray",
            'steps': [
                {'range': [0, fair_value * 0.8], 'color': '#06A77D'},
                {'range': [fair_value * 0.8, fair_value * 1.2], 'color': '#F4D35E'},
                {'range': [fair_value * 1.2, fair_value * 2], 'color': '#D62828'}],
            'threshold': {
                'line': {'color': "red", 'width': 4},
                'thickness': 0.75,
                'value': fair_value}}))
    
    fig.update_layout(height=400)
    
    if ratio < 80:
        recommendation = "üü¢ UNDERVALUED - Potential Buy"
    elif ratio > 120:
        recommendation = "üî¥ OVERVALUED - Potential Sell"
    else:
        recommendation = "üü° FAIRLY VALUED - Hold"
    
    fig.add_annotation(
        text=recommendation,
        xref="paper", yref="paper",
        x=0.5, y=-0.1,
        showarrow=False,
        font=dict(size=16, color="black", family="Arial Black")
    )
    
    return fig

# ================================
# DCF CALCULATION FUNCTIONS
# ================================

def extract_financials_unlisted(df_bs, df_pl, year_cols):
    """Extract financial metrics from Excel DataFrames"""
    num_years = min(3, len(year_cols))
    last_years = year_cols[-num_years:]
    
    financials = {
        'years': last_years,
        'revenue': [],
        'cogs': [],
        'opex': [],
        'ebitda': [],
        'depreciation': [],
        'ebit': [],
        'interest': [],
        'interest_income': [],  # Added for business classification
        'tax': [],
        'nopat': [],
        'fixed_assets': [],
        'inventory': [],
        'receivables': [],
        'payables': [],
        'cash': [],
        'equity': [],
        'st_debt': [],
        'lt_debt': [],
    }
    
    for year_col in last_years:
        # Income Statement
        revenue = get_value_from_df(df_pl, 'Net Revenue', year_col)
        cogs = get_value_from_df(df_pl, 'Cost of Materials', year_col)
        employee_exp = get_value_from_df(df_pl, 'Employee Benefit', year_col)
        other_exp = get_value_from_df(df_pl, 'Other Expenses', year_col)
        depreciation = get_value_from_df(df_pl, 'Depreciation', year_col)
        interest = get_value_from_df(df_pl, 'Finance Costs', year_col)
        interest_income = get_value_from_df(df_pl, 'Finance Income', year_col)  # For classification
        tax = get_value_from_df(df_pl, 'Income Tax', year_col)
        
        opex = employee_exp + other_exp
        ebitda = revenue - opex - cogs
        ebit = ebitda - depreciation
        pbt = ebit - interest
        pat = pbt - tax
        nopat = ebit * (1 - 0.25)  # Assuming 25% tax
        
        financials['revenue'].append(revenue)
        financials['cogs'].append(cogs)
        financials['opex'].append(opex)
        financials['ebitda'].append(ebitda)
        financials['depreciation'].append(depreciation)
        financials['ebit'].append(ebit)
        financials['interest'].append(interest)
        financials['interest_income'].append(interest_income)  # Store for classification
        financials['tax'].append(tax)
        financials['nopat'].append(nopat)
        
        # Balance Sheet
        fixed_assets = get_value_from_df(df_bs, 'Tangible Assets', year_col)
        inventory = get_value_from_df(df_bs, 'Inventories', year_col)
        receivables = get_value_from_df(df_bs, 'Trade Receivables', year_col)
        payables = get_value_from_df(df_bs, 'Trade Payables', year_col)
        cash = get_value_from_df(df_bs, 'Cash and Bank', year_col)
        equity = get_value_from_df(df_bs, 'Total Equity', year_col)
        st_debt = get_value_from_df(df_bs, 'Short Term Borrowings', year_col)
        lt_debt = get_value_from_df(df_bs, 'Long Term Borrowings', year_col)
        
        financials['fixed_assets'].append(fixed_assets)
        financials['inventory'].append(inventory)
        financials['receivables'].append(receivables)
        financials['payables'].append(payables)
        financials['cash'].append(cash)
        financials['equity'].append(equity)
        financials['st_debt'].append(st_debt)
        financials['lt_debt'].append(lt_debt)
    
    return financials

def fetch_screener_peer_data(ticker_symbol):
    """
    Fetch peer company data from Screener.in for comparative valuation
    
    Args:
        ticker_symbol: NSE/BSE ticker (without .NS/.BO suffix)
    
    Returns:
        dict with peer financial data or None
    """
    try:
        import time
        import random
        
        # Clean ticker
        ticker_clean = ticker_symbol.replace('.NS', '').replace('.BO', '')
        
        # Add delay for respectful scraping
        time.sleep(random.uniform(1.5, 3.0))
        
        screener_data = fetch_screener_financials(ticker_clean, num_years=3)
        
        if not screener_data or not screener_data.get('financials'):
            return None
        
        financials = screener_data['financials']
        shares = screener_data.get('shares', 0)
        
        # Convert to format expected by perform_comparative_valuation
        # Values are already in Lacs, need to convert to rupees for compatibility
        revenue = ensure_valid_number(financials['revenue'][0] * 100000, 0)  # Most recent year
        ebitda = ensure_valid_number(financials['ebitda'][0] * 100000, 0)
        net_income = ensure_valid_number(financials['nopat'][0] * 100000, 0)
        total_debt = ensure_valid_number((financials['st_debt'][0] + financials['lt_debt'][0]) * 100000, 0)
        cash = ensure_valid_number(financials['cash'][0] * 100000, 0)
        equity = ensure_valid_number(financials['equity'][0] * 100000, 0)
        
        eps = safe_divide(net_income, shares, default=0)
        book_value = safe_divide(equity, shares, default=0)
        
        return {
            'ticker': ticker_clean,
            'name': screener_data.get('company_name', ticker_clean),
            'price': 0,  # Screener doesn't provide live price
            'shares': shares,
            'market_cap': 0,  # Cannot calculate without price
            'revenue': revenue,
            'ebitda': ebitda,
            'net_income': net_income,
            'eps': eps,
            'book_value': book_value,
            'total_debt': total_debt,
            'cash': cash,
            'enterprise_value': 0,  # Cannot calculate without market cap
            '_source': 'screener'
        }
    
    except Exception as e:
        print(f"Error fetching Screener data for {ticker_symbol}: {e}")
        return None


def perform_comparative_valuation(target_ticker, comp_tickers_str, target_financials=None, target_shares=None, exchange_suffix="NS", projections=None, use_screener_peers=False):
    """
    Perform comparative valuation using peer multiples
    
    Args:
        target_ticker: Target company ticker
        comp_tickers_str: Comma-separated peer tickers
        target_financials: Target company financials dict
        target_shares: Target company shares outstanding
        exchange_suffix: NS or BO
        projections: DCF projections dict with 'nopat' key
        use_screener_peers: If True, fetch peer data from Screener.in instead of Yahoo Finance
    """
    try:
        comp_tickers = [t.strip() for t in comp_tickers_str.split(',') if t.strip()]
        
        if not comp_tickers:
            return None
        
        results = {
            'target': {},
            'comparables': [],
            'multiples_stats': {},
            'valuations': {},
            '_peer_source': 'screener' if use_screener_peers else 'yahoo'
        }
        
        # Get target company data
        if target_ticker:
            # Listed company
            target_stock = get_cached_ticker(get_ticker_with_exchange(target_ticker, exchange_suffix))
            target_info = target_stock.info
            target_financials_yf = target_stock.financials
            target_bs = target_stock.balance_sheet
            
            results['target'] = {
                'name': target_info.get('longName', target_ticker),
                # Robust price fetching - try multiple methods
                'current_price': target_info.get('currentPrice', 0) or target_info.get('regularMarketPrice', 0) or 0,
                'shares': target_info.get('sharesOutstanding', 0),
                'market_cap': target_info.get('marketCap', 0),
                'enterprise_value': target_info.get('enterpriseValue', 0),
                'revenue': safe_extract(target_financials_yf, 'Total Revenue', target_financials_yf.columns[0]) if 'Total Revenue' in target_financials_yf.index else 0,
                'ebitda': target_info.get('ebitda', 0),
                'net_income': safe_extract(target_financials_yf, 'Net Income', target_financials_yf.columns[0]) if 'Net Income' in target_financials_yf.index else 0,
                'book_value_per_share': target_info.get('bookValue', 0),
                'total_debt': safe_extract(target_bs, 'Long Term Debt', target_bs.columns[0]) if 'Long Term Debt' in target_bs.index else 0,
                'cash': safe_extract(target_bs, 'Cash And Cash Equivalents', target_bs.columns[0]) if 'Cash And Cash Equivalents' in target_bs.index else 0,
            }
            
            # Calculate EPS and Book Value - ALWAYS set to avoid KeyError
            if results['target']['shares'] > 0 and results['target']['net_income'] != 0:
                results['target']['eps'] = results['target']['net_income'] / results['target']['shares']
            else:
                results['target']['eps'] = 0
                
        else:
            # Unlisted company - arrays have NEWEST first, so [0] = latest
            results['target'] = {
                'name': 'Target Company (Unlisted)',
                'current_price': 0,
                'shares': target_shares,
                'market_cap': 0,
                'enterprise_value': 0,
                'revenue': target_financials['revenue'][0] * 100000,  # [0] = latest year
                'ebitda': target_financials['ebitda'][0] * 100000,
                'net_income': target_financials['nopat'][0] * 100000,  # Using NOPAT as proxy
                'book_value_per_share': 0,
                'total_debt': (target_financials['st_debt'][0] + target_financials['lt_debt'][0]) * 100000,
                'cash': target_financials['cash'][0] * 100000,
                'eps': (target_financials['nopat'][0] * 100000) / target_shares if target_shares > 0 else 0,
            }
        
        # Get comparable companies data
        comp_data = []
        
        if use_screener_peers:
            # FETCH PEER DATA FROM SCREENER.IN
            st.info(f"üåê Fetching peer data from Screener.in for {len(comp_tickers)} companies...")
            
            for idx, ticker in enumerate(comp_tickers):
                try:
                    peer_data = fetch_screener_peer_data(ticker)
                    
                    if peer_data:
                        # Calculate multiples (most will be 0 since no price/market cap)
                        pe = 0  # Cannot calculate without price
                        pb = 0  # Cannot calculate without price
                        ps = 0  # Cannot calculate without market cap
                        
                        # Can calculate EV-based multiples if we have enterprise value
                        enterprise_value = peer_data['enterprise_value']
                        ev_ebitda = safe_divide(enterprise_value, peer_data['ebitda'], default=0)
                        ev_sales = safe_divide(enterprise_value, peer_data['revenue'], default=0)
                        
                        comp_data.append({
                            'ticker': peer_data['ticker'],
                            'name': peer_data['name'],
                            'price': peer_data['price'],
                            'market_cap': peer_data['market_cap'],
                            'revenue': peer_data['revenue'],
                            'ebitda': peer_data['ebitda'],
                            'net_income': peer_data['net_income'],
                            'eps': peer_data['eps'],
                            'book_value': peer_data['book_value'],
                            'pe': pe,
                            'pb': pb,
                            'ps': ps,
                            'ev_ebitda': ev_ebitda,
                            'ev_sales': ev_sales,
                            'enterprise_value': enterprise_value,
                            'shares': peer_data['shares']
                        })
                        
                        st.success(f"‚úÖ {peer_data['name']}")
                    else:
                        st.warning(f"‚ö†Ô∏è Could not fetch data for {ticker}")
                
                except Exception as e:
                    st.warning(f"‚ö†Ô∏è Error fetching {ticker}: {str(e)}")
                    continue
            
            if not comp_data:
                st.warning("‚ö†Ô∏è Could not fetch any peer data from Screener.in. Try Yahoo Finance instead.")
                return None
        
        else:
            # FETCH PEER DATA FROM YAHOO FINANCE (ORIGINAL CODE)
            for idx, ticker in enumerate(comp_tickers):
                try:
                    # Add delay between requests to avoid rate limiting
                    if idx > 0:
                        time.sleep(random.uniform(1.0, 1.5))
                    
                    # Ticker already has suffix (.NS or .BO) from UI combination
                    comp_stock = get_cached_ticker(ticker)
                    comp_info = comp_stock.info
                    comp_financials_yf = comp_stock.financials
                    comp_bs = comp_stock.balance_sheet
                    
                    # Extract financial data
                    shares = comp_info.get('sharesOutstanding', 0)
                    # Robust price fetching - try multiple methods
                    price = comp_info.get('currentPrice', 0)
                    if not price or price == 0:
                        price = comp_info.get('regularMarketPrice', 0)
                    if not price or price == 0:
                        try:
                            hist = comp_stock.history(period='1d')
                            if not hist.empty:
                                price = hist['Close'].iloc[-1]
                        except:
                            pass
                    market_cap = comp_info.get('marketCap', 0)
                    
                    revenue = safe_extract(comp_financials_yf, 'Total Revenue', comp_financials_yf.columns[0]) if 'Total Revenue' in comp_financials_yf.index and not comp_financials_yf.empty else 0
                    ebitda = comp_info.get('ebitda', 0)
                    net_income = safe_extract(comp_financials_yf, 'Net Income', comp_financials_yf.columns[0]) if 'Net Income' in comp_financials_yf.index and not comp_financials_yf.empty else 0
                    
                    total_debt = safe_extract(comp_bs, 'Long Term Debt', comp_bs.columns[0]) if 'Long Term Debt' in comp_bs.index and not comp_bs.empty else 0
                    cash = safe_extract(comp_bs, 'Cash And Cash Equivalents', comp_bs.columns[0]) if 'Cash And Cash Equivalents' in comp_bs.index and not comp_bs.empty else 0
                    
                    book_value = comp_info.get('bookValue', 0)
                    eps = net_income / shares if shares > 0 else 0
                    
                    # Calculate multiples
                    pe = price / eps if eps > 0 else 0
                    pb = price / book_value if book_value > 0 else 0
                    ps = market_cap / revenue if revenue > 0 else 0
                    
                    enterprise_value = market_cap + total_debt - cash
                    ev_ebitda = enterprise_value / ebitda if ebitda > 0 else 0
                    ev_sales = enterprise_value / revenue if revenue > 0 else 0
                    
                    comp_data.append({
                        'ticker': ticker,
                        'name': comp_info.get('longName', ticker),
                        'price': price,
                        'market_cap': market_cap,
                        'revenue': revenue,
                        'ebitda': ebitda,
                        'net_income': net_income,
                        'eps': eps,
                        'book_value': book_value,
                        'pe': pe,
                        'pb': pb,
                        'ps': ps,
                        'ev_ebitda': ev_ebitda,
                        'ev_sales': ev_sales,
                        'enterprise_value': enterprise_value,
                        'shares': shares
                    })
                    
                except Exception as e:
                    st.warning(f"Could not fetch data for {ticker}: {str(e)}")
                    continue
        
        results['comparables'] = comp_data
        
        if not comp_data:
            st.error("No comparable company data could be fetched")
            return None
        
        # Calculate statistics for each multiple
        multiples = ['pe', 'pb', 'ps', 'ev_ebitda', 'ev_sales']
        
        for multiple in multiples:
            valid_values = [c[multiple] for c in comp_data if c.get(multiple, 0) > 0]
            
            if not valid_values:
                continue
            
            results['multiples_stats'][multiple] = {
                'average': np.mean(valid_values),
                'median': np.median(valid_values),
                'min': np.min(valid_values),
                'max': np.max(valid_values),
                'std': np.std(valid_values),
                'values': valid_values
            }
        
        # Calculate implied valuations
        target = results['target']
        valuations_summary = {}
        
        # P/E Method
        if 'pe' in results['multiples_stats'] and target['eps'] > 0:
            stats = results['multiples_stats']['pe']
            
            fair_value_avg = target['eps'] * stats['average']
            fair_value_median = target['eps'] * stats['median']
            
            valuations_summary['pe'] = {
                'method': 'Price-to-Earnings (P/E)',
                'target_metric': target['eps'],
                'metric_name': 'EPS',
                'avg_multiple': stats['average'],
                'median_multiple': stats['median'],
                'fair_value_avg': fair_value_avg,
                'fair_value_median': fair_value_median,
                'current_price': target['current_price'],
                'upside_avg': ((fair_value_avg - target['current_price']) / target['current_price'] * 100) if target['current_price'] else 0,
                'upside_median': ((fair_value_median - target['current_price']) / target['current_price'] * 100) if target['current_price'] else 0,
                'formula_avg': f"EPS √ó Avg P/E = ‚Çπ{target['eps']:.2f} √ó {stats['average']:.2f} = ‚Çπ{fair_value_avg:.2f}",
                'formula_median': f"EPS √ó Median P/E = ‚Çπ{target['eps']:.2f} √ó {stats['median']:.2f} = ‚Çπ{fair_value_median:.2f}"
            }
        
        # P/B Method
        if 'pb' in results['multiples_stats'] and target['book_value_per_share'] > 0:
            stats = results['multiples_stats']['pb']
            
            fair_value_avg = target['book_value_per_share'] * stats['average']
            fair_value_median = target['book_value_per_share'] * stats['median']
            
            valuations_summary['pb'] = {
                'method': 'Price-to-Book (P/B)',
                'target_metric': target['book_value_per_share'],
                'metric_name': 'Book Value per Share',
                'avg_multiple': stats['average'],
                'median_multiple': stats['median'],
                'fair_value_avg': fair_value_avg,
                'fair_value_median': fair_value_median,
                'current_price': target['current_price'],
                'upside_avg': ((fair_value_avg - target['current_price']) / target['current_price'] * 100) if target['current_price'] else 0,
                'upside_median': ((fair_value_median - target['current_price']) / target['current_price'] * 100) if target['current_price'] else 0,
                'formula_avg': f"BVPS √ó Avg P/B = ‚Çπ{target['book_value_per_share']:.2f} √ó {stats['average']:.2f} = ‚Çπ{fair_value_avg:.2f}",
                'formula_median': f"BVPS √ó Median P/B = ‚Çπ{target['book_value_per_share']:.2f} √ó {stats['median']:.2f} = ‚Çπ{fair_value_median:.2f}"
            }
        
        # P/S Method
        if 'ps' in results['multiples_stats'] and target['revenue'] > 0 and target['shares'] > 0:
            stats = results['multiples_stats']['ps']
            
            revenue_per_share = target['revenue'] / target['shares']
            fair_value_avg = revenue_per_share * stats['average']
            fair_value_median = revenue_per_share * stats['median']
            
            valuations_summary['ps'] = {
                'method': 'Price-to-Sales (P/S)',
                'target_metric': revenue_per_share,
                'metric_name': 'Revenue per Share',
                'avg_multiple': stats['average'],
                'median_multiple': stats['median'],
                'fair_value_avg': fair_value_avg,
                'fair_value_median': fair_value_median,
                'current_price': target['current_price'],
                'upside_avg': ((fair_value_avg - target['current_price']) / target['current_price'] * 100) if target['current_price'] else 0,
                'upside_median': ((fair_value_median - target['current_price']) / target['current_price'] * 100) if target['current_price'] else 0,
                'formula_avg': f"Revenue/Share √ó Avg P/S = ‚Çπ{revenue_per_share:.2f} √ó {stats['average']:.2f} = ‚Çπ{fair_value_avg:.2f}",
                'formula_median': f"Revenue/Share √ó Median P/S = ‚Çπ{revenue_per_share:.2f} √ó {stats['median']:.2f} = ‚Çπ{fair_value_median:.2f}"
            }
        
        # EV/EBITDA Method
        if 'ev_ebitda' in results['multiples_stats'] and target['ebitda'] > 0 and target['shares'] > 0:
            stats = results['multiples_stats']['ev_ebitda']
            
            implied_ev_avg = target['ebitda'] * stats['average']
            implied_ev_median = target['ebitda'] * stats['median']
            
            net_debt = target['total_debt'] - target['cash']
            
            equity_value_avg = implied_ev_avg - net_debt
            equity_value_median = implied_ev_median - net_debt
            
            fair_value_avg = equity_value_avg / target['shares']
            fair_value_median = equity_value_median / target['shares']
            
            valuations_summary['ev_ebitda'] = {
                'method': 'EV/EBITDA',
                'target_metric': target['ebitda'],
                'metric_name': 'EBITDA',
                'avg_multiple': stats['average'],
                'median_multiple': stats['median'],
                'implied_ev_avg': implied_ev_avg,
                'implied_ev_median': implied_ev_median,
                'total_debt': total_debt,
        'cash': cash,
        'net_debt': net_debt,
                'fair_value_avg': fair_value_avg,
                'fair_value_median': fair_value_median,
                'current_price': target['current_price'],
                'upside_avg': ((fair_value_avg - target['current_price']) / target['current_price'] * 100) if target['current_price'] else 0,
                'upside_median': ((fair_value_median - target['current_price']) / target['current_price'] * 100) if target['current_price'] else 0,
                'formula_avg': f"(EBITDA √ó Avg EV/EBITDA - Net Debt) / Shares = (‚Çπ{target['ebitda']/1e7:.2f}Cr √ó {stats['average']:.2f} - ‚Çπ{net_debt/1e7:.2f}Cr) / {target['shares']/1e7:.2f}Cr",
                'formula_median': f"(EBITDA √ó Median EV/EBITDA - Net Debt) / Shares = (‚Çπ{target['ebitda']/1e7:.2f}Cr √ó {stats['median']:.2f} - ‚Çπ{net_debt/1e7:.2f}Cr) / {target['shares']/1e7:.2f}Cr"
            }
        
        # EV/Sales Method
        if 'ev_sales' in results['multiples_stats'] and target['revenue'] > 0 and target['shares'] > 0:
            stats = results['multiples_stats']['ev_sales']
            
            implied_ev_avg = target['revenue'] * stats['average']
            implied_ev_median = target['revenue'] * stats['median']
            
            net_debt = target['total_debt'] - target['cash']
            
            equity_value_avg = implied_ev_avg - net_debt
            equity_value_median = implied_ev_median - net_debt
            
            fair_value_avg = equity_value_avg / target['shares']
            fair_value_median = equity_value_median / target['shares']
            
            valuations_summary['ev_sales'] = {
                'method': 'EV/Sales',
                'target_metric': target['revenue'],
                'metric_name': 'Revenue',
                'avg_multiple': stats['average'],
                'median_multiple': stats['median'],
                'implied_ev_avg': implied_ev_avg,
                'implied_ev_median': implied_ev_median,
                'net_debt': net_debt,
                'fair_value_avg': fair_value_avg,
                'fair_value_median': fair_value_median,
                'current_price': target['current_price'],
                'upside_avg': ((fair_value_avg - target['current_price']) / target['current_price'] * 100) if target['current_price'] else 0,
                'upside_median': ((fair_value_median - target['current_price']) / target['current_price'] * 100) if target['current_price'] else 0,
                'formula_avg': f"(Revenue √ó Avg EV/Sales - Net Debt) / Shares = (‚Çπ{target['revenue']/1e7:.2f}Cr √ó {stats['average']:.2f} - ‚Çπ{net_debt/1e7:.2f}Cr) / {target['shares']/1e7:.2f}Cr",
                'formula_median': f"(Revenue √ó Median EV/Sales - Net Debt) / Shares = (‚Çπ{target['revenue']/1e7:.2f}Cr √ó {stats['median']:.2f} - ‚Çπ{net_debt/1e7:.2f}Cr) / {target['shares']/1e7:.2f}Cr"
            }
        
        results['valuations'] = valuations_summary
        
        # Calculate 12-Month Forward P/E using EXISTING PROJECTED NOPAT (NO DUPLICATION!)
        # Forward P/E = Current Price / Future EPS
        # Future EPS = Projected Year 1 NOPAT / Current Outstanding Shares
        
        # Get current EPS for comparison
        current_eps = 0
        if target_financials and 'nopat' in target_financials:
            current_nopat = target_financials['nopat'][0] * 100000  # Most recent historical year
            current_eps = current_nopat / target['shares'] if target['shares'] > 0 else 0
        
        # Use EXISTING projections if available, otherwise calculate growth
        if projections and 'nopat' in projections and len(projections['nopat']) > 0:
            # ‚úÖ USE EXISTING PROJECTED NOPAT - NO DUPLICATION!
            projected_nopat_year1 = projections['nopat'][0] * 100000  # Year 1 projection (already calculated in DCF!)
            future_eps = projected_nopat_year1 / target['shares'] if target['shares'] > 0 else 0
            
            # Calculate growth rate for display purposes only
            if current_eps > 0 and future_eps > 0:
                growth_rate = ((future_eps / current_eps) - 1) * 100
            else:
                growth_rate = 0.0
            
            calculation_method = "Using DCF Year 1 Projected NOPAT"
            
        elif target_financials and 'nopat' in target_financials and len(target_financials['nopat']) >= 2:
            # Fallback: Calculate growth if projections not available
            recent_nopat = target_financials['nopat'][:min(3, len(target_financials['nopat']))]
            if len(recent_nopat) >= 2 and recent_nopat[-1] > 0:
                num_years = len(recent_nopat) - 1
                growth_rate = ((recent_nopat[0] / recent_nopat[-1]) ** (1/num_years) - 1) * 100
                growth_rate = max(-5, min(growth_rate, 25))
            else:
                growth_rate = 10.0
            
            future_eps = current_eps * (1 + growth_rate / 100)
            calculation_method = f"Using historical CAGR ({growth_rate:.1f}%)"
        else:
            future_eps = current_eps * 1.10  # Default 10% growth
            growth_rate = 10.0
            calculation_method = "Using default 10% growth"
        
        if 'pe' in results['multiples_stats'] and future_eps > 0:
            stats = results['multiples_stats']['pe']
            forward_fair_value_avg = future_eps * stats['average']
            forward_fair_value_median = future_eps * stats['median']
            
            results['forward_pe'] = {
                'current_eps': current_eps,
                'forward_eps': future_eps,
                'earnings_growth_rate': growth_rate,
                'fair_value_avg': forward_fair_value_avg,
                'fair_value_median': forward_fair_value_median,
                'formula_avg': f"Forward EPS √ó Avg P/E = ‚Çπ{future_eps:.2f} √ó {stats['average']:.2f} = ‚Çπ{forward_fair_value_avg:.2f}",
                'formula_median': f"Forward EPS √ó Median P/E = ‚Çπ{future_eps:.2f} √ó {stats['median']:.2f} = ‚Çπ{forward_fair_value_median:.2f}",
                'calculation_method': calculation_method,
                'calculation_note': f"12-Month Forward EPS: {calculation_method}",
                'using_dcf_projections': bool(projections)
            }
        
        return results
        
    except Exception as e:
        st.error(f"Comparative valuation error: {str(e)}")
        import traceback
        st.error(traceback.format_exc())
        return None

def calculate_working_capital_metrics(financials):
    """
    Calculate working capital days with ROBUST None/zero handling
    
    CRITICAL FIX: If inventory, receivables, or payables are None/zero for ALL years,
    exclude them from working capital calculations to prevent NaN cascade
    """
    wc_metrics = {
        'inventory_days': [],
        'debtor_days': [],
        'creditor_days': []
    }
    
    # ROBUST: Check if we have valid data for each WC component across ALL years
    has_valid_inventory = any(ensure_valid_number(financials['inventory'][i], 0) > 0 for i in range(len(financials['years'])))
    has_valid_receivables = any(ensure_valid_number(financials['receivables'][i], 0) > 0 for i in range(len(financials['years'])))
    has_valid_payables = any(ensure_valid_number(financials['payables'][i], 0) > 0 for i in range(len(financials['years'])))
    
    for i in range(len(financials['years'])):
        # ROBUST: Ensure all values are valid numbers
        revenue = ensure_valid_number(financials['revenue'][i], 0)
        cogs = ensure_valid_number(financials['cogs'][i], 0)
        inventory = ensure_valid_number(financials['inventory'][i], 0)
        receivables = ensure_valid_number(financials['receivables'][i], 0)
        payables = ensure_valid_number(financials['payables'][i], 0)
        
        # Calculate days only if we have valid denominators
        inv_days = safe_divide(inventory * 365, cogs, default=0) if has_valid_inventory and cogs > 0 else 0
        deb_days = safe_divide(receivables * 365, revenue, default=0) if has_valid_receivables and revenue > 0 else 0
        cred_days = safe_divide(payables * 365, cogs, default=0) if has_valid_payables and cogs > 0 else 0
        
        wc_metrics['inventory_days'].append(ensure_valid_number(inv_days, 0))
        wc_metrics['debtor_days'].append(ensure_valid_number(deb_days, 0))
        wc_metrics['creditor_days'].append(ensure_valid_number(cred_days, 0))
    
    # Average days - ROBUST: Only calculate if we have valid data
    if has_valid_inventory and wc_metrics['inventory_days']:
        valid_inv_days = [d for d in wc_metrics['inventory_days'] if d > 0]
        wc_metrics['avg_inv_days'] = np.mean(valid_inv_days) if valid_inv_days else 0
    else:
        wc_metrics['avg_inv_days'] = 0  # No inventory data available
    
    if has_valid_receivables and wc_metrics['debtor_days']:
        valid_deb_days = [d for d in wc_metrics['debtor_days'] if d > 0]
        wc_metrics['avg_deb_days'] = np.mean(valid_deb_days) if valid_deb_days else 0
    else:
        wc_metrics['avg_deb_days'] = 0  # No receivables data available
    
    if has_valid_payables and wc_metrics['creditor_days']:
        valid_cred_days = [d for d in wc_metrics['creditor_days'] if d > 0]
        wc_metrics['avg_cred_days'] = np.mean(valid_cred_days) if valid_cred_days else 0
    else:
        wc_metrics['avg_cred_days'] = 0  # No payables data available
    
    # Add flags to indicate which components are available
    wc_metrics['has_inventory'] = has_valid_inventory
    wc_metrics['has_receivables'] = has_valid_receivables
    wc_metrics['has_payables'] = has_valid_payables
    
    # Ensure all values are valid numbers (not NaN or inf)
    wc_metrics['avg_inv_days'] = ensure_valid_number(wc_metrics['avg_inv_days'], 0)
    wc_metrics['avg_deb_days'] = ensure_valid_number(wc_metrics['avg_deb_days'], 0)
    wc_metrics['avg_cred_days'] = ensure_valid_number(wc_metrics['avg_cred_days'], 0)
    
    return wc_metrics

def calculate_historical_capex_ratio(financials):
    """
    Calculate historical CapEx as % of Revenue for each year and return average
    This provides a more realistic CapEx projection method
    
    Returns:
        dict: {
            'capex_ratios': list of ratios for each year,
            'avg_capex_ratio': average ratio to use for projections
        }
    """
    revenues = financials['revenue']
    fixed_assets = financials['fixed_assets']
    
    capex_ratios = []
    
    # Calculate CapEx for each historical year
    # CapEx = Change in Fixed Assets + Depreciation
    for i in range(len(revenues) - 1):
        # Yahoo data is newest to oldest, so:
        # i=0 is newest year, i=1 is previous year
        newer_fa = fixed_assets[i]
        older_fa = fixed_assets[i + 1]
        revenue = revenues[i]
        
        if revenue > 0:
            # Estimate depreciation as ~5% of older FA (conservative estimate)
            estimated_dep = older_fa * 0.05
            capex = (newer_fa - older_fa) + estimated_dep
            
            # Calculate as % of revenue
            capex_ratio = (capex / revenue) * 100
            
            # Sanity check: Flag unusually high CapEx ratios but allow them
            if capex_ratio > 150.0:
                capex_ratio = 150.0  # Maximum cap at 150%
            elif capex_ratio > 50.0:
                # Just log high CapEx, don't cap it
                pass
            
            capex_ratios.append(capex_ratio)
    
    # Calculate average
    avg_capex_ratio = np.mean(capex_ratios) if capex_ratios else 5.0  # Default 5% if no data
    
    return {
        'capex_ratios': capex_ratios,
        'avg_capex_ratio': avg_capex_ratio
    }

def calculate_bank_fcfe_valuation(projections, ke, terminal_growth, num_shares):
    """
    Bank Valuation using FCFE (Free Cash Flow to Equity)
    
    CRITICAL DIFFERENCES FROM NORMAL DCF:
    - Uses Ke (Cost of Equity), NOT WACC
    - Discounts FCFE, NOT FCFF
    - Values EQUITY directly, NOT Enterprise Value
    - No debt adjustment (debt is raw material for banks)
    
    Formula:
    Equity Value = Œ£ PV(FCFE) + Terminal Value
    Value per Share = Equity Value / Shares
    
    Args:
        projections: Bank projections with 'fcfe' key
        ke: Cost of Equity % (NOT WACC!)
        terminal_growth: Terminal growth rate %
        num_shares: Outstanding shares
    """
    
    if 'fcfe' not in projections:
        return None, "‚ùå Bank projections must have 'fcfe' (Free Cash Flow to Equity)"
    
    fcfe_list = projections['fcfe']
    g = terminal_growth
    
    # Validation: Terminal growth must be < Ke
    if g >= ke:
        return None, f"‚ùå ERROR: Terminal growth ({g:.1f}%) must be less than Ke ({ke:.1f}%)"
    
    # =====================================================
    # STEP 1: Discount FCFE at Ke
    # =====================================================
    pv_fcfe = []
    for year, fcfe in enumerate(fcfe_list, start=1):
        fcfe_rupees = fcfe * 100000  # Convert from Lacs to Rupees
        pv = fcfe_rupees / ((1 + ke / 100) ** year)
        pv_fcfe.append(pv)
    
    sum_pv_fcfe = sum(pv_fcfe)
    
    # =====================================================
    # STEP 2: Terminal Value using Gordon Growth
    # =====================================================
    last_fcfe = fcfe_list[-1] * 100000  # Rupees
    fcfe_terminal = last_fcfe * (1 + g / 100)
    
    terminal_value = fcfe_terminal / ((ke / 100) - (g / 100))
    
    n = len(fcfe_list)
    pv_terminal = terminal_value / ((1 + ke / 100) ** n)
    
    # =====================================================
    # STEP 3: EQUITY VALUE (NOT Enterprise Value!)
    # =====================================================
    equity_value = sum_pv_fcfe + pv_terminal  # In Rupees
    
    # Convert to per share
    value_per_share = equity_value / num_shares if num_shares > 0 else 0
    
    # For banks, sustainable growth rate
    if 'roe' in projections:
        roe = projections.get('roe', 15)
        # Retention ratio = (PAT - FCFE) / PAT
        avg_pat = np.mean([projections['pat'][i] for i in range(len(projections['pat']))])
        avg_fcfe = np.mean(fcfe_list)
        retention_ratio = (avg_pat - avg_fcfe) / avg_pat if avg_pat > 0 else 0.5
        sustainable_growth = roe * retention_ratio
    else:
        sustainable_growth = g
    
    return {
        'valuation_method': 'Bank FCFE (Equity DCF)',
        'fair_value_per_share': value_per_share,
        'equity_value': equity_value / 100000,  # In Lacs
        'sum_pv_fcfe': sum_pv_fcfe / 100000,
        'terminal_value_pv': pv_terminal / 100000,
        'cost_of_equity': ke,
        'terminal_growth': g,
        'sustainable_growth': sustainable_growth,
        'projections': projections,
        'pv_fcfe_by_year': [pv / 100000 for pv in pv_fcfe]
    }, None

def project_financials_bank(financials, years, tax_rate, car_ratio=14.0, rwa_percentage=75.0):
    """
    Bank-Specific Financial Projection using FCFE (Free Cash Flow to Equity) Methodology
    
    CRITICAL: Banks CANNOT use FCFF/WACC. Must use FCFE/Ke.
    
    Why FCFE for banks:
    - Debt is raw material (not financing)
    - Interest expense is operating cost (not financing cost)
    - Capital Adequacy Ratio (CAR) constrains growth
    - Only equity capital matters
    
    FCFE Formula for Banks:
    FCFE = PAT - (Growth in Advances √ó RWA% √ó CAR)
    
    Where:
    - PAT = Net Profit After Tax
    - Growth in Advances = Loan book growth
    - RWA = Risk Weighted Assets ‚âà 70-80% of Advances
    - CAR = Capital Adequacy Ratio (13-15% in India)
    
    Args:
        financials: Historical financial data
        years: Projection period
        tax_rate: Tax rate %
        car_ratio: Capital Adequacy Ratio % (default 14%)
        rwa_percentage: Risk Weight % (default 75%)
    """
    
    st.info("üè¶ **Using Bank FCFE Methodology** (Not FCFF - that's wrong for banks!)")
    
    # =====================================================
    # STEP 1: Get Historical Data
    # =====================================================
    
    # Revenue = Total Revenue (Interest Income + Non-Interest Income)
    revenue_history = financials['revenue'] if 'revenue' in financials else []
    
    # Net Profit After Tax
    nopat_history = financials['nopat'] if 'nopat' in financials else []
    
    # Equity (Book Value)
    equity_history = financials['equity'] if 'equity' in financials else []
    
    if not revenue_history or not nopat_history or not equity_history:
        st.error("‚ùå Insufficient bank data for FCFE projection")
        return None, None
    
    # =====================================================
    # STEP 2: Calculate Historical Growth Rates
    # =====================================================
    
    # Revenue growth (proxy for advances growth in absence of loan book data)
    if len(revenue_history) >= 2 and revenue_history[-1] > 0 and revenue_history[0] > 0:
        num_years_hist = len(revenue_history) - 1
        revenue_growth = ((revenue_history[0] / revenue_history[-1]) ** (1/num_years_hist) - 1) * 100
        revenue_growth = max(5, min(revenue_growth, 150))  # Cap between 5-150% for banks
    else:
        revenue_growth = 12.0  # Default Indian bank growth
    
    # Historical ROE
    latest_equity = equity_history[0] * 100000  # Convert from Lacs to Rupees
    latest_pat = nopat_history[0] * 100000
    current_roe = (latest_pat / latest_equity * 100) if latest_equity > 0 else 15.0
    
    st.info(f"""
    üìä **Bank Metrics Calculated:**
    - Revenue Growth (Advances proxy): {revenue_growth:.2f}%
    - Current ROE: {current_roe:.2f}%
    - CAR Target: {car_ratio:.1f}%
    - RWA Weight: {rwa_percentage:.1f}%
    """)
    
    # =====================================================
    # STEP 3: Project PAT (Net Profit)
    # =====================================================
    
    # Use ROE √ó Book Value approach for PAT projection
    projected_pat = []
    projected_equity = []
    projected_revenue = []
    projected_fcfe = []
    
    current_revenue = revenue_history[0]
    current_equity_val = latest_equity
    
    for year in range(1, years + 1):
        # Project revenue (advances)
        current_revenue = current_revenue * (1 + revenue_growth / 100)
        projected_revenue.append(current_revenue / 100000)  # Back to Lacs
        
        # Project equity (retained earnings accumulate)
        # Equity grows with retained PAT
        projected_equity.append(current_equity_val / 100000)
        
        # Project PAT using ROE
        pat = current_equity_val * (current_roe / 100)
        projected_pat.append(pat / 100000)
        
        # =====================================================
        # STEP 4: Calculate FCFE (KEY FORMULA)
        # =====================================================
        # FCFE = PAT - Equity Required for Growth
        # Equity Required = Growth in Advances √ó RWA% √ó CAR%
        
        advances_growth_amount = current_revenue * (revenue_growth / 100)  # Rupees
        equity_required = advances_growth_amount * (rwa_percentage / 100) * (car_ratio / 100)
        
        fcfe = pat - equity_required
        projected_fcfe.append(fcfe / 100000)  # Back to Lacs
        
        # Update equity for next year (add retained earnings)
        # Retained = PAT - Dividends, approximate as PAT - FCFE
        retained = pat - fcfe
        current_equity_val += retained
    
    projections = {
        'year': list(range(1, years + 1)),
        'revenue': projected_revenue,  # Total revenue
        'pat': projected_pat,  # Net Profit After Tax
        'equity': projected_equity,  # Book Value
        'fcfe': projected_fcfe  # Free Cash Flow to Equity (THIS IS WHAT WE DISCOUNT!)
    }
    
    drivers = {
        'revenue_growth': revenue_growth,
        'roe': current_roe,
        'car_ratio': car_ratio,
        'rwa_percentage': rwa_percentage,
        'tax_rate': tax_rate
    }
    
    st.success(f"‚úÖ **Bank FCFE Projections Complete:** {years} years projected")
    st.info(f"üí° **FCFE** (not FCFF!) will be discounted at **Ke** (not WACC!) to get equity value directly")
    
    return projections, drivers

def project_financials(financials, wc_metrics, years, tax_rate, 
                      rev_growth_override, opex_margin_override, capex_ratio_override=None,
                      # NEW PARAMETERS - Complete user control
                      ebitda_margin_override=None,
                      depreciation_rate_override=None,
                      depreciation_method="Auto",
                      inventory_days_override=None,
                      debtor_days_override=None,
                      creditor_days_override=None,
                      interest_rate_override=None,
                      working_capital_pct_override=None):
    """
    INDUSTRY-GRADE Financial Projection Engine with FULL USER CONTROL
    ==================================================================
    
    Uses multiple validation layers and industry best practices:
    1. Revenue: CAGR with economic floor (user can override)
    2. Margins: Normalized with trend analysis (user can override)
    3. CapEx: % of Revenue (user can override)
    4. Working Capital: Days-based methodology (user can override each component)
    5. Depreciation: Multiple methods (user can select)
    6. Interest: Auto or user override
    7. Sanity checks at every step
    
    ALL parameters can be overridden by the user for complete control.
    """
    
    # ============================================
    # STEP 1: CALCULATE HISTORICAL CAPEX RATIO
    # ============================================
    capex_info = calculate_historical_capex_ratio(financials)
    avg_capex_ratio = capex_info['avg_capex_ratio']
    
    # Apply override if provided
    if capex_ratio_override:
        try:
            avg_capex_ratio = float(capex_ratio_override)
        except:
            pass  # Keep calculated value if override is invalid
    
    # ============================================
    # STEP 2: REVENUE GROWTH WITH INTELLIGENT FLOORS
    # ============================================
    revenues = financials['revenue']
    num_years = len(revenues) - 1
    
    if num_years > 0 and revenues[0] > 0 and revenues[-1] > 0:
        # CAGR: Start = OLDEST (last element), End = NEWEST (first element)
        start_revenue = revenues[-1]  # Oldest year
        end_revenue = revenues[0]     # Newest year
        historical_cagr = ((end_revenue / start_revenue) ** (1 / num_years) - 1) * 100
        
        # INDUSTRY PRACTICE: Apply GDP floor for low-growth companies
        # But also cap unrealistic high growth
        if historical_cagr < 4.0:
            # Blend historical with GDP floor (7% for India)
            avg_growth = (historical_cagr * 0.6) + (7.0 * 0.4)
        elif historical_cagr > 150.0:
            # Cap extremely excessive growth - maximum 150%
            avg_growth = 150.0
            try:
                st.warning(f"‚ö†Ô∏è Historical Revenue CAGR ({historical_cagr:.1f}%) capped at 150%. Original: {historical_cagr:.1f}%")
            except:
                pass
        else:
            avg_growth = historical_cagr
            if historical_cagr > 50.0:
                try:
                    st.info(f"üìä High Growth Detected: Revenue CAGR = {historical_cagr:.1f}% (using actual historical rate)")
                except:
                    pass
    else:
        avg_growth = 8.0  # Reasonable default for Indian economy
    
    if rev_growth_override:
        avg_growth = float(rev_growth_override)
    
    # ============================================
    # CRITICAL CAPEX NORMALIZATION RULE
    # ============================================
    # RULE: If CapEx/Revenue ratio > Revenue Growth rate,
    # then CapEx is consuming too much cash relative to growth
    # SOLUTION: Cap CapEx at 1/4 of revenue growth rate
    
    if avg_capex_ratio > avg_growth:
        original_capex_ratio = avg_capex_ratio
        avg_capex_ratio = avg_growth / 4.0
        
        # Log the adjustment for transparency
        import streamlit as st
        st.warning(f"‚ö†Ô∏è **CapEx Normalization Applied**")
        st.info(f"""
        üìä **Original CapEx/Revenue:** {original_capex_ratio:.2f}%
        üìà **Revenue Growth Rate:** {avg_growth:.2f}%
        
        **Issue Detected:** CapEx ratio ({original_capex_ratio:.2f}%) exceeds revenue growth ({avg_growth:.2f}%)
        
        **Action Taken:** CapEx ratio normalized to **{avg_capex_ratio:.2f}%** (1/4 of revenue growth)
        
        **Rationale:** Sustainable companies cannot indefinitely spend more on CapEx (as % of revenue) 
        than their revenue growth rate. This normalization ensures long-term financial viability.
        """)
    
    # ============================================
    # STEP 3: MARGIN ANALYSIS WITH NORMALIZATION
    # ============================================
    
    # COGS Margin - Use median to avoid outliers
    cogs_margins = []
    for i in range(len(revenues)):
        if financials['revenue'][i] > 0:
            margin = (financials['cogs'][i] / financials['revenue'][i]) * 100
            # Sanity check: COGS should be 20-85% of revenue
            if 20 <= margin <= 85:
                cogs_margins.append(margin)
    
    avg_cogs_margin = np.median(cogs_margins) if cogs_margins else 55.0
    
    # OpEx Margin - Use median and exclude outliers
    opex_margins = []
    for i in range(len(revenues)):
        if financials['revenue'][i] > 0:
            margin = (financials['opex'][i] / financials['revenue'][i]) * 100
            # Sanity check: OpEx should be 5-50% of revenue
            if 5 <= margin <= 50:
                opex_margins.append(margin)
    
    avg_opex_margin = np.median(opex_margins) if opex_margins else 15.0
    
    if opex_margin_override:
        avg_opex_margin = float(opex_margin_override)
    
    # ============================================
    # STEP 4: DEPRECIATION RATE - NORMALIZED
    # ============================================
    dep_rates = []
    for i in range(len(revenues)):
        if financials['fixed_assets'][i] > 0:
            rate = (financials['depreciation'][i] / financials['fixed_assets'][i]) * 100
            # Sanity: Depreciation typically 3-15% of FA
            if 3 <= rate <= 15:
                dep_rates.append(rate)
    
    avg_dep_rate = np.median(dep_rates) if dep_rates else 6.0
    
    # User override for depreciation rate
    if depreciation_rate_override:
        avg_dep_rate = float(depreciation_rate_override)
    
    # ============================================
    # STEP 5: INTEREST RATE ON DEBT
    # ============================================
    total_debts = [financials['st_debt'][i] + financials['lt_debt'][i] for i in range(len(revenues))]
    fin_cost_rates = []
    
    for i in range(len(revenues)):
        if total_debts[i] > 0 and financials['interest'][i] > 0:
            rate = (financials['interest'][i] / total_debts[i]) * 100
            # Sanity: Interest rate should be 4-18%
            if 4 <= rate <= 18:
                fin_cost_rates.append(rate)
    
    avg_fin_cost_rate = np.median(fin_cost_rates) if fin_cost_rates else 8.0
    
    # User override for interest rate
    if interest_rate_override:
        avg_fin_cost_rate = float(interest_rate_override)
    
    # ============================================
    # STEP 6: BALANCE SHEET GROWTH RATES
    # ============================================
    
    # Equity growth (for completeness)
    equity_values = financials['equity']
    if len(equity_values) > 1 and equity_values[-1] > 0 and equity_values[0] > 0:
        avg_equity_growth = ((equity_values[0] / equity_values[-1]) ** (1 / (len(equity_values) - 1)) - 1) * 100
        # Cap equity growth at revenue growth + 5%
        avg_equity_growth = min(avg_equity_growth, avg_growth + 5)
    else:
        avg_equity_growth = avg_growth
    
    # Debt growth - conservative assumption
    if len(total_debts) > 1 and total_debts[-1] > 0 and total_debts[0] > 0:
        historical_debt_growth = ((total_debts[0] / total_debts[-1]) ** (1 / (len(total_debts) - 1)) - 1) * 100
        # INDUSTRY PRACTICE: Debt shouldn't grow faster than revenue
        avg_debt_growth = min(historical_debt_growth, avg_growth * 0.8)
    else:
        avg_debt_growth = 0.0  # Conservative: assume no debt growth
    
    # ============================================
    # STEP 7: PROJECTIONS WITH ROBUST VALIDATION
    # ============================================
    projections = {
        'year': [],
        'revenue': [],
        'cogs': [],
        'opex': [],
        'ebitda': [],
        'depreciation': [],
        'ebit': [],
        'interest': [],
        'nopat': [],
        'fixed_assets': [],
        'equity': [],
        'debt': [],
        'wc': [],
        'delta_wc': [],
        'capex': [],
        'fcff': []
    }
    
    # Starting point - USE NEWEST YEAR DATA (index 0)
    last_revenue = revenues[0]
    last_fa = financials['fixed_assets'][0]
    last_equity = financials['equity'][0]
    last_debt = total_debts[0] if total_debts[0] > 0 else 0
    
    # CRITICAL FIX: Calculate initial working capital from most recent historical year
    # This ensures delta_wc calculations are based on actual historical WC, not zero
    last_inventory = ensure_valid_number(financials['inventory'][0], 0)
    last_receivables = ensure_valid_number(financials['receivables'][0], 0)
    last_payables = ensure_valid_number(financials['payables'][0], 0)
    last_wc = last_inventory + last_receivables - last_payables
    last_wc = ensure_valid_number(last_wc, 0)
    
    for year in range(1, years + 1):
        # ============================================
        # REVENUE PROJECTION
        # ============================================
        projected_revenue = last_revenue * (1 + avg_growth / 100)
        
        # ============================================
        # P&L PROJECTIONS
        # ============================================
        projected_cogs = projected_revenue * (avg_cogs_margin / 100)
        projected_opex = projected_revenue * (avg_opex_margin / 100)
        projected_ebitda = projected_revenue - projected_cogs - projected_opex
        
        # Sanity check: EBITDA should be positive for healthy companies
        if projected_ebitda < 0:
            # Adjust opex to maintain 5% EBITDA margin
            projected_opex = projected_revenue * 0.85 - projected_cogs
            projected_ebitda = projected_revenue * 0.15
        
        # ============================================
        # CAPEX PROJECTION (INDUSTRY METHOD)
        # ============================================
        # CapEx as % of Revenue (most reliable method)
        capex = projected_revenue * (avg_capex_ratio / 100)
        
        # Depreciation: Apply to growing FA base
        # FA will grow based on net CapEx
        projected_fa = last_fa + capex  # Will subtract depreciation next
        projected_dep = projected_fa * (avg_dep_rate / 100)
        
        # Adjust FA after depreciation
        projected_fa = projected_fa - projected_dep
        
        # ============================================
        # EBIT & NOPAT
        # ============================================
        projected_ebit = projected_ebitda - projected_dep
        projected_nopat = projected_ebit * (1 - tax_rate / 100)
        
        # ============================================
        # DEBT & INTEREST
        # ============================================
        projected_debt = last_debt * (1 + avg_debt_growth / 100) if last_debt > 0 else 0
        projected_interest = projected_debt * (avg_fin_cost_rate / 100) if projected_debt > 0 else 0
        
        # ============================================
        # WORKING CAPITAL (INDUSTRY STANDARD METHOD)
        # ROBUST: Handle cases where inventory/receivables/payables data is None
        # CRITICAL FIX: User overrides take priority over historical data
        # ============================================
        
        # Initialize WC components
        projected_inventory = 0
        projected_receivables = 0
        projected_payables = 0
        
        # Determine inventory days: USER OVERRIDE > Historical Data > 0
        if inventory_days_override and inventory_days_override > 0:
            inv_days_to_use = inventory_days_override
        elif wc_metrics.get('has_inventory', False) and wc_metrics['avg_inv_days'] > 0:
            inv_days_to_use = wc_metrics['avg_inv_days']
        else:
            inv_days_to_use = 0
        
        # Determine debtor days: USER OVERRIDE > Historical Data > 0
        if debtor_days_override and debtor_days_override > 0:
            deb_days_to_use = debtor_days_override
        elif wc_metrics.get('has_receivables', False) and wc_metrics['avg_deb_days'] > 0:
            deb_days_to_use = wc_metrics['avg_deb_days']
        else:
            deb_days_to_use = 0
        
        # Determine creditor days: USER OVERRIDE > Historical Data > 0
        if creditor_days_override and creditor_days_override > 0:
            cred_days_to_use = creditor_days_override
        elif wc_metrics.get('has_payables', False) and wc_metrics['avg_cred_days'] > 0:
            cred_days_to_use = wc_metrics['avg_cred_days']
        else:
            cred_days_to_use = 0
        
        # Calculate WC components using determined days
        if inv_days_to_use > 0:
            projected_inventory = safe_divide(projected_cogs * inv_days_to_use, 365, default=0)
        
        if deb_days_to_use > 0:
            projected_receivables = safe_divide(projected_revenue * deb_days_to_use, 365, default=0)
        
        if cred_days_to_use > 0:
            projected_payables = safe_divide(projected_cogs * cred_days_to_use, 365, default=0)
        
        # ROBUST: Ensure all WC components are valid numbers
        projected_inventory = ensure_valid_number(projected_inventory, 0)
        projected_receivables = ensure_valid_number(projected_receivables, 0)
        projected_payables = ensure_valid_number(projected_payables, 0)
        
        # Calculate projected WC
        projected_wc = projected_inventory + projected_receivables - projected_payables
        projected_wc = ensure_valid_number(projected_wc, 0)
        
        # Calculate change in WC
        delta_wc = projected_wc - last_wc
        delta_wc = ensure_valid_number(delta_wc, 0)
        
        # CRITICAL: Cap WC changes to prevent unrealistic swings
        # Industry practice: ŒîWC shouldn't exceed 20% of revenue
        max_delta_wc = projected_revenue * 0.20
        if abs(delta_wc) > max_delta_wc:
            delta_wc = max_delta_wc if delta_wc > 0 else -max_delta_wc
            projected_wc = last_wc + delta_wc
            projected_wc = ensure_valid_number(projected_wc, 0)
            delta_wc = ensure_valid_number(delta_wc, 0)
        
        # ============================================
        # FCFF CALCULATION
        # ROBUST: Ensure all components are valid numbers before calculation
        # ============================================
        projected_nopat = ensure_valid_number(projected_nopat, 0)
        projected_dep = ensure_valid_number(projected_dep, 0)
        delta_wc = ensure_valid_number(delta_wc, 0)
        capex = ensure_valid_number(capex, 0)
        
        fcff = projected_nopat + projected_dep - delta_wc - capex
        fcff = ensure_valid_number(fcff, 0)  # Ensure FCFF is never NaN
        
        # SANITY CHECK: Only intervene if FCFF is extremely negative AND components are likely invalid
        # Removed aggressive normalization that was incorrectly zeroing out legitimate working capital changes
        # Negative FCFF is acceptable for high-growth companies with working capital needs
        
        # ============================================
        # STORE PROJECTIONS
        # ============================================
        projections['year'].append(year)
        projections['revenue'].append(projected_revenue)
        projections['cogs'].append(projected_cogs)
        projections['opex'].append(projected_opex)
        projections['ebitda'].append(projected_ebitda)
        projections['depreciation'].append(projected_dep)
        projections['ebit'].append(projected_ebit)
        projections['interest'].append(projected_interest)
        projections['nopat'].append(projected_nopat)
        projections['fixed_assets'].append(projected_fa)
        projections['equity'].append(last_equity * (1 + avg_equity_growth / 100))
        projections['debt'].append(projected_debt)
        projections['wc'].append(projected_wc)
        projections['delta_wc'].append(delta_wc)
        projections['capex'].append(capex)
        projections['fcff'].append(fcff)
        
        # Update for next iteration
        last_revenue = projected_revenue
        last_fa = projected_fa
        last_equity = projections['equity'][-1]
        last_debt = projected_debt
        last_wc = projected_wc
    
    return projections, {
        'avg_growth': avg_growth,
        'avg_cogs_margin': avg_cogs_margin,
        'avg_opex_margin': avg_opex_margin,
        'avg_dep_rate': avg_dep_rate,
        'avg_fin_cost_rate': avg_fin_cost_rate,
        'avg_capex_ratio': avg_capex_ratio
    }

def calculate_wacc(financials, tax_rate, peer_tickers=None, manual_rf_rate=None):
    """Calculate WACC with proper beta calculation from peers"""
    # Cost of Equity (Ke)
    # ALWAYS use manual_rf_rate (passed from session state), never fetch
    rf = manual_rf_rate if manual_rf_rate is not None else 6.83  # Fallback to default, DON'T fetch
    rm = get_market_return()
    
    # Calculate beta from peer tickers
    beta = 1.0
    if peer_tickers and peer_tickers.strip():
        ticker_list = [t.strip() for t in peer_tickers.split(',') if t.strip()]
        betas = []
        
        for ticker in ticker_list:
            try:
                ticker_beta = get_stock_beta(ticker)
                if ticker_beta > 0:
                    betas.append(ticker_beta)
                    st.info(f"Beta for {ticker}: {ticker_beta:.3f}")
            except Exception as e:
                st.warning(f"Could not fetch beta for {ticker}: {str(e)}")
        
        if betas:
            beta = np.mean(betas)
            st.success(f"‚úÖ Average peer beta: {beta:.3f} (from {len(betas)} peers)")
        else:
            st.warning("‚ö†Ô∏è Could not calculate beta from peers, using default Œ≤=1.0")
            beta = 1.0
    else:
        st.warning("‚ö†Ô∏è No peer tickers provided, using default Œ≤=1.0")
    
    ke = rf + (beta * (rm - rf))
    
    # Cost of Debt (Kd) - USE NEWEST values (index 0)
    # Handle debt properly - could be 0 or NaN
    st_debt = financials['st_debt'][0] if financials['st_debt'][0] > 0 else 0
    lt_debt = financials['lt_debt'][0] if financials['lt_debt'][0] > 0 else 0
    total_debt = st_debt + lt_debt
    interest = financials['interest'][0]
    
    # Cost of Debt - handle zero/NaN debt
    if total_debt > 0 and interest > 0:
        kd = (interest / total_debt * 100)
    else:
        kd = 0.0  # Debt-free company has no cost of debt
    kd_after_tax = kd * (1 - tax_rate / 100)
    
    # WACC - USE NEWEST equity (index 0)
    equity = financials['equity'][0]
    total_capital = equity + total_debt
    
    # Handle weights for debt-free companies
    if total_capital > 0 and equity > 0:
        we = equity / total_capital
        wd = total_debt / total_capital if total_debt > 0 else 0.0
    else:
        we = 1.0  # 100% equity financed
        wd = 0.0  # No debt
    
    wacc = (we * ke) + (wd * kd_after_tax)
    
    return {
        'wacc': wacc,
        'ke': ke,
        'kd': kd,
        'kd_after_tax': kd_after_tax,
        'rf': rf,
        'rm': rm,
        'beta': beta,
        'we': we * 100,
        'wd': wd * 100,
        'equity': equity,
        'debt': total_debt
    }

def calculate_wacc_bank(financials, tax_rate, peer_tickers=None, manual_rf_rate=None):
    """
    Calculate WACC for BANKS/NBFCs with proper Cost of Funds methodology
    
    For banks:
    - Kd = Cost of Funds (WACF) = Interest Expended / Average Interest-Bearing Liabilities
    - Interest-bearing liabilities include: Deposits, Borrowings, Bonds, Subordinated debt
    - NOT just simple Interest/Debt ratio (that's wrong for banks!)
    """
    
    # Cost of Equity (Ke) - Same as normal companies
    # ALWAYS use manual_rf_rate (passed from session state), never fetch
    rf = manual_rf_rate if manual_rf_rate is not None else 6.83  # Fallback to default, DON'T fetch
    rm = get_market_return()
    
    beta = 1.0
    if peer_tickers and peer_tickers.strip():
        ticker_list = [t.strip() for t in peer_tickers.split(',') if t.strip()]
        betas = []
        
        for ticker in ticker_list:
            try:
                ticker_beta = get_stock_beta(ticker)
                if ticker_beta > 0:
                    betas.append(ticker_beta)
            except:
                pass
        
        if betas:
            beta = np.mean(betas)
            st.success(f"‚úÖ Average bank peer beta: {beta:.3f}")
        else:
            beta = 1.0
    
    ke = rf + (beta * (rm - rf))
    
    # Cost of Debt (Kd) - BANK METHODOLOGY
    # Kd = WACF (Weighted Average Cost of Funds)
    # Formula: Interest Expended / Average Interest-Bearing Liabilities
    
    interest_expense = financials['interest'][0]  # Total interest paid
    
    # Interest-bearing liabilities (need to calculate from balance sheet)
    # For banks: Deposits + Borrowings = primary interest-bearing liabilities
    # We approximate: Total Liabilities - Equity - Other non-interest liabilities
    
    equity_current = financials['equity'][0]
    equity_previous = financials['equity'][1] if len(financials['equity']) > 1 else equity_current
    
    # For banks, we need total assets - equity = total liabilities (most are interest-bearing)
    # Rough approximation: Use debt as proxy for interest-bearing liabilities
    st_debt_current = financials['st_debt'][0] if financials['st_debt'][0] > 0 else 0
    lt_debt_current = financials['lt_debt'][0] if financials['lt_debt'][0] > 0 else 0
    total_liabilities_current = st_debt_current + lt_debt_current
    
    st_debt_previous = financials['st_debt'][1] if len(financials['st_debt']) > 1 and financials['st_debt'][1] > 0 else st_debt_current
    lt_debt_previous = financials['lt_debt'][1] if len(financials['lt_debt']) > 1 and financials['lt_debt'][1] > 0 else lt_debt_current
    total_liabilities_previous = st_debt_previous + lt_debt_previous
    
    # Average interest-bearing liabilities
    avg_interest_bearing_liabilities = (total_liabilities_current + total_liabilities_previous) / 2
    
    # Cost of Funds (Kd for banks)
    if avg_interest_bearing_liabilities > 0 and interest_expense > 0:
        kd = (interest_expense / avg_interest_bearing_liabilities) * 100
        st.info(f"üí° **Bank Cost of Funds (Kd):** {kd:.2f}% = Interest Expended ‚Çπ{interest_expense:,.0f} / Avg Liabilities ‚Çπ{avg_interest_bearing_liabilities:,.0f}")
    else:
        kd = 5.0  # Default reasonable cost of funds for banks
        st.warning("‚ö†Ô∏è Using default bank cost of funds: 5%")
    
    # For banks, tax shield on interest is different (interest is operating expense, not financing)
    # But for WACC calculation, we still apply tax shield
    kd_after_tax = kd * (1 - tax_rate / 100)
    
    # Capital structure weights
    total_capital = equity_current + total_liabilities_current
    
    if total_capital > 0:
        we = equity_current / total_capital
        wd = total_liabilities_current / total_capital
    else:
        we = 0.20  # Banks typically have low equity weight
        wd = 0.80
    
    wacc = (we * ke) + (wd * kd_after_tax)
    
    st.success(f"‚úÖ **Bank WACC Calculated:** {wacc:.2f}% | Equity Weight: {we*100:.1f}% | Debt Weight: {wd*100:.1f}%")
    
    return {
        'wacc': wacc,
        'ke': ke,
        'kd': kd,
        'kd_after_tax': kd_after_tax,
        'rf': rf,
        'rm': rm,
        'beta': beta,
        'we': we * 100,
        'wd': wd * 100,
        'equity': equity_current,
        'debt': total_liabilities_current,
        'calculation_method': 'Bank Methodology (Cost of Funds)',
        'interest_expense': interest_expense,
        'avg_liabilities': avg_interest_bearing_liabilities
    }

def calculate_dcf_valuation(projections, wacc_details, terminal_growth, num_shares, cash_balance=0, manual_discount_rate=None):
    """
    Calculate DCF valuation with Rulebook-compliant validations and intelligent FCFF recovery
    
    Args:
        manual_discount_rate: Optional manual override for discount rate (instead of WACC)
    """
    # Use manual discount rate if provided, otherwise use WACC
    if manual_discount_rate and manual_discount_rate > 0:
        wacc = manual_discount_rate
        discount_rate_source = f"Manual Override ({manual_discount_rate:.2f}%)"
    else:
        wacc = wacc_details['wacc']
        discount_rate_source = f"WACC ({wacc:.2f}%)"
    
    g = terminal_growth
    
    # RULEBOOK SECTION 8.2: Terminal growth must be < discount rate
    if g >= wacc:
        return None, f"‚ùå HARD ERROR: Terminal growth rate must be less than discount rate (Rulebook 8.2). Current: TG={g:.1f}%, Discount={wacc:.1f}%"
    
    # CRITICAL FIX: Terminal growth should generally be lower than long-term revenue growth
    # Extract revenue growth from projections
    if len(projections['revenue']) >= 2:
        first_rev = projections['revenue'][0]
        last_rev = projections['revenue'][-1]
        num_years = len(projections['revenue']) - 1
        
        # BUGFIX: Protect against zero/negative revenues to prevent ZeroDivisionError
        if first_rev > 0 and last_rev > 0:
            implied_revenue_cagr = ((last_rev / first_rev) ** (1 / num_years) - 1) * 100
        else:
            implied_revenue_cagr = 0  # Default to 0 if revenue data is invalid
            st.warning("‚ö†Ô∏è Revenue data contains zero or negative values. Cannot calculate revenue CAGR.")
        
        # Warning if terminal growth is too close to revenue growth
        if g > implied_revenue_cagr * 0.9 and implied_revenue_cagr > 0:
            st.warning(f"‚ö†Ô∏è Terminal Growth Rate ({g:.1f}%) is very close to projected revenue CAGR ({implied_revenue_cagr:.1f}%)")
            st.info("üí° **Recommendation:** Terminal growth should typically be 40-60% of revenue growth for conservative valuations")
    
    # RULEBOOK SECTION 8.2: Check terminal year FCFF
    last_fcff = projections['fcff'][-1]
    fcff_adjusted = False
    adjustment_details = {}
    
    if last_fcff <= 0:
        # =====================================================
        # INTELLIGENT FCFF RECOVERY MECHANISM
        # =====================================================
        st.warning(f"‚ö†Ô∏è Terminal year FCFF is {last_fcff:.2f} Lacs (negative or zero)")
        st.info("üîß **Activating Intelligent FCFF Recovery Mechanism**")
        
        # Analyze all projected FCFFs to understand the issue
        all_fcffs = projections['fcff']
        positive_fcffs = [f for f in all_fcffs if f > 0]
        
        # Even if NO positive FCFFs, we can still recover using ULTRA-AGGRESSIVE strategies
        if len(positive_fcffs) == 0:
            st.error("‚ö†Ô∏è **SEVERE CASE:** All projected FCFFs are negative or zero")
            st.warning("üîß **Activating ULTRA-AGGRESSIVE Recovery Mechanisms**")
            st.caption("Using fundamental value drivers to construct viable terminal FCFF")
        
        # Calculate intelligent adjustments
        avg_positive_fcff = np.mean(positive_fcffs) if positive_fcffs else 0
        median_positive_fcff = np.median(positive_fcffs) if positive_fcffs else 0
        max_fcff = max(all_fcffs) if all_fcffs else 0
        
        # Analyze components to find best recovery path
        last_nopat = projections['nopat'][-1]
        last_dep = projections['depreciation'][-1]
        last_dwc = projections['delta_wc'][-1]
        last_capex = projections['capex'][-1]
        
        # Calculate component ratios
        revenue = projections['revenue'][-1]
        ebitda = projections['ebitda'][-1]
        
        # INTELLIGENT RECOVERY STRATEGIES
        recovery_options = []
        
        # =====================================================
        # ULTRA-AGGRESSIVE STRATEGIES (For severe cases with all negative FCFFs)
        # =====================================================
        
        # Ultra Strategy 1: Revenue-Based Proxy with Industry Margins
        if revenue > 0:
            # Assume industry-standard metrics:
            # - Operating Margin: 10-15% (conservative for mature company)
            # - Tax Rate: from WACC details or default 25%
            # - FCFF/Revenue: 5-8% for mature companies
            tax_rate = wacc_details.get('tax_rate', 25)
            
            # Conservative approach: 6% FCFF/Revenue ratio
            fcff_from_revenue = revenue * 0.06
            
            if fcff_from_revenue > 0:
                recovery_options.append({
                    'strategy': 'Revenue-Based Proxy (Ultra-Aggressive)',
                    'fcff': fcff_from_revenue,
                    'adjustments': {
                        'revenue': f'‚Çπ{revenue:.2f} Lacs',
                        'fcff_margin': '6% of Revenue (conservative industry standard)',
                        'rationale': 'Assumes normalized mature company cash conversion',
                        'calculated_fcff': f'{fcff_from_revenue:.2f} Lacs'
                    }
                })
        
        # Ultra Strategy 2: Reverse-Engineered from Growth Rate
        # If we know terminal growth rate, work backwards from sustainable metrics
        if terminal_growth > 0 and revenue > 0:
            # Sustainable FCFF = Revenue √ó Terminal Growth √ó FCFF/Growth ratio
            # Typical FCFF/Growth ratio: 2-4 (i.e., if growing at 5%, FCFF ~10-20% of revenue)
            sustainable_fcff = revenue * (terminal_growth / 100) * 2.5
            
            if sustainable_fcff > 0:
                recovery_options.append({
                    'strategy': 'Growth-Reverse-Engineered (Ultra-Aggressive)',
                    'fcff': sustainable_fcff,
                    'adjustments': {
                        'terminal_growth': f'{terminal_growth}%',
                        'revenue': f'‚Çπ{revenue:.2f} Lacs',
                        'fcff_growth_ratio': '2.5x (industry standard)',
                        'logic': f'Sustainable FCFF = Revenue √ó {terminal_growth}% √ó 2.5',
                        'calculated_fcff': f'{sustainable_fcff:.2f} Lacs'
                    }
                })
        
        # Ultra Strategy 3: NOPAT-Based with Zero CapEx/WC (Absolute Floor)
        # Assumes company transitions to asset-light model
        if last_nopat > 0:
            fcff_nopat_only = last_nopat + last_dep  # Just operating cash
            
            if fcff_nopat_only > 0:
                recovery_options.append({
                    'strategy': 'NOPAT-Based Floor (Ultra-Aggressive)',
                    'fcff': fcff_nopat_only,
                    'adjustments': {
                        'nopat': f'‚Çπ{last_nopat:.2f} Lacs',
                        'depreciation': f'‚Çπ{last_dep:.2f} Lacs',
                        'capex': 'Assumed ZERO (asset-light transition)',
                        'working_capital': 'Assumed ZERO (normalized)',
                        'calculated_fcff': f'{fcff_nopat_only:.2f} Lacs'
                    }
                })
        
        # Ultra Strategy 4: Operating Cash Flow Proxy
        # EBITDA - Tax (ignoring all capital requirements)
        if ebitda > 0:
            tax_rate = wacc_details.get('tax_rate', 25)
            operating_cash = ebitda * (1 - tax_rate / 100)
            
            if operating_cash > 0:
                recovery_options.append({
                    'strategy': 'Operating Cash Proxy (Ultra-Aggressive)',
                    'fcff': operating_cash,
                    'adjustments': {
                        'ebitda': f'‚Çπ{ebitda:.2f} Lacs',
                        'tax_rate': f'{tax_rate}%',
                        'assumption': 'Pure operating cash, no capital requirements',
                        'calculated_fcff': f'{operating_cash:.2f} Lacs'
                    }
                })
        
        # Ultra Strategy 5: Minimum Viable FCFF (Absolute Last Resort)
        # Use 3% of revenue as bare minimum sustainable cash generation
        if revenue > 0:
            min_viable_fcff = revenue * 0.03
            
            recovery_options.append({
                'strategy': 'Minimum Viable FCFF (Last Resort)',
                'fcff': min_viable_fcff,
                'adjustments': {
                    'revenue': f'‚Çπ{revenue:.2f} Lacs',
                    'fcff_margin': '3% of Revenue (bare minimum)',
                    'rationale': 'Absolute floor for sustainable operations',
                    'note': 'This is the MOST aggressive assumption',
                    'calculated_fcff': f'{min_viable_fcff:.2f} Lacs'
                }
            })
        
        # =====================================================
        # STANDARD RECOVERY STRATEGIES
        # =====================================================
        
        # Strategy 1: Reduce CapEx to sustainable level (typically 80% of depreciation)
        sustainable_capex = last_dep * 0.8
        capex_savings = last_capex - sustainable_capex if last_capex > sustainable_capex else 0
        fcff_with_capex_adj = last_nopat + last_dep - last_dwc - sustainable_capex
        if fcff_with_capex_adj > 0:
            recovery_options.append({
                'strategy': 'Reduced CapEx to Sustainable Level',
                'fcff': fcff_with_capex_adj,
                'adjustments': {
                    'capex': f'Reduced from {last_capex:.2f} to {sustainable_capex:.2f} Lacs (80% of Depreciation)',
                    'savings': f'CapEx savings: {capex_savings:.2f} Lacs'
                }
            })
        
        # Strategy 2: Normalize Working Capital (assume zero working capital change in terminal year)
        fcff_with_wc_normalization = last_nopat + last_dep - last_capex
        if fcff_with_wc_normalization > 0:
            recovery_options.append({
                'strategy': 'Normalized Working Capital',
                'fcff': fcff_with_wc_normalization,
                'adjustments': {
                    'working_capital': f'Set ŒîWC to 0 (from {last_dwc:.2f} Lacs)',
                    'assumption': 'Working capital stabilizes at terminal year'
                }
            })
        
        # Strategy 3: Combined approach (sustainable capex + normalized WC)
        fcff_combined = last_nopat + last_dep - sustainable_capex
        if fcff_combined > 0:
            recovery_options.append({
                'strategy': 'Combined: Sustainable CapEx + Normalized WC',
                'fcff': fcff_combined,
                'adjustments': {
                    'capex': f'CapEx at 80% depreciation: {sustainable_capex:.2f} Lacs',
                    'working_capital': 'ŒîWC = 0 (normalized)',
                    'combined_impact': f'Total improvement: {fcff_combined - last_fcff:.2f} Lacs'
                }
            })
        
        # Strategy 4: Use average of positive FCFFs (conservative)
        if avg_positive_fcff > 0:
            recovery_options.append({
                'strategy': 'Average of Positive Historical FCFFs',
                'fcff': avg_positive_fcff,
                'adjustments': {
                    'basis': f'Average of {len(positive_fcffs)} positive FCFF years',
                    'value': f'{avg_positive_fcff:.2f} Lacs',
                    'rationale': 'Uses historical positive cash generation as sustainable baseline'
                }
            })
        
        # Strategy 5: EBITDA-based proxy (if operations are profitable)
        if ebitda > 0:
            # Typical FCFF = ~40-60% of EBITDA for mature companies
            fcff_from_ebitda = ebitda * 0.5 * (1 - wacc_details.get('tax_rate', 25) / 100)
            if fcff_from_ebitda > 0:
                recovery_options.append({
                    'strategy': 'EBITDA-Based Proxy',
                    'fcff': fcff_from_ebitda,
                    'adjustments': {
                        'ebitda': f'{ebitda:.2f} Lacs',
                        'conversion': '50% FCFF/EBITDA ratio (industry standard)',
                        'tax_adjusted': 'Applied tax rate',
                        'calculated_fcff': f'{fcff_from_ebitda:.2f} Lacs'
                    }
                })
        
        # Select BEST recovery option
        # With ultra-aggressive strategies, we ALWAYS have at least one option
        if not recovery_options:
            # Absolute fallback - should never reach here
            # Create emergency FCFF based on revenue if available
            if revenue > 0:
                emergency_fcff = revenue * 0.02  # 2% of revenue
                recovery_options.append({
                    'strategy': 'Emergency Fallback (2% of Revenue)',
                    'fcff': emergency_fcff,
                    'adjustments': {
                        'note': 'Emergency fallback - uses minimal assumptions',
                        'revenue': f'‚Çπ{revenue:.2f} Lacs',
                        'calculated_fcff': f'{emergency_fcff:.2f} Lacs'
                    }
                })
            elif ebitda > 0:
                emergency_fcff = ebitda * 0.3  # 30% of EBITDA
                recovery_options.append({
                    'strategy': 'Emergency Fallback (30% of EBITDA)',
                    'fcff': emergency_fcff,
                    'adjustments': {
                        'note': 'Emergency fallback - uses minimal assumptions',
                        'ebitda': f'‚Çπ{ebitda:.2f} Lacs',
                        'calculated_fcff': f'{emergency_fcff:.2f} Lacs'
                    }
                })
            else:
                # Truly desperate - use 1 Lac as minimum
                recovery_options.append({
                    'strategy': 'Absolute Minimum (‚Çπ1 Lac)',
                    'fcff': 1.0,
                    'adjustments': {
                        'note': '‚ö†Ô∏è No viable metrics available - using symbolic minimum',
                        'warning': 'Valuation highly uncertain - use with extreme caution'
                    }
                })
        
        # Rank by FCFF value (higher is better, but prefer more conservative approaches)
        # Preference order: Standard strategies > Ultra-aggressive strategies
        strategy_preference = {
            # Standard strategies (highest priority)
            'Combined: Sustainable CapEx + Normalized WC': 10,
            'Reduced CapEx to Sustainable Level': 9,
            'Normalized Working Capital': 8,
            'Average of Positive Historical FCFFs': 7,
            'EBITDA-Based Proxy': 6,
            # Ultra-aggressive strategies (medium priority)
            'Revenue-Based Proxy (Ultra-Aggressive)': 5,
            'Growth-Reverse-Engineered (Ultra-Aggressive)': 4,
            'NOPAT-Based Floor (Ultra-Aggressive)': 3,
            'Operating Cash Proxy (Ultra-Aggressive)': 2,
            'Minimum Viable FCFF (Last Resort)': 1,
            # Emergency fallbacks (lowest priority)
            'Emergency Fallback (2% of Revenue)': 0.5,
            'Emergency Fallback (30% of EBITDA)': 0.4,
            'Absolute Minimum (‚Çπ1 Lac)': 0.1
        }
        
        # Sort by preference, then by FCFF value
        best_option = max(recovery_options, 
                         key=lambda x: (strategy_preference.get(x['strategy'], 0), x['fcff']))
        
        # Apply the best recovery strategy
        last_fcff = best_option['fcff']
        fcff_adjusted = True
        adjustment_details = best_option
        
        # Display the recovery strategy
        st.success(f"‚úÖ **Selected Recovery Strategy:** {best_option['strategy']}")
        st.write(f"**Adjusted Terminal FCFF:** ‚Çπ{last_fcff:.2f} Lacs (Original: ‚Çπ{projections['fcff'][-1]:.2f} Lacs)")
        
        # Add extra warning for ultra-aggressive strategies
        if 'Ultra-Aggressive' in best_option['strategy'] or 'Last Resort' in best_option['strategy'] or 'Emergency' in best_option['strategy'] or 'Absolute' in best_option['strategy']:
            st.warning("‚ö†Ô∏è **CAUTION:** This strategy uses aggressive assumptions due to severe cash flow issues")
            st.caption("üí° Recommendation: Cross-validate with alternative valuation methods (P/E, P/B, EV/EBITDA)")
        
        with st.expander("üìã View All Recovery Options Considered"):
            for i, opt in enumerate(recovery_options, 1):
                st.markdown(f"### Option {i}: {opt['strategy']}")
                st.write(f"**Resulting FCFF:** ‚Çπ{opt['fcff']:.2f} Lacs")
                st.write("**Adjustments:**")
                for key, value in opt['adjustments'].items():
                    st.write(f"  - {key.replace('_', ' ').title()}: {value}")
                if opt == best_option:
                    st.success("‚úÖ **SELECTED** (Best balance of conservatism and cash flow)")
                st.markdown("---")
        
        st.info("üí° **Note:** These adjustments reflect sustainable long-term assumptions required for terminal value calculation.")
        
        # CRITICAL: Update projections with adjusted terminal FCFF
        # This ensures all downstream calculations use the recovered value
        projections['fcff'][-1] = last_fcff
        st.caption(f"üìå Terminal year FCFF in projections updated to ‚Çπ{last_fcff:.2f} Lacs")
    
    # Present Value of FCFFs
    pv_fcffs = []
    for i, fcff in enumerate(projections['fcff']):
        year = i + 1
        pv = fcff / ((1 + wacc / 100) ** year)
        pv_fcffs.append(pv)
    
    sum_pv_fcff = sum(pv_fcffs)
    
    # CRITICAL CHECK: If sum of PV(FCFF) is negative, we need additional recovery
    # This happens when ALL or most FCFFs are negative (high growth/investment phase)
    growth_phase_adjusted = False
    original_sum_pv_fcff = sum_pv_fcff
    
    if sum_pv_fcff < 0:
        st.warning(f"‚ö†Ô∏è **Additional Issue Detected:** Sum of PV(FCFF) is negative (‚Çπ{sum_pv_fcff:.2f} Lacs)")
        st.info("üîß **Applying Growth-Phase Adjustment**: Treating as high-growth company transitioning to maturity")
        
        # For high-growth companies, we should focus entirely on terminal value
        # Set sum_pv_fcff to zero (ignore negative cash flows during growth phase)
        sum_pv_fcff = 0
        growth_phase_adjusted = True
        
        st.success(f"‚úÖ **Growth-Phase Adjustment Applied:**")
        st.write(f"   - Original Sum PV(FCFF): ‚Çπ{original_sum_pv_fcff:.2f} Lacs (negative due to growth)")
        st.write(f"   - Adjusted Sum PV(FCFF): ‚Çπ{sum_pv_fcff:.2f} Lacs (set to zero)")
        st.write(f"   - **Rationale:** Company in high-growth phase; value comes from mature cash flows")
        st.caption("üí° This is common for high-growth companies that invest heavily before generating positive cash flows")
    
    # Terminal Value (Rulebook Section 8.1)
    fcff_n_plus_1 = last_fcff * (1 + g / 100)
    terminal_value = fcff_n_plus_1 / ((wacc / 100) - (g / 100))
    
    n = len(projections['fcff'])
    pv_terminal_value = terminal_value / ((1 + wacc / 100) ** n)
    
    # Enterprise Value (in Lacs)
    enterprise_value = sum_pv_fcff + pv_terminal_value
    
    # RULEBOOK SECTION 13.1: Terminal Value sanity checks
    tv_percentage = (pv_terminal_value / enterprise_value * 100) if enterprise_value > 0 else 0
    
    if tv_percentage > 100:
        return None, f"‚ùå ERROR: Terminal Value ({tv_percentage:.1f}%) exceeds 100% of Enterprise Value (Rulebook 13.1)"
    
    # Equity Value Calculation: EV - Net Debt
    # Handle debt properly (could be 0, None, or NaN for debt-free companies)
    total_debt = wacc_details.get('debt', 0)
    if total_debt is None or (isinstance(total_debt, float) and np.isnan(total_debt)):
        total_debt = 0
    total_debt = float(total_debt) if total_debt > 0 else 0
    
    # Handle cash properly (extract from parameter, could be 0)
    cash = float(cash_balance) if cash_balance and cash_balance > 0 else 0
    
    # Net Debt = Total Debt - Cash
    # Can be negative (net cash position) if cash > debt
    net_debt = total_debt - cash
    
    # Equity Value = Enterprise Value - Net Debt
    # For debt-free companies: EV - (-cash) = EV + cash
    equity_value = enterprise_value - net_debt
    
    # Convert Equity Value from Lacs to absolute Rupees, then divide by shares
    # Equity Value is in Lacs, so multiply by 100,000 to get Rupees
    equity_value_rupees = equity_value * 100000
    fair_value_per_share = equity_value_rupees / num_shares if num_shares > 0 else 0
    
    # CRITICAL VALIDATION: Check if fair value is negative
    negative_fair_value_warning = None
    if fair_value_per_share < 0:
        negative_fair_value_warning = {
            'enterprise_value': enterprise_value,
            'total_debt': total_debt,
            'cash': cash,
            'net_debt': net_debt,
            'equity_value': equity_value,
            'num_shares': num_shares,
            'reason': []
        }
        
        # Diagnose the problem
        if enterprise_value < net_debt:
            negative_fair_value_warning['reason'].append(
                f"Enterprise Value (‚Çπ{enterprise_value:.2f} Lacs) is less than Net Debt (‚Çπ{net_debt:.2f} Lacs)"
            )
        if enterprise_value <= 0:
            negative_fair_value_warning['reason'].append(
                f"Enterprise Value is zero or negative (‚Çπ{enterprise_value:.2f} Lacs)"
            )
        if net_debt > enterprise_value * 2:
            negative_fair_value_warning['reason'].append(
                f"Net Debt (‚Çπ{net_debt:.2f} Lacs) is more than 2x Enterprise Value"
            )
    
    return {
        'pv_fcffs': pv_fcffs,
        'sum_pv_fcff': sum_pv_fcff,
        'original_sum_pv_fcff': original_sum_pv_fcff if growth_phase_adjusted else sum_pv_fcff,
        'growth_phase_adjusted': growth_phase_adjusted,
        'terminal_value': terminal_value,
        'pv_terminal_value': pv_terminal_value,
        'enterprise_value': enterprise_value,
        'total_debt': total_debt,
        'cash': cash,
        'net_debt': net_debt,
        'equity_value': equity_value,
        'equity_value_rupees': equity_value_rupees,
        'fair_value_per_share': fair_value_per_share,
        'tv_percentage': tv_percentage,
        'tv_warning': tv_percentage > 90,  # Flag for warning
        'fcff_adjusted': fcff_adjusted,
        'adjustment_details': adjustment_details if fcff_adjusted else None,
        'adjusted_terminal_fcff': last_fcff if fcff_adjusted else projections['fcff'][-1],
        'negative_fair_value_warning': negative_fair_value_warning,
        'wacc': wacc,
        'discount_rate_source': discount_rate_source
    }, None
# ================================
# MAIN UI FUNCTION
# ================================
def main():
    """Main DCF UI function - can be called from dashboard or run standalone"""
    
    # ===== INITIALIZE RISK-FREE RATE ONCE AT STARTUP =====
    # This runs ONLY ONCE when app starts, not on every interaction
    if 'rf_rate_initialized' not in st.session_state:
        st.session_state.rf_rate_initialized = True
        st.session_state.cached_rf_rate_listed = 6.83  # Default fallback
        st.session_state.cached_rf_rate_screener = 6.83  # Default fallback
        # Don't fetch automatically - let user click button when ready
    
    # Initialize session state for peer auto-fetch (Screener mode)
    if 'nse_peers_input' not in st.session_state:
        st.session_state.nse_peers_input = ''
    if 'bse_peers_input' not in st.session_state:
        st.session_state.bse_peers_input = ''
    
    st.title("üè¶ DCF Valuation Engine")
    st.markdown("**Listed + Unlisted | Excel-Integrated | Traditional WACC**")

    # Show rate limit status
    if st.session_state.yahoo_request_count > 0:
        with st.expander("üìä API Usage Status", expanded=False):
            st.caption(f"Yahoo Finance requests this hour: {st.session_state.yahoo_request_count}")
            if st.session_state.yahoo_request_count > 10:
                st.warning("‚ö†Ô∏è High API usage detected. Consider waiting if you encounter rate limits.")

    # Display current India inflation rate - USING ACTUAL DATA
    try:
        @st.cache_data(ttl=86400)  # Cache for 24 hours
        def get_india_inflation():
            """Fetch latest India CPI inflation rate from World Bank"""
            try:
                # Method 1: Try world_bank_data (pip install world-bank-data)
                try:
                    import world_bank_data as wb
                    # India CPI inflation: FP.CPI.TOTL.ZG
                    df = wb.get_series('FP.CPI.TOTL.ZG', id_or_value='id', simplify_index=True, mrv=1)
                    india_data = df[df.index.get_level_values(0) == 'IND']
                    if not india_data.empty:
                        inflation = float(india_data.iloc[0])
                        year = india_data.index.get_level_values(1)[0]
                        return inflation, f"World Bank ({year})"
                except:
                    pass
            
                # Method 2: Try pandas_datareader (RBI/FRED)
                try:
                    import pandas_datareader as pdr
                    from datetime import datetime, timedelta
                
                    # Try FRED (Federal Reserve Economic Data) - has India CPI
                    end_date = datetime.now()
                    start_date = end_date - timedelta(days=365)
                
                    # India CPI (INDCPIALLMINMEI)
                    india_cpi = pdr.get_data_fred('INDCPIALLMINMEI', start_date, end_date)
                    if not india_cpi.empty:
                        # Calculate YoY inflation from latest data
                        latest_cpi = india_cpi.iloc[-1].values[0]
                        year_ago_cpi = india_cpi.iloc[0].values[0]
                        inflation = ((latest_cpi - year_ago_cpi) / year_ago_cpi) * 100
                        date = india_cpi.index[-1].strftime('%b %Y')
                        return inflation, f"FRED ({date})"
                except:
                    pass
            
                # Method 3: Reasonable estimate from known data
                return 5.22, "Estimate (Jan 2026)"
            
            except Exception as e:
                return 5.5, "Fallback Estimate"
    
        india_inflation, inflation_source = get_india_inflation()
    
        # Display inflation banner
        col_inf1, col_inf2, col_inf3 = st.columns([1, 2, 1])
        with col_inf2:
            if "Estimate" in inflation_source or "Fallback" in inflation_source:
                st.warning(f"üìä **India CPI Inflation ({inflation_source}):** {india_inflation:.2f}% | Reference for terminal growth")
            else:
                st.success(f"üìä **India CPI Inflation ({inflation_source}):** {india_inflation:.2f}% | Reference for terminal growth")
    except Exception as e:
        pass  # Silently fail if inflation display doesn't work

    # PDF EXPORT SECTION
    # ============================================
# ============================================
    st.markdown("---")
    st.markdown("### üì• Export PDF Report")
    if 'pdf_bytes' in st.session_state and st.session_state.pdf_bytes:
        st.success("‚úÖ **PDF Report Ready!**")
        col1, col2 = st.columns([3, 1])
        with col1:
            st.download_button(
                label="üíæ Download PDF Report",
                data=st.session_state.pdf_bytes,
                file_name=f"Valuation_Report_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf",
                mime="application/pdf",
                type="primary",
                use_container_width=True,
                key="download_pdf_top"
            )
        with col2:
            if st.button("üîÑ Clear", key="clear_pdf"):
                st.session_state.pdf_bytes = None
                st.rerun()
    else:
        st.info("üí° Complete a valuation below to generate PDF report")
    st.markdown("---")

    # ================================
    # SESSION STATE INITIALIZATION
    # ================================
    if 'show_results_listed' not in st.session_state:
        st.session_state.show_results_listed = False
    if 'show_results_unlisted' not in st.session_state:
        st.session_state.show_results_unlisted = False
    if 'fetched_nse_peers' not in st.session_state:
        st.session_state.fetched_nse_peers = {}
    if 'fetched_bse_peers' not in st.session_state:
        st.session_state.fetched_bse_peers = {}
    if 'fetch_status' not in st.session_state:
        st.session_state.fetch_status = {}
    if 'current_ticker' not in st.session_state:
        st.session_state.current_ticker = None
    if 'yahoo_request_count' not in st.session_state:
        st.session_state.yahoo_request_count = 0

    # Cache status display and control
    col_cache1, col_cache2 = st.columns([3, 1])
    with col_cache1:
        cache_size = len(_TICKER_DATA_CACHE)
        if cache_size > 0:
            st.info(f"üöÄ **Rate Limit Protection Active:** {cache_size} data entries cached (prevents duplicate API calls)")
        else:
            st.info("üöÄ **Rate Limit Protection:** Caching enabled to prevent duplicate API calls")
    with col_cache2:
        if st.button("üóëÔ∏è Clear Cache", help="Clear ticker cache if you need fresh data"):
            clear_ticker_cache()
            st.success("Cache cleared!")
            st.rerun()

    mode = st.radio("Select Mode:", 
                    ["Listed Company (Yahoo Finance)", 
                     "Unlisted Company (Excel Upload)",
                     "Screener Excel Mode (Screener.in Template)"], 
                    horizontal=True)

    if mode == "Listed Company (Yahoo Finance)":
        st.subheader("üìà Listed Company Valuation")
        
        # ===== RISK-FREE RATE TICKER - AT THE TOP, ALWAYS VISIBLE =====
        st.markdown("### üèõÔ∏è Risk-Free Rate Configuration")
        st.info("üí° **Default: 6.83%**. Enter any Yahoo Finance ticker and click Fetch. Common options: ^TNX (US 10Y), ^IRX (US 3-month), or any stock/index for custom rates.")
        
        # Initialize if not exists
        if 'cached_rf_rate_listed' not in st.session_state:
            st.session_state.cached_rf_rate_listed = 6.83
        
        rf_col1, rf_col2, rf_col3 = st.columns([3, 2, 1])
        with rf_col1:
            custom_rf_ticker_listed = st.text_input(
                "Yahoo Finance Ticker for Risk-Free Rate",
                value="NIFTYGS10YR.NS",
                key='custom_rf_ticker_listed_top',
                help="Enter any ticker - will calculate CAGR for stocks/indices, or use value directly for yields. Examples: NIFTYGS10YR.NS, ^TNX, RELIANCE.NS"
            )
        with rf_col2:
            st.metric("Current RF Rate", f"{st.session_state.cached_rf_rate_listed:.2f}%")
        with rf_col3:
            st.write("")
            st.write("")
            
            # Add a counter to track button clicks
            if 'rf_fetch_click_count_listed' not in st.session_state:
                st.session_state.rf_fetch_click_count_listed = 0
            
            if st.button("üîÑ Fetch", key='refresh_rf_listed_top'):
                # Increment click counter FIRST
                st.session_state.rf_fetch_click_count_listed += 1
                
                # Store debug output in session state so it persists across reruns
                debug_output = []
                debug_output.append(f"üîÑ **FETCH BUTTON CLICKED #{st.session_state.rf_fetch_click_count_listed} - LISTED MODE**")
                debug_output.append(f"üìù Input ticker: `{custom_rf_ticker_listed}`")
                
                ticker_to_use = custom_rf_ticker_listed.strip() if custom_rf_ticker_listed.strip() else None
                debug_output.append(f"üìù Ticker to use (after strip): `{ticker_to_use}`")
                
                debug_output.append("‚è≥ Calling get_risk_free_rate()...")
                fetched_rate, fetch_debug = get_risk_free_rate(ticker_to_use)
                
                # Add all debug messages from the function
                debug_output.extend(fetch_debug)
                
                debug_output.append(f"‚úÖ Function returned: {fetched_rate}%")
                
                debug_output.append(f"üíæ Updating session state...")
                debug_output.append(f"   - Before: {st.session_state.get('cached_rf_rate_listed', 'NOT SET')}")
                st.session_state.cached_rf_rate_listed = fetched_rate
                debug_output.append(f"   - After: {st.session_state.cached_rf_rate_listed}")
                
                # Force update the manual input field
                debug_output.append(f"üîÑ Clearing manual input widget state...")
                if 'manual_rf_listed' in st.session_state:
                    del st.session_state['manual_rf_listed']
                    debug_output.append(f"   - Widget state cleared")
                else:
                    debug_output.append(f"   - Widget state was not set")
                
                # Store debug output in session state
                st.session_state.rf_fetch_debug_listed = debug_output
                st.session_state.rf_fetch_success_listed = True
                
                # Rerun to update UI
                st.rerun()
        
        # Display debug output if available (persists across reruns)
        if st.session_state.get('rf_fetch_debug_listed'):
            with st.expander("üìã Last Fetch Debug Output", expanded=True):
                for line in st.session_state.rf_fetch_debug_listed:
                    st.write(line)
                if st.session_state.get('rf_fetch_success_listed'):
                    st.success(f"‚úì Successfully updated to {st.session_state.cached_rf_rate_listed:.2f}%")
        
        st.markdown("---")
        # ===== END RF RATE CONFIG =====
        
        # DEBUG: Show session state
        with st.expander("üêõ DEBUG: Session State Values (Listed)", expanded=False):
            st.write("**Current Session State for RF Rate:**")
            st.write(f"- `cached_rf_rate_listed`: {st.session_state.get('cached_rf_rate_listed', 'NOT SET')}")
            st.write(f"- `manual_rf_listed` widget: {st.session_state.get('manual_rf_listed', 'NOT SET')}")
            st.write(f"- `custom_rf_ticker_listed_top`: {st.session_state.get('custom_rf_ticker_listed_top', 'NOT SET')}")
            st.write(f"- **Fetch button clicks**: {st.session_state.get('rf_fetch_click_count_listed', 0)}")
    
        col1, col2 = st.columns(2)
    
        with col1:
            # Exchange selection for COMPANY BEING VALUED
            exchange = st.radio("Company Exchange:", ["NSE", "BSE"], horizontal=True, help="Select exchange for the company being valued")
            exchange_suffix = "NS" if exchange == "NSE" else "BO"
        
            ticker_label = f"Enter {exchange} Ticker:"
            ticker_placeholder = "e.g., RELIANCE" if exchange == "NSE" else "e.g., RELIANCE"
            ticker = st.text_input(ticker_label, placeholder=ticker_placeholder)
        
            # Reset analysis state when ticker changes
            if ticker and ticker != st.session_state.get('current_ticker'):
                st.session_state.current_ticker = ticker
                st.session_state.show_results_listed = False
        
            # Historical and projection years
            st.markdown("**‚è±Ô∏è Time Periods**")
            hist_col1, hist_col2 = st.columns(2)
            with hist_col1:
                historical_years_listed = st.number_input(
                    "Historical Years",
                    min_value=2,
                    max_value=10,
                    value=3,
                    step=1,
                    help="Years of historical data to use for calculations"
                )
            with hist_col2:
                projection_years_listed = st.number_input(
                    "Projection Years",
                    min_value=1,
                    max_value=15,
                    value=5,
                    step=1,
                    help="Years to project into future"
                )
        
            st.markdown("---")
            st.markdown("**üìä Peer Companies (Both Exchanges)**")
        
            # Normalize ticker key (remove suffix for consistent storage)
            ticker_key = ticker.replace('.NS', '').replace('.BO', '') if ticker else ""
        
            # Initialize session state for fetch status
            if 'fetch_status' not in st.session_state:
                st.session_state.fetch_status = {}
            if 'fetched_nse_peers' not in st.session_state:
                st.session_state.fetched_nse_peers = {}
            if 'fetched_bse_peers' not in st.session_state:
                st.session_state.fetched_bse_peers = {}
        
            # Show last fetch status if available
            if ticker_key in st.session_state.fetch_status:
                status = st.session_state.fetch_status[ticker_key]
                if status['success']:
                    st.success(f"‚úÖ Last fetch: {status['nse_count']} NSE peers, {status['bse_count']} BSE peers")
                    with st.expander("üëÅÔ∏è View Peer Details"):
                        st.markdown("**NSE Peers:**")
                        # Use text_area to avoid truncation
                        nse_display = status.get('nse_display', ", ".join(status.get('nse_peers', [])))
                        if nse_display:
                            st.text_area("NSE Tickers (comma-separated)", nse_display, height=100, disabled=True, label_visibility="collapsed")
                        else:
                            st.info("None")
                    
                        st.markdown("**BSE Peers:**")
                        # Use text_area to avoid truncation
                        bse_display = status.get('bse_display', ", ".join(status.get('bse_peers', [])))
                        if bse_display:
                            st.text_area("BSE Tickers (comma-separated)", bse_display, height=100, disabled=True, label_visibility="collapsed")
                        else:
                            st.info("None")
                else:
                    st.error(f"‚ùå Last fetch failed: {status.get('error', 'Unknown error')}")
        
            # Auto-fetch peers button
            if ticker_key:
                if PEER_FETCHER_AVAILABLE:
                    if st.button("üîç Auto-Fetch Peers", help="Fetch industry peers and auto-detect their exchanges"):
                        with st.spinner("üîç Fetching and analyzing peers..."):
                            try:
                                auto_peers = []
                            
                                # Try with the actual ticker as entered (with exchange if provided)
                                full_ticker = ticker if ('.NS' in ticker or '.BO' in ticker) else f"{ticker_key}.{exchange_suffix}"
                            
                                st.info(f"üîç Fetching peers for: {full_ticker}")
                            
                                try:
                                    auto_peers = get_industry_peers(full_ticker, max_peers=20, exclude_self=True)
                                except Exception as e:
                                    st.warning(f"Peer fetcher error: {str(e)}")
                            
                                if auto_peers and len(auto_peers) > 0:
                                    st.success(f"‚úÖ Found {len(auto_peers)} potential peers")
                                    with st.expander("üìã View fetched peer tickers"):
                                        st.write(auto_peers)
                                
                                    nse_list = []
                                    bse_list = []
                                
                                    # Progress tracking
                                    progress_bar = st.progress(0)
                                    status_text = st.empty()
                                    total = len(auto_peers)
                                
                                    for idx, peer in enumerate(auto_peers):
                                        status_text.text(f"Checking {peer}... ({idx+1}/{total})")
                                        progress_bar.progress((idx + 1) / total)
                                    
                                        # Add delay between peers to prevent rate limiting
                                        if idx > 0:
                                            time.sleep(random.uniform(1.5, 2.5))  # Reduced from 2-3s since we have caching
                                    
                                        # ALWAYS check BOTH exchanges for each peer
                                        found_on_nse = False
                                        found_on_bse = False
                                    
                                        # Check NSE first (priority)
                                        try:
                                            nse_test = get_cached_ticker(f"{peer}.NS")
                                            nse_info = nse_test.info
                                            if nse_info and len(nse_info) > 5:
                                                if any(key in nse_info for key in ['regularMarketPrice', 'currentPrice', 'previousClose']):
                                                    found_on_nse = True
                                        except:
                                            pass
                                    
                                        # Only check BSE if NSE not found (optimization)
                                        if not found_on_nse:
                                            # Small delay before BSE check
                                            time.sleep(random.uniform(0.3, 0.7))
                                        
                                            # Check BSE
                                            try:
                                                bse_test = get_cached_ticker(f"{peer}.BO")
                                                bse_info = bse_test.info
                                                if bse_info and len(bse_info) > 5:
                                                    if any(key in bse_info for key in ['regularMarketPrice', 'currentPrice', 'previousClose']):
                                                        found_on_bse = True
                                            except:
                                                pass
                                    
                                        # NO DUPLICATES - NSE gets priority
                                        if found_on_nse:
                                            nse_list.append(peer)
                                        elif found_on_bse:
                                            bse_list.append(peer)
                                        else:
                                            nse_list.append(peer)
                                
                                    progress_bar.empty()
                                    status_text.empty()
                                
                                    # Store using normalized ticker key
                                    st.session_state.fetched_nse_peers[ticker_key] = ",".join(nse_list)
                                    st.session_state.fetched_bse_peers[ticker_key] = ",".join(bse_list)
                                
                                    # Store detailed status for display
                                    st.session_state.fetch_status[ticker_key] = {
                                        'success': True,
                                        'nse_count': len(nse_list),
                                        'bse_count': len(bse_list),
                                        'nse_peers': nse_list,
                                        'bse_peers': bse_list,
                                        'nse_display': ", ".join(nse_list),
                                        'bse_display': ", ".join(bse_list)
                                    }
                                
                                    # No st.rerun() - let Streamlit handle the update naturally
                                    st.success(f"‚úÖ Peers fetched successfully! NSE: {len(nse_list)}, BSE: {len(bse_list)}")
                                    st.info("üí° Peer tickers have been populated below. Click 'Fetch & Analyze' to proceed.")
                                else:
                                    st.session_state.fetch_status[ticker_key] = {
                                        'success': False,
                                        'error': 'No peers found - company may not have comparable peers or ticker is invalid'
                                    }
                                    st.warning("‚ö†Ô∏è No peers found for this ticker")
                            except Exception as e:
                                st.session_state.fetch_status[ticker_key] = {
                                    'success': False,
                                    'error': str(e)
                                }
                                st.error(f"‚ùå Error: {str(e)}")
                                import traceback
                                with st.expander("üîç Show error details"):
                                    st.code(traceback.format_exc())
                else:
                    st.info("üí° Auto-fetch not available - utils_peer_fetcher module not found. Enter peers manually below.")
        
            # Get stored peers using normalized key - this auto-populates the text boxes!
            stored_nse = st.session_state.fetched_nse_peers.get(ticker_key, "") if ticker_key else ""
            stored_bse = st.session_state.fetched_bse_peers.get(ticker_key, "") if ticker_key else ""
        
            # NSE Peers Box - AUTO-POPULATED from session state
            nse_peers = st.text_input(
                "NSE Peer Tickers (comma-separated):",
                value=stored_nse,  # ‚Üê This auto-fills from session state!
                placeholder="e.g., TCS, INFY, WIPRO",
                key='nse_peers_listed',
                help="üîç Click Auto-Fetch above to populate automatically, or enter manually"
            )
        
            # BSE Peers Box - AUTO-POPULATED from session state
            bse_peers = st.text_input(
                "BSE Peer Tickers (comma-separated):",
                value=stored_bse,  # ‚Üê This auto-fills from session state!
                placeholder="e.g., SUNPHARMA, TATAMOTORS",
                key='bse_peers_listed',
                help="Enter BSE-listed peer companies"
            )
        
            # Combine peers with their exchange suffixes
            comp_tickers_listed = ""
            if nse_peers.strip():
                nse_list = [f"{t.strip()}.NS" if '.NS' not in t and '.BO' not in t else t.strip() for t in nse_peers.split(',') if t.strip()]
                comp_tickers_listed = ",".join(nse_list)
            if bse_peers.strip():
                bse_list = [f"{t.strip()}.BO" if '.NS' not in t and '.BO' not in t else t.strip() for t in bse_peers.split(',') if t.strip()]
                if comp_tickers_listed:
                    comp_tickers_listed += "," + ",".join(bse_list)
                else:
                    comp_tickers_listed = ",".join(bse_list)
    
        with col2:
            tax_rate = st.number_input("Tax Rate (%):", min_value=0.0, max_value=100.0, value=25.0, step=0.5, key='listed_tax')
            terminal_growth = st.number_input("Terminal Growth Rate (%):", min_value=0.0, max_value=10.0, value=4.0, step=0.5, key='listed_tg')
            
            # Risk-free rate override
            st.markdown("**üèõÔ∏è Risk-Free Rate (G-Sec 10Y)**")
            
            # CRITICAL FIX: Use session state value directly, not a separate variable
            # This ensures the fetched value is actually used
            manual_rf_rate = st.number_input(
                f"Risk-Free Rate (%)",
                min_value=0.0,
                max_value=20.0,
                value=st.session_state.get('cached_rf_rate_listed', 6.83),
                step=0.1,
                key='manual_rf_listed',
                help="Auto-fetched from ticker above. You can manually edit this value."
            )
            
            # Update session state if user manually changes it
            if abs(manual_rf_rate - st.session_state.get('cached_rf_rate_listed', 6.83)) > 0.05:
                st.session_state.cached_rf_rate_listed = manual_rf_rate
                st.info(f"üí° Using custom rate: {manual_rf_rate:.2f}%")
        
            # Manual discount rate override
            st.markdown("**üí∞ Discount Rate Override (Optional)**")
            manual_discount_rate = st.number_input(
                "Manual Discount Rate (%)",
                min_value=0.0,
                max_value=50.0,
                value=0.0,
                step=0.5,
                key='manual_discount_listed',
                help="‚ö†Ô∏è Override WACC calculation. Leave at 0 to use auto-calculated WACC. Use this if you want to use a specific discount rate instead of WACC."
            )
            if manual_discount_rate > 0:
                st.info(f"üí° Using manual discount rate: {manual_discount_rate:.2f}% (Overriding WACC)")
    
        with st.expander("‚öôÔ∏è Advanced Projection Assumptions - FULL CONTROL"):
            st.info("üí° **Complete Control:** Override ANY projection parameter below. Leave at 0 or blank for auto-calculation from historical data.")
        
            st.markdown("### üìä Shares Outstanding Override")
            st.caption("For newly listed companies or when Yahoo Finance fails to fetch shares")
            manual_shares_override = st.number_input(
                "Manual Shares Outstanding (Absolute Number)",
                min_value=0,
                value=0,
                step=1000000,
                format="%d",
                help="‚ö†Ô∏è Use ONLY if Yahoo Finance fails to fetch shares. Enter absolute number (e.g., 50000000 for 5 crore shares). Leave at 0 to use auto-fetched value.",
                key="manual_shares_listed"
            )
            if manual_shares_override > 0:
                st.success(f"‚úÖ Manual override active: **{manual_shares_override:,}** shares ({manual_shares_override/10000000:.2f} Crore)")
        
            st.markdown("---")
            st.markdown("### üìä Revenue & Growth")
            col1, col2, col3 = st.columns(3)
            with col1:
                rev_growth_override_listed = st.number_input(
                    "Revenue Growth (%/year)", 
                    min_value=0.0, max_value=100.0, value=0.0, step=0.5,
                    key='listed_rev_growth',
                    help="0 = Auto from historical CAGR"
                )
            with col2:
                opex_margin_override_listed = st.number_input(
                    "Operating Expense Margin (%)", 
                    min_value=0.0, max_value=100.0, value=0.0, step=0.5,
                    key='listed_opex_margin',
                    help="0 = Auto from historical average"
                )
            with col3:
                ebitda_margin_override = st.number_input(
                    "EBITDA Margin (%)", 
                    min_value=0.0, max_value=100.0, value=0.0, step=0.5,
                    key='listed_ebitda',
                    help="0 = Calculated as Revenue - OpEx"
                )
        
            st.markdown("### üèóÔ∏è CapEx & Depreciation")
            col4, col5, col6 = st.columns(3)
            with col4:
                capex_ratio_override_listed = st.number_input(
                    "CapEx/Revenue (%)", 
                    min_value=0.0, max_value=50.0, value=0.0, step=0.5,
                    key='listed_capex_ratio',
                    help="0 = Auto from historical average"
                )
            with col5:
                depreciation_rate_override = st.number_input(
                    "Depreciation Rate (%)", 
                    min_value=0.0, max_value=30.0, value=0.0, step=0.5,
                    key='listed_dep_rate',
                    help="0 = Auto calculated"
                )
            with col6:
                depreciation_method = st.selectbox(
                    "Depreciation Method",
                    ["Auto", "% of Fixed Assets", "% of Revenue", "Absolute Value"],
                    key='listed_dep_method'
                )
        
            st.markdown("### üí∞ Working Capital Management")
            col7, col8, col9 = st.columns(3)
            with col7:
                inventory_days_override = st.number_input(
                    "Inventory Days", 
                    min_value=0.0, max_value=365.0, value=0.0, step=1.0,
                    key='listed_inv_days',
                    help="0 = Auto from historical average"
                )
            with col8:
                debtor_days_override = st.number_input(
                    "Debtor/Receivables Days", 
                    min_value=0.0, max_value=365.0, value=0.0, step=1.0,
                    key='listed_deb_days',
                    help="0 = Auto from historical average"
                )
            with col9:
                creditor_days_override = st.number_input(
                    "Creditor/Payables Days", 
                    min_value=0.0, max_value=365.0, value=0.0, step=1.0,
                    key='listed_cred_days',
                    help="0 = Auto from historical average"
                )
        
            st.markdown("### üìà Tax & Interest")
            col10, col11, col12 = st.columns(3)
            with col10:
                interest_rate_override = st.number_input(
                    "Interest Rate (%)", 
                    min_value=0.0, max_value=30.0, value=0.0, step=0.25,
                    key='listed_interest',
                    help="0 = Auto calculated from Debt"
                )
            with col11:
                tax_shield = st.checkbox(
                    "Include Tax Shield",
                    value=True,
                    key='listed_tax_shield',
                    help="Apply tax benefit on interest expense"
                )
            with col12:
                working_capital_as_pct_revenue = st.number_input(
                    "Working Capital % of Revenue", 
                    min_value=0.0, max_value=50.0, value=0.0, step=0.5,
                    key='listed_wc_pct',
                    help="0 = Calculate from Inv+Deb-Cred days"
                )
    
        # ================================
        # BANK/NBFC PARAMETERS (BEFORE FETCH)
        # ================================
        with st.expander("üè¶ Bank/NBFC Valuation Parameters (Adjust if company is a Bank/NBFC)"):
            st.info("üí° **Note:** These parameters are used ONLY if the company is classified as a Bank/NBFC/Interest-Dominant entity after analysis. Adjust them before clicking 'Fetch & Analyze'.")
        
            col1, col2, col3, col4 = st.columns(4)
        
            with col1:
                terminal_growth_bank = st.number_input(
                    "Terminal Growth (%)",
                    min_value=0.0,
                    max_value=10.0,
                    value=3.5,
                    step=0.5,
                    key='terminal_growth_bank_pre',
                    help="Long-term perpetual growth rate for banks"
                )
        
            with col2:
                projection_years_bank = st.number_input(
                    "Projection Years",
                    min_value=3,
                    max_value=15,
                    value=5,
                    step=1,
                    key='projection_years_bank_pre',
                    help="Years to project for bank valuations"
                )
        
            with col3:
                assumed_roe_bank = st.number_input(
                    "Assumed ROE (%)",
                    min_value=5.0,
                    max_value=30.0,
                    value=15.0,
                    step=1.0,
                    key='assumed_roe_bank_pre',
                    help="Return on Equity assumption (15 = auto-calculated)"
                )
        
            with col4:
                cost_of_equity_override = st.number_input(
                    "Cost of Equity (%)",
                    min_value=0.0,
                    max_value=25.0,
                    value=0.0,
                    step=0.5,
                    key='cost_of_equity_override_pre',
                    help="Override Ke (0 = use calculated CAPM)"
                )
        
            col5, col6, col7 = st.columns(3)
        
            with col5:
                div_growth_bank = st.number_input(
                    "Dividend Growth (%)",
                    min_value=-10.0,
                    max_value=20.0,
                    value=0.0,
                    step=1.0,
                    key='div_growth_bank_pre',
                    help="0 = use calculated from history"
                )
        
            with col6:
                payout_ratio_bank = st.number_input(
                    "Payout Ratio (%)",
                    min_value=0.0,
                    max_value=90.0,
                    value=0.0,
                    step=5.0,
                    key='payout_ratio_bank_pre',
                    help="0 = use calculated from history"
                )
        
            with col7:
                st.metric("Info", "0 = Auto", help="Setting 0 uses auto-calculated values from historical data")
    
        # ================================
        # DDM & RIM PARAMETERS FOR NON-BANKING COMPANIES
        # ================================
        with st.expander("üíé DDM & RIM Parameters for Non-Banking Companies"):
            st.info("üí° **Note:** These parameters apply to Dividend Discount Model and Residual Income Model for non-banking companies. Leave at 0 for auto-calculation from actual data.")
        
            st.markdown("### üí∞ Dividend Discount Model (DDM) Parameters")
            col1, col2 = st.columns(2)
        
            with col1:
                ddm_dividend_growth_override = st.number_input(
                    "Dividend Growth Rate (%)",
                    min_value=0.0,
                    max_value=30.0,
                    value=0.0,
                    step=0.5,
                    key='ddm_div_growth',
                    help="0 = Auto-calculate from historical dividend data. Used in Gordon Growth Model."
                )
        
            with col2:
                ddm_payout_ratio_override = st.number_input(
                    "Payout Ratio (%)",
                    min_value=0.0,
                    max_value=100.0,
                    value=0.0,
                    step=5.0,
                    key='ddm_payout',
                    help="0 = Auto-calculate from historical data. % of earnings paid as dividends."
                )
        
            st.markdown("### üè¢ Residual Income Model (RIM) Parameters")
            col3, col4, col5 = st.columns(3)
        
            with col3:
                rim_assumed_roe_override = st.number_input(
                    "Assumed ROE (%)",
                    min_value=0.0,
                    max_value=50.0,
                    value=0.0,
                    step=1.0,
                    key='rim_roe',
                    help="0 = Auto-calculate from historical data. Return on Equity assumption."
                )
        
            with col4:
                rim_terminal_growth_override = st.number_input(
                    "Terminal Growth (%)",
                    min_value=0.0,
                    max_value=10.0,
                    value=0.0,
                    step=0.5,
                    key='rim_terminal_growth',
                    help="0 = Use same as DCF terminal growth. Long-term perpetual growth rate."
                )
        
            with col5:
                rim_projection_years_override = st.number_input(
                    "Projection Years",
                    min_value=0,
                    max_value=15,
                    value=0,
                    step=1,
                    key='rim_proj_years',
                    help="0 = Use same as DCF projection years. Number of years to project."
                )
    
        if ticker:
            # Check if we're approaching rate limits
            if st.session_state.yahoo_request_count > 15:
                st.warning(f"""
                ‚ö†Ô∏è **High API Usage Warning**
            
                You've made {st.session_state.yahoo_request_count} Yahoo Finance requests this hour.
                You may encounter rate limits soon.
            
                **Recommendations:**
                - Wait a few minutes between requests
                - Use cached results when possible
                - Consider using Screener.in as alternative data source (checkbox below)
                """)
        
            # DATA SOURCE SELECTION CHECKBOXES
            st.markdown("---")
            st.markdown("### üìä Data Source Options")
        
            col_src1, col_src2 = st.columns(2)
        
            with col_src1:
                use_screener_data = st.checkbox(
                    "üåê Use Screener.in for company data",
                    value=False,
                    key='use_screener_data',
                    help="Fetch financial data from Screener.in instead of Yahoo Finance. Useful when Yahoo Finance is rate-limited or data is missing."
                )
                if use_screener_data:
                    st.info("üìå Company financials will be fetched from Screener.in (Indian stock screener)")
        
            with col_src2:
                use_screener_peers = st.checkbox(
                    "üîó Use Screener.in for peer data",
                    value=False,
                    key='use_screener_peers',
                    help="Fetch peer company data from Screener.in for comparative valuation. Works independently of main data source."
                )
                if use_screener_peers:
                    st.info("üìå Peer comparison data will be fetched from Screener.in")
        
            st.markdown("---")
            
            # Stock Price Comparison Feature Toggle - compact, before fetch button
            enable_stock_comparison = st.checkbox(
                "üìà Show Stock Price vs Revenue & EPS chart",
                value=False,
                key='enable_stock_comparison_listed',
                help="Compare stock price with financials (max 4 years)"
            )
        
            # INPUT CHANGE DETECTION - Reset results if key inputs change
            current_inputs_listed = {
                'ticker': ticker,
                'projection_years': projection_years_listed,
                'terminal_growth': terminal_growth,
                'tax_rate': tax_rate,
                'manual_discount': manual_discount_rate if manual_discount_rate > 0 else None
            }
            
            # Check if inputs changed
            if 'previous_inputs_listed' in st.session_state:
                if st.session_state.previous_inputs_listed != current_inputs_listed:
                    # Inputs changed - clear results
                    st.session_state.show_results_listed = False
            
            # Store current inputs
            st.session_state.previous_inputs_listed = current_inputs_listed
        
            st.markdown("---")
            st.markdown("### üéØ Ready to Run Valuation")
            st.info("üí° **Click the button below to run valuation.** Results will appear only after clicking.")
            
            col_fetch1, col_fetch2, col_fetch3 = st.columns([2, 1, 1])
        
            with col_fetch1:
                if st.button("üöÄ Fetch & Analyze", type="primary", key="fetch_analyze_listed"):
                    st.session_state.show_results_listed = True
        
            with col_fetch2:
                if st.button("üóëÔ∏è Clear Results", help="Clear current valuation results"):
                    st.session_state.show_results_listed = False
                    st.success("‚úÖ Results cleared!")
                    st.rerun()
            
            with col_fetch3:
                if st.button("üîÑ Reset Cache", help="Clear cached data to force fresh fetch"):
                    st.cache_data.clear()
                    st.success("‚úÖ Cache cleared!")
                    st.rerun()
        
            if st.session_state.get('show_results_listed', False):
                # Determine data source based on checkbox
                use_screener = st.session_state.get('use_screener_data', False)
            
            
                if use_screener:
                    # SCREENER.IN DATA SOURCE
                    with st.spinner("üåê Fetching data from Screener.in..."):
                        try:
                            # Remove exchange suffix for Screener.in
                            ticker_clean = ticker.replace('.NS', '').replace('.BO', '')
                        
                            screener_data = fetch_screener_financials(ticker_clean, num_years=5)
                        
                            if screener_data and screener_data.get('financials'):
                                st.success(f"‚úÖ Data fetched from Screener.in for {screener_data.get('company_name', ticker_clean)}")
                            
                                # Try to get current price from Screener
                                current_price = screener_data.get('current_price', 0)
                                if current_price == 0:
                                    st.info("üí° Fetching current price from Yahoo Finance...")
                                    try:
                                        import yfinance as yf
                                        yf_ticker = yf.Ticker(ticker)
                                        current_price = yf_ticker.info.get('currentPrice', 0) or yf_ticker.info.get('regularMarketPrice', 0)
                                        if current_price > 0:
                                            st.success(f"‚úÖ Current Price: ‚Çπ{current_price:.2f} (Yahoo Finance)")
                                    except:
                                        st.warning("‚ö†Ô∏è Could not fetch current price - will use 0")
                            
                                # Convert to Yahoo Finance compatible format
                                yahoo_data = {
                                    'info': {
                                        'longName': screener_data.get('company_name', ticker_clean),
                                        'currentPrice': current_price,
                                        'sharesOutstanding': screener_data.get('shares', 0)
                                    },
                                    'income_statement': pd.DataFrame(),
                                    'balance_sheet': pd.DataFrame(),
                                    'cash_flow': pd.DataFrame(),
                                    'shares': screener_data.get('shares', 0),
                                    'shares_source': 'Screener.in (derived from EPS)',
                                    '_screener_financials': screener_data['financials'],  # Stash for extraction
                                    '_data_source': 'screener'
                                }
                            
                                shares = screener_data.get('shares', 0)
                                shares_source = 'Screener.in (EPS-based)'
                                company_name = screener_data.get('company_name', ticker_clean)
                                error = None
                            
                                # BUGFIX: Validate shares from Screener and provide feedback
                                if shares > 0:
                                    st.success(f"‚úÖ Shares Outstanding fetched: **{shares:,}** from Screener.in (EPS-based)")
                                else:
                                    st.warning("‚ö†Ô∏è Screener.in returned shares outstanding = 0. This may be due to missing or incomplete EPS data.")
                                    st.info("üí° **Troubleshooting:**\n- The stock may not have sufficient historical data on Screener.in\n- Try using manual override above\n- Or uncheck 'Use Screener.in' to fetch from Yahoo Finance")
                            
                            else:
                                st.error("‚ùå Could not fetch data from Screener.in. Please check ticker symbol or try Yahoo Finance.")
                                st.stop()
                    
                        except Exception as e:
                            st.error(f"‚ùå Error fetching from Screener.in: {str(e)}")
                            st.info("üí° Try using Yahoo Finance instead (uncheck the Screener.in option)")
                            st.stop()
            
                if not use_screener:
                    # YAHOO FINANCE DATA SOURCE (DEFAULT)
                    with st.spinner("üìä Fetching data from Yahoo Finance..."):
                        yahoo_data, error = fetch_yahoo_financials(ticker, exchange_suffix)
                    
                        if error:
                            st.error(error)
                            st.info("üí° **Suggestion:** Try checking the 'Use Screener.in' option above to fetch from an alternative data source")
                            st.stop()
                    
                        shares = yahoo_data.get('shares', 0)
                        shares_source = yahoo_data.get('shares_source', 'Unknown')
                        company_name = yahoo_data['info'].get('longName', ticker)
                        yahoo_data['_data_source'] = 'yahoo'
                    
                        # BUGFIX: Add validation and fallback for shares
                        if shares > 0:
                            st.success(f"‚úÖ Shares Outstanding fetched: **{shares:,}** from {shares_source}")
                        else:
                            st.warning(f"‚ö†Ô∏è Shares outstanding = 0. Source attempted: {shares_source}")
                            # Try fallback calculation using Market Cap / Price
                            info = yahoo_data.get('info', {})
                            if 'marketCap' in info and 'currentPrice' in info:
                                market_cap = info.get('marketCap', 0)
                                current_price = info.get('currentPrice', 0)
                                if market_cap > 0 and current_price > 0:
                                    shares = int(market_cap / current_price)
                                    shares_source = "Calculated (Market Cap √∑ Current Price)"
                                    st.info(f"üîÑ Recalculated shares: **{shares:,}** using {shares_source}")
            
                # Common processing for both Screener.in and Yahoo Finance
                # Apply manual override if provided
                if manual_shares_override > 0:
                    shares = manual_shares_override
                    shares_source = "Manual Override (User Input)"
                    st.warning(f"‚ö†Ô∏è Using manual shares override: **{shares:,}** shares")
                elif shares == 0:
                    st.error("‚ùå **CRITICAL:** Could not fetch shares outstanding!")
                    st.warning("üîß **Action Required:** Please enter shares outstanding manually using the override field above.")
                    st.info("üí° **Where to find shares:**\n- Company's latest annual report\n- BSE/NSE company page\n- Screener.in or Moneycontrol")
                    st.stop()
            
                st.success(f"‚úÖ Loaded data for **{company_name}**")
            
                # Show shares with source
                shares_in_crore = shares / 10000000
                col_sh1, col_sh2 = st.columns(2)
                with col_sh1:
                    st.metric("üìä Shares Outstanding", f"{shares:,}", help=f"Source: {shares_source}")
                with col_sh2:
                    st.metric("üìä Shares (Crore)", f"{shares_in_crore:.2f} Cr")
            
                if shares_source != "Manual Override (User Input)" and shares_source != "Direct (sharesOutstanding)":
                    st.caption(f"‚ÑπÔ∏è Shares source: {shares_source}")
            
                # Extract financials with user-selected historical years
                financials = extract_financials_listed(yahoo_data, num_years=historical_years_listed)
            
                if financials is None:
                    st.error("Failed to extract financial data")
                    st.stop()
            
                # ================================
                # BUSINESS MODEL CLASSIFICATION (RULEBOOK SECTION 2)
                # ================================
                st.markdown("---")
                st.subheader("üè¢ Business Model Classification")
                
                classification = classify_business_model(
                    financials, 
                    income_stmt=yahoo_data['income_statement'], 
                    balance_sheet=yahoo_data['balance_sheet']
                )
                
                # Show classification and check if FCFF DCF is allowed
                should_stop = show_classification_warning(classification)
                
                if should_stop:
                    # For Interest-Dominant entities, run alternative valuation methods
                    st.markdown("---")
                    st.header("üè¶ Bank/NBFC Valuation Methods")
                    st.info("Running alternative valuation methods appropriate for interest-dominant entities...")
                
                    # Initialize session state for bank parameters if not exists
                    if 'bank_params_applied' not in st.session_state:
                        st.session_state.bank_params_applied = {
                            'terminal_growth_bank': 3.5,
                            'projection_years_bank': 5,
                            'assumed_roe_bank': 15.0,
                            'cost_of_equity_override': 0.0,
                            'div_growth_bank': 0.0,
                            'payout_ratio_bank': 0.0
                        }
                
                    # Use parameters set BEFORE Fetch button (from the expandable section)
                    # No need for Apply Changes button - parameters are already set!
                
                    # Get current price
                    stock = get_cached_ticker(get_ticker_with_exchange(ticker, exchange_suffix))
                    info = stock.info
                    current_price = info.get('currentPrice', 0)
                
                    # Calculate Ke for bank valuations
                    wacc_details = calculate_wacc(financials, tax_rate, peer_tickers=None, manual_rf_rate=manual_rf_rate)
                    beta = get_stock_beta(get_ticker_with_exchange(ticker, exchange_suffix), period_years=3)
                    wacc_details['beta'] = beta
                    wacc_details['ke'] = wacc_details['rf'] + (beta * (wacc_details['rm'] - wacc_details['rf']))
                
                    # Use override or calculated Ke
                    if cost_of_equity_override > 0:
                        cost_of_equity = cost_of_equity_override
                    else:
                        cost_of_equity = wacc_details['ke']
                
                    # Run all bank valuation methods with user parameters
                    ri_model = calculate_residual_income_model(
                        financials, shares, cost_of_equity,
                        terminal_growth=terminal_growth_bank,
                        projection_years=projection_years_bank,
                        assumed_roe=assumed_roe_bank if assumed_roe_bank != 15 else None
                    )
                
                    ddm_model = calculate_dividend_discount_model(
                        financials, shares, cost_of_equity,
                        ticker=ticker,
                        div_growth_override=div_growth_bank if div_growth_bank != 0 else None,
                        payout_ratio_override=payout_ratio_bank if payout_ratio_bank != 0 else None
                    )
                
                    pb_roe_model = calculate_pb_roe_valuation(
                        financials, shares, cost_of_equity,
                        assumed_roe=assumed_roe_bank if assumed_roe_bank != 15 else None
                    )
                
                    rel_val = calculate_relative_valuation(ticker, financials, shares, peer_tickers=comp_tickers_listed, exchange_suffix=exchange_suffix)
                
                    # ============================================
                    # BANK DCF OPTION - Using NII as Revenue
                    # ============================================
                    st.markdown("---")
                    st.markdown("### üè¶ **DCF for Banks (Experimental)**")
                    st.info("üí° **New Feature:** Calculate DCF for banks by treating Net Interest Income (NII) as revenue and interest expense as operating cost.")
                
                    bank_dcf_result = None
                    with st.expander("üîß Enable DCF for Banks", expanded=False):
                        st.markdown("""
                        **How Bank DCF Works:**
                        - ‚úÖ Revenue = Net Interest Income (Interest Income - Interest Expense)
                        - ‚úÖ Operating Expenses = Non-Interest Expenses
                        - ‚úÖ WACC = Bank-specific calculation (Cost of Funds methodology)
                        - ‚úÖ FCFF = NII - OpEx - Taxes (no CapEx/WC for banks)
                        """)
                    
                        if st.checkbox("Calculate Bank DCF", key="enable_bank_dcf"):
                            try:
                                st.info("üè¶ Calculating Bank FCFE Valuation (Equity DCF)...")
                            
                                # Get Ke (Cost of Equity) - NOT WACC!
                                # Calculate using CAPM
                                wacc_details_temp = calculate_wacc(financials, tax_rate, peer_tickers=comp_tickers_listed)
                                ke = wacc_details_temp['ke']  # Use only Ke, ignore WACC
                            
                                st.info(f"üí° Using **Ke = {ke:.2f}%** (NOT WACC - banks use equity DCF)")
                            
                                # Advanced bank parameters
                                with st.expander("üîß Bank DCF Parameters"):
                                    col_bank1, col_bank2 = st.columns(2)
                                    with col_bank1:
                                        car_ratio = st.number_input(
                                            "Capital Adequacy Ratio (CAR) %",
                                            min_value=10.0, max_value=20.0, value=14.0, step=0.5,
                                            help="Target CAR - usually 13-15% in India"
                                        )
                                    with col_bank2:
                                        rwa_pct = st.number_input(
                                            "Risk Weight %",
                                            min_value=50.0, max_value=100.0, value=75.0, step=5.0,
                                            help="RWA as % of advances - usually 70-80%"
                                        )
                            
                                # Use bank-specific projections (FCFE)
                                projections_bank, drivers_bank = project_financials_bank(
                                    financials, 
                                    projection_years_bank, 
                                    tax_rate,
                                    car_ratio=car_ratio,
                                    rwa_percentage=rwa_pct
                                )
                            
                                if projections_bank:
                                    # Use bank FCFE valuation (NOT regular DCF!)
                                    bank_dcf_result, error_bank = calculate_bank_fcfe_valuation(
                                        projections_bank,
                                        ke,  # Cost of Equity, NOT WACC
                                        terminal_growth_bank,
                                        shares
                                    )
                                
                                    if error_bank:
                                        st.error(f"‚ùå Bank FCFE Error: {error_bank}")
                                    elif bank_dcf_result:
                                        st.success(f"‚úÖ **Bank FCFE Fair Value:** ‚Çπ{bank_dcf_result['fair_value_per_share']:.2f}")
                                    
                                        # Display key metrics
                                        col_dcf1, col_dcf2, col_dcf3, col_dcf4 = st.columns(4)
                                        with col_dcf1:
                                            st.metric("FCFE Fair Value", f"‚Çπ{bank_dcf_result['fair_value_per_share']:.2f}")
                                        with col_dcf2:
                                            st.metric("Cost of Equity (Ke)", f"{ke:.2f}%")
                                        with col_dcf3:
                                            upside = ((bank_dcf_result['fair_value_per_share'] - current_price) / current_price * 100) if current_price > 0 else 0
                                            st.metric("Upside/Downside", f"{upside:+.1f}%")
                                        with col_dcf4:
                                            st.metric("Sustainable Growth", f"{bank_dcf_result.get('sustainable_growth', 0):.2f}%")
                                    
                                        # Show projection drivers
                                        st.markdown("---")
                                        st.markdown("**üìä Bank FCFE Model Inputs:**")
                                        col_input1, col_input2 = st.columns(2)
                                        with col_input1:
                                            st.write(f"‚Ä¢ **Revenue Growth:** {drivers_bank['revenue_growth']:.2f}%")
                                            st.write(f"‚Ä¢ **ROE:** {drivers_bank['roe']:.2f}%")
                                            st.write(f"‚Ä¢ **CAR Target:** {drivers_bank['car_ratio']:.1f}%")
                                        with col_input2:
                                            st.write(f"‚Ä¢ **RWA Weight:** {drivers_bank['rwa_percentage']:.1f}%")
                                            st.write(f"‚Ä¢ **Tax Rate:** {drivers_bank['tax_rate']:.1f}%")
                                            st.write(f"‚Ä¢ **Terminal Growth:** {terminal_growth_bank:.2f}%")
                                    
                                        st.info("""
                                        üí° **FCFE Formula Used:**
                                    
                                        FCFE = PAT - (Growth in Advances √ó RWA% √ó CAR%)
                                    
                                        - Discounted at **Ke** (not WACC)
                                        - Values **Equity directly** (not Enterprise Value)
                                        - No debt adjustment (debt is raw material)
                                        """)
                                else:
                                    st.error("‚ùå Could not generate bank projections")
                                
                            except Exception as e:
                                st.error(f"‚ùå Bank DCF calculation failed: {str(e)}")
                                import traceback
                                with st.expander("üîç Error Details"):
                                    st.code(traceback.format_exc())
                
                    # Display results
                    st.markdown("---")
                    st.success("‚úÖ Bank Valuation Complete!")
                
                    # Current Price & Fair Value Display
                    col_price1, col_price2 = st.columns(2)
                    with col_price1:
                        st.metric("üìä Current Market Price", f"‚Çπ {current_price:.2f}")
                
                    # Collect all fair values (initialize list first!)
                    fair_values = []
                    if ri_model:
                        fair_values.append(ri_model['value_per_share'])
                    if ddm_model:
                        fair_values.append(ddm_model['value_per_share'])
                    if pb_roe_model:
                        fair_values.append(pb_roe_model['value_per_share'])
                    if rel_val:
                        fair_values.append(rel_val['avg_fair_value'])
                    if bank_dcf_result:
                        fair_values.append(bank_dcf_result['fair_value_per_share'])
                
                    avg_fair_value = np.mean(fair_values) if fair_values else 0
                
                    with col_price2:
                        st.metric("üéØ Average Fair Value", f"‚Çπ {avg_fair_value:.2f}",
                                 delta=f"{((avg_fair_value - current_price) / current_price * 100):.1f}%")
                
                    # Price vs Value Gauge
                    if avg_fair_value > 0:
                        st.plotly_chart(create_price_vs_value_gauge(current_price, avg_fair_value), 
                                      use_container_width=True)
                
                    # Valuation Methods Tabs
                    tabs_list = [
                        "üìä Summary", 
                        "üè¢ Residual Income", 
                        "üí∞ Dividend Discount",
                        "üìà P/B with ROE",
                        "üîÑ Relative Valuation"
                    ]
                
                    # Add Stock Comparison tab if enabled
                    stock_comp_data = None
                    enable_stock_comparison_state = st.session_state.get('enable_stock_comparison_listed', False)
                    
                    if enable_stock_comparison_state and STOCK_COMPARISON_AVAILABLE:
                        tabs_list.append("üìà Stock vs Financials")
                
                    # Add Bank DCF tab if calculated
                    if bank_dcf_result:
                        tabs_list.append("üè¶ Bank DCF")
                    
                    # Create tabs dynamically
                    tabs = st.tabs(tabs_list)
                    tab_idx = 0
                    tab1 = tabs[tab_idx]; tab_idx += 1
                    tab2 = tabs[tab_idx]; tab_idx += 1
                    tab3 = tabs[tab_idx]; tab_idx += 1
                    tab4 = tabs[tab_idx]; tab_idx += 1
                    tab5 = tabs[tab_idx]; tab_idx += 1
                    tab_stock = tabs[tab_idx] if enable_stock_comparison_state and STOCK_COMPARISON_AVAILABLE else None
                    if tab_stock: tab_idx += 1
                    tab6 = tabs[tab_idx] if bank_dcf_result else None
                
                    with tab1:
                        st.subheader("Valuation Summary - All Methods")
                    
                        # Comparison chart
                        valuations_dict = {}
                        if ri_model:
                            valuations_dict['Residual Income Model'] = ri_model
                        if ddm_model:
                            valuations_dict['Dividend Discount Model'] = ddm_model
                        if pb_roe_model:
                            valuations_dict['P/B with ROE'] = pb_roe_model
                        if rel_val:
                            valuations_dict['Relative Valuation'] = {'value_per_share': rel_val['avg_fair_value']}
                    
                        if valuations_dict:
                            st.plotly_chart(create_bank_valuation_comparison_chart(valuations_dict), 
                                          use_container_width=True)
                    
                        # Summary table
                        summary_data = []
                        if ri_model:
                            summary_data.append(['Residual Income Model', f"‚Çπ{ri_model['value_per_share']:.2f}", 
                                               f"{((ri_model['value_per_share'] - current_price) / current_price * 100):.1f}%"])
                        if ddm_model:
                            summary_data.append(['Dividend Discount Model', f"‚Çπ{ddm_model['value_per_share']:.2f}",
                                               f"{((ddm_model['value_per_share'] - current_price) / current_price * 100):.1f}%"])
                        if pb_roe_model:
                            summary_data.append(['P/B with ROE Analysis', f"‚Çπ{pb_roe_model['value_per_share']:.2f}",
                                               f"{((pb_roe_model['value_per_share'] - current_price) / current_price * 100):.1f}%"])
                        if rel_val:
                            summary_data.append(['Relative Valuation (Avg)', f"‚Çπ{rel_val['avg_fair_value']:.2f}",
                                               f"{((rel_val['avg_fair_value'] - current_price) / current_price * 100):.1f}%"])
                    
                        summary_df = pd.DataFrame(summary_data, columns=['Method', 'Fair Value', 'Upside/Downside'])
                        st.dataframe(summary_df, use_container_width=True, hide_index=True)
                    
                        # Historical charts
                        st.markdown("---")
                        st.subheader("üìà Historical Financial Analysis")
                        st.plotly_chart(create_historical_financials_chart(financials), use_container_width=True)
                
                    with tab2:
                        if ri_model:
                            st.subheader("Residual Income Model")
                            st.write(f"**Fair Value per Share:** ‚Çπ{ri_model['value_per_share']:.2f}")
                        
                            col1, col2, col3 = st.columns(3)
                            with col1:
                                st.metric("Current Book Value", f"‚Çπ{ri_model['current_book_value']:,.0f}")
                            with col2:
                                st.metric("ROE", f"{ri_model['roe']:.2f}%")
                            with col3:
                                st.metric("BV Growth Rate", f"{ri_model.get('bv_growth', 10):.1f}%",
                                         help="Historical book value growth rate")
                        
                            # Growth assumptions
                            st.markdown("---")
                            st.subheader("Growth Assumptions")
                            col4, col5 = st.columns(2)
                            with col4:
                                st.info(f"**Projection Phase Growth:** {ri_model.get('bv_growth', 10):.1f}%")
                                if ri_model.get('historical_bv_growth'):
                                    hist_growth = ri_model['historical_bv_growth']
                                    st.caption(f"Based on historical growth: {', '.join([f'{g:.1f}%' for g in hist_growth])}")
                                else:
                                    st.caption("Using default 10% growth")
                        
                            with col5:
                                st.info(f"**Terminal Growth:** {ri_model.get('terminal_growth', 4):.1f}%")
                                st.caption("Assumes gradual decline to sustainable long-term rate")
                        
                            # Projections table
                            st.markdown("---")
                            st.subheader("5-Year Residual Income Projections")
                            proj_df = pd.DataFrame(ri_model['projections'])
                            proj_df['book_value'] = proj_df['book_value'].apply(lambda x: f"‚Çπ{x:,.0f}")
                            proj_df['net_income'] = proj_df['net_income'].apply(lambda x: f"‚Çπ{x:,.0f}")
                            proj_df['residual_income'] = proj_df['residual_income'].apply(lambda x: f"‚Çπ{x:,.0f}")
                            proj_df['pv_ri'] = proj_df['pv_ri'].apply(lambda x: f"‚Çπ{x:,.0f}")
                            st.dataframe(proj_df, use_container_width=True)
                        
                            # Value breakdown
                            st.markdown("---")
                            st.subheader("Valuation Breakdown")
                            breakdown_df = pd.DataFrame({
                                'Component': ['Current Book Value', 'PV of 5Y Residual Income', 'PV of Terminal Value', 'Total Equity Value'],
                                'Value (‚Çπ)': [
                                    f"‚Çπ{ri_model['current_book_value']:,.0f}",
                                    f"‚Çπ{ri_model['sum_pv_ri']:,.0f}",
                                    f"‚Çπ{ri_model['terminal_ri_pv']:,.0f}",
                                    f"‚Çπ{ri_model['total_equity_value']:,.0f}"
                                ]
                            })
                            st.table(breakdown_df)
                        
                            # Formula explanation
                            with st.expander("üìñ Residual Income Formula"):
                                st.latex(r"RI = NI - (K_e \times BV)")
                                st.latex(r"Value = BV_0 + \sum_{t=1}^{n} \frac{RI_t}{(1+K_e)^t} + \frac{TV}{(1+K_e)^n}")
                                st.write("Where:")
                                st.write("- RI = Residual Income")
                                st.write("- NI = Net Income")
                                st.write(f"- K‚Çë = Cost of Equity = {cost_of_equity:.2f}%")
                                st.write("- BV = Book Value of Equity")
                                st.write("- TV = Terminal Value")
                        else:
                            st.warning("Residual Income Model calculation failed")
                
                    with tab3:
                        if ddm_model:
                            st.subheader("Dividend Discount Model")
                        
                            # Show data source
                            if ddm_model.get('using_actual_data'):
                                st.success("‚úÖ Using actual historical dividend data")
                            else:
                                st.info("‚ÑπÔ∏è Using estimated dividend data (no historical dividends found)")
                        
                            st.write(f"**Fair Value per Share:** ‚Çπ{ddm_model['value_per_share']:.2f}")
                        
                            col1, col2 = st.columns(2)
                            with col1:
                                st.metric("Current DPS", f"‚Çπ{ddm_model['current_dps']:.2f}")
                                st.metric("Payout Ratio", f"{ddm_model['payout_ratio']:.1f}%")
                        
                            with col2:
                                st.metric("Dividend Growth", f"{ddm_model['dividend_growth']:.1f}%")
                                st.metric("Next Year DPS (D1)", f"‚Çπ{ddm_model['next_year_dps']:.2f}")
                        
                            # Historical dividends if available
                            if ddm_model.get('historical_dividends'):
                                st.markdown("---")
                                st.subheader("Historical Dividends (Annual)")
                                hist_divs = ddm_model['historical_dividends']
                                years_range = list(range(len(hist_divs), 0, -1))
                                hist_df = pd.DataFrame({
                                    'Year': [f"T-{y}" for y in years_range],
                                    'Dividend (‚Çπ)': hist_divs
                                })
                                st.dataframe(hist_df, use_container_width=True, hide_index=True)
                        
                            st.markdown("---")
                            st.subheader("5-Year Dividend Projection")
                        
                            # Dividend projections
                            div_df = pd.DataFrame(ddm_model['projections'])
                            div_df['dividend'] = div_df['dividend'].apply(lambda x: f"‚Çπ{x:.2f}")
                            div_df['pv_dividend'] = div_df['pv_dividend'].apply(lambda x: f"‚Çπ{x:.2f}")
                            st.dataframe(div_df, use_container_width=True)
                        
                            # DDM formula explanation
                            with st.expander("üìñ DDM Formula & Assumptions"):
                                st.latex(r"Value = \frac{D_1}{K_e - g}")
                                st.write(f"Where:")
                                st.write(f"- D‚ÇÅ = Next year dividend = ‚Çπ{ddm_model['next_year_dps']:.2f}")
                                st.write(f"- K‚Çë = Cost of Equity = {cost_of_equity:.2f}%")
                                st.write(f"- g = Dividend Growth Rate = {ddm_model['dividend_growth']:.2f}%")
                        else:
                            st.warning("DDM calculation failed or not applicable (cost of equity ‚â§ growth rate)")
                
                    with tab4:
                        if pb_roe_model:
                            st.subheader("P/B with ROE Analysis")
                            st.write(f"**Fair Value per Share:** ‚Çπ{pb_roe_model['value_per_share']:.2f}")
                            st.write(f"**Book Value per Share:** ‚Çπ{pb_roe_model['book_value_per_share']:.2f}")
                            st.write(f"**ROE:** {pb_roe_model['roe']:.2f}%")
                            st.write(f"**Cost of Equity:** {pb_roe_model['cost_of_equity']:.2f}%")
                            st.write(f"**Fair P/B Ratio:** {pb_roe_model['fair_pb_ratio']:.2f}x")
                        
                            # Historical ROE chart
                            roe_years = financials['years']
                            fig = go.Figure()
                            fig.add_trace(go.Scatter(x=roe_years, y=pb_roe_model['historical_roe'],
                                                    mode='lines+markers', name='ROE',
                                                    line=dict(color='#06A77D', width=3)))
                            fig.add_hline(y=pb_roe_model['cost_of_equity'], line_dash="dash",
                                         annotation_text=f"Cost of Equity: {pb_roe_model['cost_of_equity']:.2f}%")
                            fig.update_layout(title="Historical ROE Trend", xaxis_title="Year",
                                            yaxis_title="ROE %", height=400)
                            st.plotly_chart(fig, use_container_width=True)
                        else:
                            st.warning("P/B ROE model calculation failed")
                
                    with tab5:
                        if rel_val:
                            st.subheader("Relative Valuation (Peer-Based)")
                        
                            # Show peer data summary
                            if rel_val.get('peer_count', 0) > 0:
                                st.success(f"‚úÖ Analyzed {rel_val['peer_count']} peer companies")
                            else:
                                st.warning("‚ö†Ô∏è Using default market averages (add peer tickers for better accuracy)")
                        
                            # Main metrics
                            col1, col2, col3 = st.columns(3)
                            with col1:
                                st.metric("Current P/E", f"{rel_val['current_pe']:.2f}x")
                                st.metric("Peer Avg P/E", f"{rel_val['sector_avg_pe']:.2f}x")
                                st.caption(f"Range: {rel_val.get('sector_low_pe', 0):.1f}x - {rel_val.get('sector_high_pe', 0):.1f}x")
                        
                            with col2:
                                st.metric("Current P/B", f"{rel_val['current_pb']:.2f}x")
                                st.metric("Peer Avg P/B", f"{rel_val['sector_avg_pb']:.2f}x")
                                st.caption(f"Range: {rel_val.get('sector_low_pb', 0):.1f}x - {rel_val.get('sector_high_pb', 0):.1f}x")
                        
                            with col3:
                                st.metric("Fair Value (P/E)", f"‚Çπ{rel_val['fair_value_pe_based']:.2f}")
                                st.metric("Fair Value (P/B)", f"‚Çπ{rel_val['fair_value_pb_based']:.2f}")
                        
                            st.markdown("---")
                        
                            # Valuation ranges
                            st.subheader("Valuation Range Analysis")
                            col4, col5, col6 = st.columns(3)
                            with col4:
                                st.metric("Conservative", f"‚Çπ{rel_val.get('conservative_value', 0):.2f}",
                                         help="Based on 25th percentile peer P/E")
                            with col5:
                                st.metric("Fair Value", f"‚Çπ{rel_val['avg_fair_value']:.2f}",
                                         help="Average of P/E and P/B based valuations")
                            with col6:
                                st.metric("Aggressive", f"‚Çπ{rel_val.get('aggressive_value', 0):.2f}",
                                         help="Based on 75th percentile peer P/E")
                        
                            # Peer comparison table
                            if rel_val.get('peer_data') and len(rel_val['peer_data']) > 0:
                                st.markdown("---")
                                st.subheader("Peer Comparison")
                                peer_df = pd.DataFrame(rel_val['peer_data'])
                                peer_df.columns = ['Ticker', 'Price (‚Çπ)', 'P/E', 'P/B']
                                st.dataframe(peer_df, use_container_width=True, hide_index=True)
                        
                            # Interpretation
                            st.markdown("---")
                            st.subheader("Interpretation")
                            pe_premium = ((rel_val['current_pe'] - rel_val['sector_avg_pe']) / rel_val['sector_avg_pe'] * 100) if rel_val['sector_avg_pe'] > 0 else 0
                            pb_premium = ((rel_val['current_pb'] - rel_val['sector_avg_pb']) / rel_val['sector_avg_pb'] * 100) if rel_val['sector_avg_pb'] > 0 else 0
                        
                            if pe_premium > 20:
                                st.warning(f"üìä **P/E Analysis:** Trading at {pe_premium:.1f}% premium to peers. May be overvalued unless justified by superior growth.")
                            elif pe_premium < -20:
                                st.success(f"üìä **P/E Analysis:** Trading at {abs(pe_premium):.1f}% discount to peers. Potential undervaluation.")
                            else:
                                st.info(f"üìä **P/E Analysis:** Trading in line with peers ({pe_premium:+.1f}% premium).")
                        
                            if pb_premium > 20:
                                st.warning(f"üìà **P/B Analysis:** Trading at {pb_premium:.1f}% premium to peers. May indicate high growth expectations.")
                            elif pb_premium < -20:
                                st.success(f"üìà **P/B Analysis:** Trading at {abs(pb_premium):.1f}% discount to peers. Potential value opportunity.")
                            else:
                                st.info(f"üìà **P/B Analysis:** Trading in line with peers ({pb_premium:+.1f}% premium).")
                        else:
                            st.warning("Relative valuation calculation failed")
                
                    # Stock Price Comparison Tab (if enabled)
                    if tab_stock and enable_stock_comparison_state and STOCK_COMPARISON_AVAILABLE:
                        with tab_stock:
                            st.subheader("üìà Stock Price vs Revenue & EPS Analysis")
                            
                            with st.spinner("Fetching stock price data..."):
                                try:
                                    # Determine years to fetch (max 4 for Yahoo Finance)
                                    years_to_fetch = min(historical_years_listed, 4)
                                    
                                    # Fetch comparison data
                                    stock_comp_data = get_stock_comparison_data_listed(
                                        ticker=full_ticker,
                                        company_name=company_name,
                                        financials=financials,
                                        num_years=years_to_fetch
                                    )
                                    
                                    if stock_comp_data and stock_comp_data['chart_fig']:
                                        st.plotly_chart(stock_comp_data['chart_fig'], use_container_width=True)
                                        
                                        # Show data tables in expanders
                                        with st.expander("üìä View Raw Data"):
                                            col1, col2, col3 = st.columns(3)
                                            
                                            with col1:
                                                st.markdown("**Revenue Data**")
                                                if stock_comp_data['revenue_df'] is not None:
                                                    st.dataframe(stock_comp_data['revenue_df'], hide_index=True)
                                                else:
                                                    st.info("No revenue data available")
                                            
                                            with col2:
                                                st.markdown("**EPS Data**")
                                                if stock_comp_data['eps_df'] is not None:
                                                    st.dataframe(stock_comp_data['eps_df'], hide_index=True)
                                                else:
                                                    st.info("No EPS data available")
                                            
                                            with col3:
                                                st.markdown("**Stock Price Summary**")
                                                if stock_comp_data['stock_prices_df'] is not None:
                                                    price_df = stock_comp_data['stock_prices_df']
                                                    st.metric("Latest Price", f"‚Çπ{price_df['Close'].iloc[-1]:.2f}")
                                                    st.metric("Period Return", f"{((price_df['Close'].iloc[-1] - price_df['Close'].iloc[0]) / price_df['Close'].iloc[0] * 100):.2f}%")
                                                    st.metric("Major Changes", f"{price_df['is_major'].sum()}" if 'is_major' in price_df.columns else "N/A")
                                                else:
                                                    st.info("No stock price data available")
                                        
                                        st.info(f"üí° Chart shows {years_to_fetch} years of data (Yahoo Finance limit: 4 years max)")
                                    else:
                                        st.warning("Could not generate stock comparison chart. Check if ticker and financial data are available.")
                                
                                except Exception as e:
                                    st.error(f"Error generating stock comparison: {str(e)}")
                
                    # Bank DCF Tab (if calculated)
                    if bank_dcf_result:
                        with tab6:
                            st.subheader("üè¶ Bank FCFE Valuation (Equity DCF)")
                            st.success("‚úÖ Using proper FCFE methodology - NOT FCFF!")
                        
                            # Fair value display
                            col_dcf1, col_dcf2, col_dcf3, col_dcf4 = st.columns(4)
                            with col_dcf1:
                                st.metric("Fair Value/Share", f"‚Çπ{bank_dcf_result['fair_value_per_share']:.2f}",
                                        delta=f"{((bank_dcf_result['fair_value_per_share'] - current_price) / current_price * 100):.1f}%" if current_price > 0 else None)
                            with col_dcf2:
                                st.metric("Equity Value", f"‚Çπ{bank_dcf_result['equity_value']:,.0f} Lacs")
                            with col_dcf3:
                                st.metric("Cost of Equity (Ke)", f"{bank_dcf_result['cost_of_equity']:.2f}%")
                            with col_dcf4:
                                st.metric("Terminal Growth", f"{bank_dcf_result['terminal_growth']:.2f}%")
                        
                            st.markdown("---")
                        
                            # Show projections
                            st.markdown("**üìà Bank FCFE Projection Details**")
                            col_proj1, col_proj2 = st.columns(2)
                        
                            with col_proj1:
                                st.markdown("**Model Inputs:**")
                                st.write(f"‚Ä¢ Revenue Growth: {drivers_bank['revenue_growth']:.2f}%")
                                st.write(f"‚Ä¢ ROE: {drivers_bank['roe']:.2f}%")
                                st.write(f"‚Ä¢ CAR Target: {drivers_bank['car_ratio']:.1f}%")
                                st.write(f"‚Ä¢ RWA Weight: {drivers_bank['rwa_percentage']:.1f}%")
                        
                            with col_proj2:
                                st.markdown("**Valuation Components:**")
                                st.write(f"‚Ä¢ PV of FCFE (5Y): ‚Çπ{bank_dcf_result['sum_pv_fcfe']:,.0f} Lacs")
                                st.write(f"‚Ä¢ PV of Terminal: ‚Çπ{bank_dcf_result['terminal_value_pv']:,.0f} Lacs")
                                st.write(f"‚Ä¢ Sustainable Growth: {bank_dcf_result['sustainable_growth']:.2f}%")
                        
                            st.markdown("---")
                        
                            # Show FCFE projections
                            if 'fcfe' in projections_bank:
                                st.markdown("**üí∞ FCFE Projections (Free Cash Flow to Equity)**")
                            
                                # Create detailed projection table
                                proj_data = {
                                    'Year': projections_bank['year'],
                                    'Revenue': [f"‚Çπ{rev:,.0f}" for rev in projections_bank['revenue']],
                                    'PAT': [f"‚Çπ{pat:,.0f}" for pat in projections_bank['pat']],
                                    'FCFE': [f"‚Çπ{fcfe:,.0f}" for fcfe in projections_bank['fcfe']],
                                }
                            
                                # Add PV of FCFE
                                if 'pv_fcfe_by_year' in bank_dcf_result:
                                    proj_data['PV of FCFE'] = [f"‚Çπ{pv:,.0f}" for pv in bank_dcf_result['pv_fcfe_by_year']]
                            
                                proj_df = pd.DataFrame(proj_data)
                                st.dataframe(proj_df, use_container_width=True, hide_index=True)
                            
                                st.caption("üí° FCFE = PAT - Equity Required for Growth")
                        
                            st.markdown("---")
                            st.markdown("**üí° FCFE Methodology for Banks:**")
                            st.info("""
                            **Why FCFE (not FCFF):**
                            - For banks, debt is RAW MATERIAL (not financing)
                            - Interest expense is OPERATING COST (not financing cost)
                            - Cannot separate enterprise value from equity value
                            - Must value equity directly using FCFE
                        
                            **FCFE Formula:**
                            ```
                            FCFE = PAT - Equity Required for Growth
                        
                            Where:
                            Equity Required = Growth in Advances √ó RWA% √ó CAR%
                            ```
                        
                            **Discounting:**
                            - Uses Ke (Cost of Equity) - NOT WACC
                            - Values equity directly - NOT enterprise value
                            - No debt adjustment needed
                        
                            **Capital Adequacy:**
                            - Banks need equity capital to support loan growth
                            - Retained earnings fund asset expansion
                            - Only excess cash after capital requirements = FCFE
                            """)
                
                    st.stop()
            
                # ================================
                # STANDARD FCFF DCF FOR NON-BANK COMPANIES
                # ================================
                else:
                    st.markdown("---")
                
                # Calculate WC metrics
                wc_metrics = calculate_working_capital_metrics(financials)
                
                # Show Working Capital Data Status
                wc_status_parts = []
                if wc_metrics.get('has_inventory', False) or (inventory_days_override and inventory_days_override > 0):
                    inv_source = f"User Override: {inventory_days_override} days" if inventory_days_override and inventory_days_override > 0 else f"Historical: {wc_metrics['avg_inv_days']:.1f} days"
                    wc_status_parts.append(f"‚úÖ Inventory ({inv_source})")
                else:
                    wc_status_parts.append("‚ö†Ô∏è Inventory (No data)")
                
                if wc_metrics.get('has_receivables', False) or (debtor_days_override and debtor_days_override > 0):
                    deb_source = f"User Override: {debtor_days_override} days" if debtor_days_override and debtor_days_override > 0 else f"Historical: {wc_metrics['avg_deb_days']:.1f} days"
                    wc_status_parts.append(f"‚úÖ Debtors ({deb_source})")
                else:
                    wc_status_parts.append("‚ö†Ô∏è Debtors (No data)")
                
                if wc_metrics.get('has_payables', False) or (creditor_days_override and creditor_days_override > 0):
                    cred_source = f"User Override: {creditor_days_override} days" if creditor_days_override and creditor_days_override > 0 else f"Historical: {wc_metrics['avg_cred_days']:.1f} days"
                    wc_status_parts.append(f"‚úÖ Creditors ({cred_source})")
                else:
                    wc_status_parts.append("‚ö†Ô∏è Creditors (No data)")
                
                st.info(f"üîç **Working Capital Projection Status:** {' | '.join(wc_status_parts)}")
                
                # Project financials with ALL user overrides
                projections, drivers = project_financials(
                    financials, wc_metrics, projection_years_listed, tax_rate,
                    rev_growth_override_listed, opex_margin_override_listed, capex_ratio_override_listed,
                    # Pass all advanced user controls
                    ebitda_margin_override=ebitda_margin_override if ebitda_margin_override > 0 else None,
                    depreciation_rate_override=depreciation_rate_override if depreciation_rate_override > 0 else None,
                    depreciation_method=depreciation_method,
                    inventory_days_override=inventory_days_override if inventory_days_override > 0 else None,
                    debtor_days_override=debtor_days_override if debtor_days_override > 0 else None,
                    creditor_days_override=creditor_days_override if creditor_days_override > 0 else None,
                    interest_rate_override=interest_rate_override if interest_rate_override > 0 else None,
                    working_capital_pct_override=working_capital_as_pct_revenue if working_capital_as_pct_revenue > 0 else None
                )
                
                # Calculate WACC (beta of the company itself)
                st.info("Calculating beta for the stock...")
                # Pass ticker WITH exchange suffix for proper beta calculation
                full_ticker = get_ticker_with_exchange(ticker, exchange_suffix)
                beta = get_stock_beta(full_ticker, period_years=3)
                st.success(f"‚úÖ Beta calculated: {beta:.3f}")
                
                st.info(f"üèõÔ∏è Risk-Free Rate (India 10Y G-Sec): {manual_rf_rate:.2f}%")
                wacc_details = calculate_wacc(financials, tax_rate, peer_tickers=None, manual_rf_rate=manual_rf_rate)
                wacc_details['beta'] = beta  # Override with actual stock beta
                # Recalculate Ke and WACC with actual beta
                wacc_details['ke'] = wacc_details['rf'] + (beta * (wacc_details['rm'] - wacc_details['rf']))
                wacc_details['wacc'] = (wacc_details['we']/100 * wacc_details['ke']) + (wacc_details['wd']/100 * wacc_details['kd_after_tax'])
                
                # DCF Valuation
                # Extract cash balance
                cash_balance = financials['cash'][0] if financials['cash'][0] > 0 else 0
                
                valuation, error = calculate_dcf_valuation(
                    projections, wacc_details, terminal_growth, shares, cash_balance,
                    manual_discount_rate=manual_discount_rate if manual_discount_rate > 0 else None
                )
                
                if error:
                    st.error(error)
                    st.stop()
                
                # ================================
                # GET CURRENT PRICE EARLY (before PDF generation)
                # ================================
                current_price = 0
                try:
                    stock = get_cached_ticker(get_ticker_with_exchange(ticker, exchange_suffix))
                    info = stock.info
                    # Try multiple methods to get current price (Phase 1 approach)
                    current_price = info.get('currentPrice', 0)
                    if not current_price or current_price == 0:
                        current_price = info.get('regularMarketPrice', 0)
                    if not current_price or current_price == 0:
                        # Try getting from recent history
                        try:
                            hist = stock.history(period='1d')
                            if not hist.empty:
                                current_price = hist['Close'].iloc[-1]
                        except:
                            pass
                except Exception as e:
                    st.warning(f"Could not fetch current price: {e}")
                    current_price = 0
                
                # ================================
                # ADDITIONAL VALUATION MODELS FOR NON-BANKING COMPANIES
                # ================================
                
                # Calculate Dividend Discount Model (DDM)
                st.info("Calculating Dividend Discount Model...")
                ddm_result = calculate_dividend_discount_model(
                    financials, shares, wacc_details['ke'], 
                    ticker=ticker,
                    div_growth_override=ddm_dividend_growth_override if ddm_dividend_growth_override > 0 else None,
                    payout_ratio_override=ddm_payout_ratio_override if ddm_payout_ratio_override > 0 else None,
                    dcf_projections=projections  # ‚úÖ PASS EXISTING PROJECTIONS - NO DUPLICATION!
                )
                if ddm_result and isinstance(ddm_result, dict) and 'value_per_share' in ddm_result:
                    st.success(f"‚úÖ DDM Fair Value: ‚Çπ{ddm_result['value_per_share']:.2f}")
                    if ddm_result.get('using_dcf_projections'):
                        st.caption("üí° Using DCF projected NOPAT for dividend projections")
                else:
                    st.warning("‚ö†Ô∏è DDM not applicable (company may not pay dividends)")
                
                # Calculate Residual Income Model (RIM)
                st.info("Calculating Residual Income Model...")
                # Use user overrides or defaults
                rim_terminal_g = rim_terminal_growth_override if rim_terminal_growth_override > 0 else terminal_growth
                rim_proj_years = rim_projection_years_override if rim_projection_years_override > 0 else projection_years_listed
                rim_roe = rim_assumed_roe_override if rim_assumed_roe_override > 0 else None
                
                rim_result = calculate_residual_income_model(
                    financials, shares, wacc_details['ke'], 
                    terminal_growth=rim_terminal_g, 
                    projection_years=rim_proj_years,
                    assumed_roe=rim_roe,
                    dcf_projections=projections  # ‚úÖ PASS EXISTING PROJECTIONS - NO DUPLICATION!
                )
                if rim_result and isinstance(rim_result, dict) and 'value_per_share' in rim_result:
                    st.success(f"‚úÖ RIM Fair Value: ‚Çπ{rim_result['value_per_share']:.2f}")
                    if rim_result.get('using_dcf_projections'):
                        st.caption("üí° Using DCF projected NOPAT as Net Income")
                else:
                    st.warning("‚ö†Ô∏è RIM calculation returned incomplete results")
                    if rim_result and isinstance(rim_result, dict):
                        st.caption(f"RIM result keys: {list(rim_result.keys())}")
                
                # ================================
                # DISPLAY RESULTS (SAME AS UNLISTED)
                # ================================
                
                st.success("‚úÖ Valuation Complete!")
                
                # AUTO-GENERATE PDF
                try:
                    all_fair_values = {'DCF': valuation['fair_value_per_share']}
                
                    # Add DDM if calculated
                    if ddm_result and ddm_result.get('value_per_share', 0) > 0:
                        all_fair_values['DDM'] = ddm_result['value_per_share']
                
                    # Add RIM if calculated
                    if rim_result and rim_result.get('value_per_share', 0) > 0:
                        all_fair_values['Residual Income'] = rim_result['value_per_share']
                
                    # Add comparative valuations
                    if 'comp_results' in locals() and comp_results:
                        for method, val_data in comp_results.get('valuations', {}).items():
                            if val_data.get('fair_value_avg', 0) > 0:
                                all_fair_values[method.upper().replace('_', ' ')] = val_data['fair_value_avg']
                
                    pdf_path = export_to_pdf({
                        'company_name': company_name,
                        'ticker': ticker,
                        'financials': financials,
                        'dcf_results': valuation,
                        'fair_values': all_fair_values,
                        'current_price': current_price if 'current_price' in locals() else 0,
                        'peer_data': pd.DataFrame(),
                        'comp_results': comp_results if 'comp_results' in locals() else None
                    })
                
                    with open(pdf_path, "rb") as f:
                        st.session_state.pdf_bytes = f.read()
                
                    st.toast("‚úÖ PDF Generated! Scroll to top to download", icon="üì•")
                except Exception as e:
                    st.error(f"PDF Generation Error: {str(e)}")
                
                # Calculate comparative valuation EARLY for Forward P/E display
                comp_results = None
                if comp_tickers_listed and comp_tickers_listed.strip():
                    try:
                        # Get use_screener_peers flag from session state
                        use_screener_for_peers = st.session_state.get('use_screener_peers', False)
                    
                        comp_results = perform_comparative_valuation(
                            ticker, 
                            comp_tickers_listed, 
                            financials, 
                            shares, 
                            exchange_suffix, 
                            projections=projections,
                            use_screener_peers=use_screener_for_peers
                        )
                    except Exception as e:
                        st.warning(f"Could not calculate comparative valuation: {str(e)}")
                
                # Key Metrics with Current Price and P/E
                # Get current P/E if available
                current_pe = info.get('trailingPE', 0) if 'info' in locals() else 0
                current_eps = (financials['nopat'][0] * 100000) / shares if shares > 0 and financials['nopat'][0] > 0 else 0
                
                st.markdown("### üìä Key Valuation Metrics")
                
                col1, col2, col3, col4, col5, col6 = st.columns(6)
                with col1:
                    st.metric("üìä Current Price", f"‚Çπ {current_price:.2f}")
                with col2:
                    st.metric("üéØ Fair Value (DCF)", f"‚Çπ {valuation['fair_value_per_share']:.2f}",
                             delta=f"{((valuation['fair_value_per_share'] - current_price) / current_price * 100):.1f}%")
                with col3:
                    st.metric("Current P/E", f"{current_pe:.2f}x" if current_pe > 0 else "N/A")
                with col4:
                    st.metric("Current EPS", f"‚Çπ {current_eps:.2f}" if current_eps > 0 else "N/A")
                with col5:
                    st.metric("WACC", f"{wacc_details['wacc']:.2f}%")
                with col6:
                    st.metric("Terminal Growth", f"{terminal_growth:.1f}%")
                
                # Forward P/E Display (if available)
                if 'comp_results' in locals() and comp_results and 'forward_pe' in comp_results:
                    st.markdown("---")
                    st.markdown("### üìÖ 12-Month Forward P/E Valuation")
                
                    fpe = comp_results['forward_pe']
                
                    col_fpe1, col_fpe2, col_fpe3, col_fpe4, col_fpe5 = st.columns(5)
                
                    with col_fpe1:
                        st.metric("Current EPS", f"‚Çπ{fpe.get('current_eps', 0):.2f}")
                
                    with col_fpe2:
                        st.metric("Forward EPS (12M)", f"‚Çπ{fpe['forward_eps']:.2f}",
                                delta=f"+{fpe.get('earnings_growth_rate', 0):.1f}%",
                                help="Projected EPS for next 12 months")
                
                    with col_fpe3:
                        st.metric("Peer Avg P/E", f"{comp_results['multiples_stats']['pe']['average']:.2f}x" if 'pe' in comp_results['multiples_stats'] else "N/A")
                
                    with col_fpe4:
                        st.metric("Forward Fair Value (Avg)", f"‚Çπ{fpe['fair_value_avg']:.2f}",
                                delta=f"{((fpe['fair_value_avg'] - current_price) / current_price * 100):.1f}%" if current_price > 0 else None)
                
                    with col_fpe5:
                        st.metric("Forward Fair Value (Median)", f"‚Çπ{fpe['fair_value_median']:.2f}",
                                delta=f"{((fpe['fair_value_median'] - current_price) / current_price * 100):.1f}%" if current_price > 0 else None)
                
                    st.caption(f"üí° {fpe.get('calculation_note', 'Forward EPS projected using historical growth')}")
                
                st.markdown("---")
                
                # Price vs Value Gauge
                if valuation['fair_value_per_share'] > 0:
                    st.plotly_chart(create_price_vs_value_gauge(current_price, valuation['fair_value_per_share']), 
                                  use_container_width=True)
                
                # COMPARISON BAR CHART - All Valuation Methods vs Current Price
                with st.expander("üìä **All Valuation Methods Comparison**", expanded=False):
                    st.markdown("### Fair Value Comparison Across All Methods")
                
                    # Collect all available fair values
                    valuation_methods = []
                    fair_values = []
                    colors = []
                
                    # DCF
                    if valuation and valuation.get('fair_value_per_share', 0) > 0:
                        valuation_methods.append('DCF (FCFF)')
                        fair_values.append(valuation['fair_value_per_share'])
                        colors.append('#1f77b4')  # Blue
                
                    # DDM
                    if ddm_result and ddm_result.get('value_per_share', 0) > 0 and not ddm_result.get('error'):
                        valuation_methods.append('DDM (Gordon)')
                        fair_values.append(ddm_result['value_per_share'])
                        colors.append('#2ca02c')  # Green
                
                    # RIM
                    if rim_result and rim_result.get('value_per_share', 0) > 0 and not rim_result.get('error'):
                        valuation_methods.append('RIM')
                        fair_values.append(rim_result['value_per_share'])
                        colors.append('#ff7f0e')  # Orange
                
                    # Forward P/E
                    if comp_results and 'forward_pe' in comp_results:
                        fpe = comp_results['forward_pe']
                        if fpe.get('fair_value_avg', 0) > 0:
                            valuation_methods.append('Forward P/E (Avg)')
                            fair_values.append(fpe['fair_value_avg'])
                            colors.append('#d62728')  # Red
                    
                        if fpe.get('fair_value_median', 0) > 0:
                            valuation_methods.append('Forward P/E (Median)')
                            fair_values.append(fpe['fair_value_median'])
                            colors.append('#9467bd')  # Purple
                
                    # Comparative Valuation methods
                    if comp_results and 'valuations' in comp_results:
                        for method, val_data in comp_results['valuations'].items():
                            if val_data.get('fair_value_avg', 0) > 0:
                                method_name = val_data.get('method', method).replace('_', ' ').title()
                                valuation_methods.append(f"{method_name} (Avg)")
                                fair_values.append(val_data['fair_value_avg'])
                                colors.append('#8c564b')  # Brown
                
                    # Create comparison chart if we have data
                    if valuation_methods and current_price > 0 and len(fair_values) > 0:
                    
                        fig = go.Figure()
                    
                        # Add bars for each valuation method
                        fig.add_trace(go.Bar(
                            y=valuation_methods,
                            x=fair_values,
                            marker_color=colors,
                            text=[f"‚Çπ{v:.2f}" for v in fair_values],
                            textposition='outside',
                            orientation='h',
                            name='Fair Value',
                            hovertemplate='<b>%{y}</b><br>Fair Value: ‚Çπ%{x:.2f}<extra></extra>'
                        ))
                    
                        # Add red line for current price
                        fig.add_vline(
                            x=current_price,
                            line_dash="dash",
                            line_color="red",
                            line_width=3,
                            annotation_text=f"Current Price: ‚Çπ{current_price:.2f}",
                            annotation_position="top right"
                        )
                    
                        # Add average fair value line
                        avg_fair_value = np.mean(fair_values)
                        fig.add_vline(
                            x=avg_fair_value,
                            line_dash="dot",
                            line_color="green",
                            line_width=2,
                            annotation_text=f"Average: ‚Çπ{avg_fair_value:.2f}",
                            annotation_position="bottom right"
                        )
                    
                        fig.update_layout(
                            title={
                                'text': "Fair Value Comparison: All Methods vs Current Price",
                                'x': 0.5,
                                'xanchor': 'center'
                            },
                            xaxis_title="Fair Value (‚Çπ)",
                            yaxis_title="Valuation Method",
                            height=max(400, len(valuation_methods) * 60),
                            showlegend=False,
                            xaxis=dict(gridcolor='lightgray', zeroline=True),
                            yaxis=dict(autorange="reversed"),  # Top to bottom
                            plot_bgcolor='rgba(240,240,240,0.5)'
                        )
                    
                        st.plotly_chart(fig, use_container_width=True)
                    
                        # Summary statistics
                        st.markdown("### üìä Summary Statistics")
                        col_stat1, col_stat2, col_stat3, col_stat4 = st.columns(4)
                    
                        with col_stat1:
                            st.metric("Average Fair Value", f"‚Çπ{avg_fair_value:.2f}",
                                    delta=f"{((avg_fair_value - current_price) / current_price * 100):.1f}%" if current_price > 0 else None)
                    
                        with col_stat2:
                            median_fv = np.median(fair_values)
                            st.metric("Median Fair Value", f"‚Çπ{median_fv:.2f}",
                                    delta=f"{((median_fv - current_price) / current_price * 100):.1f}%" if current_price > 0 else None)
                    
                        with col_stat3:
                            max_fv = max(fair_values)
                            min_fv = min(fair_values)
                            st.metric("Range", f"‚Çπ{min_fv:.2f} - ‚Çπ{max_fv:.2f}")
                    
                        with col_stat4:
                            consensus = "üî¥ Overvalued" if current_price > avg_fair_value else "üü¢ Undervalued"
                            deviation = abs((avg_fair_value - current_price) / current_price * 100)
                            st.metric("Consensus", consensus)
                            st.caption(f"Deviation: {deviation:.1f}%")
                    
                        # Methods breakdown
                        st.markdown("---")
                        st.markdown("**üí° Interpretation:**")
                        if current_price < min_fv:
                            st.success(f"‚úÖ **Strong Buy Signal**: Current price (‚Çπ{current_price:.2f}) is below ALL valuation methods. Significant upside potential.")
                        elif current_price > max_fv:
                            st.error(f"‚ö†Ô∏è **Overvalued**: Current price (‚Çπ{current_price:.2f}) exceeds ALL valuation methods. Consider taking profits.")
                        elif current_price < avg_fair_value:
                            st.info(f"üìà **Undervalued**: Current price (‚Çπ{current_price:.2f}) is below average fair value. Potential upside of {((avg_fair_value - current_price) / current_price * 100):.1f}%")
                        else:
                            st.warning(f"üìâ **Fairly Valued to Overvalued**: Current price (‚Çπ{current_price:.2f}) is at or above average fair value.")
                
                    else:
                        st.info("üí° Complete more valuation methods to see comprehensive comparison chart")
                
                # Tabs for detailed output - make dynamic
                tab_list_nonbank = [
                    "üìä Historical Analysis",
                    "üìà Projections",
                    "üí∞ FCF Working",
                    "üéØ WACC Breakdown",
                    "üèÜ Valuation Summary",
                    "üìâ Sensitivity Analysis",
                    "üìÅ Comparative Valuation",
                    "üè¢ Peer Comparison",
                    "üí∞ Dividend Discount Model",
                    "üè¢ Residual Income Model",
                    "‚öôÔ∏è Assumptions & Parameters"
                ]
                
                # Add Stock Comparison tab if enabled
                enable_stock_comparison_state = st.session_state.get('enable_stock_comparison_listed', False)
                if enable_stock_comparison_state and STOCK_COMPARISON_AVAILABLE:
                    tab_list_nonbank.append("üìà Stock vs Financials")
                
                tabs_nonbank = st.tabs(tab_list_nonbank)
                tab1 = tabs_nonbank[0]
                tab2 = tabs_nonbank[1]
                tab3 = tabs_nonbank[2]
                tab4 = tabs_nonbank[3]
                tab5 = tabs_nonbank[4]
                tab6 = tabs_nonbank[5]
                tab7 = tabs_nonbank[6]
                tab8 = tabs_nonbank[7]
                tab9 = tabs_nonbank[8]
                tab10 = tabs_nonbank[9]
                tab11 = tabs_nonbank[10]
                tab_stock_nonbank = tabs_nonbank[11] if len(tabs_nonbank) > 11 else None
                
                with tab1:
                    st.subheader("üìä Comprehensive Historical Financial Analysis")
                
                    # Use advanced charting function
                    st.plotly_chart(create_historical_financials_chart(financials), use_container_width=True)
                
                    # Data tables below charts
                    with st.expander("üìã View Raw Data Tables"):
                        st.subheader("Historical Financials (Last 3 Years)")
                    
                        hist_df = pd.DataFrame({
                            'Year': [str(y) for y in financials['years']],
                            'Revenue': financials['revenue'],
                            'Operating Expenses': financials['opex'],
                            'EBITDA': financials['ebitda'],
                            'Depreciation': financials['depreciation'],
                            'EBIT': financials['ebit'],
                            'Interest': financials['interest'],
                            'Tax': financials['tax'],
                            'NOPAT': financials['nopat']
                        })
                    
                        numeric_cols = hist_df.select_dtypes(include=[np.number]).columns.tolist()
                        format_dict = {col: '{:.2f}' for col in numeric_cols}
                        st.dataframe(hist_df.style.format(format_dict), use_container_width=True)
                    
                        st.subheader("Balance Sheet Metrics")
                        bs_df = pd.DataFrame({
                            'Year': [str(y) for y in financials['years']],
                            'Fixed Assets': financials['fixed_assets'],
                            'Inventory': financials['inventory'],
                            'Receivables': financials['receivables'],
                            'Payables': financials['payables'],
                            'Equity': financials['equity'],
                            'ST Debt': financials['st_debt'],
                            'LT Debt': financials['lt_debt']
                        })
                        numeric_cols = bs_df.select_dtypes(include=[np.number]).columns.tolist()
                        format_dict = {col: '{:.2f}' for col in numeric_cols}
                        st.dataframe(bs_df.style.format(format_dict), use_container_width=True)
                    
                        st.subheader("Working Capital Days")
                        wc_df = pd.DataFrame({
                            'Year': [str(y) for y in financials['years']],
                            'Inventory Days': wc_metrics['inventory_days'],
                            'Debtor Days': wc_metrics['debtor_days'],
                            'Creditor Days': wc_metrics['creditor_days']
                        })
                        numeric_cols = wc_df.select_dtypes(include=[np.number]).columns.tolist()
                        format_dict = {col: '{:.2f}' for col in numeric_cols}
                        st.dataframe(wc_df.style.format(format_dict), use_container_width=True)
                    
                        st.info(f"**Average Working Capital Days:** Inventory: {wc_metrics['avg_inv_days']:.1f} | Debtors: {wc_metrics['avg_deb_days']:.1f} | Creditors: {wc_metrics['avg_cred_days']:.1f}")
                
                with tab2:
                    st.subheader(f"üìà Projected Financials ({projection_years_listed} Years)")
                
                    # Use advanced charting function
                    st.plotly_chart(create_fcff_projection_chart(projections), use_container_width=True)
                
                    # Data table below
                    with st.expander("üìã View Projection Data Table"):
                        proj_df = pd.DataFrame({
                            'Year': [str(y) for y in projections['year']],
                            'Revenue': projections['revenue'],
                            'EBITDA': projections['ebitda'],
                            'Depreciation': projections['depreciation'],
                            'EBIT': projections['ebit'],
                            'NOPAT': projections['nopat'],
                            'Capex': projections['capex'],
                            'Œî WC': projections['delta_wc'],
                            'FCFF': projections['fcff']
                        })
                        numeric_cols = proj_df.select_dtypes(include=[np.number]).columns.tolist()
                        format_dict = {col: '{:.2f}' for col in numeric_cols}
                        st.dataframe(proj_df.style.format(format_dict), use_container_width=True)
                
                    st.info(f"**Key Drivers:** Revenue Growth: {drivers['avg_growth']:.2f}% | Opex Margin: {drivers['avg_opex_margin']:.2f}% | CapEx/Revenue: {drivers['avg_capex_ratio']:.2f}% | Depreciation Rate: {drivers['avg_dep_rate']:.2f}%")
                
                with tab3:
                    st.subheader("Free Cash Flow Working")
                
                    fcff_df = pd.DataFrame({
                        'Year': [str(y) for y in projections['year']],
                        'NOPAT': projections['nopat'],
                        '+ Depreciation': projections['depreciation'],
                        '- Œî WC': projections['delta_wc'],
                        '- Capex': projections['capex'],
                        '= FCFF': projections['fcff'],
                        'Discount Factor': [(1 + wacc_details['wacc']/100)**(-y) for y in projections['year']],
                        'PV(FCFF)': valuation['pv_fcffs']
                    })
                    numeric_cols = fcff_df.select_dtypes(include=[np.number]).columns.tolist()
                    format_dict = {col: '{:.4f}' for col in numeric_cols}
                    st.dataframe(fcff_df.style.format(format_dict), use_container_width=True)
                
                    st.metric("Sum of PV(FCFF)", f"‚Çπ {valuation['sum_pv_fcff']:.2f} Lacs")
                
                with tab4:
                    st.subheader("üéØ WACC Calculation & Breakdown")
                
                    # Advanced WACC breakdown chart
                    st.plotly_chart(create_wacc_breakdown_chart(wacc_details), use_container_width=True)
                
                    st.markdown("---")
                
                    col1, col2 = st.columns(2)
                
                    with col1:
                        st.markdown("**Cost of Equity (Ke)**")
                        st.write(f"Risk-free Rate (Rf): **{wacc_details['rf']:.2f}%**")
                        st.write(f"Market Return (Rm): **{wacc_details['rm']:.2f}%**")
                        st.write(f"Beta (Œ≤) - {ticker}: **{wacc_details['beta']:.3f}**")
                        st.write(f"Ke = Rf + Œ≤ √ó (Rm - Rf)")
                        st.write(f"Ke = {wacc_details['rf']:.2f}% + {wacc_details['beta']:.3f} √ó ({wacc_details['rm']:.2f}% - {wacc_details['rf']:.2f}%)")
                        st.write(f"**Ke = {wacc_details['ke']:.2f}%**")
                
                    with col2:
                        st.markdown("**Cost of Debt (Kd)**")
                        st.write(f"Interest Expense: **‚Çπ {financials['interest'][0]:.2f} Lacs**")
                        st.write(f"Total Debt: **‚Çπ {wacc_details['debt']:.2f} Lacs**")
                        st.write(f"Kd (pre-tax) = {wacc_details['kd']:.2f}%")
                        st.write(f"Tax Rate = {tax_rate}%")
                        st.write(f"**Kd (after-tax) = {wacc_details['kd_after_tax']:.2f}%**")
                
                    st.markdown("---")
                    st.markdown("**WACC Calculation**")
                
                    col3, col4 = st.columns(2)
                    with col3:
                        st.write(f"Equity (E): **‚Çπ {wacc_details['equity']:.2f} Lacs** ({wacc_details['we']:.2f}%)")
                        st.write(f"Debt (D): **‚Çπ {wacc_details['debt']:.2f} Lacs** ({wacc_details['wd']:.2f}%)")
                        st.write(f"Total Capital (V): **‚Çπ {wacc_details['equity'] + wacc_details['debt']:.2f} Lacs**")
                
                    with col4:
                        st.write(f"WACC = (E/V √ó Ke) + (D/V √ó Kd √ó (1-Tax))")
                        st.write(f"WACC = ({wacc_details['we']:.2f}% √ó {wacc_details['ke']:.2f}%) + ({wacc_details['wd']:.2f}% √ó {wacc_details['kd_after_tax']:.2f}%)")
                        st.write(f"**WACC = {wacc_details['wacc']:.2f}%**")
                
                with tab5:
                    st.subheader("üèÜ DCF Valuation Summary")
                
                    # Show FCFF adjustment notice if applicable
                    if valuation.get('fcff_adjusted', False):
                        st.warning("‚ö†Ô∏è **FCFF Adjustment Applied**")
                        adj_details = valuation.get('adjustment_details', {})
                        st.write(f"**Strategy Used:** {adj_details.get('strategy', 'N/A')}")
                        st.write(f"**Original Terminal FCFF:** ‚Çπ{projections['fcff'][-1]:.2f} Lacs")
                        st.write(f"**Adjusted Terminal FCFF:** ‚Çπ{valuation['adjusted_terminal_fcff']:.2f} Lacs")
                        st.caption("üìå Adjustment details shown during valuation run above")
                        st.markdown("---")
                
                    # Waterfall chart showing value buildup
                    st.plotly_chart(create_waterfall_chart(valuation), use_container_width=True)
                
                    st.markdown("### Terminal Value Calculation")
                
                    # Use adjusted FCFF if available
                    terminal_fcff = valuation.get('adjusted_terminal_fcff', projections['fcff'][-1])
                
                    st.write(f"FCFF (Year {projection_years_listed}): **‚Çπ {terminal_fcff:.2f} Lacs**")
                    if valuation.get('fcff_adjusted', False):
                        st.caption(f"(Original: ‚Çπ{projections['fcff'][-1]:.2f} Lacs - Adjusted for sustainability)")
                
                    st.write(f"Terminal Growth Rate (g): **{terminal_growth}%**")
                    st.write(f"FCFF (Year {projection_years_listed + 1}) = FCFF{projection_years_listed} √ó (1 + g)")
                    st.write(f"FCFF (Year {projection_years_listed + 1}) = ‚Çπ {terminal_fcff:.2f} √ó (1 + {terminal_growth/100})")
                    st.write(f"FCFF (Year {projection_years_listed + 1}) = **‚Çπ {terminal_fcff * (1 + terminal_growth/100):.2f} Lacs**")
                
                    st.write(f"\nTerminal Value = FCFF{projection_years_listed + 1} / (WACC - g)")
                    st.write(f"Terminal Value = ‚Çπ {projections['fcff'][-1] * (1 + terminal_growth/100):.2f} / ({wacc_details['wacc']:.2f}% - {terminal_growth}%)")
                    st.write(f"**Terminal Value = ‚Çπ {valuation['terminal_value']:.2f} Lacs**")
                
                    st.write(f"\nPV(Terminal Value) = TV / (1 + WACC)^{projection_years_listed}")
                    st.write(f"**PV(Terminal Value) = ‚Çπ {valuation['pv_terminal_value']:.2f} Lacs**")
                
                    st.markdown("---")
                    st.markdown("### Enterprise Value")

                    # Show growth phase adjustment if applied
                    if valuation.get('growth_phase_adjusted', False):
                        st.info("üìä **Growth-Phase Company:** Sum of PV(FCFF) was adjusted from negative to zero")
                        st.caption(f"Original: ‚Çπ{valuation['original_sum_pv_fcff']:.2f} Lacs ‚Üí Adjusted: ‚Çπ{valuation['sum_pv_fcff']:.2f} Lacs")
                
                    ev_df = pd.DataFrame({
                        'Component': ['Sum of PV(FCFF)', 'PV(Terminal Value)', 'Enterprise Value'],
                        'Value (‚Çπ Lacs)': [
                            valuation['sum_pv_fcff'],
                            valuation['pv_terminal_value'],
                            valuation['enterprise_value']
                        ]
                    })
                    st.dataframe(ev_df.style.format({'Value (‚Çπ Lacs)': '{:.2f}'}), use_container_width=True)
                
                    tv_pct = valuation['tv_percentage']
                    if tv_pct > 90:
                        st.warning(f"‚ö†Ô∏è Terminal Value represents {tv_pct:.1f}% of Enterprise Value (>90% is high)")
                    else:
                        st.info(f"Terminal Value represents {tv_pct:.1f}% of Enterprise Value")
                
                    st.markdown("---")
                    st.markdown("### Equity Value & Fair Value per Share")
                
                    equity_calc_df = pd.DataFrame({
                        'Item': ['Enterprise Value', 'Less: Total Debt', 'Add: Cash & Equivalents', '= Net Debt', 'Equity Value', 'Equity Value (‚Çπ)', 'Number of Shares', 'Fair Value per Share'],
                        'Value': [
                            f"‚Çπ {valuation['enterprise_value']:.2f} Lacs",
                            f"‚Çπ {valuation['total_debt']:.2f} Lacs",
                            f"‚Çπ {valuation['cash']:.2f} Lacs",
                            f"‚Çπ {valuation['net_debt']:.2f} Lacs",
                            f"‚Çπ {valuation['equity_value']:.2f} Lacs",
                            f"‚Çπ {valuation['equity_value_rupees']:,.0f}",
                            f"{shares:,}" if 'shares' in locals() else f"{num_shares:,}",
                            f"‚Çπ {valuation['fair_value_per_share']:.2f}"
                        ]
                    })
                    st.table(equity_calc_df)
                
                    st.success(f"### üéØ Fair Value per Share (DCF): ‚Çπ {valuation['fair_value_per_share']:.2f}")
                
                    # Show all valuation methods if available
                    if (ddm_result and ddm_result.get('value_per_share', 0) > 0) or (rim_result and rim_result.get('value_per_share', 0) > 0):
                        st.markdown("---")
                        st.markdown("### üíé Additional Valuation Methods")
                    
                        all_methods = [
                            ("DCF (FCFF)", valuation['fair_value_per_share'])
                        ]
                    
                        if ddm_result and ddm_result.get('value_per_share', 0) > 0:
                            all_methods.append(("DDM (Gordon)", ddm_result['value_per_share']))
                    
                        if rim_result and rim_result.get('value_per_share', 0) > 0:
                            all_methods.append(("Residual Income", rim_result['value_per_share']))
                    
                        # Create comparison table
                        methods_df = pd.DataFrame(all_methods, columns=['Method', 'Fair Value'])
                        methods_df['Fair Value'] = methods_df['Fair Value'].apply(lambda x: f"‚Çπ{x:.2f}")
                        methods_df['Upside/Downside'] = [
                            f"{((v - current_price) / current_price * 100):.1f}%" if current_price > 0 else "N/A"
                            for _, v in all_methods
                        ]
                    
                        st.dataframe(methods_df, use_container_width=True, hide_index=True)
                    
                        avg_all = np.mean([v for _, v in all_methods])
                        st.info(f"üìä **Average Fair Value (All Methods):** ‚Çπ{avg_all:.2f}")
                    
                        if current_price > 0:
                            upside_avg = ((avg_all - current_price) / current_price * 100)
                            st.metric("Overall Upside/Downside", f"{upside_avg:+.1f}%")
                
                with tab6:
                    st.subheader("üìâ Advanced Sensitivity Analysis")
                
                    wacc_range = np.arange(max(1.0, wacc_details['wacc'] - 3), wacc_details['wacc'] + 3.5, 0.5)
                    g_range = np.arange(max(1.0, terminal_growth - 2), min(terminal_growth + 3, wacc_details['wacc'] - 1), 0.5)
                
                    if len(g_range) == 0:
                        g_range = np.array([terminal_growth])
                
                    # Interactive heatmap
                    st.plotly_chart(create_sensitivity_heatmap(projections, wacc_range, g_range, shares),
                                  use_container_width=True)
                
                    # Traditional table below
                    with st.expander("üìã View Sensitivity Data Table"):
                        sensitivity_data = []
                    
                        for w in wacc_range:
                            row_data = {'WACC ‚Üí': f"{w:.1f}%"}
                            for g_val in g_range:
                                if g_val >= w - 0.1:  # Need at least 0.1% gap
                                    row_data[f"g={g_val:.1f}%"] = "N/A"
                                else:
                                    try:
                                        fcff_n_plus_1 = projections['fcff'][-1] * (1 + g_val / 100)
                                        tv = fcff_n_plus_1 / ((w / 100) - (g_val / 100))
                                        pv_tv = tv / ((1 + w / 100) ** projection_years_listed)
                                        ev = valuation['sum_pv_fcff'] + pv_tv
                                        eq_val = ev - valuation['net_debt']
                                        eq_val_rupees = eq_val * 100000
                                        fv = eq_val_rupees / shares if shares > 0 else 0
                                        row_data[f"g={g_val:.1f}%"] = f"‚Çπ{fv:.2f}"
                                    except:
                                        row_data[f"g={g_val:.1f}%"] = "Error"
                            sensitivity_data.append(row_data)
                    
                        sensitivity_df = pd.DataFrame(sensitivity_data)
                        st.dataframe(sensitivity_df, use_container_width=True)
                    
                        st.caption("Sensitivity table shows Fair Value per Share for different WACC and terminal growth rate combinations")
                
                with tab7:
                    st.subheader("üîç Comparative (Relative) Valuation")
                
                    if comp_tickers_listed:
                        # Use already calculated comp_results from top of page
                        if not comp_results:
                            with st.spinner("Fetching comparable companies data..."):
                                comp_results = perform_comparative_valuation(ticker, comp_tickers_listed, financials, shares, exchange_suffix, projections=projections)
                    
                        if comp_results:
                            # Show comparables table
                            st.markdown("### Comparable Companies")
                            comp_df = pd.DataFrame(comp_results['comparables'])
                            if not comp_df.empty:
                                display_comp_df = comp_df[['ticker', 'name', 'price', 'pe', 'pb', 'ps', 'ev_ebitda', 'ev_sales']]
                                st.dataframe(display_comp_df.style.format({
                                    'price': '‚Çπ{:.2f}',
                                    'pe': '{:.2f}x',
                                    'pb': '{:.2f}x',
                                    'ps': '{:.2f}x',
                                    'ev_ebitda': '{:.2f}x',
                                    'ev_sales': '{:.2f}x'
                                }), use_container_width=True)
                        
                            # Show multiples statistics
                            st.markdown("### Peer Multiples Statistics")
                            for multiple, stats in comp_results['multiples_stats'].items():
                                with st.expander(f"üìä {multiple.upper()} - Avg: {stats['average']:.2f}x, Median: {stats['median']:.2f}x"):
                                    st.write(f"**Range:** {stats['min']:.2f}x - {stats['max']:.2f}x")
                                    st.write(f"**Std Dev:** {stats['std']:.2f}x")
                                    st.write(f"**Peer Values:** {', '.join([f'{v:.2f}x' for v in stats['values']])}")
                        
                            # Show implied valuations
                            st.markdown("### Implied Fair Values")
                        
                            all_avg_values = []
                            all_median_values = []
                        
                            for method_key, val_data in comp_results['valuations'].items():
                                st.markdown(f"#### {val_data['method']}")
                            
                                col1, col2 = st.columns(2)
                            
                                with col1:
                                    st.markdown("**Using Average Multiple:**")
                                    st.write(val_data['formula_avg'])
                                    st.metric("Fair Value (Avg)", f"‚Çπ{val_data['fair_value_avg']:.2f}", 
                                            f"{val_data['upside_avg']:.1f}%" if val_data['current_price'] else None)
                                    all_avg_values.append(val_data['fair_value_avg'])
                            
                                with col2:
                                    st.markdown("**Using Median Multiple:**")
                                    st.write(val_data['formula_median'])
                                    st.metric("Fair Value (Median)", f"‚Çπ{val_data['fair_value_median']:.2f}",
                                            f"{val_data['upside_median']:.1f}%" if val_data['current_price'] else None)
                                    all_median_values.append(val_data['fair_value_median'])
                            
                                st.markdown("---")
                        
                            # Forward P/E if available
                            if 'forward_pe' in comp_results:
                                st.markdown("#### üìÖ 12-Month Forward P/E Valuation")
                                fpe = comp_results['forward_pe']
                            
                                st.info(f"**{fpe.get('calculation_note', '12-Month Forward EPS projection')}**")
                            
                                col1, col2, col3 = st.columns(3)
                                with col1:
                                    st.metric("Current EPS", f"‚Çπ{fpe.get('current_eps', 0):.2f}")
                                with col2:
                                    st.metric("Forward EPS (12M)", f"‚Çπ{fpe['forward_eps']:.2f}",
                                            delta=f"{fpe.get('earnings_growth_rate', 0):.1f}%",
                                            help="Projected EPS for next 12 months")
                                with col3:
                                    st.metric("Growth Rate", f"{fpe.get('earnings_growth_rate', 0):.1f}%")
                            
                                st.markdown("---")
                            
                                col1, col2 = st.columns(2)
                                with col1:
                                    st.write(fpe['formula_avg'])
                                    st.metric("Forward Fair Value (Avg)", f"‚Çπ{fpe['fair_value_avg']:.2f}")
                                with col2:
                                    st.write(fpe['formula_median'])
                                    st.metric("Forward Fair Value (Median)", f"‚Çπ{fpe['fair_value_median']:.2f}")
                            
                                st.markdown("---")
                            # Summary statistics
                            if all_avg_values and all_median_values:
                                st.markdown("### üìà Comparative Valuation Summary")
                            
                                col1, col2, col3 = st.columns(3)
                            
                                with col1:
                                    st.metric("Average (All Methods)", f"‚Çπ{np.mean(all_avg_values):.2f}")
                                    st.metric("Median (All Methods)", f"‚Çπ{np.median(all_median_values):.2f}")
                            
                                with col2:
                                    st.metric("Min Fair Value", f"‚Çπ{min(all_avg_values + all_median_values):.2f}")
                                    st.metric("Max Fair Value", f"‚Çπ{max(all_avg_values + all_median_values):.2f}")
                            
                                with col3:
                                    if valuation['fair_value_per_share'] > 0:
                                        st.metric("DCF Fair Value", f"‚Çπ{valuation['fair_value_per_share']:.2f}")
                                        combined_avg = (np.mean(all_avg_values) + valuation['fair_value_per_share']) / 2
                                        st.metric("DCF + Comp Avg", f"‚Çπ{combined_avg:.2f}")
                        
                            # VISUAL ANALYSIS - CHARTS
                            st.markdown("---")
                            st.markdown("### üìä Visual Analysis")
                        
                            comp_df_charts = pd.DataFrame(comp_results['comparables'])
                            if not comp_df_charts.empty:
                                from plotly.subplots import make_subplots
                            
                                # Chart 1: Multiples Comparison (4-in-1)
                                fig1 = make_subplots(rows=2, cols=2, subplot_titles=('P/E Ratio', 'P/B Ratio', 'EV/EBITDA', 'P/S Ratio'))
                                fig1.add_trace(go.Bar(x=comp_df_charts['name'], y=comp_df_charts['pe'], marker_color='steelblue', showlegend=False), row=1, col=1)
                                if 'pe' in comp_results['multiples_stats']:
                                    fig1.add_hline(y=comp_results['multiples_stats']['pe']['average'], line_dash="dash", line_color="red", row=1, col=1)
                                fig1.add_trace(go.Bar(x=comp_df_charts['name'], y=comp_df_charts['pb'], marker_color='lightcoral', showlegend=False), row=1, col=2)
                                if 'pb' in comp_results['multiples_stats']:
                                    fig1.add_hline(y=comp_results['multiples_stats']['pb']['average'], line_dash="dash", line_color="red", row=1, col=2)
                                fig1.add_trace(go.Bar(x=comp_df_charts['name'], y=comp_df_charts['ev_ebitda'], marker_color='mediumseagreen', showlegend=False), row=2, col=1)
                                if 'ev_ebitda' in comp_results['multiples_stats']:
                                    fig1.add_hline(y=comp_results['multiples_stats']['ev_ebitda']['average'], line_dash="dash", line_color="red", row=2, col=1)
                                fig1.add_trace(go.Bar(x=comp_df_charts['name'], y=comp_df_charts['ps'], marker_color='plum', showlegend=False), row=2, col=2)
                                if 'ps' in comp_results['multiples_stats']:
                                    fig1.add_hline(y=comp_results['multiples_stats']['ps']['average'], line_dash="dash", line_color="red", row=2, col=2)
                                fig1.update_layout(height=700, title_text="Peer Valuation Multiples Comparison")
                                fig1.update_xaxes(tickangle=-45)
                                st.plotly_chart(fig1, use_container_width=True)
                            
                                # Chart 2: Financial Metrics
                                fig2 = go.Figure()
                                fig2.add_trace(go.Bar(name='Revenue', x=comp_df_charts['name'], y=comp_df_charts['revenue']/1e7, marker_color='steelblue'))
                                fig2.add_trace(go.Bar(name='EBITDA', x=comp_df_charts['name'], y=comp_df_charts['ebitda']/1e7, marker_color='lightcoral'))
                                fig2.add_trace(go.Bar(name='Net Income', x=comp_df_charts['name'], y=comp_df_charts['net_income']/1e7, marker_color='mediumseagreen'))
                                fig2.update_layout(title="Financial Metrics (‚Çπ Crores)", barmode='group', height=500, xaxis_tickangle=-45)
                                st.plotly_chart(fig2, use_container_width=True)
                            
                                # Chart 3: Implied Valuations
                                if comp_results['valuations']:
                                    val_methods = [m.upper().replace('_', ' ') for m in comp_results['valuations'].keys()]
                                    val_avg = [d['fair_value_avg'] for d in comp_results['valuations'].values()]
                                    val_median = [d['fair_value_median'] for d in comp_results['valuations'].values()]
                                    fig3 = go.Figure()
                                    fig3.add_trace(go.Bar(name='Avg', x=val_methods, y=val_avg, marker_color='steelblue'))
                                    fig3.add_trace(go.Bar(name='Median', x=val_methods, y=val_median, marker_color='lightcoral'))
                                    if comp_results['target'].get('current_price', 0) > 0:
                                        fig3.add_hline(y=comp_results['target']['current_price'], line_dash="dash", line_color="green", annotation_text="Current")
                                    if valuation.get('fair_value_per_share', 0) > 0:
                                        fig3.add_hline(y=valuation['fair_value_per_share'], line_dash="dot", line_color="purple", annotation_text="DCF")
                                    fig3.update_layout(title="Implied Valuations", barmode='group', yaxis_title="Price (‚Çπ)", height=500)
                                    st.plotly_chart(fig3, use_container_width=True)
                
                    else:
                        st.info("Enter comparable tickers above to see relative valuation")
                with tab8:
                    st.subheader("üè¢ Advanced Peer Comparison Dashboard")
                
                    if comp_tickers_listed:
                        from peer_comparison_charts import create_peer_comparison_dashboard
                        create_peer_comparison_dashboard(ticker, comp_tickers_listed)
                    else:
                        st.info("üí° Click 'Auto-Fetch Peers' button above or enter peer tickers manually to see detailed peer comparison with 3D visualizations")
                
                with tab9:
                    st.subheader("üí∞ Dividend Discount Model (DDM)")
                    st.caption("Gordon Growth Model for dividend-paying companies")
                
                    # DDM Section
                    if ddm_result and ddm_result.get('value_per_share', 0) > 0:
                        # Top metrics
                        col1, col2, col3, col4 = st.columns(4)
                        with col1:
                            st.metric("Fair Value/Share", f"‚Çπ{ddm_result.get('value_per_share', 0):.2f}",
                                    delta=f"{((ddm_result.get('value_per_share', 0) - current_price) / current_price * 100):.1f}%" if current_price > 0 else None)
                        with col2:
                            st.metric("Current DPS", f"‚Çπ{ddm_result.get('current_dps', 0):.2f}")
                        with col3:
                            st.metric("Dividend Growth", f"{ddm_result.get('dividend_growth', 0):.1f}%")
                        with col4:
                            st.metric("Payout Ratio", f"{ddm_result.get('payout_ratio', 0):.1f}%")
                    
                        st.markdown("---")
                    
                        col_a, col_b = st.columns(2)
                    
                        with col_a:
                            st.markdown("**üìä Model Parameters**")
                            st.write(f"‚Ä¢ **Current Dividend per Share:** ‚Çπ{ddm_result.get('current_dps', 0):.2f}")
                            st.write(f"‚Ä¢ **Next Year DPS (D1):** ‚Çπ{ddm_result.get('next_year_dps', 0):.2f}")
                            st.write(f"‚Ä¢ **Dividend Growth Rate:** {ddm_result.get('dividend_growth', 0):.2f}%")
                            st.write(f"‚Ä¢ **Required Return (Ke):** {ddm_result.get('required_return', 0):.2f}%")
                            st.write(f"‚Ä¢ **Payout Ratio:** {ddm_result.get('payout_ratio', 0):.1f}%")
                        
                            if ddm_result.get('using_actual_data'):
                                st.success("‚úÖ Using actual dividend history from market data")
                            else:
                                st.info("‚ÑπÔ∏è Using estimated dividends based on earnings")
                    
                        with col_b:
                            st.markdown("**üìê Gordon Growth Model Formula**")
                            st.markdown("")  # Spacing
                        
                            # Larger, more readable formula
                            st.latex(r"\huge P_0 = \frac{D_1}{r - g}")
                        
                            st.markdown("")  # Spacing
                            st.markdown("**Where:**")
                        
                            # Better formatted definitions with boxes
                            st.markdown(f"""
                            - **P‚ÇÄ** (Fair Value per Share) = **‚Çπ{ddm_result.get('value_per_share', 0):.2f}**
                            - **D‚ÇÅ** (Next Year Dividend) = ‚Çπ{ddm_result.get('next_year_dps', 0):.2f}
                            - **r** (Required Return/Cost of Equity) = {ddm_result.get('required_return', 0):.2f}%
                            - **g** (Dividend Growth Rate) = {ddm_result.get('dividend_growth', 0):.2f}%
                            """)
                        
                            st.markdown("")  # Spacing
                            calc_fv = ddm_result.get('next_year_dps', 0) / ((ddm_result.get('required_return', 10) - ddm_result.get('dividend_growth', 5)) / 100) if (ddm_result.get('required_return', 10) - ddm_result.get('dividend_growth', 5)) > 0 else 0
                        
                            st.markdown("**üí° Calculation:**")
                            st.code(f"""
    Fair Value = D‚ÇÅ / (r - g)
               = ‚Çπ{ddm_result.get('next_year_dps', 0):.2f} / ({ddm_result.get('required_return', 0):.2f}% - {ddm_result.get('dividend_growth', 0):.2f}%)
               = ‚Çπ{ddm_result.get('next_year_dps', 0):.2f} / {(ddm_result.get('required_return', 10) - ddm_result.get('dividend_growth', 5)):.2f}%
               = ‚Çπ{calc_fv:.2f}
                                """, language="text")
                        
                            # 5-year dividend projections
                            if 'projections' in ddm_result and ddm_result.get('projections'):
                                st.markdown("---")
                                st.markdown("**üìà 5-Year Dividend Projections**")
                            
                                proj_df = pd.DataFrame(ddm_result['projections'])
                            
                                # Handle source column if present
                                if 'source' in proj_df.columns:
                                    proj_display = pd.DataFrame({
                                        'Year': proj_df['year'],
                                        'Dividend per Share': proj_df['dividend'].apply(lambda x: f"‚Çπ{x:.2f}"),
                                        'PV of Dividend': proj_df['pv_dividend'].apply(lambda x: f"‚Çπ{x:.2f}"),
                                        'Source': proj_df['source']
                                    })
                                else:
                                    proj_display = pd.DataFrame({
                                        'Year': proj_df['year'],
                                        'Dividend per Share': proj_df['dividend'].apply(lambda x: f"‚Çπ{x:.2f}"),
                                        'PV of Dividend': proj_df['pv_dividend'].apply(lambda x: f"‚Çπ{x:.2f}")
                                    })
                            
                                st.dataframe(proj_display, use_container_width=True, hide_index=True)
                        
                            if ddm_result.get('historical_dividends'):
                                st.markdown("---")
                                st.markdown("**üìä Historical Dividends**")
                                hist_divs = ddm_result['historical_dividends']
                                st.write(f"Recent dividend history: {', '.join([f'‚Çπ{d:.2f}' for d in hist_divs])}")
                        
                            # When to use DDM
                            st.markdown("---")
                            st.markdown("**üí° When to use DDM:**")
                            st.write("‚úÖ Company pays regular dividends")
                            st.write("‚úÖ Stable dividend payout history")
                            st.write("‚úÖ Mature, established companies")
                            st.write("‚úÖ Predictable dividend growth")
                    
                    elif ddm_result == None or not ddm_result.get('value_per_share'):
                        st.warning("‚ö†Ô∏è DDM not applicable for this company")
                        if ddm_result and 'reason' in ddm_result:
                            st.info(f"**Reason:** {ddm_result['reason']}")
                        st.caption("DDM requires the company to pay regular dividends. For non-dividend paying companies, focus on DCF and RIM models.")
                    
                        st.markdown("---")
                        st.markdown("**üí° When DDM is applicable:**")
                        st.write("‚úÖ Company has a history of paying dividends")
                        st.write("‚úÖ Dividend payments are stable and predictable")
                        st.write("‚úÖ Company is in mature growth stage")
                        st.write("‚úÖ Required return > dividend growth rate")
                
                    with tab10:
                        st.subheader("üè¢ Residual Income Model (RIM)")
                        st.caption("Equity valuation based on book value and excess returns - FULL DISCLOSURE")
                    
                        # RIM Section with COMPLETE transparency
                        if rim_result and rim_result.get('value_per_share', 0) > 0:
                            # Top-level result
                            st.success(f"### üéØ Fair Value per Share (RIM): ‚Çπ{rim_result['value_per_share']:.2f}")
                            
                            # Top metrics
                            col1, col2, col3, col4 = st.columns(4)
                            with col1:
                                st.metric("Fair Value/Share", f"‚Çπ{rim_result['value_per_share']:.2f}",
                                        delta=f"{((rim_result['value_per_share'] - current_price) / current_price * 100):.1f}%" if current_price > 0 else None)
                            with col2:
                                st.metric("Book Value/Share", f"‚Çπ{rim_result['book_value_per_share']:.2f}")
                            with col3:
                                st.metric("ROE", f"{rim_result['roe']:.2f}%")
                            with col4:
                                st.metric("BV Growth", f"{rim_result.get('bv_growth', 10):.1f}%")
                        
                            st.markdown("---")
                        
                            # SECTION 1: FORMULAS
                            st.markdown("### üìê RIM Formulas & Methodology")
                            
                            col_formula1, col_formula2 = st.columns(2)
                            
                            with col_formula1:
                                st.markdown("**Main Valuation Formula:**")
                                st.latex(r"P_0 = BV_0 + \sum_{t=1}^{n} \frac{RI_t}{(1+r)^t} + \frac{TV}{(1+r)^n}")
                                
                                st.markdown("**Residual Income Formula:**")
                                st.latex(r"RI_t = NI_t - (r \times BV_{t-1})")
                                
                                st.markdown("**Terminal Value Formula:**")
                                st.latex(r"TV = \frac{RI_n \times (1+g)}{r - g}")
                            
                            with col_formula2:
                                st.markdown("**Where:**")
                                st.markdown(f"""
                                - **P‚ÇÄ** = Fair Value per Share = **‚Çπ{rim_result['value_per_share']:.2f}**
                                - **BV‚ÇÄ** = Current Book Value = **‚Çπ{rim_result['current_book_value']:,.0f}**
                                - **RI** = Residual Income (excess return)
                                - **NI** = Net Income (projected)
                                - **r** = Cost of Equity = **{rim_result['cost_of_equity']:.2f}%**
                                - **g** = Terminal Growth = **{rim_result['terminal_growth']:.2f}%**
                                - **n** = Projection Period = **5 years**
                                """)
                            
                            st.markdown("---")
                            
                            # SECTION 2: INPUT PARAMETERS
                            st.markdown("### üìä Input Parameters & Assumptions")
                            
                            col_input1, col_input2, col_input3 = st.columns(3)
                            
                            with col_input1:
                                st.markdown("**Current State:**")
                                st.write(f"‚Ä¢ Book Value (Total): ‚Çπ{rim_result['current_book_value']:,.0f}")
                                st.write(f"‚Ä¢ Book Value/Share: ‚Çπ{rim_result['book_value_per_share']:.2f}")
                                st.write(f"‚Ä¢ Current EPS: ‚Çπ{rim_result['current_eps']:.2f}")
                                st.write(f"‚Ä¢ Number of Shares: {shares:,.0f}")
                            
                            with col_input2:
                                st.markdown("**Profitability:**")
                                st.write(f"‚Ä¢ Return on Equity: {rim_result['roe']:.2f}%")
                                st.write(f"‚Ä¢ Cost of Equity: {rim_result['cost_of_equity']:.2f}%")
                                excess_return = rim_result['roe'] - rim_result['cost_of_equity']
                                st.write(f"‚Ä¢ Excess Return: {excess_return:.2f}%")
                            
                            with col_input3:
                                st.markdown("**Growth Rates:**")
                                st.write(f"‚Ä¢ Book Value Growth: {rim_result.get('bv_growth', 10):.2f}%")
                                st.write(f"‚Ä¢ Terminal Growth: {rim_result['terminal_growth']:.2f}%")
                                if rim_result.get('using_dcf_projections'):
                                    st.info("‚úÖ Using DCF Projected NOPAT")
                                else:
                                    st.info("üìä Using ROE-based projection")
                            
                            st.markdown("---")
                            
                            # SECTION 3: YEAR-BY-YEAR PROJECTIONS WITH VISUALS
                            st.markdown("### üìà Year-by-Year Residual Income Projections")
                            
                            projections = rim_result.get('projections', [])
                            if projections and len(projections) > 0:
                                # Create detailed projection table
                                proj_data = []
                                years_list = []
                                bv_list = []
                                ni_list = []
                                ri_list = []
                                pv_ri_list = []
                                
                                for proj in projections:
                                    year = proj.get('year', 0)
                                    bv_year = proj.get('book_value', 0) / 100000  # Convert to Lacs
                                    ni_year = proj.get('net_income', 0) / 100000
                                    ri_year = proj.get('residual_income', 0) / 100000
                                    pv_ri_year = proj.get('pv_ri', 0) / 100000
                                    req_return = bv_year * rim_result['cost_of_equity'] / 100
                                    
                                    years_list.append(f"Year {year}")
                                    bv_list.append(bv_year)
                                    ni_list.append(ni_year)
                                    ri_list.append(ri_year)
                                    pv_ri_list.append(pv_ri_year)
                                    
                                    proj_data.append({
                                        'Year': f"Year {year}",
                                        'Book Value (‚Çπ Lacs)': f"{bv_year:.2f}",
                                        'Net Income (‚Çπ Lacs)': f"{ni_year:.2f}",
                                        'Required Return (‚Çπ Lacs)': f"{req_return:.2f}",
                                        'Residual Income (‚Çπ Lacs)': f"{ri_year:.2f}",
                                        'Discount Factor': f"{1 / ((1 + rim_result['cost_of_equity']/100) ** year):.4f}",
                                        'PV of RI (‚Çπ Lacs)': f"{pv_ri_year:.2f}"
                                    })
                                
                                # Display table
                                proj_df = pd.DataFrame(proj_data)
                                st.dataframe(proj_df, use_container_width=True, hide_index=True)
                                
                                # VISUAL 1: Residual Income by Year (Bar Chart)
                                st.markdown("#### üìä Visual: Residual Income by Year")
                                
                                fig_ri = go.Figure()
                                
                                # Residual Income bars
                                colors = ['#06A77D' if ri > 0 else '#E63946' for ri in ri_list]
                                fig_ri.add_trace(go.Bar(
                                    x=years_list,
                                    y=ri_list,
                                    name='Residual Income',
                                    marker_color=colors,
                                    text=[f"‚Çπ{ri:.2f}" for ri in ri_list],
                                    textposition='outside'
                                ))
                                
                                fig_ri.update_layout(
                                    title="Residual Income by Year (‚Çπ Lacs)",
                                    xaxis_title="Year",
                                    yaxis_title="Residual Income (‚Çπ Lacs)",
                                    height=320,
                                    showlegend=False,
                                    hovermode='x unified'
                                )
                                
                                st.plotly_chart(fig_ri, use_container_width=True)
                                
                                # VISUAL 2: Present Value Contribution (Bar Chart)
                                st.markdown("#### üí∞ Visual: Present Value Contributions")
                                
                                fig_pv = go.Figure()
                                
                                fig_pv.add_trace(go.Bar(
                                    x=years_list,
                                    y=pv_ri_list,
                                    name='PV of RI',
                                    marker_color='#2E86AB',
                                    text=[f"‚Çπ{pv:.2f}" for pv in pv_ri_list],
                                    textposition='outside'
                                ))
                                
                                fig_pv.update_layout(
                                    title="Present Value Contribution by Year (‚Çπ Lacs)",
                                    xaxis_title="Year",
                                    yaxis_title="PV of RI (‚Çπ Lacs)",
                                    height=320,
                                    showlegend=False
                                )
                                
                                st.plotly_chart(fig_pv, use_container_width=True)
                            else:
                                st.warning("‚ö†Ô∏è No year-by-year projection data available.")
                                
                                # Show ACTUAL calculations for ALL years
                                st.markdown("### üî¢ Detailed Calculations for Each Year")
                                
                                for idx, proj in enumerate(projections):
                                    year = proj.get('year', 0)
                                    bv_year = proj.get('book_value', 0) / 100000
                                    ni_year = proj.get('net_income', 0) / 100000
                                    req_ret_year = bv_year * rim_result['cost_of_equity'] / 100
                                    ri_year = proj.get('residual_income', 0) / 100000
                                    pv_ri_year = proj.get('pv_ri', 0) / 100000
                                    discount_factor = 1 / ((1 + rim_result['cost_of_equity']/100) ** year)
                                    
                                    with st.expander(f"üìä Year {year} Calculation", expanded=(idx==0)):
                                        st.code(f"""
YEAR {year} CALCULATIONS:
{'='*60}

Step 1: Book Value
    Book Value (Year {year}) = ‚Çπ{bv_year:.2f} Lacs

Step 2: Net Income  
    Net Income (Year {year}) = ‚Çπ{ni_year:.2f} Lacs

Step 3: Required Return (Equity Charge)
    Required Return = Book Value √ó Cost of Equity
                   = ‚Çπ{bv_year:.2f} √ó {rim_result['cost_of_equity']:.2f}%
                   = ‚Çπ{req_ret_year:.2f} Lacs

Step 4: Residual Income (Excess Profit)
    Residual Income = Net Income - Required Return
                   = ‚Çπ{ni_year:.2f} - ‚Çπ{req_ret_year:.2f}
                   = ‚Çπ{ri_year:.2f} Lacs
    
    {'‚úÖ CREATING VALUE' if ri_year > 0 else '‚ö†Ô∏è DESTROYING VALUE'} - {'Company earns more than required return' if ri_year > 0 else 'Company earns less than required return'}

Step 5: Present Value (Discount to Today)
    Discount Factor = 1 / (1 + Ke)^{year}
                   = 1 / (1 + {rim_result['cost_of_equity']/100:.4f})^{year}
                   = {discount_factor:.4f}
    
    PV of RI = Residual Income √ó Discount Factor
            = ‚Çπ{ri_year:.2f} √ó {discount_factor:.4f}
            = ‚Çπ{pv_ri_year:.2f} Lacs

{'='*60}
CONTRIBUTION TO FAIR VALUE: ‚Çπ{pv_ri_year:.2f} Lacs
                                        """, language="text")
                                
                                # TERMINAL VALUE CALCULATION
                                st.markdown("### üéØ Terminal Value Calculation")
                                
                                if len(projections) > 0:
                                    last_proj = projections[-1]
                                    last_ri = last_proj.get('residual_income', 0) / 100000
                                    
                                    st.code(f"""
TERMINAL VALUE (Beyond Year 5):
{'='*60}

Step 1: Terminal Year Residual Income
    RI (Year 5) = ‚Çπ{last_ri:.2f} Lacs

Step 2: Grow at Terminal Growth Rate
    RI (Year 6) = RI (Year 5) √ó (1 + g)
                = ‚Çπ{last_ri:.2f} √ó (1 + {rim_result['terminal_growth']/100:.4f})
                = ‚Çπ{last_ri * (1 + rim_result['terminal_growth']/100):.2f} Lacs

Step 3: Perpetuity Value (Gordon Growth Model)
    Terminal Value = RI (Year 6) / (Ke - g)
                  = ‚Çπ{last_ri * (1 + rim_result['terminal_growth']/100):.2f} / ({rim_result['cost_of_equity']:.2f}% - {rim_result['terminal_growth']:.2f}%)
                  = ‚Çπ{last_ri * (1 + rim_result['terminal_growth']/100):.2f} / {rim_result['cost_of_equity'] - rim_result['terminal_growth']:.2f}%
                  = ‚Çπ{(last_ri * (1 + rim_result['terminal_growth']/100)) / ((rim_result['cost_of_equity'] - rim_result['terminal_growth']) / 100):.2f} Lacs

Step 4: Discount to Present Value
    Discount Factor = 1 / (1 + Ke)^5
                   = 1 / (1 + {rim_result['cost_of_equity']/100:.4f})^5
                   = {1 / ((1 + rim_result['cost_of_equity']/100) ** 5):.4f}
    
    PV of Terminal Value = TV √ó Discount Factor
                        = ‚Çπ{(last_ri * (1 + rim_result['terminal_growth']/100)) / ((rim_result['cost_of_equity'] - rim_result['terminal_growth']) / 100):.2f} √ó {1 / ((1 + rim_result['cost_of_equity']/100) ** 5):.4f}
                        = ‚Çπ{rim_result['terminal_ri_pv'] / 100000:.2f} Lacs

{'='*60}
TERMINAL VALUE CONTRIBUTION: ‚Çπ{rim_result['terminal_ri_pv'] / 100000:.2f} Lacs
                                    """, language="text")
                            
                            st.markdown("---")
                            
                            # SECTION 4: VALUE BUILD-UP WITH VISUALS
                            st.markdown("### üí∞ Fair Value Build-Up")
                            
                            pv_ri_per_share = (rim_result['sum_pv_ri'] / shares) if shares > 0 else 0
                            tv_per_share = (rim_result['terminal_ri_pv'] / shares) if shares > 0 else 0
                            
                            # VISUAL: Waterfall Chart for Value Build-up
                            st.markdown("#### üìä Visual: Fair Value Waterfall")
                            
                            fig_waterfall = go.Figure(go.Waterfall(
                                name="Fair Value",
                                orientation="v",
                                measure=["absolute", "relative", "relative", "total"],
                                x=["Book Value<br>per Share", "PV of RI<br>(Years 1-5)", "Terminal<br>Value", "Fair Value<br>per Share"],
                                textposition="outside",
                                text=[f"‚Çπ{rim_result['book_value_per_share']:.2f}", f"‚Çπ{pv_ri_per_share:.2f}", f"‚Çπ{tv_per_share:.2f}", f"‚Çπ{rim_result['value_per_share']:.2f}"],
                                y=[rim_result['book_value_per_share'], pv_ri_per_share, tv_per_share, rim_result['value_per_share']],
                                connector={"line": {"color": "rgb(63, 63, 63)"}},
                                decreasing={"marker": {"color": "#E63946"}},
                                increasing={"marker": {"color": "#06A77D"}},
                                totals={"marker": {"color": "#2E86AB"}}
                            ))
                            
                            fig_waterfall.update_layout(
                                title="Fair Value Build-Up (‚Çπ per Share)",
                                showlegend=False,
                                height=380,
                                yaxis_title="Value (‚Çπ)"
                            )
                            
                            st.plotly_chart(fig_waterfall, use_container_width=True)
                            
                            # VISUAL: Pie Chart for Value Composition
                            st.markdown("#### ü•ß Visual: Value Composition")
                            
                            # Check if we have negative components
                            has_negative = (pv_ri_per_share < 0 or tv_per_share < 0)
                            
                            if has_negative:
                                st.warning("‚ö†Ô∏è **Note:** Company has negative residual income (destroying value). Pie chart shows absolute values for visualization.")
                                
                                # Use absolute values for pie chart
                                fig_pie = go.Figure(data=[go.Pie(
                                    labels=['Book Value', 'PV of RI (5Y)', 'Terminal Value'],
                                    values=[abs(rim_result['book_value_per_share']), abs(pv_ri_per_share), abs(tv_per_share)],
                                    marker=dict(colors=['#2E86AB', '#E63946', '#F4D35E']),
                                    textinfo='label+percent',
                                    texttemplate='<b>%{label}</b><br>%{percent}',
                                    hovertemplate='<b>%{label}</b><br>‚Çπ%{value:.2f} (absolute)<extra></extra>'
                                )])
                                
                                fig_pie.update_layout(
                                    title="Value Components (Absolute Values)",
                                    height=380
                                )
                                
                                st.plotly_chart(fig_pie, use_container_width=True)
                                
                                # Show actual signed values in a bar chart instead
                                st.markdown("#### üìä Visual: Signed Value Components")
                                
                                fig_bar = go.Figure()
                                
                                components = ['Book Value', 'PV of RI (5Y)', 'Terminal Value', 'Total']
                                values = [rim_result['book_value_per_share'], pv_ri_per_share, tv_per_share, rim_result['value_per_share']]
                                colors_signed = ['#2E86AB', '#E63946' if pv_ri_per_share < 0 else '#06A77D', 
                                                '#E63946' if tv_per_share < 0 else '#06A77D',
                                                '#E63946' if rim_result['value_per_share'] < 0 else '#06A77D']
                                
                                fig_bar.add_trace(go.Bar(
                                    x=components,
                                    y=values,
                                    marker_color=colors_signed,
                                    text=[f"‚Çπ{v:.2f}" for v in values],
                                    textposition='outside'
                                ))
                                
                                fig_bar.update_layout(
                                    title="Fair Value Components (Actual Signed Values)",
                                    yaxis_title="Value per Share (‚Çπ)",
                                    height=320,
                                    showlegend=False
                                )
                                
                                st.plotly_chart(fig_bar, use_container_width=True)
                                
                            else:
                                # All positive - normal pie chart
                                fig_pie = go.Figure(data=[go.Pie(
                                    labels=['Book Value', 'PV of RI (5Y)', 'Terminal Value'],
                                    values=[rim_result['book_value_per_share'], pv_ri_per_share, tv_per_share],
                                    marker=dict(colors=['#2E86AB', '#06A77D', '#F4D35E']),
                                    textinfo='label+percent+value',
                                    texttemplate='<b>%{label}</b><br>%{percent}<br>‚Çπ%{value:.2f}',
                                    hovertemplate='<b>%{label}</b><br>‚Çπ%{value:.2f}<br>%{percent}<extra></extra>'
                                )])
                                
                                fig_pie.update_layout(
                                    title="Fair Value Composition (‚Çπ per Share)",
                                    height=380
                                )
                                
                                st.plotly_chart(fig_pie, use_container_width=True)
                                
                                st.plotly_chart(fig_pie, use_container_width=True)
                            
                            # Detailed breakdown in columns
                            col_buildup1, col_buildup2 = st.columns([2, 1])
                            
                            with col_buildup1:
                                st.code(f"""
FAIR VALUE CALCULATION (Per Share Basis):
==========================================

Starting Point:
    Current Book Value/Share              = ‚Çπ{rim_result['book_value_per_share']:.2f}

Add: Present Value of Residual Income (Years 1-5)
    Sum of PV(RI) per Share               = ‚Çπ{pv_ri_per_share:.2f}

Add: Terminal Value
    PV of Terminal RI per Share           = ‚Çπ{tv_per_share:.2f}

==========================================
FAIR VALUE PER SHARE                      = ‚Çπ{rim_result['value_per_share']:.2f}
                                """, language="text")
                            
                            with col_buildup2:
                                st.markdown("**Total Equity Value:**")
                                st.write(f"‚Ä¢ Book Value: ‚Çπ{rim_result['current_book_value']/100000:.2f} Lacs")
                                st.write(f"‚Ä¢ PV of RI (5Y): ‚Çπ{rim_result['sum_pv_ri']/100000:.2f} Lacs")
                                st.write(f"‚Ä¢ Terminal Value: ‚Çπ{rim_result['terminal_ri_pv']/100000:.2f} Lacs")
                                st.write(f"‚Ä¢ **Total**: ‚Çπ{rim_result['total_equity_value']/100000:.2f} Lacs")
                                st.write("")
                                st.write(f"√∑ Shares: {shares:,.0f}")
                                st.success(f"**= ‚Çπ{rim_result['value_per_share']:.2f} per share**")
                            
                            st.markdown("---")
                            
                            # SECTION 5: KEY INSIGHTS
                            st.markdown("### üí° Key Insights")
                            
                            col_insight1, col_insight2 = st.columns(2)
                            
                            with col_insight1:
                                st.markdown("**Value Composition:**")
                                total_value = rim_result['book_value_per_share'] + pv_ri_per_share + tv_per_share
                                if total_value > 0:
                                    bv_pct = (rim_result['book_value_per_share'] / total_value) * 100
                                    ri_pct = (pv_ri_per_share / total_value) * 100
                                    tv_pct = (tv_per_share / total_value) * 100
                                    
                                    st.write(f"‚Ä¢ Book Value: {bv_pct:.1f}% (‚Çπ{rim_result['book_value_per_share']:.2f})")
                                    st.write(f"‚Ä¢ RI (5 Years): {ri_pct:.1f}% (‚Çπ{pv_ri_per_share:.2f})")
                                    st.write(f"‚Ä¢ Terminal Value: {tv_pct:.1f}% (‚Çπ{tv_per_share:.2f})")
                            
                            with col_insight2:
                                st.markdown("**Economic Profit:**")
                                if rim_result['roe'] > rim_result['cost_of_equity']:
                                    st.success(f"‚úÖ Creating Value: ROE ({rim_result['roe']:.2f}%) > Ke ({rim_result['cost_of_equity']:.2f}%)")
                                    st.write(f"‚Ä¢ Excess return: {excess_return:.2f}%")
                                elif rim_result['roe'] < rim_result['cost_of_equity']:
                                    st.error(f"‚ö†Ô∏è Destroying Value: ROE ({rim_result['roe']:.2f}%) < Ke ({rim_result['cost_of_equity']:.2f}%)")
                                    st.write(f"‚Ä¢ Value deficit: {excess_return:.2f}%")
                                else:
                                    st.info(f"Earning exactly required return: ROE = Ke = {rim_result['roe']:.2f}%")
                            
                            st.markdown("---")
                            st.caption("üìò **RIM Model Note:** Residual Income Model values companies based on their ability to generate returns above the cost of equity. Positive residual income indicates value creation.")
                        
                            # 5-year RI projections
                            if 'projections' in rim_result and rim_result['projections']:
                                st.markdown("---")
                                st.markdown("**üìà 5-Year Residual Income Projections**")
                            
                                proj_df = pd.DataFrame(rim_result['projections'])
                            
                                # Build display with source column if available
                                display_dict = {
                                    'Year': proj_df['year'],
                                    'Book Value': proj_df['book_value'].apply(lambda x: f"‚Çπ{x:,.0f}"),
                                    'Net Income': proj_df['net_income'].apply(lambda x: f"‚Çπ{x:,.0f}"),
                                    'Residual Income': proj_df['residual_income'].apply(lambda x: f"‚Çπ{x:,.0f}"),
                                    'PV of RI': proj_df['pv_ri'].apply(lambda x: f"‚Çπ{x:,.0f}")
                                }
                            
                                if 'source' in proj_df.columns:
                                    display_dict['Source'] = proj_df['source']
                            
                                proj_display = pd.DataFrame(display_dict)
                            
                                st.dataframe(proj_display, use_container_width=True, hide_index=True)
                            
                                st.caption(f"Sum of PV(RI): ‚Çπ{rim_result['sum_pv_ri']:,.0f} | Terminal Value PV: ‚Çπ{rim_result['terminal_ri_pv']:,.0f}")
                        
                            # When to use RIM
                            st.markdown("---")
                            st.markdown("**üí° When to use RIM:**")
                            st.write("‚úÖ Any company with positive equity")
                            st.write("‚úÖ Banks & financial institutions")
                            st.write("‚úÖ Asset-intensive businesses")
                            st.write("‚úÖ Companies that retain earnings")
                    
                        else:
                            # RIM calculation failed - show detailed error
                            st.error("‚ö†Ô∏è RIM Not Applicable for This Company")
                        
                            if rim_result and rim_result.get('error'):
                                st.markdown("### üîç Diagnostic Information")
                            
                                col_err1, col_err2 = st.columns(2)
                            
                                with col_err1:
                                    st.markdown("**‚ùå Reason:**")
                                    st.write(rim_result.get('reason', 'Unknown error'))
                                
                                    if 'technical_details' in rim_result:
                                        with st.expander("üîß Technical Details"):
                                            st.code(rim_result['technical_details'])
                            
                                with col_err2:
                                    st.markdown("**üí° Suggestion:**")
                                    st.info(rim_result.get('suggestion', 'Use DCF or other valuation method'))
                            else:
                                st.warning("RIM calculation returned no results")
                        
                            st.markdown("---")
                            st.markdown("### ‚úÖ When RIM Works Best")
                            col_req1, col_req2 = st.columns(2)
                        
                            with col_req1:
                                st.markdown("**Requirements:**")
                                st.write("‚úÖ Positive book value (equity)")
                                st.write("‚úÖ Positive and stable ROE")
                                st.write("‚úÖ Profitable company")
                                st.write("‚úÖ Complete financial data")
                        
                            with col_req2:
                                st.markdown("**Best for:**")
                                st.write("üè¶ Banks & financial institutions")
                                st.write("üèóÔ∏è Asset-intensive businesses")
                                st.write("üìà Mature, stable companies")
                                st.write("üí∞ Companies retaining earnings")
                        
                            st.markdown("---")
                            st.markdown("### üéØ Alternative Valuation Methods")
                            st.write("Since RIM is not applicable, please use:")
                            st.write("‚Ä¢ **DCF (Tab 1)** - Primary valuation method")
                            st.write("‚Ä¢ **Comparative Valuation (Tab 7)** - P/E, P/B multiples")
                            st.write("‚Ä¢ **DDM (Tab 9)** - If company pays dividends")
                
                with tab11:
                    st.subheader("‚öôÔ∏è Assumptions & Parameters")
                    
                    # Compact Data Config
                    years_display = [str(y).replace('_', '') for y in financials['years']]
                    years_range = f"{years_display[-1]} - {years_display[0]}" if len(years_display) > 1 else years_display[0]
                    
                    col1, col2, col3 = st.columns(3)
                    with col1:
                        st.metric("üìÖ Historical Period", f"{len(financials['years'])} years", delta=years_range, delta_color="off")
                    with col2:
                        st.metric("üîÆ Projection Years", f"{projection_years_listed} years")
                    with col3:
                        st.metric("üè¢ Shares", f"{shares:,}")
                    
                    st.markdown("---")
                    
                    # WACC Bar Chart
                    st.markdown("### üí∞ WACC Components")
                    fig_wacc = go.Figure()
                    fig_wacc.add_trace(go.Bar(
                        x=['WACC', 'Ke', 'Kd', 'Terminal g'],
                        y=[wacc_details['wacc'], wacc_details['ke'], wacc_details['kd_after_tax'], terminal_growth],
                        marker=dict(color=['#667eea', '#764ba2', '#f093fb', '#f5576c']),
                        text=[f"{wacc_details['wacc']:.2f}%", f"{wacc_details['ke']:.2f}%", 
                              f"{wacc_details['kd_after_tax']:.2f}%", f"{terminal_growth:.2f}%"],
                        textposition='auto'
                    ))
                    fig_wacc.update_layout(height=220, margin=dict(l=20, r=20, t=10, b=20), showlegend=False,
                                          yaxis=dict(title='%'), paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)')
                    st.plotly_chart(fig_wacc, use_container_width=True)
                    
                    col_a, col_b = st.columns(2)
                    with col_a:
                        rf_display = f"{wacc_details['rf']:.2f}%"
                        if abs(wacc_details['rf'] - 6.75) < 0.1:
                            rf_display += " (Fallback)"
                        st.caption(f"**Tax:** {tax_rate:.2f}% | **Beta:** {wacc_details['beta']:.3f} | **Rf:** {rf_display}")
                    with col_b:
                        st.caption(f"**Equity Wt:** {wacc_details['we']:.1f}% | **Debt Wt:** {wacc_details['wd']:.1f}%")
                    
                    st.markdown("---")
                    
                    # Projection Drivers - Compact Expander
                    st.markdown("### üìä Projection Drivers")
                    with st.expander("View Parameters", expanded=False):
                        if drivers:
                            drivers_data = [
                                ['Revenue Growth', f"{drivers.get('revenue_growth_rate', 0):.2f}%", 'Historical CAGR'],
                                ['EBITDA Margin', f"{drivers.get('ebitda_margin', 0):.2f}%", 'Historical Avg'],
                                ['CapEx/Revenue', f"{drivers.get('capex_ratio', 0):.2f}%", 'Historical Avg']
                            ]
                            if wc_metrics.get('avg_inv_days', 0) > 0:
                                drivers_data.append(['Inventory Days', f"{wc_metrics['avg_inv_days']:.0f}", 'Historical'])
                            if wc_metrics.get('avg_deb_days', 0) > 0:
                                drivers_data.append(['Debtor Days', f"{wc_metrics['avg_deb_days']:.0f}", 'Historical'])
                            if wc_metrics.get('avg_cred_days', 0) > 0:
                                drivers_data.append(['Creditor Days', f"{wc_metrics['avg_cred_days']:.0f}", 'Historical'])
                            
                            st.dataframe(pd.DataFrame(drivers_data, columns=['Parameter', 'Value', 'Source']),
                                       use_container_width=True, hide_index=True, height=200)
                
                # Stock Price Comparison Tab (if enabled)
                if tab_stock_nonbank and enable_stock_comparison_state and STOCK_COMPARISON_AVAILABLE:
                    with tab_stock_nonbank:
                        st.subheader("üìà Stock Price vs Revenue & EPS Analysis")
                        
                        with st.spinner("Fetching stock price data..."):
                            try:
                                # Determine years to fetch (max 4 for Yahoo Finance)
                                years_to_fetch = min(historical_years_listed, 4)
                                
                                # Fetch comparison data
                                stock_comp_data = get_stock_comparison_data_listed(
                                    ticker=full_ticker,
                                    company_name=company_name,
                                    financials=financials,
                                    num_years=years_to_fetch
                                )
                                
                                if stock_comp_data and stock_comp_data['chart_fig']:
                                    st.plotly_chart(stock_comp_data['chart_fig'], use_container_width=True)
                                    
                                    # Show data tables in expanders
                                    with st.expander("üìä View Raw Data"):
                                        col1, col2, col3 = st.columns(3)
                                        
                                        with col1:
                                            st.markdown("**Revenue Data**")
                                            if stock_comp_data['revenue_df'] is not None:
                                                st.dataframe(stock_comp_data['revenue_df'], hide_index=True)
                                            else:
                                                st.info("No revenue data available")
                                        
                                        with col2:
                                            st.markdown("**EPS Data**")
                                            if stock_comp_data['eps_df'] is not None:
                                                st.dataframe(stock_comp_data['eps_df'], hide_index=True)
                                            else:
                                                st.info("No EPS data available")
                                        
                                        with col3:
                                            st.markdown("**Stock Price Summary**")
                                            if stock_comp_data['stock_prices_df'] is not None:
                                                price_df = stock_comp_data['stock_prices_df']
                                                st.metric("Latest Price", f"‚Çπ{price_df['Close'].iloc[-1]:.2f}")
                                                st.metric("Period Return", f"{((price_df['Close'].iloc[-1] - price_df['Close'].iloc[0]) / price_df['Close'].iloc[0] * 100):.2f}%")
                                                st.metric("Major Changes", f"{price_df['is_major'].sum()}" if 'is_major' in price_df.columns else "N/A")
                                            else:
                                                st.info("No stock price data available")
                                    
                                    st.info(f"üí° Chart shows {years_to_fetch} years of data (Yahoo Finance limit: 4 years max)")
                                else:
                                    st.warning("Could not generate stock comparison chart. Check if ticker and financial data are available.")
                            
                            except Exception as e:
                                st.error(f"Error generating stock comparison: {str(e)}")
    
    
    elif mode == "Unlisted Company (Excel Upload)":
        st.subheader("üìÑ Unlisted Company Valuation")
        
        # ===== RISK-FREE RATE TICKER - AT THE TOP, ALWAYS VISIBLE =====
        st.markdown("### üèõÔ∏è Risk-Free Rate Configuration")
        st.info("üí° **Default: 6.83%**. Enter any Yahoo Finance ticker and click Fetch. Common options: ^TNX (US 10Y), ^IRX (US 3-month), or any stock/index for custom rates.")
        
        # Initialize if not exists
        if 'cached_rf_rate_unlisted' not in st.session_state:
            st.session_state.cached_rf_rate_unlisted = 6.83
        
        rf_col1, rf_col2, rf_col3 = st.columns([3, 2, 1])
        with rf_col1:
            custom_rf_ticker_unlisted = st.text_input(
                "Yahoo Finance Ticker for Risk-Free Rate",
                value="NIFTYGS10YR.NS",
                key='custom_rf_ticker_unlisted_top',
                help="Enter any ticker - will calculate CAGR for stocks/indices, or use value directly for yields. Examples: NIFTYGS10YR.NS, ^TNX, RELIANCE.NS"
            )
        with rf_col2:
            st.metric("Current RF Rate", f"{st.session_state.cached_rf_rate_unlisted:.2f}%")
        with rf_col3:
            st.write("")
            st.write("")
            
            # Add a counter to track button clicks
            if 'rf_fetch_click_count_unlisted' not in st.session_state:
                st.session_state.rf_fetch_click_count_unlisted = 0
            
            if st.button("üîÑ Fetch", key='refresh_rf_unlisted_top'):
                # Increment click counter FIRST
                st.session_state.rf_fetch_click_count_unlisted += 1
                
                # Store debug output in session state so it persists across reruns
                debug_output = []
                debug_output.append(f"üîÑ **FETCH BUTTON CLICKED #{st.session_state.rf_fetch_click_count_unlisted} - UNLISTED MODE**")
                debug_output.append(f"üìù Input ticker: `{custom_rf_ticker_unlisted}`")
                
                ticker_to_use = custom_rf_ticker_unlisted.strip() if custom_rf_ticker_unlisted.strip() else None
                debug_output.append(f"üìù Ticker to use (after strip): `{ticker_to_use}`")
                
                debug_output.append("‚è≥ Calling get_risk_free_rate()...")
                fetched_rate, fetch_debug = get_risk_free_rate(ticker_to_use)
                
                # Add all debug messages from the function
                debug_output.extend(fetch_debug)
                
                debug_output.append(f"‚úÖ Function returned: {fetched_rate}%")
                
                debug_output.append(f"üíæ Updating session state...")
                debug_output.append(f"   - Before: {st.session_state.get('cached_rf_rate_unlisted', 'NOT SET')}")
                st.session_state.cached_rf_rate_unlisted = fetched_rate
                debug_output.append(f"   - After: {st.session_state.cached_rf_rate_unlisted}")
                
                # Force update the manual input field
                debug_output.append(f"üîÑ Clearing manual input widget state...")
                if 'manual_rf_unlisted' in st.session_state:
                    del st.session_state['manual_rf_unlisted']
                    debug_output.append(f"   - Widget state cleared")
                else:
                    debug_output.append(f"   - Widget state was not set")
                
                # Store debug output in session state
                st.session_state.rf_fetch_debug_unlisted = debug_output
                st.session_state.rf_fetch_success_unlisted = True
                
                # Rerun to update UI
                st.rerun()
        
        # Display debug output if available (persists across reruns)
        if st.session_state.get('rf_fetch_debug_unlisted'):
            with st.expander("üìã Last Fetch Debug Output", expanded=True):
                for line in st.session_state.rf_fetch_debug_unlisted:
                    st.write(line)
                if st.session_state.get('rf_fetch_success_unlisted'):
                    st.success(f"‚úì Successfully updated to {st.session_state.cached_rf_rate_unlisted:.2f}%")
        
        st.markdown("---")
        # ===== END RF RATE CONFIG =====
        
        # DEBUG: Show session state
        with st.expander("üêõ DEBUG: Session State Values (Unlisted)", expanded=False):
            st.write("**Current Session State for RF Rate:**")
            st.write(f"- `cached_rf_rate_unlisted`: {st.session_state.get('cached_rf_rate_unlisted', 'NOT SET')}")
            st.write(f"- `manual_rf_unlisted` widget: {st.session_state.get('manual_rf_unlisted', 'NOT SET')}")
            st.write(f"- `custom_rf_ticker_unlisted_top`: {st.session_state.get('custom_rf_ticker_unlisted_top', 'NOT SET')}")
            st.write(f"- **Fetch button clicks**: {st.session_state.get('rf_fetch_click_count_unlisted', 0)}")
    
        # Template download section
        st.markdown("#### üì• Download Excel Template")
        st.caption("Use this template to enter your company's financial data")
    
        # Create template in memory
        from io import BytesIO
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl import Workbook
    
        def create_template():
            wb = Workbook()
            if 'Sheet' in wb.sheetnames:
                del wb['Sheet']
        
            # Balance Sheet
            ws_bs = wb.create_sheet('BalanceSheet')
            ws_bs['A1'] = 'BALANCE SHEET'
            ws_bs['B1'] = 23
            ws_bs['C1'] = 24
            ws_bs['D1'] = 25
        
            for cell in ['A1', 'B1', 'C1', 'D1']:
                ws_bs[cell].font = Font(bold=True, color='FFFFFF')
                ws_bs[cell].fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
                ws_bs[cell].alignment = Alignment(horizontal='center')
        
            bs_items = [
                'Equity and Liabilities', 'Equity', 'Share Capital', 'Reserves and Surplus', 'Other Equity',
                'Total Equity', 'Liabilities', 'Non-current Liabilities', 'Long Term Borrowings',
                'Net Deferred Tax Liabilities', 'Other Long Term Liabilities', 'Long Term Provisions',
                'Total Non-current Liabilities', 'Current Liabilities', 'Short Term Borrowings',
                'Trade Payables', 'Other Current Liabilities', 'Short Term Provisions',
                'Total Current Liabilities', 'Total Equity and Liabilities', '',
                'Assets', 'Non-current Assets', 'Tangible Assets', 'Intangible Assets',
                'Capital Work in Progress', 'Non-current Investments', 'Long Term Loans and Advances',
                'Other Non-current Assets', 'Total Non-current Assets', 'Current Assets',
                'Inventories', 'Trade Receivables', 'Cash and Bank Balances',
                'Short Term Loans and Advances', 'Other Current Assets', 'Total Current Assets', 'Total Assets'
            ]
        
            row = 2
            for item in bs_items:
                ws_bs[f'A{row}'] = item
                if item in ['Equity and Liabilities', 'Equity', 'Liabilities', 'Non-current Liabilities',
                            'Current Liabilities', 'Assets', 'Non-current Assets', 'Current Assets']:
                    ws_bs[f'A{row}'].font = Font(bold=True)
                row += 1
        
            ws_bs.column_dimensions['A'].width = 35
            ws_bs.column_dimensions['B'].width = 15
            ws_bs.column_dimensions['C'].width = 15
            ws_bs.column_dimensions['D'].width = 15
        
            # Profit & Loss
            ws_pl = wb.create_sheet('Profit&Loss')
            ws_pl['A1'] = 'PROFIT & LOSS'
            ws_pl['B1'] = 23
            ws_pl['C1'] = 24
            ws_pl['D1'] = 25
        
            for cell in ['A1', 'B1', 'C1', 'D1']:
                ws_pl[cell].font = Font(bold=True, color='FFFFFF')
                ws_pl[cell].fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
                ws_pl[cell].alignment = Alignment(horizontal='center')
        
            pl_items = [
                'Net Revenue', 'Revenue Growth', 'Operating Cost', 'Cost of Materials Consumed',
                'Purchases of Stock-in-trade', 'Changes in Inventories / Finished Goods',
                'Employee Benefit Expense', 'Other Expenses', 'Total Operating Cost', 'EBITDA',
                'Other Income', 'Depreciation and Amortization Expense', 'Profit Before Interest and Tax',
                'Finance Costs', 'Profit Before Tax and Exceptional Items Before Tax',
                'Exceptional Items Before Tax', 'Profit Before Tax', 'Income Tax',
                'Profit for the Period from Continuing Operations'
            ]
        
            row = 2
            for item in pl_items:
                ws_pl[f'A{row}'] = item
                if item in ['Net Revenue', 'Total Operating Cost', 'EBITDA', 'Profit Before Interest and Tax',
                            'Profit Before Tax', 'Profit for the Period from Continuing Operations']:
                    ws_pl[f'A{row}'].font = Font(bold=True)
                row += 1
        
            ws_pl.column_dimensions['A'].width = 50
            ws_pl.column_dimensions['B'].width = 15
            ws_pl.column_dimensions['C'].width = 15
            ws_pl.column_dimensions['D'].width = 15
        
            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            return buffer.getvalue()
    
        template_data = create_template()
    
        st.download_button(
            label="üì• Download Excel Template",
            data=template_data,
            file_name="Financials_Template.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            help="Download this template and fill in your company's financial data"
        )
    
        st.markdown("---")
    
        col1, col2 = st.columns(2)
    
        with col1:
            company_name = st.text_input("Company Name:")
            excel_file = st.file_uploader("Upload Financial Excel Template", type=['xlsx', 'xls'])
        
            st.markdown("---")
            st.markdown("**üìä Peer Companies (Both Exchanges)**")
        
            # NSE Peers Box
            nse_peers = st.text_input(
                "NSE Peer Tickers (comma-separated):",
                placeholder="e.g., RELIANCE, TCS, INFY",
                key='nse_peers_unlisted',
                help="Enter NSE-listed peer companies for beta calculation"
            )
        
            # BSE Peers Box
            bse_peers = st.text_input(
                "BSE Peer Tickers (comma-separated):",
                placeholder="e.g., RELIANCE, TCS, INFY",
                key='bse_peers_unlisted',
                help="Enter BSE-listed peer companies for beta calculation"
            )
        
            # Combine peers with their exchange suffixes
            peer_tickers = ""
            if nse_peers.strip():
                nse_list = [f"{t.strip()}.NS" if '.NS' not in t and '.BO' not in t else t.strip() for t in nse_peers.split(',') if t.strip()]
                peer_tickers = ",".join(nse_list)
            if bse_peers.strip():
                bse_list = [f"{t.strip()}.BO" if '.NS' not in t and '.BO' not in t else t.strip() for t in bse_peers.split(',') if t.strip()]
                if peer_tickers:
                    peer_tickers += "," + ",".join(bse_list)
                else:
                    peer_tickers = ",".join(bse_list)
    
        with col2:
            num_shares = st.number_input("Number of Shares Outstanding:", min_value=1, value=100, step=1)
            tax_rate = st.number_input("Tax Rate (%):", min_value=0.0, max_value=100.0, value=25.0, step=0.5)
            terminal_growth = st.number_input("Terminal Growth Rate (%):", min_value=0.0, max_value=10.0, value=4.0, step=0.5)
            
            # Risk-free rate manual input
            st.markdown("**üèõÔ∏è Risk-Free Rate (G-Sec 10Y)**")
            manual_rf_rate = st.number_input(
                f"Risk-Free Rate (%)",
                min_value=0.0,
                max_value=20.0,
                value=st.session_state.get('cached_rf_rate_unlisted', 6.83),
                step=0.1,
                key='manual_rf_unlisted',
                help="Auto-fetched from ticker above. You can manually edit this value."
            )
            
            # Update session state if user manually changes it
            if abs(manual_rf_rate - st.session_state.get('cached_rf_rate_unlisted', 6.83)) > 0.05:
                st.session_state.cached_rf_rate_unlisted = manual_rf_rate
                st.info(f"üí° Using custom rate: {manual_rf_rate:.2f}%")
    
        with st.expander("‚öôÔ∏è Advanced Projection Assumptions - FULL CONTROL"):
            st.info("üí° **Complete Control:** Override ANY projection parameter below. Leave at 0 or blank for auto-calculation from historical data.")
        
            st.markdown("### üìä Revenue & Growth")
            col1, col2, col3 = st.columns(3)
            with col1:
                rev_growth_override_unlisted = st.number_input(
                    "Revenue Growth (%/year)", 
                    min_value=0.0, max_value=100.0, value=0.0, step=0.5,
                    key='unlisted_rev_growth',
                    help="0 = Auto from historical CAGR"
                )
            with col2:
                opex_margin_override_unlisted = st.number_input(
                    "Operating Expense Margin (%)", 
                    min_value=0.0, max_value=100.0, value=0.0, step=0.5,
                    key='unlisted_opex_margin',
                    help="0 = Auto from historical average"
                )
            with col3:
                ebitda_margin_override_unlisted = st.number_input(
                    "EBITDA Margin (%)", 
                    min_value=0.0, max_value=100.0, value=0.0, step=0.5,
                    key='unlisted_ebitda',
                    help="0 = Calculated as Revenue - OpEx"
                )
        
            st.markdown("### üèóÔ∏è CapEx & Depreciation")
            col4, col5, col6 = st.columns(3)
            with col4:
                capex_ratio_override_unlisted = st.number_input(
                    "CapEx/Revenue (%)", 
                    min_value=0.0, max_value=50.0, value=0.0, step=0.5,
                    key='unlisted_capex_ratio',
                    help="0 = Auto from historical average"
                )
            with col5:
                depreciation_rate_override_unlisted = st.number_input(
                    "Depreciation Rate (%)", 
                    min_value=0.0, max_value=30.0, value=0.0, step=0.5,
                    key='unlisted_dep_rate',
                    help="0 = Auto calculated"
                )
            with col6:
                depreciation_method_unlisted = st.selectbox(
                    "Depreciation Method",
                    ["Auto", "% of Fixed Assets", "% of Revenue", "Absolute Value"],
                    key='unlisted_dep_method'
                )
        
            st.markdown("### üí∞ Working Capital Management")
            col7, col8, col9 = st.columns(3)
            with col7:
                inventory_days_override_unlisted = st.number_input(
                    "Inventory Days", 
                    min_value=0.0, max_value=365.0, value=0.0, step=1.0,
                    key='unlisted_inv_days',
                    help="0 = Auto from historical average"
                )
            with col8:
                debtor_days_override_unlisted = st.number_input(
                    "Debtor/Receivables Days", 
                    min_value=0.0, max_value=365.0, value=0.0, step=1.0,
                    key='unlisted_deb_days',
                    help="0 = Auto from historical average"
                )
            with col9:
                creditor_days_override_unlisted = st.number_input(
                    "Creditor/Payables Days", 
                    min_value=0.0, max_value=365.0, value=0.0, step=1.0,
                    key='unlisted_cred_days',
                    help="0 = Auto from historical average"
                )
        
            st.markdown("### üìà Tax & Interest")
            col10, col11, col12 = st.columns(3)
            with col10:
                interest_rate_override_unlisted = st.number_input(
                    "Interest Rate (%)", 
                    min_value=0.0, max_value=30.0, value=0.0, step=0.25,
                    key='unlisted_interest',
                    help="0 = Auto calculated from Debt"
                )
            with col11:
                working_capital_as_pct_revenue_unlisted = st.number_input(
                    "Working Capital % of Revenue", 
                    min_value=0.0, max_value=50.0, value=0.0, step=0.5,
                    key='unlisted_wc_pct',
                    help="0 = Calculate from Inv+Deb-Cred days"
                )
            with col12:
                projection_years = st.number_input(
                    "Projection Years", 
                    min_value=3, max_value=15, value=5, step=1,
                    key='unlisted_proj_years',
                    help="Number of years to project forward"
                )
    
        if excel_file and company_name and num_shares:
            # INPUT CHANGE DETECTION - Reset results if key inputs change
            current_inputs_unlisted = {
                'excel_file': excel_file.name if excel_file else None,
                'company_name': company_name,
                'num_shares': num_shares,
                'terminal_growth': terminal_growth,
                'tax_rate': tax_rate_input,
                'wacc': wacc_input if wacc_input > 0 else None
            }
            
            # Check if inputs changed
            if 'previous_inputs_unlisted' in st.session_state:
                if st.session_state.previous_inputs_unlisted != current_inputs_unlisted:
                    # Inputs changed - clear results
                    st.session_state.show_results_unlisted = False
            
            # Store current inputs
            st.session_state.previous_inputs_unlisted = current_inputs_unlisted
            
            st.markdown("---")
            st.markdown("### üéØ Ready to Run Valuation")
            st.info("üí° **Click the button below to run valuation.** Results will appear only after clicking.")
            
            col_run1, col_run2 = st.columns([2, 1])
            
            with col_run1:
                if st.button("üöÄ Run DCF Valuation", type="primary", key="run_dcf_btn"):
                    st.session_state.show_results_unlisted = True
            
            with col_run2:
                if st.button("üóëÔ∏è Clear Results", help="Clear current valuation results", key="clear_unlisted_results"):
                    st.session_state.show_results_unlisted = False
                    st.success("‚úÖ Results cleared!")
                    st.rerun()
        
            if st.session_state.get('show_results_unlisted', False):
                with st.spinner("Processing..."):
                    # Parse Excel
                    df_bs, df_pl = parse_excel_to_dataframes(excel_file)
                
                    if df_bs is None or df_pl is None:
                        st.error("Failed to parse Excel file")
                        st.stop()
                
                    # Detect year columns
                    year_cols = detect_year_columns(df_bs)
                
                    if len(year_cols) < 3:
                        st.error("Need at least 3 years of historical data")
                        st.stop()
                
                    st.success(f"‚úÖ Loaded {len(year_cols)} years of data")
                
                    # Extract financials
                    financials = extract_financials_unlisted(df_bs, df_pl, year_cols)
                
                    # ================================
                    # BUSINESS MODEL CLASSIFICATION (RULEBOOK SECTION 2)
                    # ================================
                    st.markdown("---")
                    st.subheader("üè¢ Business Model Classification")
                
                    classification = classify_business_model(financials, income_stmt=None, balance_sheet=None)
                
                    # Show classification and check if FCFF DCF is allowed
                    should_stop = show_classification_warning(classification)
                
                    if should_stop:
                        st.stop()
                
                    st.markdown("---")
                
                    # Calculate WC metrics
                    wc_metrics = calculate_working_capital_metrics(financials)
                    
                    # Show Working Capital Data Status
                    wc_status_parts = []
                    if wc_metrics.get('has_inventory', False) or (inventory_days_override_unlisted and inventory_days_override_unlisted > 0):
                        inv_source = f"User Override: {inventory_days_override_unlisted} days" if inventory_days_override_unlisted and inventory_days_override_unlisted > 0 else f"Historical: {wc_metrics['avg_inv_days']:.1f} days"
                        wc_status_parts.append(f"‚úÖ Inventory ({inv_source})")
                    else:
                        wc_status_parts.append("‚ö†Ô∏è Inventory (No data)")
                    
                    if wc_metrics.get('has_receivables', False) or (debtor_days_override_unlisted and debtor_days_override_unlisted > 0):
                        deb_source = f"User Override: {debtor_days_override_unlisted} days" if debtor_days_override_unlisted and debtor_days_override_unlisted > 0 else f"Historical: {wc_metrics['avg_deb_days']:.1f} days"
                        wc_status_parts.append(f"‚úÖ Debtors ({deb_source})")
                    else:
                        wc_status_parts.append("‚ö†Ô∏è Debtors (No data)")
                    
                    if wc_metrics.get('has_payables', False) or (creditor_days_override_unlisted and creditor_days_override_unlisted > 0):
                        cred_source = f"User Override: {creditor_days_override_unlisted} days" if creditor_days_override_unlisted and creditor_days_override_unlisted > 0 else f"Historical: {wc_metrics['avg_cred_days']:.1f} days"
                        wc_status_parts.append(f"‚úÖ Creditors ({cred_source})")
                    else:
                        wc_status_parts.append("‚ö†Ô∏è Creditors (No data)")
                    
                    st.info(f"üîç **Working Capital Projection Status:** {' | '.join(wc_status_parts)}")
                
                    # Project financials
                    projections, drivers = project_financials(
                        financials, wc_metrics, projection_years, tax_rate,
                        rev_growth_override_unlisted if rev_growth_override_unlisted > 0 else None, 
                        opex_margin_override_unlisted if opex_margin_override_unlisted > 0 else None, 
                        capex_ratio_override_unlisted if capex_ratio_override_unlisted > 0 else None,
                        # Pass all advanced user controls
                        ebitda_margin_override=ebitda_margin_override_unlisted if ebitda_margin_override_unlisted > 0 else None,
                        depreciation_rate_override=depreciation_rate_override_unlisted if depreciation_rate_override_unlisted > 0 else None,
                        depreciation_method=depreciation_method_unlisted,
                        inventory_days_override=inventory_days_override_unlisted if inventory_days_override_unlisted > 0 else None,
                        debtor_days_override=debtor_days_override_unlisted if debtor_days_override_unlisted > 0 else None,
                        creditor_days_override=creditor_days_override_unlisted if creditor_days_override_unlisted > 0 else None,
                        interest_rate_override=interest_rate_override_unlisted if interest_rate_override_unlisted > 0 else None,
                        working_capital_pct_override=working_capital_as_pct_revenue_unlisted if working_capital_as_pct_revenue_unlisted > 0 else None
                    )
                
                    # Calculate WACC (unlisted companies use auto-fetched risk-free rate)
                    wacc_details = calculate_wacc(financials, tax_rate, peer_tickers=peer_tickers, manual_rf_rate=manual_rf_rate)
                
                    # DCF Valuation
                    # Extract cash balance
                    cash_balance = financials['cash'][0] if financials['cash'][0] > 0 else 0
                
                    valuation, error = calculate_dcf_valuation(
                        projections, wacc_details, terminal_growth, num_shares, cash_balance
                    )
                
                    if error:
                        st.error(error)
                        st.stop()
                
                    # ================================
                    # DISPLAY RESULTS
                    # ================================
                
                    st.success("‚úÖ Valuation Complete!")
                
                    # AUTO-GENERATE PDF
                    try:
                        all_fair_values = {'DCF': valuation['fair_value_per_share']}
                        if 'comp_results' in locals() and comp_results:
                            for method, val_data in comp_results.get('valuations', {}).items():
                                if val_data.get('fair_value_avg', 0) > 0:
                                    all_fair_values[method.upper().replace('_', ' ')] = val_data['fair_value_avg']
                    
                        pdf_path = export_to_pdf({
                            'company_name': company_name,
                            'ticker': 'UNLISTED',
                            'financials': financials,
                            'dcf_results': valuation,
                            'fair_values': all_fair_values,
                            'current_price': 0,
                            'peer_data': pd.DataFrame(),
                            'comp_results': comp_results if 'comp_results' in locals() else None
                        })
                    
                        with open(pdf_path, "rb") as f:
                            st.session_state.pdf_bytes = f.read()
                    
                        st.toast("‚úÖ PDF Generated! Scroll to top to download", icon="üì•")
                    except Exception as e:
                        st.error(f"PDF Generation Error: {str(e)}")
                
                
                    # Key Metrics
                    col1, col2, col3, col4 = st.columns(4)
                    with col1:
                        st.metric("Enterprise Value", f"‚Çπ {valuation['enterprise_value']:.2f} Lacs")
                    with col2:
                        st.metric("Equity Value", f"‚Çπ {valuation['equity_value']:.2f} Lacs")
                    with col3:
                        st.metric("Fair Value/Share", f"‚Çπ {valuation['fair_value_per_share']:.2f}")
                    with col4:
                        st.metric("WACC", f"{wacc_details['wacc']:.2f}%")
                
                    # Tabs for detailed output
                    tab1, tab2, tab3, tab4, tab5, tab6, tab7, tab8 = st.tabs([
                        "üìä Historical Financials",
                        "üìà Projections",
                        "üí∞ FCF Working",
                        "üéØ WACC Calculation",
                        "üèÜ Valuation Summary",
                        "üîç Comparative Valuation",
                        "üè¢ Peer Comparison",
                        "‚öôÔ∏è Assumptions & Parameters"
                    ])
                
                    with tab1:
                        st.subheader("Historical Financials (Last 3 Years)")
                    
                        hist_df = pd.DataFrame({
                            'Year': [str(y) for y in financials['years']],
                            'Revenue': financials['revenue'],
                            'Operating Expenses': financials['opex'],
                            'EBITDA': financials['ebitda'],
                            'Depreciation': financials['depreciation'],
                            'EBIT': financials['ebit'],
                            'Interest': financials['interest'],
                            'Tax': financials['tax'],
                            'NOPAT': financials['nopat']
                        })
                    
                        # Format numeric columns only
                        numeric_cols = hist_df.select_dtypes(include=[np.number]).columns.tolist()
                        format_dict = {col: '{:.2f}' for col in numeric_cols}
                        st.dataframe(hist_df.style.format(format_dict), use_container_width=True)
                    
                        st.subheader("Balance Sheet Metrics")
                        bs_df = pd.DataFrame({
                            'Year': [str(y) for y in financials['years']],
                            'Fixed Assets': financials['fixed_assets'],
                            'Inventory': financials['inventory'],
                            'Receivables': financials['receivables'],
                            'Payables': financials['payables'],
                            'Equity': financials['equity'],
                            'ST Debt': financials['st_debt'],
                            'LT Debt': financials['lt_debt']
                        })
                        numeric_cols = bs_df.select_dtypes(include=[np.number]).columns.tolist()
                        format_dict = {col: '{:.2f}' for col in numeric_cols}
                        st.dataframe(bs_df.style.format(format_dict), use_container_width=True)
                    
                        st.subheader("Working Capital Days")
                        wc_df = pd.DataFrame({
                            'Year': [str(y) for y in financials['years']],
                            'Inventory Days': wc_metrics['inventory_days'],
                            'Debtor Days': wc_metrics['debtor_days'],
                            'Creditor Days': wc_metrics['creditor_days']
                        })
                        numeric_cols = wc_df.select_dtypes(include=[np.number]).columns.tolist()
                        format_dict = {col: '{:.2f}' for col in numeric_cols}
                        st.dataframe(wc_df.style.format(format_dict), use_container_width=True)
                    
                        st.info(f"**Average Working Capital Days:** Inventory: {wc_metrics['avg_inv_days']:.1f} | Debtors: {wc_metrics['avg_deb_days']:.1f} | Creditors: {wc_metrics['avg_cred_days']:.1f}")
                
                    with tab2:
                        st.subheader(f"Projected Financials ({projection_years} Years)")
                    
                        proj_df = pd.DataFrame({
                            'Year': [str(y) for y in projections['year']],
                            'Revenue': projections['revenue'],
                            'EBITDA': projections['ebitda'],
                            'Depreciation': projections['depreciation'],
                            'EBIT': projections['ebit'],
                            'NOPAT': projections['nopat'],
                            'Capex': projections['capex'],
                            'Œî WC': projections['delta_wc'],
                            'FCFF': projections['fcff']
                        })
                        numeric_cols = proj_df.select_dtypes(include=[np.number]).columns.tolist()
                        format_dict = {col: '{:.2f}' for col in numeric_cols}
                        st.dataframe(proj_df.style.format(format_dict), use_container_width=True)
                    
                        st.info(f"**Key Drivers:** Revenue Growth: {drivers['avg_growth']:.2f}% | Opex Margin: {drivers['avg_opex_margin']:.2f}% | CapEx/Revenue: {drivers['avg_capex_ratio']:.2f}% | Depreciation Rate: {drivers['avg_dep_rate']:.2f}%")
                
                    with tab3:
                        st.subheader("Free Cash Flow Working")
                    
                        fcff_df = pd.DataFrame({
                            'Year': [str(y) for y in projections['year']],
                            'NOPAT': projections['nopat'],
                            '+ Depreciation': projections['depreciation'],
                            '- Œî WC': projections['delta_wc'],
                            '- Capex': projections['capex'],
                            '= FCFF': projections['fcff'],
                            'Discount Factor': [(1 + wacc_details['wacc']/100)**(-y) for y in projections['year']],
                            'PV(FCFF)': valuation['pv_fcffs']
                        })
                        numeric_cols = fcff_df.select_dtypes(include=[np.number]).columns.tolist()
                        format_dict = {col: '{:.4f}' for col in numeric_cols}
                        st.dataframe(fcff_df.style.format(format_dict), use_container_width=True)
                    
                        st.metric("Sum of PV(FCFF)", f"‚Çπ {valuation['sum_pv_fcff']:.2f} Lacs")
                
                    with tab4:
                        st.subheader("WACC Calculation")
                    
                        col1, col2 = st.columns(2)
                    
                        with col1:
                            st.markdown("**Cost of Equity (Ke)**")
                            st.write(f"Risk-free Rate (Rf): **{wacc_details['rf']:.2f}%**")
                            st.write(f"Market Return (Rm): **{wacc_details['rm']:.2f}%**")
                            st.write(f"Beta (Œ≤): **{wacc_details['beta']:.2f}**")
                            st.write(f"Ke = Rf + Œ≤ √ó (Rm - Rf)")
                            st.write(f"Ke = {wacc_details['rf']:.2f}% + {wacc_details['beta']:.2f} √ó ({wacc_details['rm']:.2f}% - {wacc_details['rf']:.2f}%)")
                            st.write(f"**Ke = {wacc_details['ke']:.2f}%**")
                    
                        with col2:
                            st.markdown("**Cost of Debt (Kd)**")
                            st.write(f"Interest Expense: **‚Çπ {financials['interest'][0]:.2f} Lacs**")
                            st.write(f"Total Debt: **‚Çπ {wacc_details['debt']:.2f} Lacs**")
                            st.write(f"Kd (pre-tax) = {wacc_details['kd']:.2f}%")
                            st.write(f"Tax Rate = {tax_rate}%")
                            st.write(f"**Kd (after-tax) = {wacc_details['kd_after_tax']:.2f}%**")
                    
                        st.markdown("---")
                        st.markdown("**WACC Calculation**")
                    
                        col3, col4 = st.columns(2)
                        with col3:
                            st.write(f"Equity (E): **‚Çπ {wacc_details['equity']:.2f} Lacs** ({wacc_details['we']:.2f}%)")
                            st.write(f"Debt (D): **‚Çπ {wacc_details['debt']:.2f} Lacs** ({wacc_details['wd']:.2f}%)")
                            st.write(f"Total Capital (V): **‚Çπ {wacc_details['equity'] + wacc_details['debt']:.2f} Lacs**")
                    
                        with col4:
                            st.write(f"WACC = (E/V √ó Ke) + (D/V √ó Kd √ó (1-Tax))")
                            st.write(f"WACC = ({wacc_details['we']:.2f}% √ó {wacc_details['ke']:.2f}%) + ({wacc_details['wd']:.2f}% √ó {wacc_details['kd_after_tax']:.2f}%)")
                            st.write(f"**WACC = {wacc_details['wacc']:.2f}%**")
                
                    with tab5:
                        st.subheader("DCF Valuation Summary")
                    
                        # Show FCFF adjustment notice if applicable
                        if valuation.get('fcff_adjusted', False):
                            st.warning("‚ö†Ô∏è **FCFF Adjustment Applied**")
                            adj_details = valuation.get('adjustment_details', {})
                            st.write(f"**Strategy Used:** {adj_details.get('strategy', 'N/A')}")
                            st.write(f"**Original Terminal FCFF:** ‚Çπ{projections['fcff'][-1]:.2f} Lacs")
                            st.write(f"**Adjusted Terminal FCFF:** ‚Çπ{valuation['adjusted_terminal_fcff']:.2f} Lacs")
                            st.caption("üìå Adjustment details shown during valuation run above")
                            st.markdown("---")
                    
                        st.markdown("### Terminal Value Calculation")
                    
                        # Use adjusted FCFF if available
                        terminal_fcff = valuation.get('adjusted_terminal_fcff', projections['fcff'][-1])
                    
                        st.write(f"FCFF (Year {projection_years}): **‚Çπ {terminal_fcff:.2f} Lacs**")
                        if valuation.get('fcff_adjusted', False):
                            st.caption(f"(Original: ‚Çπ{projections['fcff'][-1]:.2f} Lacs - Adjusted for sustainability)")
                    
                        st.write(f"Terminal Growth Rate (g): **{terminal_growth}%**")
                        st.write(f"FCFF (Year {projection_years + 1}) = FCFF{projection_years} √ó (1 + g)")
                        st.write(f"FCFF (Year {projection_years + 1}) = ‚Çπ {terminal_fcff:.2f} √ó (1 + {terminal_growth/100})")
                        st.write(f"FCFF (Year {projection_years + 1}) = **‚Çπ {terminal_fcff * (1 + terminal_growth/100):.2f} Lacs**")
                    
                        st.write(f"\nTerminal Value = FCFF{projection_years + 1} / (WACC - g)")
                        st.write(f"Terminal Value = ‚Çπ {projections['fcff'][-1] * (1 + terminal_growth/100):.2f} / ({wacc_details['wacc']:.2f}% - {terminal_growth}%)")
                        st.write(f"**Terminal Value = ‚Çπ {valuation['terminal_value']:.2f} Lacs**")
                    
                        st.write(f"\nPV(Terminal Value) = TV / (1 + WACC)^{projection_years}")
                        st.write(f"**PV(Terminal Value) = ‚Çπ {valuation['pv_terminal_value']:.2f} Lacs**")
                    
                        st.markdown("---")
                        st.markdown("### Enterprise Value")

                        # Show growth phase adjustment if applied
                        if valuation.get('growth_phase_adjusted', False):
                            st.info("üìä **Growth-Phase Company:** Sum of PV(FCFF) was adjusted from negative to zero")
                            st.caption(f"Original: ‚Çπ{valuation['original_sum_pv_fcff']:.2f} Lacs ‚Üí Adjusted: ‚Çπ{valuation['sum_pv_fcff']:.2f} Lacs")
                    
                        ev_df = pd.DataFrame({
                            'Component': ['Sum of PV(FCFF)', 'PV(Terminal Value)', 'Enterprise Value'],
                            'Value (‚Çπ Lacs)': [
                                valuation['sum_pv_fcff'],
                                valuation['pv_terminal_value'],
                                valuation['enterprise_value']
                            ]
                        })
                        st.dataframe(ev_df.style.format({'Value (‚Çπ Lacs)': '{:.2f}'}), use_container_width=True)
                    
                        tv_pct = valuation['tv_percentage']
                        if tv_pct > 90:
                            st.warning(f"‚ö†Ô∏è Terminal Value represents {tv_pct:.1f}% of Enterprise Value (>90% is high)")
                        else:
                            st.info(f"Terminal Value represents {tv_pct:.1f}% of Enterprise Value")
                    
                        st.markdown("---")
                        st.markdown("### Equity Value & Fair Value per Share")
                    
                        equity_calc_df = pd.DataFrame({
                            'Item': ['Enterprise Value', 'Less: Total Debt', 'Add: Cash & Equivalents', '= Net Debt', 'Equity Value', 'Equity Value (‚Çπ)', 'Number of Shares', 'Fair Value per Share'],
                            'Value': [
                                f"‚Çπ {valuation['enterprise_value']:.2f} Lacs",
                                f"‚Çπ {valuation['total_debt']:.2f} Lacs",
                                f"‚Çπ {valuation['cash']:.2f} Lacs",
                                f"‚Çπ {valuation['net_debt']:.2f} Lacs",
                                f"‚Çπ {valuation['equity_value']:.2f} Lacs",
                                f"‚Çπ {valuation['equity_value_rupees']:,.0f}",
                                f"{shares:,}" if 'shares' in locals() else f"{num_shares:,}",
                                f"‚Çπ {valuation['fair_value_per_share']:.2f}"
                            ]
                        })
                        st.table(equity_calc_df)
                    
                        st.success(f"### üéØ Fair Value per Share: ‚Çπ {valuation['fair_value_per_share']:.2f}")
                    
                        # Sensitivity Analysis
                        st.markdown("---")
                        st.subheader("üìä Sensitivity Analysis")
                    
                        wacc_range = np.arange(max(1.0, wacc_details['wacc'] - 3), wacc_details['wacc'] + 3.5, 0.5)
                        g_range = np.arange(max(1.0, terminal_growth - 2), min(terminal_growth + 3, wacc_details['wacc'] - 1), 0.5)
                    
                        if len(g_range) == 0:
                            g_range = np.array([terminal_growth])
                    
                        sensitivity_data = []
                    
                        for w in wacc_range:
                            row_data = {'WACC ‚Üí': f"{w:.1f}%"}
                            for g_val in g_range:
                                if g_val >= w - 0.1:  # Need at least 0.1% gap
                                    row_data[f"g={g_val:.1f}%"] = "N/A"
                                else:
                                    try:
                                        fcff_n_plus_1 = projections['fcff'][-1] * (1 + g_val / 100)
                                        tv = fcff_n_plus_1 / ((w / 100) - (g_val / 100))
                                        pv_tv = tv / ((1 + w / 100) ** projection_years)
                                        ev = valuation['sum_pv_fcff'] + pv_tv
                                        eq_val = ev - valuation['net_debt']
                                        eq_val_rupees = eq_val * 100000
                                        fv = eq_val_rupees / num_shares if num_shares > 0 else 0
                                        row_data[f"g={g_val:.1f}%"] = f"‚Çπ{fv:.2f}"
                                    except:
                                        row_data[f"g={g_val:.1f}%"] = "Error"
                            sensitivity_data.append(row_data)
                    
                        sensitivity_df = pd.DataFrame(sensitivity_data)
                        st.dataframe(sensitivity_df, use_container_width=True)
                    
                        st.caption("Sensitivity table shows Fair Value per Share for different WACC and terminal growth rate combinations")
                
                    with tab6:
                        st.subheader("üîç Comparative (Relative) Valuation")
                    
                        if peer_tickers and peer_tickers.strip():
                            with st.spinner("Fetching comparable companies data..."):
                                # Note: Unlisted companies don't have session state checkboxes
                                # For now, use Yahoo Finance (default behavior)
                                comp_results = perform_comparative_valuation(
                                    None, 
                                    peer_tickers, 
                                    financials, 
                                    num_shares, 
                                    "NS",
                                    use_screener_peers=False
                                )
                        
                            if comp_results:
                                # Show comparables table
                                st.markdown("### Comparable Companies")
                                comp_df = pd.DataFrame(comp_results['comparables'])
                                if not comp_df.empty:
                                    display_comp_df = comp_df[['ticker', 'name', 'price', 'pe', 'pb', 'ps', 'ev_ebitda', 'ev_sales']]
                                    st.dataframe(display_comp_df.style.format({
                                        'price': '‚Çπ{:.2f}',
                                        'pe': '{:.2f}x',
                                        'pb': '{:.2f}x',
                                        'ps': '{:.2f}x',
                                        'ev_ebitda': '{:.2f}x',
                                        'ev_sales': '{:.2f}x'
                                    }), use_container_width=True)
                            
                                # Show multiples statistics
                                st.markdown("### Peer Multiples Statistics")
                                for multiple, stats in comp_results['multiples_stats'].items():
                                    with st.expander(f"üìä {multiple.upper()} - Avg: {stats['average']:.2f}x, Median: {stats['median']:.2f}x"):
                                        st.write(f"**Range:** {stats['min']:.2f}x - {stats['max']:.2f}x")
                                        st.write(f"**Std Dev:** {stats['std']:.2f}x")
                                        st.write(f"**Peer Values:** {', '.join([f'{v:.2f}x' for v in stats['values']])}")
                            
                                # Show implied valuations
                                st.markdown("### Implied Fair Values")
                            
                                all_avg_values = []
                                all_median_values = []
                            
                                for method_key, val_data in comp_results['valuations'].items():
                                    st.markdown(f"#### {val_data['method']}")
                                
                                    col1, col2 = st.columns(2)
                                
                                    with col1:
                                        st.markdown("**Using Average Multiple:**")
                                        st.write(val_data['formula_avg'])
                                        st.metric("Fair Value (Avg)", f"‚Çπ{val_data['fair_value_avg']:.2f}")
                                        all_avg_values.append(val_data['fair_value_avg'])
                                
                                    with col2:
                                        st.markdown("**Using Median Multiple:**")
                                        st.write(val_data['formula_median'])
                                        st.metric("Fair Value (Median)", f"‚Çπ{val_data['fair_value_median']:.2f}")
                                        all_median_values.append(val_data['fair_value_median'])
                                
                                    st.markdown("---")
                            
                                # Summary statistics
                                if all_avg_values and all_median_values:
                                    st.markdown("### üìà Comparative Valuation Summary")
                                
                                    col1, col2, col3 = st.columns(3)
                                
                                    with col1:
                                        st.metric("Average (All Methods)", f"‚Çπ{np.mean(all_avg_values):.2f}")
                                        st.metric("Median (All Methods)", f"‚Çπ{np.median(all_median_values):.2f}")
                                
                                    with col2:
                                        st.metric("Min Fair Value", f"‚Çπ{min(all_avg_values + all_median_values):.2f}")
                                        st.metric("Max Fair Value", f"‚Çπ{max(all_avg_values + all_median_values):.2f}")
                                
                                    with col3:
                                        if valuation['fair_value_per_share'] > 0:
                                            st.metric("DCF Fair Value", f"‚Çπ{valuation['fair_value_per_share']:.2f}")
                                            combined_avg = (np.mean(all_avg_values) + valuation['fair_value_per_share']) / 2
                                            st.metric("DCF + Comp Avg", f"‚Çπ{combined_avg:.2f}")
                            else:
                                st.warning("Could not fetch comparable companies data")
                        else:
                            st.info("üí° Enter peer tickers above (e.g., 'RELIANCE, TATASTEEL') to see comparative valuation based on peer multiples")
                
                    with tab7:
                        st.subheader("üè¢ Advanced Peer Comparison Dashboard")
                    
                        if peer_tickers and peer_tickers.strip():
                            try:
                                from peer_comparison_charts import create_peer_comparison_dashboard
                                import inspect
                            
                                # Check if the function supports unlisted_data parameter
                                sig = inspect.signature(create_peer_comparison_dashboard)
                                supports_unlisted = 'unlisted_data' in sig.parameters
                            
                                if supports_unlisted:
                                    # Module supports unlisted - create full comparison
                                    synthetic_ticker = company_name.replace(' ', '_').upper()[:10] if company_name else "UNLISTED"
                                    create_peer_comparison_dashboard(synthetic_ticker, peer_tickers, unlisted_data=financials)
                                else:
                                    # Module doesn't support unlisted companies
                                    st.warning("‚ö†Ô∏è **Advanced Peer Comparison Not Available for Unlisted Companies**")
                                    st.info("""
                                    The current peer comparison dashboard module only supports listed companies.
                                
                                    **However, you still have comprehensive peer analysis available:**
                                    - **Tab 6 (Comparative Valuation)** provides detailed peer multiples and valuations
                                    - Shows peer P/E, P/B, P/S, EV/EBITDA ratios
                                    - Calculates implied fair values based on peer averages
                                    - Includes peer statistics and comparisons
                                
                                    üí° **For full 3D visualizations:** The peer comparison module needs to be updated 
                                    to support unlisted company data.
                                    """)
                                    
                            except ImportError:
                                st.warning("‚ö†Ô∏è **Peer Comparison Dashboard Module Not Installed**")
                                st.info("""
                                The `peer_comparison_charts.py` module is not available.
                            
                                **Don't worry!** You still have comprehensive peer analysis in **Tab 6 (Comparative Valuation)** which includes:
                                - Detailed peer company metrics
                                - Multiple valuation methods (P/E, P/B, P/S, EV/EBITDA)
                                - Implied fair values based on peer averages
                                - Full peer statistics and comparisons
                                """)
                            except Exception as e:
                                st.error(f"‚ö†Ô∏è **Error Loading Peer Comparison Dashboard**")
                                st.write(f"Error details: {str(e)}")
                                st.info("üí° **Alternative:** Use **Tab 6 (Comparative Valuation)** for comprehensive peer analysis with multiples-based valuation")
                        else:
                            st.info("üí° Enter peer tickers above (e.g., 'RELIANCE, TATASTEEL') to see detailed peer comparison with 3D visualizations")
                    
                    with tab8:
                        st.subheader("‚öôÔ∏è Assumptions & Parameters")
                        
                        # Compact metrics
                        years_display = [str(y).replace('_', '') for y in financials['years']]
                        years_range = f"{years_display[-1]} - {years_display[0]}" if len(years_display) > 1 else years_display[0]
                        
                        col1, col2, col3 = st.columns(3)
                        with col1:
                            st.metric("üìÖ Historical Period", f"{len(financials['years'])} years", delta=years_range, delta_color="off")
                        with col2:
                            st.metric("üîÆ Projection Years", f"{projection_years} years")
                        with col3:
                            st.metric("üè¢ Shares", f"{num_shares:,}")
                        
                        st.markdown("---")
                        
                        # WACC Bar Chart
                        st.markdown("### üí∞ WACC Components")
                        fig_wacc = go.Figure()
                        fig_wacc.add_trace(go.Bar(
                            x=['WACC', 'Ke', 'Kd', 'Terminal g'],
                            y=[wacc_details['wacc'], wacc_details['ke'], wacc_details['kd_after_tax'], terminal_growth],
                            marker=dict(color=['#667eea', '#764ba2', '#f093fb', '#f5576c']),
                            text=[f"{wacc_details['wacc']:.2f}%", f"{wacc_details['ke']:.2f}%", 
                                  f"{wacc_details['kd_after_tax']:.2f}%", f"{terminal_growth:.2f}%"],
                            textposition='auto'
                        ))
                        fig_wacc.update_layout(height=220, margin=dict(l=20, r=20, t=10, b=20), showlegend=False,
                                              yaxis=dict(title='%'), paper_bgcolor='rgba(0,0,0,0)', plot_bgcolor='rgba(0,0,0,0)')
                        st.plotly_chart(fig_wacc, use_container_width=True)
                        
                        col_a, col_b = st.columns(2)
                        with col_a:
                            st.caption(f"**Tax:** {tax_rate:.2f}% | **Beta:** {wacc_details['beta']:.3f} | **Rf:** {wacc_details['rf']:.2f}%")
                        with col_b:
                            st.caption(f"**Equity Wt:** {wacc_details['we']:.1f}% | **Debt Wt:** {wacc_details['wd']:.1f}%")
                        
                        st.markdown("---")
                        
                        # Compact Drivers
                        st.markdown("### üìä Projection Drivers")
                        with st.expander("View Parameters", expanded=False):
                            drivers_data = [
                                ['Revenue Growth', f"{drivers['avg_growth']:.2f}%", 'Historical CAGR'],
                                ['EBITDA Margin', f"{drivers['avg_ebitda_margin']:.2f}%", 'Historical Avg'],
                                ['CapEx/Revenue', f"{drivers['avg_capex_ratio']:.2f}%", 'Historical Avg'],
                                ['Inventory Days', f"{wc_metrics['avg_inv_days']:.0f}", 'Historical'],
                                ['Debtor Days', f"{wc_metrics['avg_deb_days']:.0f}", 'Historical'],
                                ['Creditor Days', f"{wc_metrics['avg_cred_days']:.0f}", 'Historical']
                            ]
                            st.dataframe(pd.DataFrame(drivers_data, columns=['Parameter', 'Value', 'Source']),
                                       use_container_width=True, hide_index=True, height=200)




    elif mode == "Screener Excel Mode (Screener.in Template)":
        
        # ===== RISK-FREE RATE TICKER - AT THE TOP, ALWAYS VISIBLE =====
        st.markdown("### üèõÔ∏è Risk-Free Rate Configuration")
        st.info("üí° **Default: 6.83%**. Enter any Yahoo Finance ticker and click Fetch. Common options: ^TNX (US 10Y), ^IRX (US 3-month), or any stock/index for custom rates.")
        
        # Initialize if not exists
        if 'cached_rf_rate_screener' not in st.session_state:
            st.session_state.cached_rf_rate_screener = 6.83
        
        rf_col1, rf_col2, rf_col3 = st.columns([3, 2, 1])
        with rf_col1:
            custom_rf_ticker_screener = st.text_input(
                "Yahoo Finance Ticker for Risk-Free Rate",
                value="NIFTYGS10YR.NS",
                key='custom_rf_ticker_screener_top',
                help="Enter any ticker - will calculate CAGR for stocks/indices, or use value directly for yields. Examples: NIFTYGS10YR.NS, ^TNX, RELIANCE.NS"
            )
        with rf_col2:
            st.metric("Current RF Rate", f"{st.session_state.cached_rf_rate_screener:.2f}%")
        with rf_col3:
            st.write("")
            st.write("")
            
            # Add a counter to track button clicks
            if 'rf_fetch_click_count_screener' not in st.session_state:
                st.session_state.rf_fetch_click_count_screener = 0
            
            if st.button("üîÑ Fetch", key='refresh_rf_screener_top'):
                # Increment click counter FIRST
                st.session_state.rf_fetch_click_count_screener += 1
                
                # Store debug output in session state so it persists across reruns
                debug_output = []
                debug_output.append(f"üîÑ **FETCH BUTTON CLICKED #{st.session_state.rf_fetch_click_count_screener} - SCREENER MODE**")
                debug_output.append(f"üìù Input ticker: `{custom_rf_ticker_screener}`")
                
                ticker_to_use = custom_rf_ticker_screener.strip() if custom_rf_ticker_screener.strip() else None
                debug_output.append(f"üìù Ticker to use (after strip): `{ticker_to_use}`")
                
                debug_output.append("‚è≥ Calling get_risk_free_rate()...")
                fetched_rate, fetch_debug = get_risk_free_rate(ticker_to_use)
                
                # Add all debug messages from the function
                debug_output.extend(fetch_debug)
                
                debug_output.append(f"‚úÖ Function returned: {fetched_rate}%")
                
                debug_output.append(f"üíæ Updating session state...")
                debug_output.append(f"   - Before: {st.session_state.get('cached_rf_rate_screener', 'NOT SET')}")
                st.session_state.cached_rf_rate_screener = fetched_rate
                debug_output.append(f"   - After: {st.session_state.cached_rf_rate_screener}")
                
                # Force update the manual input field
                debug_output.append(f"üîÑ Clearing manual input widget state...")
                if 'manual_rf_screener' in st.session_state:
                    del st.session_state['manual_rf_screener']
                    debug_output.append(f"   - Widget state cleared")
                else:
                    debug_output.append(f"   - Widget state was not set")
                
                # Store debug output in session state
                st.session_state.rf_fetch_debug_screener = debug_output
                st.session_state.rf_fetch_success_screener = True
                
                # Rerun to update UI
                st.rerun()
        
        # Display debug output if available (persists across reruns)
        if st.session_state.get('rf_fetch_debug_screener'):
            with st.expander("üìã Last Fetch Debug Output", expanded=True):
                for line in st.session_state.rf_fetch_debug_screener:
                    st.write(line)
                if st.session_state.get('rf_fetch_success_screener'):
                    st.success(f"‚úì Successfully updated to {st.session_state.cached_rf_rate_screener:.2f}%")
        
        st.markdown("---")
        # ===== END RF RATE CONFIG =====
        
        # DEBUG: Show session state
        with st.expander("üêõ DEBUG: Session State Values (Screener)", expanded=False):
            st.write("**Current Session State for RF Rate:**")
            st.write(f"- `cached_rf_rate_screener`: {st.session_state.get('cached_rf_rate_screener', 'NOT SET')}")
            st.write(f"- `manual_rf_screener` widget: {st.session_state.get('manual_rf_screener', 'NOT SET')}")
            st.write(f"- `custom_rf_ticker_screener_top`: {st.session_state.get('custom_rf_ticker_screener_top', 'NOT SET')}")
            st.write(f"- **Fetch button clicks**: {st.session_state.get('rf_fetch_click_count_screener', 0)}")
        
        # Add template download button
        st.markdown("#### üì• Download Screener Template")
        st.caption("Download the pre-configured template that matches Screener.in format")
        
        try:
            with open('Screener_template.xlsx', 'rb') as f:
                template_bytes = f.read()
            
            st.download_button(
                label="üì• Download Screener Template",
                data=template_bytes,
                file_name="Screener_Template.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                help="Download this template matching Screener.in export format"
            )
        except FileNotFoundError:
            st.warning("‚ö†Ô∏è Screener_template.xlsx not found in the current directory")
        
        if not SCREENER_MODE_AVAILABLE:
            st.error(f"‚ùå Screener Mode module not available: {SCREENER_MODE_ERROR}")
            st.info("Please ensure screener_excel_mode.py is in the same directory as this file.")
            st.stop()
        
        st.markdown("---")
        
        col1, col2 = st.columns(2)
        
        with col1:
            company_name_screener = st.text_input("Company Name:", key='screener_company_name')
            
            # NEW: Ticker input for Yahoo Finance data (current price + beta)
            col_ticker1, col_ticker2 = st.columns([3, 1])
            with col_ticker1:
                ticker_symbol_screener = st.text_input(
                    "Stock Ticker Symbol (for current price & beta):",
                    placeholder="e.g., RELIANCE, TCS, HDFCBANK",
                    key='screener_ticker',
                    help="Enter stock ticker to fetch current market price and beta from Yahoo Finance"
                )
            with col_ticker2:
                exchange_screener = st.selectbox(
                    "Exchange",
                    ["NS", "BO"],
                    key='screener_exchange',
                    help="NS=NSE, BO=BSE"
                )
            
            # Integrated Auto Download or Manual Upload Section
            if AUTO_DOWNLOAD_AVAILABLE:
                use_auto_dl, excel_file_screener = integrate_with_existing_upload_section(
                    cookies_path="screener_cookies.pkl"
                )
            else:
                # Fallback to manual upload only
                st.markdown("---")
                st.markdown("### üìä Screener Excel Data Source")
                excel_file_screener = st.file_uploader(
                    "Upload Screener Excel Template", 
                    type=['xlsx', 'xls'],
                    key='screener_excel_upload',
                    help="Upload Excel file with 'Balance Sheet' and 'Profit and Loss Account' sheets"
                )
                if not AUTO_DOWNLOAD_AVAILABLE:
                    st.info("üí° Auto-download feature not available. Please upload Excel manually.")

            
            st.markdown("---")
            st.markdown("**üìä Peer Companies (Both Exchanges)**")
            
            # Auto-fetch button ABOVE the input boxes (so we can set state before inputs are created)
            if ticker_symbol_screener and ticker_symbol_screener.strip():
                col_btn1, col_btn2 = st.columns([1, 3])
                with col_btn1:
                    if st.button("üîç Auto-Fetch Peers", key='screener_auto_fetch_peers', help="Automatically fetch industry peer companies", use_container_width=True):
                        with st.spinner("Fetching peers from Yahoo Finance..."):
                            try:
                                # Construct full ticker with exchange
                                full_ticker = f"{ticker_symbol_screener.strip().upper()}.{exchange_screener}"
                                st.info(f"üîç Searching peers for {full_ticker}...")
                                
                                if PEER_FETCHER_AVAILABLE:
                                    # Fetch peers
                                    peers_list = get_industry_peers(full_ticker, max_peers=10, exclude_self=True)
                                    
                                    if peers_list:
                                        # Update the text input session state keys directly BEFORE widgets are created
                                        if exchange_screener == "NS":
                                            st.session_state.nse_peers_input = ','.join(peers_list)
                                            st.session_state.bse_peers_input = ''  # Clear other exchange
                                        else:  # BO
                                            st.session_state.bse_peers_input = ','.join(peers_list)
                                            st.session_state.nse_peers_input = ''  # Clear other exchange
                                        
                                        # Force rerun to show peers immediately
                                        st.rerun()
                                    else:
                                        st.warning("‚ö†Ô∏è No peers found. Please enter manually.")
                                else:
                                    st.error("‚ùå Peer fetcher module not available (PEER_FETCHER_AVAILABLE = False). Please enter peers manually.")
                            except Exception as e:
                                st.error(f"‚ùå Error fetching peers: {str(e)}")
                                import traceback
                                st.code(traceback.format_exc())
                with col_btn2:
                    st.caption("üí° Click to automatically fetch industry peers from Yahoo Finance")
            
            # NSE Peers Box
            nse_peers_screener = st.text_input(
                "NSE Peer Tickers (comma-separated):",
                placeholder="e.g., RELIANCE, TCS, INFY",
                key='nse_peers_input',
                help="Enter NSE-listed peer companies for beta calculation. Auto-populated if you clicked 'Auto-Fetch Peers'."
            )
            
            # BSE Peers Box
            bse_peers_screener = st.text_input(
                "BSE Peer Tickers (comma-separated):",
                placeholder="e.g., RELIANCE, TCS, INFY",
                key='bse_peers_input',
                help="Enter BSE-listed peer companies for beta calculation. Auto-populated if you clicked 'Auto-Fetch Peers'."
            )
            
            # Combine peers with their exchange suffixes
            peer_tickers_screener = ""
            if nse_peers_screener.strip():
                nse_list = [f"{t.strip()}.NS" if '.NS' not in t and '.BO' not in t else t.strip() 
                           for t in nse_peers_screener.split(',') if t.strip()]
                peer_tickers_screener = ",".join(nse_list)
            if bse_peers_screener.strip():
                bse_list = [f"{t.strip()}.BO" if '.NS' not in t and '.BO' not in t else t.strip() 
                           for t in bse_peers_screener.split(',') if t.strip()]
                if peer_tickers_screener:
                    peer_tickers_screener += "," + ",".join(bse_list)
                else:
                    peer_tickers_screener = ",".join(bse_list)
            
            st.markdown("---")
            st.markdown("**üéØ Data & Projection Configuration**")
            
            # Historical Years and Projection Years - PROMINENTLY PLACED
            col_years1, col_years2 = st.columns(2)
            with col_years1:
                historical_years_screener = st.number_input(
                    "üìä Historical Years to Use",
                    min_value=3,
                    max_value=15,
                    value=5,
                    step=1,
                    key='screener_historical_years',
                    help="Select how many years of historical data to use from your Excel file (e.g., if Excel has 10 years, you can use 3, 5, or all 10)"
                )
            with col_years2:
                projection_years_screener = st.number_input(
                    "üìà Projection Years",
                    min_value=3,
                    max_value=15,
                    value=5,
                    step=1,
                    key='screener_proj_years',
                    help="Number of years to project into the future"
                )
        
        with col2:
            tax_rate_screener = st.number_input(
                "Tax Rate (%):", 
                min_value=0.0, 
                max_value=100.0, 
                value=25.0, 
                step=0.5,
                key='screener_tax'
            )
            terminal_growth_screener = st.number_input(
                "Terminal Growth Rate (%):", 
                min_value=0.0, 
                max_value=20.0, 
                value=4.0, 
                step=0.5,
                key='screener_terminal',
                help="Long-term perpetual growth rate. Should be <= long-term GDP growth + inflation (typically 4-8% for India)"
            )
            
            # Risk-free rate override
            st.markdown("**üèõÔ∏è Risk-Free Rate (G-Sec 10Y)**")
            
            # CRITICAL FIX: Use session state value directly
            manual_rf_rate_screener = st.number_input(
                f"Risk-Free Rate (%)",
                min_value=0.0,
                max_value=20.0,
                value=st.session_state.get('cached_rf_rate_screener', 6.83),
                step=0.1,
                key='manual_rf_screener',
                help="Auto-fetched from ticker above. You can manually edit this value."
            )
            
            # Update session state if user manually changes it
            if abs(manual_rf_rate_screener - st.session_state.get('cached_rf_rate_screener', 6.83)) > 0.05:
                st.session_state.cached_rf_rate_screener = manual_rf_rate_screener
                st.info(f"üí° Using custom rate: {manual_rf_rate_screener:.2f}%")
            
            # Manual discount rate override
            st.markdown("---")
            st.markdown("**üí∞ Discount Rate Override (Optional)**")
            manual_discount_rate_screener = st.number_input(
                "Manual Discount Rate Override (%):",
                min_value=0.0,
                max_value=50.0,
                value=0.0,
                step=0.5,
                key='manual_discount_screener',
                help="‚ö†Ô∏è Override WACC calculation. Leave at 0 to use auto-calculated WACC. Use this if you want to use a specific discount rate instead of WACC."
            )
            if manual_discount_rate_screener > 0:
                st.info(f"üí° Using manual discount rate: {manual_discount_rate_screener:.2f}% (Overriding WACC)")
            
            # Valuation models to run
            st.markdown("**üéØ Valuation Models**")
            run_dcf_screener = st.checkbox("DCF (FCFF)", value=True, key='screener_dcf')
            run_ddm_screener = st.checkbox("DDM (Dividend Discount)", value=True, key='screener_ddm')
            run_rim_screener = st.checkbox("RIM (Residual Income)", value=True, key='screener_rim')
            run_comp_screener = st.checkbox("Comparative Valuation", value=True, key='screener_comp')
        
        with st.expander("‚öôÔ∏è Advanced Projection Assumptions - FULL CONTROL"):
            st.info("üí° **Complete Control:** Override ANY projection parameter below. Leave at 0 or blank for auto-calculation from historical data.")
            
            st.markdown("### üìä Revenue & Growth")
            col1, col2, col3 = st.columns(3)
            with col1:
                rev_growth_override_screener = st.number_input(
                    "Revenue Growth (%/year)", 
                    min_value=0.0, max_value=200.0, value=0.0, step=0.5,
                    key='screener_rev_growth',
                    help="0 = Auto from historical CAGR. Override to use custom growth rate."
                )
            with col2:
                opex_margin_override_screener = st.number_input(
                    "Operating Expense Margin (%)", 
                    min_value=0.0, max_value=100.0, value=0.0, step=0.5,
                    key='screener_opex_margin',
                    help="0 = Auto from historical average"
                )
            with col3:
                ebitda_margin_override_screener = st.number_input(
                    "EBITDA Margin (%)", 
                    min_value=0.0, max_value=100.0, value=0.0, step=0.5,
                    key='screener_ebitda',
                    help="0 = Calculated as Revenue - OpEx"
                )
            
            st.markdown("### üèóÔ∏è CapEx & Depreciation")
            col4, col5, col6 = st.columns(3)
            with col4:
                capex_ratio_override_screener = st.number_input(
                    "CapEx/Revenue (%)", 
                    min_value=0.0, max_value=200.0, value=0.0, step=0.5,
                    key='screener_capex_ratio',
                    help="0 = Auto from historical average. Override for custom CapEx assumptions."
                )
            with col5:
                depreciation_rate_override_screener = st.number_input(
                    "Depreciation Rate (%)", 
                    min_value=0.0, max_value=30.0, value=0.0, step=0.5,
                    key='screener_dep_rate',
                    help="0 = Auto calculated"
                )
            with col6:
                depreciation_method_screener = st.selectbox(
                    "Depreciation Method",
                    ["Auto", "% of Fixed Assets", "% of Revenue", "Absolute Value"],
                    key='screener_dep_method'
                )
            
            st.markdown("### üí∞ Working Capital Management")
            col7, col8, col9 = st.columns(3)
            with col7:
                inventory_days_override_screener = st.number_input(
                    "Inventory Days", 
                    min_value=0.0, max_value=365.0, value=0.0, step=1.0,
                    key='screener_inv_days',
                    help="0 = Auto from historical average"
                )
            with col8:
                debtor_days_override_screener = st.number_input(
                    "Debtor/Receivables Days", 
                    min_value=0.0, max_value=365.0, value=0.0, step=1.0,
                    key='screener_deb_days',
                    help="0 = Auto from historical average"
                )
            with col9:
                creditor_days_override_screener = st.number_input(
                    "Creditor/Payables Days", 
                    min_value=0.0, max_value=365.0, value=0.0, step=1.0,
                    key='screener_cred_days',
                    help="0 = Auto from historical average"
                )
            
            st.markdown("### üìà Tax & Interest")
            col10, col11, col12 = st.columns(3)
            with col10:
                interest_rate_override_screener = st.number_input(
                    "Interest Rate (%)", 
                    min_value=0.0, max_value=30.0, value=0.0, step=0.25,
                    key='screener_interest',
                    help="0 = Auto calculated from Debt"
                )
            with col11:
                working_capital_as_pct_revenue_screener = st.number_input(
                    "Working Capital % of Revenue", 
                    min_value=0.0, max_value=50.0, value=0.0, step=0.5,
                    key='screener_wc_pct',
                    help="0 = Calculate from Inv+Deb-Cred days"
                )
        
        # ================================
        # DDM & RIM PARAMETERS FOR SCREENER MODE
        # ================================
        with st.expander("üíé DDM & RIM Parameters (Leave at 0 for Auto-Calculation)"):
            st.info("üí° **Note:** These parameters apply to Dividend Discount Model and Residual Income Model. Leave at 0 for auto-calculation from actual data.")
        
            st.markdown("### üí∞ Dividend Discount Model (DDM) Parameters")
            col_ddm1, col_ddm2, col_ddm3 = st.columns(3)
        
            with col_ddm1:
                ddm_dividend_growth_screener = st.number_input(
                    "Dividend Growth Rate (%)",
                    min_value=0.0,
                    max_value=200.0,
                    value=0.0,
                    step=0.5,
                    key='screener_ddm_div_growth',
                    help="0 = Auto-calculate from historical dividend data. Used in Gordon Growth Model."
                )
        
            with col_ddm2:
                ddm_required_return_screener = st.number_input(
                    "Required Return / Cost of Equity (%)",
                    min_value=5.0,
                    max_value=30.0,
                    value=12.0,
                    step=0.5,
                    key='screener_ddm_required_return',
                    help="Discount rate for DDM. Usually = Cost of Equity from CAPM."
                )
            
            with col_ddm3:
                ddm_payout_ratio_screener = st.number_input(
                    "Payout Ratio (%)",
                    min_value=0.0,
                    max_value=100.0,
                    value=0.0,
                    step=5.0,
                    key='screener_ddm_payout',
                    help="0 = Auto-calculate from historical data. % of earnings paid as dividends."
                )
        
            st.markdown("### üè¢ Residual Income Model (RIM) Parameters")
            col_rim1, col_rim2, col_rim3, col_rim4 = st.columns(4)
        
            with col_rim1:
                rim_required_return_screener = st.number_input(
                    "Required Return (%)",
                    min_value=5.0,
                    max_value=30.0,
                    value=12.0,
                    step=0.5,
                    key='screener_rim_required_return',
                    help="Discount rate for RIM. Usually = Cost of Equity."
                )
        
            with col_rim2:
                rim_assumed_roe_screener = st.number_input(
                    "Assumed ROE (%)",
                    min_value=0.0,
                    max_value=50.0,
                    value=0.0,
                    step=1.0,
                    key='screener_rim_roe',
                    help="0 = Auto-calculate from historical data. Return on Equity assumption."
                )
        
            with col_rim3:
                rim_terminal_growth_screener = st.number_input(
                    "Terminal Growth (%)",
                    min_value=0.0,
                    max_value=20.0,
                    value=0.0,
                    step=0.5,
                    key='screener_rim_terminal_growth',
                    help="0 = Use same as DCF terminal growth. Long-term perpetual growth rate."
                )
        
            with col_rim4:
                rim_projection_years_screener = st.number_input(
                    "Projection Years",
                    min_value=0,
                    max_value=15,
                    value=0,
                    step=1,
                    key='screener_rim_proj_years',
                    help="0 = Use same as DCF projection years (5). Number of years to project."
                )
        
        # Run valuation button
        if excel_file_screener and company_name_screener:
            # Stock Price Comparison Feature Toggle - compact, before run button
            enable_stock_comparison_screener = st.checkbox(
                "üìà Show Stock Price vs Revenue & EPS chart",
                value=False,
                key='screener_stock_comparison',
                help="Compare stock price with financials (max 10 years)"
            )
            
            # INPUT CHANGE DETECTION - Reset results if key inputs change
            current_inputs_screener = {
                'excel_file': excel_file_screener.name if excel_file_screener and hasattr(excel_file_screener, 'name') else None,
                'company_name': company_name_screener,
                'terminal_growth': terminal_growth_screener,
                'tax_rate': tax_rate_screener,
                'manual_discount': manual_discount_rate_screener if manual_discount_rate_screener > 0 else None,
                'rev_growth': rev_growth_override_screener
            }
            
            # Check if inputs changed
            if 'previous_inputs_screener' in st.session_state:
                if st.session_state.previous_inputs_screener != current_inputs_screener:
                    # Inputs changed - clear results
                    st.session_state.show_results_screener = False
            
            # Store current inputs
            st.session_state.previous_inputs_screener = current_inputs_screener
            
            st.markdown("---")
            st.markdown("### üéØ Ready to Run Valuation")
            st.info("üí° **Click the button below to run valuation.** Results will appear only after clicking.")
            
            col_run1, col_run2 = st.columns([2, 1])
            
            with col_run1:
                if st.button("üöÄ Run Screener Mode Valuation", type="primary", key="run_screener_dcf_btn"):
                    st.session_state.show_results_screener = True
            
            with col_run2:
                if st.button("üóëÔ∏è Clear Results", help="Clear current valuation results", key="clear_screener_results"):
                    st.session_state.show_results_screener = False
                    st.success("‚úÖ Results cleared!")
                    st.rerun()
            
            if st.session_state.get('show_results_screener', False):
                with st.spinner("Processing Screener Excel..."):
                    # Parse Excel using Screener mode parser
                    df_bs_screener, df_pl_screener = parse_screener_excel_to_dataframes(excel_file_screener)
                    
                    if df_bs_screener is None or df_pl_screener is None:
                        st.error("Failed to parse Screener Excel file")
                        st.stop()
                    
                    # Detect year columns
                    year_cols_screener_all = detect_screener_year_columns(df_bs_screener)
                    
                    if len(year_cols_screener_all) < 3:
                        st.error("Need at least 3 years of historical data in Excel file")
                        st.stop()
                    
                    st.success(f"‚úÖ Excel contains {len(year_cols_screener_all)} years of data: {', '.join([y.replace('_', '') for y in year_cols_screener_all])}")
                    
                    # Limit to user-selected historical years
                    if historical_years_screener > len(year_cols_screener_all):
                        st.warning(f"‚ö†Ô∏è Requested {historical_years_screener} years but Excel has only {len(year_cols_screener_all)} years. Using all available data.")
                        year_cols_screener = year_cols_screener_all
                    else:
                        year_cols_screener = year_cols_screener_all[-historical_years_screener:]
                        st.info(f"üìä Using {len(year_cols_screener)} most recent years as selected: {', '.join([y.replace('_', '') for y in year_cols_screener])}")
                    
                    # Try to get shares from Excel
                    shares_from_excel = get_screener_shares_outstanding(df_bs_screener, year_cols_screener[-1])
                    if shares_from_excel > 0:
                        num_shares_screener = shares_from_excel
                        st.success(f"‚úÖ Auto-detected {num_shares_screener:,} shares from Excel")
                    else:
                        # Prompt user to enter shares manually
                        st.warning("‚ö†Ô∏è Could not auto-detect shares outstanding from Excel")
                        num_shares_screener = st.number_input(
                            "Please enter Number of Shares Outstanding:",
                            min_value=1,
                            value=100000,
                            step=1000,
                            key='screener_shares_manual',
                            help="Enter the total number of shares outstanding for the company"
                        )
                        if num_shares_screener == 100000:
                            st.error("‚ùå Please enter the actual number of shares outstanding (default 100,000 is just a placeholder)")
                            st.stop()
                    
                    # Extract financials using Screener extraction
                    financials_screener = extract_screener_financials(df_bs_screener, df_pl_screener, year_cols_screener)
                    
                    # Auto-fill shares if available from Excel
                    if financials_screener.get('num_shares') and financials_screener['num_shares'] > 0:
                        num_shares_screener = financials_screener['num_shares']
                        st.success(f"‚úÖ Confirmed shares from Excel: {num_shares_screener:,} shares")
                    
                    # Add num_shares to financials dict
                    financials_screener['num_shares'] = num_shares_screener
                    
                    # Display financial summary
                    display_screener_financial_summary(financials_screener)
                    
                    # ================================
                    # FETCH TICKER DATA (Current Price & Beta)
                    # ================================
                    st.markdown("---")
                    st.subheader("üìà Market Data from Yahoo Finance")
                    
                    current_price_screener = 0.0
                    beta_screener = 1.0
                    
                    if ticker_symbol_screener and ticker_symbol_screener.strip():
                        ticker_data = fetch_ticker_data_for_screener(ticker_symbol_screener, exchange_screener)
                        
                        if ticker_data and not ticker_data.get('error'):
                            current_price_screener = ticker_data['current_price']
                            beta_screener = ticker_data['beta']
                            
                            # Only display current price (beta not needed for Screener mode)
                            st.metric("üìä Current Market Price", f"‚Çπ{current_price_screener:.2f}")
                        else:
                            st.warning(f"‚ö†Ô∏è Could not fetch ticker data: {ticker_data.get('error', 'Unknown error')}")
                            st.info("üí° Continuing with default beta (1.0) for WACC calculation")
                    else:
                        st.info("üí° No ticker provided - using default beta (1.0) for WACC calculation. Current price comparison will not be available.")
                    
                    # ================================
                    # BUSINESS MODEL CLASSIFICATION
                    # ================================
                    st.markdown("---")
                    st.subheader("üè¢ Business Model Classification")
                    
                    classification = classify_business_model(financials_screener, income_stmt=None, balance_sheet=None)
                    
                    # Show classification and check if FCFF DCF is allowed
                    should_stop = show_classification_warning(classification)
                    
                    if should_stop and run_dcf_screener:
                        st.warning("‚ö†Ô∏è DCF valuation skipped due to business model classification. Other models will still run.")
                        run_dcf_screener = False
                    
                    st.markdown("---")
                    
                    # Calculate WC metrics
                    wc_metrics = calculate_working_capital_metrics(financials_screener)
                    
                    # Show Working Capital Data Status
                    wc_status_parts = []
                    if wc_metrics.get('has_inventory', False) or (inventory_days_override_screener and inventory_days_override_screener > 0):
                        inv_source = f"User Override: {inventory_days_override_screener} days" if inventory_days_override_screener and inventory_days_override_screener > 0 else f"Historical: {wc_metrics['avg_inv_days']:.1f} days"
                        wc_status_parts.append(f"‚úÖ Inventory ({inv_source})")
                    else:
                        wc_status_parts.append("‚ö†Ô∏è Inventory (No data)")
                    
                    if wc_metrics.get('has_receivables', False) or (debtor_days_override_screener and debtor_days_override_screener > 0):
                        deb_source = f"User Override: {debtor_days_override_screener} days" if debtor_days_override_screener and debtor_days_override_screener > 0 else f"Historical: {wc_metrics['avg_deb_days']:.1f} days"
                        wc_status_parts.append(f"‚úÖ Debtors ({deb_source})")
                    else:
                        wc_status_parts.append("‚ö†Ô∏è Debtors (No data)")
                    
                    if wc_metrics.get('has_payables', False) or (creditor_days_override_screener and creditor_days_override_screener > 0):
                        cred_source = f"User Override: {creditor_days_override_screener} days" if creditor_days_override_screener and creditor_days_override_screener > 0 else f"Historical: {wc_metrics['avg_cred_days']:.1f} days"
                        wc_status_parts.append(f"‚úÖ Creditors ({cred_source})")
                    else:
                        wc_status_parts.append("‚ö†Ô∏è Creditors (No data)")
                    
                    st.info(f"üîç **Working Capital Projection Status:** {' | '.join(wc_status_parts)}")
                    
                    # Initialize results containers
                    dcf_results_screener = None
                    ddm_results_screener = None
                    rim_results_screener = None
                    comp_val_results_screener = None
                    projections_screener = None
                    
                    # ================================
                    # RUN DCF VALUATION
                    # ================================
                    if run_dcf_screener:
                        st.markdown("---")
                        st.subheader("üí∞ DCF (FCFF) Valuation")
                        
                        # Project financials (same as unlisted mode)
                        projections_screener, drivers_screener = project_financials(
                            financials_screener, wc_metrics, projection_years_screener, tax_rate_screener,
                            rev_growth_override_screener if rev_growth_override_screener > 0 else None,
                            opex_margin_override_screener if opex_margin_override_screener > 0 else None,
                            capex_ratio_override_screener if capex_ratio_override_screener > 0 else None,
                            ebitda_margin_override=ebitda_margin_override_screener if ebitda_margin_override_screener > 0 else None,
                            depreciation_rate_override=depreciation_rate_override_screener if depreciation_rate_override_screener > 0 else None,
                            depreciation_method=depreciation_method_screener,
                            inventory_days_override=inventory_days_override_screener if inventory_days_override_screener > 0 else None,
                            debtor_days_override=debtor_days_override_screener if debtor_days_override_screener > 0 else None,
                            creditor_days_override=creditor_days_override_screener if creditor_days_override_screener > 0 else None,
                            interest_rate_override=interest_rate_override_screener if interest_rate_override_screener > 0 else None,
                            working_capital_pct_override=working_capital_as_pct_revenue_screener if working_capital_as_pct_revenue_screener > 0 else None
                        )
                        
                        # Calculate WACC (using peer companies for beta)
                        wacc_details = calculate_wacc(financials_screener, tax_rate_screener / 100, peer_tickers=peer_tickers_screener, manual_rf_rate=manual_rf_rate_screener)
                        
                        # Display risk-free rate being used
                        st.info(f"üèõÔ∏è Risk-Free Rate (India 10Y G-Sec): {manual_rf_rate_screener:.2f}%")
                        
                        # Extract cash balance (index [0] is NEWEST)
                        cash_balance = financials_screener['cash'][0] if financials_screener['cash'][0] > 0 else 0
                        
                        # Calculate DCF
                        valuation, error = calculate_dcf_valuation(
                            projections_screener, wacc_details, terminal_growth_screener, num_shares_screener, cash_balance,
                            manual_discount_rate=manual_discount_rate_screener if manual_discount_rate_screener > 0 else None
                        )
                        
                        if error:
                            st.error(error)
                            st.stop()
                        
                        # Store results
                        dcf_results_screener = valuation
                        
                        # Display DCF results
                        st.success("‚úÖ DCF Valuation Complete!")
                        
                        # Key Metrics
                        col1, col2, col3, col4 = st.columns(4)
                        with col1:
                            st.metric("Enterprise Value", f"‚Çπ {dcf_results_screener['enterprise_value']:.2f} Lacs")
                        with col2:
                            st.metric("Equity Value", f"‚Çπ {dcf_results_screener['equity_value']:.2f} Lacs")
                        with col3:
                            st.metric("Fair Value/Share", f"‚Çπ {dcf_results_screener['fair_value_per_share']:.2f}")
                        with col4:
                            st.metric("WACC", f"{wacc_details['wacc']:.2f}%")
                    
                    # ================================
                    # RUN DDM VALUATION
                    # ================================
                    if run_ddm_screener:
                        # Use user-provided parameters or defaults
                        ddm_growth = ddm_dividend_growth_screener / 100 if ddm_dividend_growth_screener > 0 else 0.05
                        ddm_req_return = ddm_required_return_screener / 100
                        
                        ddm_results_screener = calculate_screener_ddm_valuation(
                            financials_screener,
                            num_shares_screener,
                            required_return=ddm_req_return,
                            growth_rate=ddm_growth
                        )
                    
                    # ================================
                    # RUN RIM VALUATION
                    # ================================
                    if run_rim_screener:
                        # Use user-provided parameters or defaults
                        rim_req_return = rim_required_return_screener / 100
                        rim_proj_yrs = rim_projection_years_screener if rim_projection_years_screener > 0 else 5
                        rim_term_growth = rim_terminal_growth_screener / 100 if rim_terminal_growth_screener > 0 else terminal_growth_screener / 100
                        rim_roe = rim_assumed_roe_screener / 100 if rim_assumed_roe_screener > 0 else None
                        
                        rim_results_screener = calculate_screener_rim_valuation(
                            financials_screener,
                            num_shares_screener,
                            required_return=rim_req_return,
                            projection_years=rim_proj_yrs,
                            terminal_growth=rim_term_growth,
                            assumed_roe=rim_roe
                        )
                    
                    # ================================
                    # RUN COMPARATIVE VALUATION
                    # ================================
                    if run_comp_screener and peer_tickers_screener:
                        comp_val_results_screener = perform_comparative_valuation(
                            target_ticker=None,  # Unlisted
                            comp_tickers_str=peer_tickers_screener,
                            target_financials=financials_screener,
                            target_shares=num_shares_screener,
                            exchange_suffix="NS",
                            projections=projections_screener if run_dcf_screener else None,
                            use_screener_peers=False  # Use Yahoo Finance for peers
                        )
                    
                    # ================================
                    # TABBED DISPLAY - CONSISTENT WITH UNLISTED MODE
                    # ================================
                    st.markdown("---")
                    
                    # ================================
                    # CURRENT PRICE VS FAIR VALUES COMPARISON
                    # ================================
                    if current_price_screener > 0:
                        st.subheader("üí∞ Current Price vs Fair Values")
                        
                        # Collect all fair values
                        fair_values_dict = {}
                        
                        if dcf_results_screener and dcf_results_screener.get('fair_value_per_share', 0) > 0:
                            fair_values_dict['DCF (FCFF)'] = dcf_results_screener['fair_value_per_share']
                        
                        if ddm_results_screener and ddm_results_screener.get('status') == 'Success':
                            fair_values_dict['DDM (Gordon)'] = ddm_results_screener.get('value_per_share', 0)
                        
                        if rim_results_screener and rim_results_screener.get('status') == 'Success':
                            fair_values_dict['RIM (Residual Income)'] = rim_results_screener.get('value_per_share', 0)
                        
                        if comp_val_results_screener and comp_val_results_screener.get('valuations'):
                            valuations = comp_val_results_screener['valuations']
                            all_avg_values = [v.get('fair_value_avg', 0) for v in valuations.values()]
                            if all_avg_values:
                                fair_values_dict['Comparative Valuation (Avg)'] = np.mean(all_avg_values)
                        
                        if fair_values_dict:
                            # Create comparison DataFrame
                            comparison_data = []
                            for method, fair_value in fair_values_dict.items():
                                upside = ((fair_value - current_price_screener) / current_price_screener * 100) if current_price_screener > 0 else 0
                                comparison_data.append({
                                    'Valuation Method': method,
                                    'Fair Value (‚Çπ)': fair_value,
                                    'Current Price (‚Çπ)': current_price_screener,
                                    'Upside/Downside (%)': upside
                                })
                            
                            comparison_df = pd.DataFrame(comparison_data)
                            
                            # Display comparison table
                            st.dataframe(
                                comparison_df.style.format({
                                    'Fair Value (‚Çπ)': '‚Çπ{:.2f}',
                                    'Current Price (‚Çπ)': '‚Çπ{:.2f}',
                                    'Upside/Downside (%)': '{:+.1f}%'
                                }).applymap(
                                    lambda x: 'background-color: #d4edda' if isinstance(x, (int, float)) and x > 0 else ('background-color: #f8d7da' if isinstance(x, (int, float)) and x < 0 else ''),
                                    subset=['Upside/Downside (%)']
                                ),
                                use_container_width=True,
                                hide_index=True
                            )
                            
                            # Calculate and display average
                            avg_fair_value = np.mean(list(fair_values_dict.values()))
                            avg_upside = ((avg_fair_value - current_price_screener) / current_price_screener * 100)
                            
                            col_avg1, col_avg2, col_avg3 = st.columns(3)
                            with col_avg1:
                                st.metric("Current Market Price", f"‚Çπ{current_price_screener:.2f}")
                            with col_avg2:
                                st.metric("Average Fair Value (All Methods)", f"‚Çπ{avg_fair_value:.2f}")
                            with col_avg3:
                                st.metric("Average Upside/Downside", f"{avg_upside:+.1f}%", 
                                         delta=f"‚Çπ{avg_fair_value - current_price_screener:+.2f}")
                            
                            # Add Gauge Chart - Only if fair value is valid (positive and reasonable)
                            if avg_fair_value > 0:
                                gauge_chart = create_price_vs_value_gauge(current_price_screener, avg_fair_value)
                                if gauge_chart:
                                    st.plotly_chart(gauge_chart, use_container_width=True)
                            else:
                                st.error("‚ö†Ô∏è **Invalid Fair Value**: The calculated fair value is negative or zero, which indicates a problem with the valuation inputs or methodology. Please review your financial data and assumptions.")
                        else:
                            st.info("No valid fair values calculated yet")
                        
                        st.markdown("---")
                    
                    # Build tab list based on what was run
                    tab_list = ["üìä Historical Financials"]
                    if run_dcf_screener:
                        tab_list.extend(["üìã Assumptions & Inputs", "üìà Projections", "üí∞ FCF Working", "üéØ WACC Calculation", "üèÜ DCF Summary", "üìâ Sensitivity Analysis"])
                    if run_ddm_screener:
                        tab_list.append("üí∏ DDM Valuation")
                    if run_rim_screener:
                        tab_list.append("üìö RIM Valuation")
                    if run_comp_screener and peer_tickers_screener:
                        tab_list.append("üìä Comparative Valuation")
                    
                    # Add Stock Comparison tab if enabled
                    enable_stock_comparison_screener_state = st.session_state.get('screener_stock_comparison', False)
                    if enable_stock_comparison_screener_state and STOCK_COMPARISON_AVAILABLE:
                        tab_list.append("üìà Stock vs Financials")
                    
                    tabs = st.tabs(tab_list)
                    tab_idx = 0
                    
                    # Tab 1: Historical Financials
                    with tabs[tab_idx]:
                        st.subheader("Historical Financials")
                        
                        # Add Historical Financials Chart with error handling
                        try:
                            # Pass reverse_years=True for chronological display (Screener data is newest-first internally)
                            st.plotly_chart(create_historical_financials_chart(financials_screener, reverse_years=True), use_container_width=True)
                        except Exception as e:
                            st.error(f"Historical chart error: {str(e)}")
                        
                        st.markdown("---")
                        st.markdown("### Income Statement")
                        
                        # Reverse data for chronological display (oldest to newest)
                        years_display = list(reversed(financials_screener['years']))
                        
                        hist_df = pd.DataFrame({
                            'Year': [str(y) for y in years_display],
                            'Revenue': list(reversed(financials_screener['revenue'])),
                            'Operating Expenses': list(reversed(financials_screener['opex'])),
                            'EBITDA': list(reversed(financials_screener['ebitda'])),
                            'Depreciation': list(reversed(financials_screener['depreciation'])),
                            'EBIT': list(reversed(financials_screener['ebit'])),
                            'Interest': list(reversed(financials_screener['interest'])),
                            'Tax': list(reversed(financials_screener['tax'])),
                            'NOPAT': list(reversed(financials_screener['nopat']))
                        })
                        numeric_cols = hist_df.select_dtypes(include=[np.number]).columns.tolist()
                        format_dict = {col: '{:.2f}' for col in numeric_cols}
                        st.dataframe(hist_df.style.format(format_dict), use_container_width=True)
                        
                        st.markdown("---")
                        st.subheader("Balance Sheet Metrics")
                        bs_df = pd.DataFrame({
                            'Year': [str(y) for y in years_display],
                            'Fixed Assets': list(reversed(financials_screener['fixed_assets'])),
                            'Inventory': list(reversed(financials_screener['inventory'])),
                            'Receivables': list(reversed(financials_screener['receivables'])),
                            'Payables': list(reversed(financials_screener['payables'])),
                            'Equity': list(reversed(financials_screener['equity'])),
                            'ST Debt': list(reversed(financials_screener['st_debt'])),
                            'LT Debt': list(reversed(financials_screener['lt_debt']))
                        })
                        numeric_cols = bs_df.select_dtypes(include=[np.number]).columns.tolist()
                        format_dict = {col: '{:.2f}' for col in numeric_cols}
                        st.dataframe(bs_df.style.format(format_dict), use_container_width=True)
                    
                    tab_idx += 1
                    
                    # Tab 2: Assumptions & Inputs (NEW - for transparency)
                    if run_dcf_screener:
                        with tabs[tab_idx]:
                            st.subheader("üìã All Assumptions, Inputs & Parameters Used")
                            st.info("üí° This tab shows ALL parameters, assumptions, and ratios used in the valuation for complete transparency.")
                            
                            # Section 1: Data Configuration - COMPACT & CLEAN
                            st.markdown("### üéØ Data Configuration")
                            
                            # Compact metrics row
                            years_display = [str(y).replace('_', '') for y in financials_screener['years']]
                            years_range = f"{years_display[-1]} - {years_display[0]}" if len(years_display) > 1 else years_display[0]
                            
                            col1, col2, col3 = st.columns(3)
                            with col1:
                                st.metric("üìÖ Historical Years", f"{len(financials_screener['years'])} years", delta=years_range, delta_color="off")
                            with col2:
                                st.metric("üîÆ Projection Period", f"{projection_years_screener} years", delta="Future", delta_color="off")
                            with col3:
                                st.metric("üè¢ Shares Outstanding", f"{num_shares_screener:,}")
                            
                            st.markdown("---")
                            
                            # Section 2: WACC Components - COMPACT VISUAL CHART
                            st.markdown("### üí∞ WACC & Discount Rate")
                            
                            # Create compact WACC breakdown chart
                            fig_wacc = go.Figure()
                            
                            fig_wacc.add_trace(go.Bar(
                                x=['WACC', 'Cost of Equity', 'Cost of Debt', 'Terminal Growth'],
                                y=[wacc_details['wacc'], wacc_details['ke'], wacc_details['kd_after_tax'], terminal_growth_screener],
                                marker=dict(color=['#667eea', '#764ba2', '#f093fb', '#f5576c']),
                                text=[f"{wacc_details['wacc']:.2f}%", f"{wacc_details['ke']:.2f}%", 
                                      f"{wacc_details['kd_after_tax']:.2f}%", f"{terminal_growth_screener:.2f}%"],
                                textposition='auto',
                                textfont=dict(size=14, color='white')
                            ))
                            
                            fig_wacc.update_layout(
                                height=250,
                                margin=dict(l=20, r=20, t=20, b=20),
                                showlegend=False,
                                paper_bgcolor='rgba(0,0,0,0)',
                                plot_bgcolor='rgba(0,0,0,0)',
                                yaxis=dict(title='Rate (%)', gridcolor='rgba(128,128,128,0.2)'),
                                xaxis=dict(title='')
                            )
                            
                            st.plotly_chart(fig_wacc, use_container_width=True)
                            
                            # Compact key assumptions in 2 columns
                            col_a, col_b = st.columns(2)
                            with col_a:
                                st.caption("**Tax Rate:** " + f"{tax_rate_screener:.2f}% | **Beta:** {wacc_details['beta']:.3f}")
                                st.caption("**Risk-Free Rate:** " + f"{wacc_details['rf']:.2f}% | **Market Return:** {wacc_details['rm']:.2f}%")
                            with col_b:
                                st.caption("**Equity Weight:** " + f"{wacc_details['we']:.2f}% | **Debt Weight:** {wacc_details['wd']:.2f}%")
                                st.caption("**Cost of Equity:** " + f"{wacc_details['ke']:.2f}% | **Cost of Debt:** {wacc_details['kd_after_tax']:.2f}%")
                            
                            st.markdown("---")
                            
                            # Section 3: Capital Structure - VISUAL PIE CHART
                            st.markdown("### üèõÔ∏è Capital Structure (Latest Year)")
                            
                            equity = financials_screener['equity'][0]
                            total_debt = financials_screener['st_debt'][0] + financials_screener['lt_debt'][0]
                            cash = financials_screener['cash'][0]
                            net_debt = total_debt - cash
                            
                            fig_cap = go.Figure(data=[go.Pie(
                                labels=['Equity', 'Debt', 'Cash'],
                                values=[equity, total_debt, cash],
                                hole=.4,
                                marker=dict(colors=['#06A77D', '#D62828', '#F77F00']),
                                textinfo='label+percent',
                                textfont=dict(size=12)
                            )])
                            
                            fig_cap.update_layout(
                                height=300,
                                margin=dict(l=20, r=20, t=20, b=20),
                                showlegend=True,
                                legend=dict(orientation="h", yanchor="bottom", y=-0.2, xanchor="center", x=0.5)
                            )
                            
                            st.plotly_chart(fig_cap, use_container_width=True)
                            
                            col_cap1, col_cap2, col_cap3 = st.columns(3)
                            with col_cap1:
                                st.metric("Equity", f"‚Çπ{equity:,.0f} L")
                            with col_cap2:
                                st.metric("Total Debt", f"‚Çπ{total_debt:,.0f} L")
                            with col_cap3:
                                st.metric("Net Debt", f"‚Çπ{net_debt:,.0f} L", delta="Cash adjusted")
                            
                            st.markdown("---")
                            
                            # Section 4: Projection Drivers - COMPACT TABLE
                            st.markdown("### üìä Projection Drivers & Ratios")
                            
                            with st.expander("üìã View All Projection Parameters", expanded=False):
                                drivers_data = []
                                if drivers_screener:
                                    drivers_data.append(['Revenue Growth', f"{drivers_screener['avg_growth']:.2f}%", 'User Override' if rev_growth_override_screener > 0 else 'Historical CAGR'])
                                    drivers_data.append(['OpEx Margin', f"{drivers_screener['avg_opex_margin']:.2f}%", 'User Override' if opex_margin_override_screener > 0 else 'Historical Avg'])
                                    drivers_data.append(['CapEx/Revenue', f"{drivers_screener['avg_capex_ratio']:.2f}%", 'User Override' if capex_ratio_override_screener > 0 else 'Historical Avg'])
                                    drivers_data.append(['Depreciation', f"{'Auto' if depreciation_rate_override_screener == 0 else depreciation_rate_override_screener}", depreciation_method_screener])
                                    
                                    if inventory_days_override_screener > 0 or wc_metrics.get('avg_inv_days', 0) > 0:
                                        drivers_data.append(['Inventory Days', f"{inventory_days_override_screener if inventory_days_override_screener > 0 else wc_metrics.get('avg_inv_days', 0):.0f}", 'Override' if inventory_days_override_screener > 0 else 'Historical'])
                                    if debtor_days_override_screener > 0 or wc_metrics.get('avg_deb_days', 0) > 0:
                                        drivers_data.append(['Debtor Days', f"{debtor_days_override_screener if debtor_days_override_screener > 0 else wc_metrics.get('avg_deb_days', 0):.0f}", 'Override' if debtor_days_override_screener > 0 else 'Historical'])
                                    if creditor_days_override_screener > 0 or wc_metrics.get('avg_cred_days', 0) > 0:
                                        drivers_data.append(['Creditor Days', f"{creditor_days_override_screener if creditor_days_override_screener > 0 else wc_metrics.get('avg_cred_days', 0):.0f}", 'Override' if creditor_days_override_screener > 0 else 'Historical'])
                                
                                if drivers_data:
                                    drivers_compact = pd.DataFrame(drivers_data, columns=['Parameter', 'Value', 'Source'])
                                    st.dataframe(drivers_compact, use_container_width=True, hide_index=True, height=250)

                            
                            # Section 4: Capital Structure
                            st.markdown("### üè¶ Capital Structure (Latest Year)")
                            capital_df = pd.DataFrame({
                                'Item': [
                                    'Equity',
                                    'Short-term Debt',
                                    'Long-term Debt',
                                    'Total Debt',
                                    'Cash & Equivalents',
                                    'Net Debt',
                                    'Total Capital'
                                ],
                                'Value (‚Çπ Lacs)': [
                                    f"{financials_screener['equity'][0]:.2f}",
                                    f"{financials_screener['st_debt'][0]:.2f}",
                                    f"{financials_screener['lt_debt'][0]:.2f}",
                                    f"{financials_screener['st_debt'][0] + financials_screener['lt_debt'][0]:.2f}",
                                    f"{financials_screener['cash'][0]:.2f}",
                                    f"{(financials_screener['st_debt'][0] + financials_screener['lt_debt'][0]) - financials_screener['cash'][0]:.2f}",
                                    f"{financials_screener['equity'][0] + financials_screener['st_debt'][0] + financials_screener['lt_debt'][0]:.2f}"
                                ]
                            })
                            st.dataframe(capital_df, use_container_width=True, hide_index=True)
                            
                            st.markdown("---")
                            
                            # Section 5: Valuation Model Selection
                            st.markdown("### üéØ Valuation Models Executed")
                            models_run = []
                            if run_dcf_screener:
                                models_run.append("‚úÖ DCF (FCFF) - Discounted Cash Flow")
                            if run_ddm_screener:
                                models_run.append("‚úÖ DDM - Dividend Discount Model")
                            if run_rim_screener:
                                models_run.append("‚úÖ RIM - Residual Income Model")
                            if run_comp_screener:
                                models_run.append("‚úÖ Comparative Valuation - Peer Multiples")
                            
                            for model in models_run:
                                st.write(model)
                            
                            st.success("üí° **Transparency Note:** All assumptions and inputs used in this valuation are documented above for your review and validation.")
                    
                    tab_idx += 1
                    
                    # Tab 2: Projections (if DCF was run)
                    if run_dcf_screener:
                        with tabs[tab_idx]:
                            st.subheader(f"Projected Financials ({projection_years_screener} Years)")
                            
                            # Add Projection Chart with error handling
                            try:
                                st.plotly_chart(create_fcff_projection_chart(projections_screener), use_container_width=True)
                            except Exception as e:
                                st.error(f"Chart error: {str(e)}")
                            
                            st.markdown("---")
                            
                            proj_df = pd.DataFrame({
                                'Year': [f"Year {y}" for y in projections_screener['year']],
                                'Revenue': projections_screener['revenue'],
                                'EBITDA': projections_screener['ebitda'],
                                'EBIT': projections_screener['ebit'],
                                'NOPAT': projections_screener['nopat'],
                                'CapEx': projections_screener['capex'],
                                'Œî WC': projections_screener['delta_wc'],
                                'FCFF': projections_screener['fcff']
                            })
                            numeric_cols = proj_df.select_dtypes(include=[np.number]).columns.tolist()
                            format_dict = {col: '{:.2f}' for col in numeric_cols}
                            st.dataframe(proj_df.style.format(format_dict), use_container_width=True)
                            
                            st.info(f"**Key Drivers:** Revenue Growth: {drivers_screener['avg_growth']:.2f}% | Opex Margin: {drivers_screener['avg_opex_margin']:.2f}% | CapEx/Revenue: {drivers_screener['avg_capex_ratio']:.2f}%")
                        
                        tab_idx += 1
                        
                        # Tab 3: FCF Working
                        with tabs[tab_idx]:
                            st.subheader("Free Cash Flow Working")
                            
                            fcf_df = pd.DataFrame({
                                'Year': [f"Year {y}" for y in projections_screener['year']],
                                'NOPAT': projections_screener['nopat'],
                                '(+) Depreciation': projections_screener['depreciation'],
                                '(-) ŒîWC': projections_screener['delta_wc'],
                                '(-) CapEx': projections_screener['capex'],
                                'FCFF': projections_screener['fcff'],
                                'PV(FCFF)': dcf_results_screener['pv_fcffs']
                            })
                            numeric_cols = fcf_df.select_dtypes(include=[np.number]).columns.tolist()
                            format_dict = {col: '{:.2f}' for col in numeric_cols}
                            st.dataframe(fcf_df.style.format(format_dict), use_container_width=True)
                            
                            st.info(f"**Sum of PV(FCFF):** ‚Çπ {dcf_results_screener['sum_pv_fcff']:.2f} Lacs")
                        
                        tab_idx += 1
                        
                        # Tab 4: WACC Calculation
                        with tabs[tab_idx]:
                            st.subheader("üéØ WACC Calculation & Breakdown")
                            
                            # Add WACC Breakdown Chart with error handling
                            try:
                                st.plotly_chart(create_wacc_breakdown_chart(wacc_details), use_container_width=True)
                            except Exception as e:
                                st.error(f"WACC chart error: {str(e)}")
                            
                            st.markdown("---")
                            
                            # WACC Components
                            col1, col2 = st.columns(2)
                            
                            with col1:
                                st.markdown("#### Cost of Equity (Ke)")
                                ke_df = pd.DataFrame({
                                    'Component': ['Risk-Free Rate', 'Beta', 'Market Return', 'Market Risk Premium', 'Cost of Equity'],
                                    'Value': [
                                        f"{wacc_details['rf']:.2f}%",
                                        f"{wacc_details['beta']:.3f}",
                                        f"{wacc_details['rm']:.2f}%",
                                        f"{wacc_details['rm'] - wacc_details['rf']:.2f}%",
                                        f"{wacc_details['ke']:.2f}%"
                                    ]
                                })
                                st.table(ke_df)
                            
                            with col2:
                                st.markdown("#### Cost of Debt (Kd)")
                                kd_df = pd.DataFrame({
                                    'Component': ['Total Debt', 'Interest Expense', 'Cost of Debt (Pre-tax)', 'Tax Rate', 'Cost of Debt (After-tax)'],
                                    'Value': [
                                        f"‚Çπ {wacc_details['debt']:.2f} Lacs",
                                        f"‚Çπ {financials_screener['interest'][0]:.2f} Lacs",
                                        f"{wacc_details['kd']:.2f}%",
                                        f"{tax_rate_screener:.2f}%",
                                        f"{wacc_details['kd_after_tax']:.2f}%"
                                    ]
                                })
                                st.table(kd_df)
                            
                            st.markdown("#### WACC Calculation")
                            wacc_calc_df = pd.DataFrame({
                                'Component': ['Equity', 'Debt', 'Total Capital', 'Weight of Equity', 'Weight of Debt', 'WACC'],
                                'Value': [
                                    f"‚Çπ {wacc_details['equity']:.2f} Lacs",
                                    f"‚Çπ {wacc_details['debt']:.2f} Lacs",
                                    f"‚Çπ {wacc_details['equity'] + wacc_details['debt']:.2f} Lacs",
                                    f"{wacc_details['we']:.2f}%",
                                    f"{wacc_details['wd']:.2f}%",
                                    f"{wacc_details['wacc']:.2f}%"
                                ]
                            })
                            st.table(wacc_calc_df)
                            
                            st.success(f"**WACC = (We √ó Ke) + (Wd √ó Kd) = ({wacc_details['we']:.2f}% √ó {wacc_details['ke']:.2f}%) + ({wacc_details['wd']:.2f}% √ó {wacc_details['kd_after_tax']:.2f}%) = {wacc_details['wacc']:.2f}%**")
                        
                        tab_idx += 1
                        
                        # Tab 5: DCF Summary
                        with tabs[tab_idx]:
                            st.subheader("üèÜ DCF Valuation Summary")
                            
                            # CRITICAL: Check for negative fair value and display warning
                            if dcf_results_screener.get('negative_fair_value_warning'):
                                warning = dcf_results_screener['negative_fair_value_warning']
                                st.error("üö® **CRITICAL ISSUE: Negative Fair Value Detected!**")
                                st.error(f"**Fair Value per Share: ‚Çπ{dcf_results_screener['fair_value_per_share']:.2f}**")
                                
                                st.markdown("### ‚ö†Ô∏è Problem Diagnosis:")
                                for reason in warning['reason']:
                                    st.warning(f"‚Ä¢ {reason}")
                                
                                st.markdown("### üìä Breakdown:")
                                problem_df = pd.DataFrame({
                                    'Metric': [
                                        'Enterprise Value',
                                        'Total Debt',
                                        'Cash & Equivalents',
                                        'Net Debt (Debt - Cash)',
                                        'Equity Value (EV - Net Debt)',
                                        'Number of Shares',
                                        'Fair Value per Share'
                                    ],
                                    'Value': [
                                        f"‚Çπ{warning['enterprise_value']:.2f} Lacs",
                                        f"‚Çπ{warning['total_debt']:.2f} Lacs",
                                        f"‚Çπ{warning['cash']:.2f} Lacs",
                                        f"‚Çπ{warning['net_debt']:.2f} Lacs",
                                        f"‚Çπ{warning['equity_value']:.2f} Lacs",
                                        f"{warning['num_shares']:,}",
                                        f"‚Çπ{dcf_results_screener['fair_value_per_share']:.2f}"
                                    ]
                                })
                                st.dataframe(problem_df, use_container_width=True, hide_index=True)
                                
                                st.markdown("### üí° Possible Solutions:")
                                st.info("""
                                **1. Reduce Debt Assumptions:** If debt is too high, reduce debt levels or reclassify some debt.
                                
                                **2. Check Historical Data:** Verify that your Excel data is correct (especially debt and equity).
                                
                                **3. Increase Terminal Growth:** A higher terminal growth rate will increase terminal value and enterprise value.
                                
                                **4. Review WACC:** A lower WACC (discount rate) will increase present values.
                                
                                **5. Improve Projections:** Ensure FCFFs are positive and growing in projection years.
                                
                                **6. Use Alternative Models:** Consider DDM, RIM, or Comparative Valuation instead of DCF for this company.
                                """)
                                
                                st.markdown("---")
                            
                            # Add Waterfall Chart with error handling
                            try:
                                st.plotly_chart(create_waterfall_chart(dcf_results_screener), use_container_width=True)
                            except Exception as e:
                                st.error(f"Waterfall chart error: {str(e)}")
                            
                            st.markdown("---")
                            
                            st.markdown("### Enterprise Value Build-up")
                            ev_df = pd.DataFrame({
                                'Component': ['Sum of PV(FCFF)', 'PV(Terminal Value)', 'Enterprise Value'],
                                'Value (‚Çπ Lacs)': [
                                    dcf_results_screener['sum_pv_fcff'],
                                    dcf_results_screener['pv_terminal_value'],
                                    dcf_results_screener['enterprise_value']
                                ]
                            })
                            st.dataframe(ev_df.style.format({'Value (‚Çπ Lacs)': '{:.2f}'}), use_container_width=True)
                            
                            st.markdown("---")
                            st.markdown("### Equity Value Calculation")
                            equity_df = pd.DataFrame({
                                'Component': ['Enterprise Value', '(-) Total Debt', '(+) Cash', 'Equity Value', 'Number of Shares', 'Fair Value per Share'],
                                'Value': [
                                    f"‚Çπ {dcf_results_screener['enterprise_value']:.2f} Lacs",
                                    f"‚Çπ {dcf_results_screener['total_debt']:.2f} Lacs",
                                    f"‚Çπ {dcf_results_screener['cash']:.2f} Lacs",
                                    f"‚Çπ {dcf_results_screener['equity_value']:.2f} Lacs",
                                    f"{num_shares_screener:,.0f}",
                                    f"‚Çπ {dcf_results_screener['fair_value_per_share']:.2f}"
                                ]
                            })
                            st.table(equity_df)
                            
                            st.info(f"**Terminal Value % of EV:** {dcf_results_screener['tv_percentage']:.1f}%")
                        
                        tab_idx += 1
                        
                        # Tab 6: Sensitivity Analysis (NEW)
                        with tabs[tab_idx]:
                            st.subheader("üìâ Sensitivity Analysis")
                            
                            # Create WACC and growth ranges
                            wacc_range = np.arange(max(1.0, wacc_details['wacc'] - 3), wacc_details['wacc'] + 3.5, 0.5)
                            g_range = np.arange(max(1.0, terminal_growth_screener - 2), min(terminal_growth_screener + 3, wacc_details['wacc'] - 1), 0.5)
                            
                            if len(g_range) == 0:
                                g_range = np.array([terminal_growth_screener])
                            
                            # Add Sensitivity Heatmap with error handling
                            try:
                                st.plotly_chart(
                                    create_sensitivity_heatmap(projections_screener, wacc_range, g_range, num_shares_screener),
                                    use_container_width=True
                                )
                            except Exception as e:
                                st.error(f"Sensitivity chart error: {str(e)}")
                            
                            st.info("üí° **How to Read:** Each cell shows the fair value per share for different combinations of WACC and terminal growth rate. Darker green = higher valuation, darker red = lower valuation.")
                        
                        tab_idx += 1
                    
                    # Tab: DDM Valuation
                    if run_ddm_screener:
                        with tabs[tab_idx]:
                            st.subheader("üí∏ Dividend Discount Model (DDM)")
                            if ddm_results_screener:
                                display_screener_ddm_results(ddm_results_screener)
                            else:
                                st.warning("DDM results not available")
                        tab_idx += 1
                    
                    # Tab: RIM Valuation
                    if run_rim_screener:
                        with tabs[tab_idx]:
                            st.subheader("üìö Residual Income Model (RIM)")
                            if rim_results_screener:
                                display_screener_rim_results(rim_results_screener)
                            else:
                                st.warning("RIM results not available")
                        tab_idx += 1
                    
                    # Tab: Comparative Valuation
                    if run_comp_screener and peer_tickers_screener:
                        with tabs[tab_idx]:
                            st.subheader("üìä Comparative Valuation (Peer Multiples)")
                            
                            if comp_val_results_screener:
                                # Show comparables table
                                st.markdown("### Comparable Companies")
                                comp_df = pd.DataFrame(comp_val_results_screener['comparables'])
                                if not comp_df.empty:
                                    display_comp_df = comp_df[['ticker', 'name', 'price', 'pe', 'pb', 'ps', 'ev_ebitda', 'ev_sales']]
                                    st.dataframe(display_comp_df.style.format({
                                        'price': '‚Çπ{:.2f}',
                                        'pe': '{:.2f}x',
                                        'pb': '{:.2f}x',
                                        'ps': '{:.2f}x',
                                        'ev_ebitda': '{:.2f}x',
                                        'ev_sales': '{:.2f}x'
                                    }), use_container_width=True)
                                
                                # Show multiples statistics
                                st.markdown("### Peer Multiples Statistics")
                                for multiple, stats in comp_val_results_screener['multiples_stats'].items():
                                    with st.expander(f"üìä {multiple.upper()} - Avg: {stats['average']:.2f}x, Median: {stats['median']:.2f}x"):
                                        st.write(f"**Range:** {stats['min']:.2f}x - {stats['max']:.2f}x")
                                        st.write(f"**Std Dev:** {stats['std']:.2f}x")
                                        st.write(f"**Peer Values:** {', '.join([f'{v:.2f}x' for v in stats['values']])}")
                                
                                # Show implied valuations
                                st.markdown("### Implied Fair Values")
                                
                                all_avg_values = []
                                all_median_values = []
                                
                                for method_key, val_data in comp_val_results_screener['valuations'].items():
                                    st.markdown(f"#### {val_data['method']}")
                                    
                                    col1, col2 = st.columns(2)
                                    
                                    with col1:
                                        st.markdown("**Using Average Multiple:**")
                                        st.write(val_data['formula_avg'])
                                        st.metric("Fair Value (Avg)", f"‚Çπ{val_data['fair_value_avg']:.2f}")
                                        all_avg_values.append(val_data['fair_value_avg'])
                                    
                                    with col2:
                                        st.markdown("**Using Median Multiple:**")
                                        st.write(val_data['formula_median'])
                                        st.metric("Fair Value (Median)", f"‚Çπ{val_data['fair_value_median']:.2f}")
                                        all_median_values.append(val_data['fair_value_median'])
                                    
                                    st.markdown("---")
                                
                                # Summary statistics
                                if all_avg_values and all_median_values:
                                    st.markdown("### üìà Comparative Valuation Summary")
                                    
                                    col1, col2, col3 = st.columns(3)
                                    
                                    with col1:
                                        st.metric("Average (All Methods)", f"‚Çπ{np.mean(all_avg_values):.2f}")
                                        st.metric("Median (All Methods)", f"‚Çπ{np.median(all_median_values):.2f}")
                                    
                                    with col2:
                                        st.metric("Min Fair Value", f"‚Çπ{min(all_avg_values + all_median_values):.2f}")
                                        st.metric("Max Fair Value", f"‚Çπ{max(all_avg_values + all_median_values):.2f}")
                                    
                                    with col3:
                                        if dcf_results_screener and dcf_results_screener['fair_value_per_share'] > 0:
                                            st.metric("DCF Fair Value", f"‚Çπ{dcf_results_screener['fair_value_per_share']:.2f}")
                                            combined_avg = (np.mean(all_avg_values) + dcf_results_screener['fair_value_per_share']) / 2
                                            st.metric("DCF + Comp Avg", f"‚Çπ{combined_avg:.2f}")
                            else:
                                st.warning("Could not fetch comparable companies data")
                    
                    # Stock Price Comparison Tab (if enabled)
                    if enable_stock_comparison_screener_state and STOCK_COMPARISON_AVAILABLE:
                        with tabs[tab_idx]:
                            st.subheader("üìà Stock Price vs Revenue & EPS Analysis")
                            
                            if ticker_symbol_screener and ticker_symbol_screener.strip():
                                with st.spinner("Fetching stock price data..."):
                                    try:
                                        # Construct full ticker
                                        full_ticker_screener = f"{ticker_symbol_screener.strip().upper()}.{exchange_screener}"
                                        
                                        # Determine years to fetch (max 10 for Screener)
                                        years_to_fetch = min(historical_years_screener, 10)
                                        
                                        # Fetch comparison data
                                        stock_comp_data_screener = get_stock_comparison_data_screener(
                                            ticker=full_ticker_screener,
                                            company_name=company_name_screener,
                                            balance_sheet_df=df_bs_screener,
                                            pnl_df=df_pl_screener,
                                            num_years=years_to_fetch
                                        )
                                        
                                        if stock_comp_data_screener and stock_comp_data_screener['chart_fig']:
                                            st.plotly_chart(stock_comp_data_screener['chart_fig'], use_container_width=True)
                                            
                                            # Show data tables in expanders
                                            with st.expander("üìä View Raw Data"):
                                                col1, col2, col3 = st.columns(3)
                                                
                                                with col1:
                                                    st.markdown("**Revenue Data**")
                                                    if stock_comp_data_screener['revenue_df'] is not None:
                                                        st.dataframe(stock_comp_data_screener['revenue_df'], hide_index=True)
                                                    else:
                                                        st.info("No revenue data available")
                                                
                                                with col2:
                                                    st.markdown("**EPS Data**")
                                                    if stock_comp_data_screener['eps_df'] is not None:
                                                        st.dataframe(stock_comp_data_screener['eps_df'], hide_index=True)
                                                    else:
                                                        st.info("No EPS data available")
                                                
                                                with col3:
                                                    st.markdown("**Stock Price Summary**")
                                                    if stock_comp_data_screener['stock_prices_df'] is not None:
                                                        price_df = stock_comp_data_screener['stock_prices_df']
                                                        st.metric("Latest Price", f"‚Çπ{price_df['Close'].iloc[-1]:.2f}")
                                                        st.metric("Period Return", f"{((price_df['Close'].iloc[-1] - price_df['Close'].iloc[0]) / price_df['Close'].iloc[0] * 100):.2f}%")
                                                        st.metric("Major Changes", f"{price_df['is_major'].sum()}" if 'is_major' in price_df.columns else "N/A")
                                                    else:
                                                        st.info("No stock price data available")
                                            
                                            st.info(f"üí° Chart shows {years_to_fetch} years of data (Screener Excel mode supports up to 10 years)")
                                        else:
                                            st.warning("Could not generate stock comparison chart. Check if ticker and Excel data are valid.")
                                    
                                    except Exception as e:
                                        st.error(f"Error generating stock comparison: {str(e)}")
                            else:
                                st.warning("Please enter a stock ticker symbol to view price comparison")
                        
                        tab_idx += 1
                    
                    # ================================
                    # SUMMARY & DOWNLOAD
                    # ================================
                    st.markdown("---")
                    st.subheader("üì• Download Comprehensive Valuation Report")
                    
                    # Generate Excel report
                    excel_buffer = generate_screener_valuation_excel(
                        company_name_screener,
                        financials_screener,
                        dcf_results_screener,
                        ddm_results_screener,
                        rim_results_screener,
                        comp_val_results_screener,
                        None  # peer_comparison
                    )
                    
                    st.download_button(
                        label="üìä Download Excel Report",
                        data=excel_buffer,
                        file_name=f"{company_name_screener}_Screener_Valuation_{datetime.now().strftime('%Y%m%d')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                    
                    st.success("‚úÖ Screener Mode Valuation Complete!")


        # Footer
        st.markdown("---")
        st.caption("üí° **Note:** All values in ‚Çπ Lacs unless specified otherwise | Built with traditional DCF methodology")

# Run main function when executed directly
if __name__ == "__main__":
    main()
